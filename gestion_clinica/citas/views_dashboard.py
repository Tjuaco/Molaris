from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Count, Sum, Q, Avg, F
from django.utils import timezone
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Cita
from pacientes.models import Cliente
from inventario.models import Insumo, MovimientoInsumo
from personal.models import Perfil
from evaluaciones.models import Evaluacion

# ========== VISTA PRINCIPAL DEL DASHBOARD ==========
@login_required
def dashboard_reportes(request):
    """Vista principal del dashboard con estadísticas y gráficos"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para acceder al dashboard.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    # Obtener fechas para filtros
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    fin_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    inicio_año = hoy.replace(month=1, day=1)
    
    # ===== ESTADÍSTICAS GENERALES =====
    total_citas = Cita.objects.count()
    total_clientes = Cliente.objects.count()
    total_personal = Perfil.objects.filter(activo=True).count()
    total_insumos = Insumo.objects.count()
    
    # Citas por estado
    citas_disponibles = Cita.objects.filter(estado='disponible').count()
    citas_reservadas = Cita.objects.filter(estado='reservada').count()
    citas_completadas = Cita.objects.filter(estado='completada').count()
    citas_canceladas = Cita.objects.filter(estado='cancelada').count()
    
    # Citas del mes
    citas_mes = Cita.objects.filter(
        fecha_hora__date__gte=inicio_mes,
        fecha_hora__date__lte=fin_mes
    ).count()
    
    citas_completadas_mes = Cita.objects.filter(
        estado='completada',
        fecha_hora__date__gte=inicio_mes,
        fecha_hora__date__lte=fin_mes
    ).count()
    
    # Insumos con stock bajo
    insumos_bajo_stock = Insumo.objects.filter(
        cantidad_actual__lte=F('cantidad_minima')
    ).count()
    
    # Dentistas más activos
    dentistas_activos = Perfil.objects.filter(
        rol='dentista',
        activo=True
    ).annotate(
        num_citas=Count('citas_asignadas')
    ).order_by('-num_citas')[:5]
    
    # Clientes nuevos este mes
    clientes_nuevos_mes = Cliente.objects.filter(
        fecha_registro__date__gte=inicio_mes,
        fecha_registro__date__lte=fin_mes
    ).count()
    
    # Tipos de consulta más frecuentes
    tipos_consulta = Cita.objects.exclude(
        tipo_consulta__isnull=True
    ).exclude(
        tipo_consulta=''
    ).values('tipo_consulta').annotate(
        total=Count('id')
    ).order_by('-total')[:5]
    
    # Citas por día de la semana (últimos 30 días)
    hace_30_dias = hoy - timedelta(days=30)
    citas_por_dia = []
    for i in range(7):
        dia = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo'][i]
        count = Cita.objects.filter(
            fecha_hora__date__gte=hace_30_dias,
            fecha_hora__week_day=i+2  # Django week_day: 1=domingo, 2=lunes...
        ).count()
        citas_por_dia.append({'dia': dia, 'total': count})
    
    # Citas por mes (últimos 6 meses)
    citas_por_mes = []
    for i in range(6):
        mes_fecha = hoy - timedelta(days=30*i)
        inicio = mes_fecha.replace(day=1)
        fin = (inicio + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        count = Cita.objects.filter(
            fecha_hora__date__gte=inicio,
            fecha_hora__date__lte=fin
        ).count()
        citas_por_mes.insert(0, {
            'mes': inicio.strftime('%B'),
            'total': count
        })
    
    # ===== ESTADÍSTICAS DE EVALUACIONES =====
    total_evaluaciones = Evaluacion.objects.filter(estado__in=['pendiente', 'revisada']).count()
    evaluaciones_pendientes = Evaluacion.objects.filter(estado='pendiente').count()
    evaluaciones_mes = Evaluacion.objects.filter(
        fecha_creacion__date__gte=inicio_mes,
        fecha_creacion__date__lte=fin_mes
    ).count()
    
    # Promedio de calificación
    promedio_evaluaciones = Evaluacion.objects.filter(
        estado__in=['pendiente', 'revisada']
    ).aggregate(Avg('estrellas'))['estrellas__avg'] or 0
    
    # Distribución de estrellas
    distribucion_evaluaciones = []
    for i in range(1, 6):
        count = Evaluacion.objects.filter(estrellas=i, estado__in=['pendiente', 'revisada']).count()
        distribucion_evaluaciones.append({'estrellas': i, 'total': count})
    
    # Últimas evaluaciones
    ultimas_evaluaciones = Evaluacion.objects.filter(
        estado__in=['pendiente', 'revisada']
    ).select_related('cliente').order_by('-fecha_creacion')[:5]
    
    context = {
        'perfil': perfil,
        'es_admin': True,
        # Estadísticas generales
        'total_citas': total_citas,
        'total_clientes': total_clientes,
        'total_personal': total_personal,
        'total_insumos': total_insumos,
        # Citas
        'citas_disponibles': citas_disponibles,
        'citas_reservadas': citas_reservadas,
        'citas_completadas': citas_completadas,
        'citas_canceladas': citas_canceladas,
        'citas_mes': citas_mes,
        'citas_completadas_mes': citas_completadas_mes,
        # Insumos
        'insumos_bajo_stock': insumos_bajo_stock,
        # Dentistas
        'dentistas_activos': dentistas_activos,
        # Clientes
        'clientes_nuevos_mes': clientes_nuevos_mes,
        # Gráficos
        'tipos_consulta': list(tipos_consulta),
        'citas_por_dia': citas_por_dia,
        'citas_por_mes': citas_por_mes,
        # Fechas
        'mes_actual': inicio_mes.strftime('%B %Y'),
        # Evaluaciones
        'total_evaluaciones': total_evaluaciones,
        'evaluaciones_pendientes': evaluaciones_pendientes,
        'evaluaciones_mes': evaluaciones_mes,
        'promedio_evaluaciones': round(promedio_evaluaciones, 1),
        'distribucion_evaluaciones': distribucion_evaluaciones,
        'ultimas_evaluaciones': ultimas_evaluaciones,
    }
    
    return render(request, 'citas/dashboard/dashboard_reportes.html', context)


# ========== EXPORTACIÓN A EXCEL - CITAS ==========
@login_required
def exportar_excel_citas(request):
    """Exporta todas las citas a Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Citas"
    
    # Estilos
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['ID', 'Fecha', 'Hora', 'Tipo Consulta', 'Estado', 'Paciente', 'Email', 'Teléfono', 'Dentista', 'Notas']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    citas = Cita.objects.all().order_by('-fecha_hora')
    for row_num, cita in enumerate(citas, 2):
        ws.cell(row=row_num, column=1, value=cita.id)
        ws.cell(row=row_num, column=2, value=cita.fecha_hora.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=3, value=cita.fecha_hora.strftime('%H:%M'))
        ws.cell(row=row_num, column=4, value=cita.tipo_consulta or 'N/A')
        ws.cell(row=row_num, column=5, value=cita.get_estado_display())
        ws.cell(row=row_num, column=6, value=cita.paciente_nombre or 'Sin asignar')
        ws.cell(row=row_num, column=7, value=cita.paciente_email or '')
        ws.cell(row=row_num, column=8, value=cita.paciente_telefono or '')
        ws.cell(row=row_num, column=9, value=cita.dentista.nombre_completo if cita.dentista else 'Sin asignar')
        ws.cell(row=row_num, column=10, value=cita.notas or '')
        
        # Aplicar bordes
        for col in range(1, 11):
            ws.cell(row=row_num, column=col).border = border
    
    # Ajustar anchos de columna
    column_widths = [8, 12, 10, 20, 12, 25, 30, 15, 25, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=citas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response


# ========== EXPORTACIÓN A EXCEL - CLIENTES ==========
@login_required
def exportar_excel_clientes(request):
    """Exporta todos los clientes a Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Estilos
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['ID', 'Nombre Completo', 'Email', 'Teléfono', 'Fecha Registro', 'Activo', 'Dentista Asignado', 'Total Citas', 'Notas']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    clientes = Cliente.objects.annotate(num_citas=Count('citas')).order_by('nombre_completo')
    for row_num, cliente in enumerate(clientes, 2):
        ws.cell(row=row_num, column=1, value=cliente.id)
        ws.cell(row=row_num, column=2, value=cliente.nombre_completo)
        ws.cell(row=row_num, column=3, value=cliente.email)
        ws.cell(row=row_num, column=4, value=cliente.telefono)
        ws.cell(row=row_num, column=5, value=cliente.fecha_registro.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row_num, column=6, value='Sí' if cliente.activo else 'No')
        ws.cell(row=row_num, column=7, value=cliente.dentista_asignado.nombre_completo if cliente.dentista_asignado else 'Sin asignar')
        ws.cell(row=row_num, column=8, value=cliente.num_citas)
        ws.cell(row=row_num, column=9, value=cliente.notas or '')
        
        for col in range(1, 10):
            ws.cell(row=row_num, column=col).border = border
    
    column_widths = [8, 30, 35, 15, 20, 10, 25, 12, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=clientes_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response


# ========== EXPORTACIÓN A EXCEL - INSUMOS ==========
@login_required
def exportar_excel_insumos(request):
    """Exporta todos los insumos a Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Insumos"
    
    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID', 'Nombre', 'Categoría', 'Cantidad Actual', 'Cantidad Mínima', 'Unidad', 'Precio Unitario', 'Estado', 'Proveedor', 'Ubicación']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    insumos = Insumo.objects.all().order_by('nombre')
    for row_num, insumo in enumerate(insumos, 2):
        ws.cell(row=row_num, column=1, value=insumo.id)
        ws.cell(row=row_num, column=2, value=insumo.nombre)
        ws.cell(row=row_num, column=3, value=insumo.get_categoria_display())
        ws.cell(row=row_num, column=4, value=insumo.cantidad_actual)
        ws.cell(row=row_num, column=5, value=insumo.cantidad_minima)
        ws.cell(row=row_num, column=6, value=insumo.unidad_medida)
        ws.cell(row=row_num, column=7, value=float(insumo.precio_unitario) if insumo.precio_unitario else 0)
        ws.cell(row=row_num, column=8, value=insumo.get_estado_display())
        ws.cell(row=row_num, column=9, value=insumo.proveedor or '')
        ws.cell(row=row_num, column=10, value=insumo.ubicacion or '')
        
        for col in range(1, 11):
            ws.cell(row=row_num, column=col).border = border
    
    column_widths = [8, 30, 20, 15, 15, 12, 15, 15, 25, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=insumos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response


# ========== EXPORTACIÓN A EXCEL - FINANZAS ==========
@login_required
def exportar_excel_finanzas(request):
    """Exporta reporte financiero a Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    
    wb = Workbook()
    
    # Hoja 1: Resumen Financiero
    ws1 = wb.active
    ws1.title = "Resumen"
    
    header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=14)
    
    ws1['A1'] = 'REPORTE FINANCIERO - CLÍNICA DENTAL'
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A1:D1')
    
    ws1['A2'] = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws1['A2'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A2:D2')
    
    # Estadísticas
    ws1['A4'] = 'ESTADÍSTICA'
    ws1['A4'].fill = header_fill
    ws1['A4'].font = header_font
    ws1['B4'] = 'VALOR'
    ws1['B4'].fill = header_fill
    ws1['B4'].font = header_font
    
    ws1['A5'] = 'Total de Citas Completadas'
    ws1['B5'] = Cita.objects.filter(estado='completada').count()
    
    ws1['A6'] = 'Total de Clientes Activos'
    ws1['B6'] = Cliente.objects.filter(activo=True).count()
    
    ws1['A7'] = 'Valor Total en Insumos'
    total_insumos = sum([
        (insumo.cantidad_actual * float(insumo.precio_unitario or 0))
        for insumo in Insumo.objects.all()
    ])
    ws1['B7'] = f'${total_insumos:,.2f}'
    
    ws1['A8'] = 'Citas Este Mes'
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    ws1['B8'] = Cita.objects.filter(fecha_hora__date__gte=inicio_mes).count()
    
    # Ajustar anchos
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=finanzas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response


from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Count, Sum, Q, Avg, F
from django.utils import timezone
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Cita
from pacientes.models import Cliente
from inventario.models import Insumo, MovimientoInsumo
from personal.models import Perfil
from evaluaciones.models import Evaluacion

# ========== VISTA PRINCIPAL DEL DASHBOARD ==========
@login_required
def dashboard_reportes(request):
    """Vista principal del dashboard con estadísticas y gráficos"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para acceder al dashboard.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    # Obtener fechas para filtros
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    fin_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    inicio_año = hoy.replace(month=1, day=1)
    
    # ===== ESTADÍSTICAS GENERALES =====
    total_citas = Cita.objects.count()
    total_clientes = Cliente.objects.count()
    total_personal = Perfil.objects.filter(activo=True).count()
    total_insumos = Insumo.objects.count()
    
    # Citas por estado
    citas_disponibles = Cita.objects.filter(estado='disponible').count()
    citas_reservadas = Cita.objects.filter(estado='reservada').count()
    citas_completadas = Cita.objects.filter(estado='completada').count()
    citas_canceladas = Cita.objects.filter(estado='cancelada').count()
    
    # Citas del mes
    citas_mes = Cita.objects.filter(
        fecha_hora__date__gte=inicio_mes,
        fecha_hora__date__lte=fin_mes
    ).count()
    
    citas_completadas_mes = Cita.objects.filter(
        estado='completada',
        fecha_hora__date__gte=inicio_mes,
        fecha_hora__date__lte=fin_mes
    ).count()
    
    # Insumos con stock bajo
    insumos_bajo_stock = Insumo.objects.filter(
        cantidad_actual__lte=F('cantidad_minima')
    ).count()
    
    # Dentistas más activos
    dentistas_activos = Perfil.objects.filter(
        rol='dentista',
        activo=True
    ).annotate(
        num_citas=Count('citas_asignadas')
    ).order_by('-num_citas')[:5]
    
    # Clientes nuevos este mes
    clientes_nuevos_mes = Cliente.objects.filter(
        fecha_registro__date__gte=inicio_mes,
        fecha_registro__date__lte=fin_mes
    ).count()
    
    # Tipos de consulta más frecuentes
    tipos_consulta = Cita.objects.exclude(
        tipo_consulta__isnull=True
    ).exclude(
        tipo_consulta=''
    ).values('tipo_consulta').annotate(
        total=Count('id')
    ).order_by('-total')[:5]
    
    # Citas por día de la semana (últimos 30 días)
    hace_30_dias = hoy - timedelta(days=30)
    citas_por_dia = []
    for i in range(7):
        dia = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo'][i]
        count = Cita.objects.filter(
            fecha_hora__date__gte=hace_30_dias,
            fecha_hora__week_day=i+2  # Django week_day: 1=domingo, 2=lunes...
        ).count()
        citas_por_dia.append({'dia': dia, 'total': count})
    
    # Citas por mes (últimos 6 meses)
    citas_por_mes = []
    for i in range(6):
        mes_fecha = hoy - timedelta(days=30*i)
        inicio = mes_fecha.replace(day=1)
        fin = (inicio + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        count = Cita.objects.filter(
            fecha_hora__date__gte=inicio,
            fecha_hora__date__lte=fin
        ).count()
        citas_por_mes.insert(0, {
            'mes': inicio.strftime('%B'),
            'total': count
        })
    
    # ===== ESTADÍSTICAS DE EVALUACIONES =====
    total_evaluaciones = Evaluacion.objects.filter(estado__in=['pendiente', 'revisada']).count()
    evaluaciones_pendientes = Evaluacion.objects.filter(estado='pendiente').count()
    evaluaciones_mes = Evaluacion.objects.filter(
        fecha_creacion__date__gte=inicio_mes,
        fecha_creacion__date__lte=fin_mes
    ).count()
    
    # Promedio de calificación
    promedio_evaluaciones = Evaluacion.objects.filter(
        estado__in=['pendiente', 'revisada']
    ).aggregate(Avg('estrellas'))['estrellas__avg'] or 0
    
    # Distribución de estrellas
    distribucion_evaluaciones = []
    for i in range(1, 6):
        count = Evaluacion.objects.filter(estrellas=i, estado__in=['pendiente', 'revisada']).count()
        distribucion_evaluaciones.append({'estrellas': i, 'total': count})
    
    # Últimas evaluaciones
    ultimas_evaluaciones = Evaluacion.objects.filter(
        estado__in=['pendiente', 'revisada']
    ).select_related('cliente').order_by('-fecha_creacion')[:5]
    
    context = {
        'perfil': perfil,
        'es_admin': True,
        # Estadísticas generales
        'total_citas': total_citas,
        'total_clientes': total_clientes,
        'total_personal': total_personal,
        'total_insumos': total_insumos,
        # Citas
        'citas_disponibles': citas_disponibles,
        'citas_reservadas': citas_reservadas,
        'citas_completadas': citas_completadas,
        'citas_canceladas': citas_canceladas,
        'citas_mes': citas_mes,
        'citas_completadas_mes': citas_completadas_mes,
        # Insumos
        'insumos_bajo_stock': insumos_bajo_stock,
        # Dentistas
        'dentistas_activos': dentistas_activos,
        # Clientes
        'clientes_nuevos_mes': clientes_nuevos_mes,
        # Gráficos
        'tipos_consulta': list(tipos_consulta),
        'citas_por_dia': citas_por_dia,
        'citas_por_mes': citas_por_mes,
        # Fechas
        'mes_actual': inicio_mes.strftime('%B %Y'),
        # Evaluaciones
        'total_evaluaciones': total_evaluaciones,
        'evaluaciones_pendientes': evaluaciones_pendientes,
        'evaluaciones_mes': evaluaciones_mes,
        'promedio_evaluaciones': round(promedio_evaluaciones, 1),
        'distribucion_evaluaciones': distribucion_evaluaciones,
        'ultimas_evaluaciones': ultimas_evaluaciones,
    }
    
    return render(request, 'citas/dashboard/dashboard_reportes.html', context)


# ========== EXPORTACIÓN A EXCEL - CITAS ==========
@login_required
def exportar_excel_citas(request):
    """Exporta todas las citas a Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Citas"
    
    # Estilos
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['ID', 'Fecha', 'Hora', 'Tipo Consulta', 'Estado', 'Paciente', 'Email', 'Teléfono', 'Dentista', 'Notas']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    citas = Cita.objects.all().order_by('-fecha_hora')
    for row_num, cita in enumerate(citas, 2):
        ws.cell(row=row_num, column=1, value=cita.id)
        ws.cell(row=row_num, column=2, value=cita.fecha_hora.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=3, value=cita.fecha_hora.strftime('%H:%M'))
        ws.cell(row=row_num, column=4, value=cita.tipo_consulta or 'N/A')
        ws.cell(row=row_num, column=5, value=cita.get_estado_display())
        ws.cell(row=row_num, column=6, value=cita.paciente_nombre or 'Sin asignar')
        ws.cell(row=row_num, column=7, value=cita.paciente_email or '')
        ws.cell(row=row_num, column=8, value=cita.paciente_telefono or '')
        ws.cell(row=row_num, column=9, value=cita.dentista.nombre_completo if cita.dentista else 'Sin asignar')
        ws.cell(row=row_num, column=10, value=cita.notas or '')
        
        # Aplicar bordes
        for col in range(1, 11):
            ws.cell(row=row_num, column=col).border = border
    
    # Ajustar anchos de columna
    column_widths = [8, 12, 10, 20, 12, 25, 30, 15, 25, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=citas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response


# ========== EXPORTACIÓN A EXCEL - CLIENTES ==========
@login_required
def exportar_excel_clientes(request):
    """Exporta todos los clientes a Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Estilos
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['ID', 'Nombre Completo', 'Email', 'Teléfono', 'Fecha Registro', 'Activo', 'Dentista Asignado', 'Total Citas', 'Notas']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    clientes = Cliente.objects.annotate(num_citas=Count('citas')).order_by('nombre_completo')
    for row_num, cliente in enumerate(clientes, 2):
        ws.cell(row=row_num, column=1, value=cliente.id)
        ws.cell(row=row_num, column=2, value=cliente.nombre_completo)
        ws.cell(row=row_num, column=3, value=cliente.email)
        ws.cell(row=row_num, column=4, value=cliente.telefono)
        ws.cell(row=row_num, column=5, value=cliente.fecha_registro.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row_num, column=6, value='Sí' if cliente.activo else 'No')
        ws.cell(row=row_num, column=7, value=cliente.dentista_asignado.nombre_completo if cliente.dentista_asignado else 'Sin asignar')
        ws.cell(row=row_num, column=8, value=cliente.num_citas)
        ws.cell(row=row_num, column=9, value=cliente.notas or '')
        
        for col in range(1, 10):
            ws.cell(row=row_num, column=col).border = border
    
    column_widths = [8, 30, 35, 15, 20, 10, 25, 12, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=clientes_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response


# ========== EXPORTACIÓN A EXCEL - INSUMOS ==========
@login_required
def exportar_excel_insumos(request):
    """Exporta todos los insumos a Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Insumos"
    
    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID', 'Nombre', 'Categoría', 'Cantidad Actual', 'Cantidad Mínima', 'Unidad', 'Precio Unitario', 'Estado', 'Proveedor', 'Ubicación']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    insumos = Insumo.objects.all().order_by('nombre')
    for row_num, insumo in enumerate(insumos, 2):
        ws.cell(row=row_num, column=1, value=insumo.id)
        ws.cell(row=row_num, column=2, value=insumo.nombre)
        ws.cell(row=row_num, column=3, value=insumo.get_categoria_display())
        ws.cell(row=row_num, column=4, value=insumo.cantidad_actual)
        ws.cell(row=row_num, column=5, value=insumo.cantidad_minima)
        ws.cell(row=row_num, column=6, value=insumo.unidad_medida)
        ws.cell(row=row_num, column=7, value=float(insumo.precio_unitario) if insumo.precio_unitario else 0)
        ws.cell(row=row_num, column=8, value=insumo.get_estado_display())
        ws.cell(row=row_num, column=9, value=insumo.proveedor or '')
        ws.cell(row=row_num, column=10, value=insumo.ubicacion or '')
        
        for col in range(1, 11):
            ws.cell(row=row_num, column=col).border = border
    
    column_widths = [8, 30, 20, 15, 15, 12, 15, 15, 25, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=insumos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response


# ========== EXPORTACIÓN A EXCEL - FINANZAS ==========
@login_required
def exportar_excel_finanzas(request):
    """Exporta reporte financiero a Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    
    wb = Workbook()
    
    # Hoja 1: Resumen Financiero
    ws1 = wb.active
    ws1.title = "Resumen"
    
    header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=14)
    
    ws1['A1'] = 'REPORTE FINANCIERO - CLÍNICA DENTAL'
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A1:D1')
    
    ws1['A2'] = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws1['A2'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A2:D2')
    
    # Estadísticas
    ws1['A4'] = 'ESTADÍSTICA'
    ws1['A4'].fill = header_fill
    ws1['A4'].font = header_font
    ws1['B4'] = 'VALOR'
    ws1['B4'].fill = header_fill
    ws1['B4'].font = header_font
    
    ws1['A5'] = 'Total de Citas Completadas'
    ws1['B5'] = Cita.objects.filter(estado='completada').count()
    
    ws1['A6'] = 'Total de Clientes Activos'
    ws1['B6'] = Cliente.objects.filter(activo=True).count()
    
    ws1['A7'] = 'Valor Total en Insumos'
    total_insumos = sum([
        (insumo.cantidad_actual * float(insumo.precio_unitario or 0))
        for insumo in Insumo.objects.all()
    ])
    ws1['B7'] = f'${total_insumos:,.2f}'
    
    ws1['A8'] = 'Citas Este Mes'
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    ws1['B8'] = Cita.objects.filter(fecha_hora__date__gte=inicio_mes).count()
    
    # Ajustar anchos
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=finanzas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response


from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Count, Sum, Q, Avg, F
from django.utils import timezone
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Cita
from pacientes.models import Cliente
from inventario.models import Insumo, MovimientoInsumo
from personal.models import Perfil
from evaluaciones.models import Evaluacion

# ========== VISTA PRINCIPAL DEL DASHBOARD ==========
@login_required
def dashboard_reportes(request):
    """Vista principal del dashboard con estadísticas y gráficos"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para acceder al dashboard.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    # Obtener fechas para filtros
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    fin_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    inicio_año = hoy.replace(month=1, day=1)
    
    # ===== ESTADÍSTICAS GENERALES =====
    total_citas = Cita.objects.count()
    total_clientes = Cliente.objects.count()
    total_personal = Perfil.objects.filter(activo=True).count()
    total_insumos = Insumo.objects.count()
    
    # Citas por estado
    citas_disponibles = Cita.objects.filter(estado='disponible').count()
    citas_reservadas = Cita.objects.filter(estado='reservada').count()
    citas_completadas = Cita.objects.filter(estado='completada').count()
    citas_canceladas = Cita.objects.filter(estado='cancelada').count()
    
    # Citas del mes
    citas_mes = Cita.objects.filter(
        fecha_hora__date__gte=inicio_mes,
        fecha_hora__date__lte=fin_mes
    ).count()
    
    citas_completadas_mes = Cita.objects.filter(
        estado='completada',
        fecha_hora__date__gte=inicio_mes,
        fecha_hora__date__lte=fin_mes
    ).count()
    
    # Insumos con stock bajo
    insumos_bajo_stock = Insumo.objects.filter(
        cantidad_actual__lte=F('cantidad_minima')
    ).count()
    
    # Dentistas más activos
    dentistas_activos = Perfil.objects.filter(
        rol='dentista',
        activo=True
    ).annotate(
        num_citas=Count('citas_asignadas')
    ).order_by('-num_citas')[:5]
    
    # Clientes nuevos este mes
    clientes_nuevos_mes = Cliente.objects.filter(
        fecha_registro__date__gte=inicio_mes,
        fecha_registro__date__lte=fin_mes
    ).count()
    
    # Tipos de consulta más frecuentes
    tipos_consulta = Cita.objects.exclude(
        tipo_consulta__isnull=True
    ).exclude(
        tipo_consulta=''
    ).values('tipo_consulta').annotate(
        total=Count('id')
    ).order_by('-total')[:5]
    
    # Citas por día de la semana (últimos 30 días)
    hace_30_dias = hoy - timedelta(days=30)
    citas_por_dia = []
    for i in range(7):
        dia = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo'][i]
        count = Cita.objects.filter(
            fecha_hora__date__gte=hace_30_dias,
            fecha_hora__week_day=i+2  # Django week_day: 1=domingo, 2=lunes...
        ).count()
        citas_por_dia.append({'dia': dia, 'total': count})
    
    # Citas por mes (últimos 6 meses)
    citas_por_mes = []
    for i in range(6):
        mes_fecha = hoy - timedelta(days=30*i)
        inicio = mes_fecha.replace(day=1)
        fin = (inicio + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        count = Cita.objects.filter(
            fecha_hora__date__gte=inicio,
            fecha_hora__date__lte=fin
        ).count()
        citas_por_mes.insert(0, {
            'mes': inicio.strftime('%B'),
            'total': count
        })
    
    # ===== ESTADÍSTICAS DE EVALUACIONES =====
    total_evaluaciones = Evaluacion.objects.filter(estado__in=['pendiente', 'revisada']).count()
    evaluaciones_pendientes = Evaluacion.objects.filter(estado='pendiente').count()
    evaluaciones_mes = Evaluacion.objects.filter(
        fecha_creacion__date__gte=inicio_mes,
        fecha_creacion__date__lte=fin_mes
    ).count()
    
    # Promedio de calificación
    promedio_evaluaciones = Evaluacion.objects.filter(
        estado__in=['pendiente', 'revisada']
    ).aggregate(Avg('estrellas'))['estrellas__avg'] or 0
    
    # Distribución de estrellas
    distribucion_evaluaciones = []
    for i in range(1, 6):
        count = Evaluacion.objects.filter(estrellas=i, estado__in=['pendiente', 'revisada']).count()
        distribucion_evaluaciones.append({'estrellas': i, 'total': count})
    
    # Últimas evaluaciones
    ultimas_evaluaciones = Evaluacion.objects.filter(
        estado__in=['pendiente', 'revisada']
    ).select_related('cliente').order_by('-fecha_creacion')[:5]
    
    context = {
        'perfil': perfil,
        'es_admin': True,
        # Estadísticas generales
        'total_citas': total_citas,
        'total_clientes': total_clientes,
        'total_personal': total_personal,
        'total_insumos': total_insumos,
        # Citas
        'citas_disponibles': citas_disponibles,
        'citas_reservadas': citas_reservadas,
        'citas_completadas': citas_completadas,
        'citas_canceladas': citas_canceladas,
        'citas_mes': citas_mes,
        'citas_completadas_mes': citas_completadas_mes,
        # Insumos
        'insumos_bajo_stock': insumos_bajo_stock,
        # Dentistas
        'dentistas_activos': dentistas_activos,
        # Clientes
        'clientes_nuevos_mes': clientes_nuevos_mes,
        # Gráficos
        'tipos_consulta': list(tipos_consulta),
        'citas_por_dia': citas_por_dia,
        'citas_por_mes': citas_por_mes,
        # Fechas
        'mes_actual': inicio_mes.strftime('%B %Y'),
        # Evaluaciones
        'total_evaluaciones': total_evaluaciones,
        'evaluaciones_pendientes': evaluaciones_pendientes,
        'evaluaciones_mes': evaluaciones_mes,
        'promedio_evaluaciones': round(promedio_evaluaciones, 1),
        'distribucion_evaluaciones': distribucion_evaluaciones,
        'ultimas_evaluaciones': ultimas_evaluaciones,
    }
    
    return render(request, 'citas/dashboard/dashboard_reportes.html', context)


# ========== EXPORTACIÓN A EXCEL - CITAS ==========
@login_required
def exportar_excel_citas(request):
    """Exporta todas las citas a Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Citas"
    
    # Estilos
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['ID', 'Fecha', 'Hora', 'Tipo Consulta', 'Estado', 'Paciente', 'Email', 'Teléfono', 'Dentista', 'Notas']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    citas = Cita.objects.all().order_by('-fecha_hora')
    for row_num, cita in enumerate(citas, 2):
        ws.cell(row=row_num, column=1, value=cita.id)
        ws.cell(row=row_num, column=2, value=cita.fecha_hora.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=3, value=cita.fecha_hora.strftime('%H:%M'))
        ws.cell(row=row_num, column=4, value=cita.tipo_consulta or 'N/A')
        ws.cell(row=row_num, column=5, value=cita.get_estado_display())
        ws.cell(row=row_num, column=6, value=cita.paciente_nombre or 'Sin asignar')
        ws.cell(row=row_num, column=7, value=cita.paciente_email or '')
        ws.cell(row=row_num, column=8, value=cita.paciente_telefono or '')
        ws.cell(row=row_num, column=9, value=cita.dentista.nombre_completo if cita.dentista else 'Sin asignar')
        ws.cell(row=row_num, column=10, value=cita.notas or '')
        
        # Aplicar bordes
        for col in range(1, 11):
            ws.cell(row=row_num, column=col).border = border
    
    # Ajustar anchos de columna
    column_widths = [8, 12, 10, 20, 12, 25, 30, 15, 25, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=citas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response


# ========== EXPORTACIÓN A EXCEL - CLIENTES ==========
@login_required
def exportar_excel_clientes(request):
    """Exporta todos los clientes a Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Estilos
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['ID', 'Nombre Completo', 'Email', 'Teléfono', 'Fecha Registro', 'Activo', 'Dentista Asignado', 'Total Citas', 'Notas']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    clientes = Cliente.objects.annotate(num_citas=Count('citas')).order_by('nombre_completo')
    for row_num, cliente in enumerate(clientes, 2):
        ws.cell(row=row_num, column=1, value=cliente.id)
        ws.cell(row=row_num, column=2, value=cliente.nombre_completo)
        ws.cell(row=row_num, column=3, value=cliente.email)
        ws.cell(row=row_num, column=4, value=cliente.telefono)
        ws.cell(row=row_num, column=5, value=cliente.fecha_registro.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row_num, column=6, value='Sí' if cliente.activo else 'No')
        ws.cell(row=row_num, column=7, value=cliente.dentista_asignado.nombre_completo if cliente.dentista_asignado else 'Sin asignar')
        ws.cell(row=row_num, column=8, value=cliente.num_citas)
        ws.cell(row=row_num, column=9, value=cliente.notas or '')
        
        for col in range(1, 10):
            ws.cell(row=row_num, column=col).border = border
    
    column_widths = [8, 30, 35, 15, 20, 10, 25, 12, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=clientes_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response


# ========== EXPORTACIÓN A EXCEL - INSUMOS ==========
@login_required
def exportar_excel_insumos(request):
    """Exporta todos los insumos a Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Insumos"
    
    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID', 'Nombre', 'Categoría', 'Cantidad Actual', 'Cantidad Mínima', 'Unidad', 'Precio Unitario', 'Estado', 'Proveedor', 'Ubicación']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    insumos = Insumo.objects.all().order_by('nombre')
    for row_num, insumo in enumerate(insumos, 2):
        ws.cell(row=row_num, column=1, value=insumo.id)
        ws.cell(row=row_num, column=2, value=insumo.nombre)
        ws.cell(row=row_num, column=3, value=insumo.get_categoria_display())
        ws.cell(row=row_num, column=4, value=insumo.cantidad_actual)
        ws.cell(row=row_num, column=5, value=insumo.cantidad_minima)
        ws.cell(row=row_num, column=6, value=insumo.unidad_medida)
        ws.cell(row=row_num, column=7, value=float(insumo.precio_unitario) if insumo.precio_unitario else 0)
        ws.cell(row=row_num, column=8, value=insumo.get_estado_display())
        ws.cell(row=row_num, column=9, value=insumo.proveedor or '')
        ws.cell(row=row_num, column=10, value=insumo.ubicacion or '')
        
        for col in range(1, 11):
            ws.cell(row=row_num, column=col).border = border
    
    column_widths = [8, 30, 20, 15, 15, 12, 15, 15, 25, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=insumos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response


# ========== EXPORTACIÓN A EXCEL - FINANZAS ==========
@login_required
def exportar_excel_finanzas(request):
    """Exporta reporte financiero a Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    
    wb = Workbook()
    
    # Hoja 1: Resumen Financiero
    ws1 = wb.active
    ws1.title = "Resumen"
    
    header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=14)
    
    ws1['A1'] = 'REPORTE FINANCIERO - CLÍNICA DENTAL'
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A1:D1')
    
    ws1['A2'] = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws1['A2'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A2:D2')
    
    # Estadísticas
    ws1['A4'] = 'ESTADÍSTICA'
    ws1['A4'].fill = header_fill
    ws1['A4'].font = header_font
    ws1['B4'] = 'VALOR'
    ws1['B4'].fill = header_fill
    ws1['B4'].font = header_font
    
    ws1['A5'] = 'Total de Citas Completadas'
    ws1['B5'] = Cita.objects.filter(estado='completada').count()
    
    ws1['A6'] = 'Total de Clientes Activos'
    ws1['B6'] = Cliente.objects.filter(activo=True).count()
    
    ws1['A7'] = 'Valor Total en Insumos'
    total_insumos = sum([
        (insumo.cantidad_actual * float(insumo.precio_unitario or 0))
        for insumo in Insumo.objects.all()
    ])
    ws1['B7'] = f'${total_insumos:,.2f}'
    
    ws1['A8'] = 'Citas Este Mes'
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    ws1['B8'] = Cita.objects.filter(fecha_hora__date__gte=inicio_mes).count()
    
    # Ajustar anchos
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=finanzas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response


from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Count, Sum, Q, Avg, F
from django.utils import timezone
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Cita
from pacientes.models import Cliente
from inventario.models import Insumo, MovimientoInsumo
from personal.models import Perfil
from evaluaciones.models import Evaluacion

# ========== VISTA PRINCIPAL DEL DASHBOARD ==========
@login_required
def dashboard_reportes(request):
    """Vista principal del dashboard con estadísticas y gráficos"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para acceder al dashboard.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    # Obtener fechas para filtros
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    fin_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    inicio_año = hoy.replace(month=1, day=1)
    
    # ===== ESTADÍSTICAS GENERALES =====
    total_citas = Cita.objects.count()
    total_clientes = Cliente.objects.count()
    total_personal = Perfil.objects.filter(activo=True).count()
    total_insumos = Insumo.objects.count()
    
    # Citas por estado
    citas_disponibles = Cita.objects.filter(estado='disponible').count()
    citas_reservadas = Cita.objects.filter(estado='reservada').count()
    citas_completadas = Cita.objects.filter(estado='completada').count()
    citas_canceladas = Cita.objects.filter(estado='cancelada').count()
    
    # Citas del mes
    citas_mes = Cita.objects.filter(
        fecha_hora__date__gte=inicio_mes,
        fecha_hora__date__lte=fin_mes
    ).count()
    
    citas_completadas_mes = Cita.objects.filter(
        estado='completada',
        fecha_hora__date__gte=inicio_mes,
        fecha_hora__date__lte=fin_mes
    ).count()
    
    # Insumos con stock bajo
    insumos_bajo_stock = Insumo.objects.filter(
        cantidad_actual__lte=F('cantidad_minima')
    ).count()
    
    # Dentistas más activos
    dentistas_activos = Perfil.objects.filter(
        rol='dentista',
        activo=True
    ).annotate(
        num_citas=Count('citas_asignadas')
    ).order_by('-num_citas')[:5]
    
    # Clientes nuevos este mes
    clientes_nuevos_mes = Cliente.objects.filter(
        fecha_registro__date__gte=inicio_mes,
        fecha_registro__date__lte=fin_mes
    ).count()
    
    # Tipos de consulta más frecuentes
    tipos_consulta = Cita.objects.exclude(
        tipo_consulta__isnull=True
    ).exclude(
        tipo_consulta=''
    ).values('tipo_consulta').annotate(
        total=Count('id')
    ).order_by('-total')[:5]
    
    # Citas por día de la semana (últimos 30 días)
    hace_30_dias = hoy - timedelta(days=30)
    citas_por_dia = []
    for i in range(7):
        dia = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo'][i]
        count = Cita.objects.filter(
            fecha_hora__date__gte=hace_30_dias,
            fecha_hora__week_day=i+2  # Django week_day: 1=domingo, 2=lunes...
        ).count()
        citas_por_dia.append({'dia': dia, 'total': count})
    
    # Citas por mes (últimos 6 meses)
    citas_por_mes = []
    for i in range(6):
        mes_fecha = hoy - timedelta(days=30*i)
        inicio = mes_fecha.replace(day=1)
        fin = (inicio + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        count = Cita.objects.filter(
            fecha_hora__date__gte=inicio,
            fecha_hora__date__lte=fin
        ).count()
        citas_por_mes.insert(0, {
            'mes': inicio.strftime('%B'),
            'total': count
        })
    
    # ===== ESTADÍSTICAS DE EVALUACIONES =====
    total_evaluaciones = Evaluacion.objects.filter(estado__in=['pendiente', 'revisada']).count()
    evaluaciones_pendientes = Evaluacion.objects.filter(estado='pendiente').count()
    evaluaciones_mes = Evaluacion.objects.filter(
        fecha_creacion__date__gte=inicio_mes,
        fecha_creacion__date__lte=fin_mes
    ).count()
    
    # Promedio de calificación
    promedio_evaluaciones = Evaluacion.objects.filter(
        estado__in=['pendiente', 'revisada']
    ).aggregate(Avg('estrellas'))['estrellas__avg'] or 0
    
    # Distribución de estrellas
    distribucion_evaluaciones = []
    for i in range(1, 6):
        count = Evaluacion.objects.filter(estrellas=i, estado__in=['pendiente', 'revisada']).count()
        distribucion_evaluaciones.append({'estrellas': i, 'total': count})
    
    # Últimas evaluaciones
    ultimas_evaluaciones = Evaluacion.objects.filter(
        estado__in=['pendiente', 'revisada']
    ).select_related('cliente').order_by('-fecha_creacion')[:5]
    
    context = {
        'perfil': perfil,
        'es_admin': True,
        # Estadísticas generales
        'total_citas': total_citas,
        'total_clientes': total_clientes,
        'total_personal': total_personal,
        'total_insumos': total_insumos,
        # Citas
        'citas_disponibles': citas_disponibles,
        'citas_reservadas': citas_reservadas,
        'citas_completadas': citas_completadas,
        'citas_canceladas': citas_canceladas,
        'citas_mes': citas_mes,
        'citas_completadas_mes': citas_completadas_mes,
        # Insumos
        'insumos_bajo_stock': insumos_bajo_stock,
        # Dentistas
        'dentistas_activos': dentistas_activos,
        # Clientes
        'clientes_nuevos_mes': clientes_nuevos_mes,
        # Gráficos
        'tipos_consulta': list(tipos_consulta),
        'citas_por_dia': citas_por_dia,
        'citas_por_mes': citas_por_mes,
        # Fechas
        'mes_actual': inicio_mes.strftime('%B %Y'),
        # Evaluaciones
        'total_evaluaciones': total_evaluaciones,
        'evaluaciones_pendientes': evaluaciones_pendientes,
        'evaluaciones_mes': evaluaciones_mes,
        'promedio_evaluaciones': round(promedio_evaluaciones, 1),
        'distribucion_evaluaciones': distribucion_evaluaciones,
        'ultimas_evaluaciones': ultimas_evaluaciones,
    }
    
    return render(request, 'citas/dashboard/dashboard_reportes.html', context)


# ========== EXPORTACIÓN A EXCEL - CITAS ==========
@login_required
def exportar_excel_citas(request):
    """Exporta todas las citas a Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Citas"
    
    # Estilos
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['ID', 'Fecha', 'Hora', 'Tipo Consulta', 'Estado', 'Paciente', 'Email', 'Teléfono', 'Dentista', 'Notas']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    citas = Cita.objects.all().order_by('-fecha_hora')
    for row_num, cita in enumerate(citas, 2):
        ws.cell(row=row_num, column=1, value=cita.id)
        ws.cell(row=row_num, column=2, value=cita.fecha_hora.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=3, value=cita.fecha_hora.strftime('%H:%M'))
        ws.cell(row=row_num, column=4, value=cita.tipo_consulta or 'N/A')
        ws.cell(row=row_num, column=5, value=cita.get_estado_display())
        ws.cell(row=row_num, column=6, value=cita.paciente_nombre or 'Sin asignar')
        ws.cell(row=row_num, column=7, value=cita.paciente_email or '')
        ws.cell(row=row_num, column=8, value=cita.paciente_telefono or '')
        ws.cell(row=row_num, column=9, value=cita.dentista.nombre_completo if cita.dentista else 'Sin asignar')
        ws.cell(row=row_num, column=10, value=cita.notas or '')
        
        # Aplicar bordes
        for col in range(1, 11):
            ws.cell(row=row_num, column=col).border = border
    
    # Ajustar anchos de columna
    column_widths = [8, 12, 10, 20, 12, 25, 30, 15, 25, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=citas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response


# ========== EXPORTACIÓN A EXCEL - CLIENTES ==========
@login_required
def exportar_excel_clientes(request):
    """Exporta todos los clientes a Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Estilos
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['ID', 'Nombre Completo', 'Email', 'Teléfono', 'Fecha Registro', 'Activo', 'Dentista Asignado', 'Total Citas', 'Notas']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    clientes = Cliente.objects.annotate(num_citas=Count('citas')).order_by('nombre_completo')
    for row_num, cliente in enumerate(clientes, 2):
        ws.cell(row=row_num, column=1, value=cliente.id)
        ws.cell(row=row_num, column=2, value=cliente.nombre_completo)
        ws.cell(row=row_num, column=3, value=cliente.email)
        ws.cell(row=row_num, column=4, value=cliente.telefono)
        ws.cell(row=row_num, column=5, value=cliente.fecha_registro.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row_num, column=6, value='Sí' if cliente.activo else 'No')
        ws.cell(row=row_num, column=7, value=cliente.dentista_asignado.nombre_completo if cliente.dentista_asignado else 'Sin asignar')
        ws.cell(row=row_num, column=8, value=cliente.num_citas)
        ws.cell(row=row_num, column=9, value=cliente.notas or '')
        
        for col in range(1, 10):
            ws.cell(row=row_num, column=col).border = border
    
    column_widths = [8, 30, 35, 15, 20, 10, 25, 12, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=clientes_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response


# ========== EXPORTACIÓN A EXCEL - INSUMOS ==========
@login_required
def exportar_excel_insumos(request):
    """Exporta todos los insumos a Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Insumos"
    
    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID', 'Nombre', 'Categoría', 'Cantidad Actual', 'Cantidad Mínima', 'Unidad', 'Precio Unitario', 'Estado', 'Proveedor', 'Ubicación']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    insumos = Insumo.objects.all().order_by('nombre')
    for row_num, insumo in enumerate(insumos, 2):
        ws.cell(row=row_num, column=1, value=insumo.id)
        ws.cell(row=row_num, column=2, value=insumo.nombre)
        ws.cell(row=row_num, column=3, value=insumo.get_categoria_display())
        ws.cell(row=row_num, column=4, value=insumo.cantidad_actual)
        ws.cell(row=row_num, column=5, value=insumo.cantidad_minima)
        ws.cell(row=row_num, column=6, value=insumo.unidad_medida)
        ws.cell(row=row_num, column=7, value=float(insumo.precio_unitario) if insumo.precio_unitario else 0)
        ws.cell(row=row_num, column=8, value=insumo.get_estado_display())
        ws.cell(row=row_num, column=9, value=insumo.proveedor or '')
        ws.cell(row=row_num, column=10, value=insumo.ubicacion or '')
        
        for col in range(1, 11):
            ws.cell(row=row_num, column=col).border = border
    
    column_widths = [8, 30, 20, 15, 15, 12, 15, 15, 25, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=insumos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response


# ========== EXPORTACIÓN A EXCEL - FINANZAS ==========
@login_required
def exportar_excel_finanzas(request):
    """Exporta reporte financiero a Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    
    wb = Workbook()
    
    # Hoja 1: Resumen Financiero
    ws1 = wb.active
    ws1.title = "Resumen"
    
    header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=14)
    
    ws1['A1'] = 'REPORTE FINANCIERO - CLÍNICA DENTAL'
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A1:D1')
    
    ws1['A2'] = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws1['A2'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A2:D2')
    
    # Estadísticas
    ws1['A4'] = 'ESTADÍSTICA'
    ws1['A4'].fill = header_fill
    ws1['A4'].font = header_font
    ws1['B4'] = 'VALOR'
    ws1['B4'].fill = header_fill
    ws1['B4'].font = header_font
    
    ws1['A5'] = 'Total de Citas Completadas'
    ws1['B5'] = Cita.objects.filter(estado='completada').count()
    
    ws1['A6'] = 'Total de Clientes Activos'
    ws1['B6'] = Cliente.objects.filter(activo=True).count()
    
    ws1['A7'] = 'Valor Total en Insumos'
    total_insumos = sum([
        (insumo.cantidad_actual * float(insumo.precio_unitario or 0))
        for insumo in Insumo.objects.all()
    ])
    ws1['B7'] = f'${total_insumos:,.2f}'
    
    ws1['A8'] = 'Citas Este Mes'
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    ws1['B8'] = Cita.objects.filter(fecha_hora__date__gte=inicio_mes).count()
    
    # Ajustar anchos
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=finanzas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    
    return response

