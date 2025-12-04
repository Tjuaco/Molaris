from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Count, Q, F, Sum, Avg
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from django.contrib.auth.views import LoginView
from django.contrib.auth import logout, authenticate
from django.contrib.auth.models import User
from django.urls import reverse_lazy, reverse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.cache import cache
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_protect
from django.utils.decorators import method_decorator
import logging
import time

# PDF generation imports
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from io import BytesIO

from .models import Cita, TipoServicio, HorarioDentista
from personal.models import Perfil
from pacientes.models import Cliente
from inventario.models import Insumo, MovimientoInsumo
from historial_clinico.models import Odontograma, EstadoDiente, Radiografia, PlanTratamiento, FaseTratamiento, ItemTratamiento, PagoTratamiento, DocumentoCliente, ConsentimientoInformado, PlantillaConsentimiento
from django.db.models import Prefetch
from proveedores.models import Proveedor, SolicitudInsumo
from evaluaciones.models import Evaluacion
from finanzas.models import IngresoManual, EgresoManual
# Importar el modelo de clientes de la app cuentas
try:
    from cuentas.models import PerfilCliente
except ImportError:
    PerfilCliente = None
from .forms import RegistroTrabajadorForm, PerfilForm
# Notificaciones ahora se envían a través de mensajeria_service (WhatsApp + SMS)
from .helpers_planes import verificar_permiso_plan_tratamiento, obtener_clientes_permitidos
from .models_auditoria import registrar_auditoria
from .validaciones import (
    validar_email_cliente,
    validar_rut_cliente,
    validar_telefono_cliente,
    validar_username_disponible,
    validar_datos_cliente_completos
)
from django.db import transaction
import re
import logging
import json

logger = logging.getLogger(__name__)


def normalizar_telefono_chileno(telefono):
    """
    Normaliza un número de teléfono chileno al formato +56912345678
    Prioriza el formato de 8 dígitos (los últimos 8 del número celular).
    Acepta varios formatos de entrada:
    - 12345678 (8 dígitos - formato preferido, se agrega 9 y código país)
    - 912345678 (9 dígitos con 9 inicial)
    - +56912345678 (formato completo)
    - 56912345678 (con código país pero sin +)
    
    Retorna el número normalizado como +56912345678 o None si no es válido
    """
    if not telefono:
        return None
    
    # Eliminar espacios, guiones, paréntesis y otros caracteres
    telefono_limpio = re.sub(r'[\s\-\(\)\.]', '', str(telefono).strip())
    
    # Si empieza con +, eliminarlo para procesar
    if telefono_limpio.startswith('+'):
        telefono_limpio = telefono_limpio[1:]
    
    # Si empieza con 0, eliminarlo (formato nacional antiguo)
    if telefono_limpio.startswith('0'):
        telefono_limpio = telefono_limpio[1:]
    
    # Validar que solo contenga dígitos
    if not telefono_limpio.isdigit():
        return None
    
    # CASO PREFERIDO: 8 dígitos (solo los últimos 8 del número celular)
    # Se agrega el 9 inicial y el código de país 56
    if len(telefono_limpio) == 8:
        return f"+569{telefono_limpio}"
    
    # CASO 2: Ya tiene código de país 56 y empieza con 9 (+56912345678 o 56912345678)
    if telefono_limpio.startswith('569') and len(telefono_limpio) == 11:
        return f"+{telefono_limpio}"
    
    # CASO 3: Solo tiene 9 dígitos empezando con 9 (912345678)
    if telefono_limpio.startswith('9') and len(telefono_limpio) == 9:
        return f"+56{telefono_limpio}"
    
    # CASO 4: Tiene código 56 seguido de 9 dígitos (56912345678)
    if telefono_limpio.startswith('56') and len(telefono_limpio) == 11:
        # Verificar que el tercer dígito sea 9 (celular)
        if telefono_limpio[2] == '9':
            return f"+{telefono_limpio}"
    
    # Si no coincide con ningún formato válido, retornar None
    return None

# Importar vistas del dashboard
from .views_dashboard import (
    dashboard_reportes,
    exportar_excel_citas,
    exportar_excel_clientes,
    exportar_excel_insumos,
    exportar_excel_finanzas
)

# Página de inicio (redirige al login)
def inicio(request):
    return redirect('login')

# Login personalizado para trabajadores con protección de seguridad
@method_decorator(never_cache, name='dispatch')
class TrabajadorLoginView(LoginView):
    template_name = 'citas/auth/login.html'
    
    def dispatch(self, *args, **kwargs):
        """Aplicar protección de rate limiting"""
        # Obtener IP del cliente
        ip_address = self.get_client_ip()
        
        # Verificar rate limiting (máximo 5 intentos en 15 minutos)
        cache_key = f'login_attempts_{ip_address}'
        attempts = cache.get(cache_key, 0)
        
        if attempts >= 5:
            # Bloquear por 15 minutos
            messages.error(self.request, '⚠️ Demasiados intentos fallidos. Por favor, espera 15 minutos antes de intentar nuevamente.')
            logger.warning(f'Login bloqueado por rate limiting - IP: {ip_address}')
            return self.render_to_response(self.get_context_data())
        
        return super().dispatch(*args, **kwargs)
    
    def get_client_ip(self):
        """Obtener la IP real del cliente"""
        x_forwarded_for = self.request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = self.request.META.get('REMOTE_ADDR')
        return ip
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Asegurar que el formulario esté en el contexto
        if 'form' not in context:
            from django.contrib.auth.forms import AuthenticationForm
            context['form'] = AuthenticationForm()
        return context
    
    def form_invalid(self, form):
        """Personalizar mensajes de error de autenticación con seguridad mejorada"""
        # Obtener IP del cliente
        ip_address = self.get_client_ip()
        
        # Incrementar contador de intentos fallidos
        cache_key = f'login_attempts_{ip_address}'
        attempts = cache.get(cache_key, 0) + 1
        cache.set(cache_key, attempts, 900)  # 15 minutos
        
        # Log del intento fallido (sin revelar información sensible)
        username_attempt = self.request.POST.get('username', '').strip()[:20]  # Limitar longitud
        logger.warning(f'Intento de login fallido - IP: {ip_address}, Usuario: {username_attempt}, Intentos: {attempts}')
        
        # Limpiar mensajes anteriores
        storage = messages.get_messages(self.request)
        storage.used = True
        
        # Mensaje genérico que no revela si el usuario existe o no (protección contra user enumeration)
        # Siempre usar el mismo tiempo de respuesta para evitar timing attacks
        time.sleep(0.1)  # Pequeño delay para normalizar tiempo de respuesta
        
        if attempts >= 5:
            messages.error(self.request, '⚠️ Demasiados intentos fallidos. Tu acceso ha sido temporalmente bloqueado por 15 minutos.')
        else:
            messages.error(self.request, '❌ Usuario o contraseña incorrectos. Por favor, verifica tus credenciales.')
            if attempts >= 3:
                messages.warning(self.request, f'⚠️ Advertencia: {5 - attempts} intentos restantes antes del bloqueo temporal.')
        
        return super().form_invalid(form)
    
    def form_valid(self, form):
        """Limpiar contador de intentos fallidos al iniciar sesión exitosamente"""
        ip_address = self.get_client_ip()
        cache_key = f'login_attempts_{ip_address}'
        cache.delete(cache_key)
        
        # Log de login exitoso
        logger.info(f'Login exitoso - Usuario: {form.get_user().username}, IP: {ip_address}')
        
        # Registrar en auditoría
        try:
            perfil = Perfil.objects.get(user=form.get_user())
            registrar_auditoria(
                usuario=perfil,
                accion='login',
                modulo='sistema',
                descripcion=f'Inicio de sesión exitoso: {perfil.nombre_completo}',
                detalles=f'Usuario: {form.get_user().username}',
                request=self.request
            )
        except Perfil.DoesNotExist:
            pass  # Si no hay perfil, no registrar (puede ser un usuario sin perfil)
        
        return super().form_valid(form)

    def get_success_url(self):
        # Redirigir según rol del trabajador
        try:
            perfil = Perfil.objects.get(user=self.request.user)
            if perfil.activo:
                # Si es dentista, redirigir al dashboard de dentista
                if perfil.es_dentista():
                    return reverse_lazy('dashboard_dentista')
                # Si es administrativo, redirigir al panel de trabajador
                else:
                    return reverse_lazy('panel_trabajador')
            else:
                messages.error(self.request, 'Tu cuenta está desactivada. Contacta al administrador.')
                return reverse_lazy('login')
        except Perfil.DoesNotExist:
            messages.error(self.request, 'No tienes un perfil de trabajador válido.')
            return reverse_lazy('login')

# Panel trabajador (recepción/dentista)
@login_required
def panel_trabajador(request):
    # Redirigir a la nueva vista de citas del día con navbar lateral
    # Mantener compatibilidad con parámetro tab para redireccionar correctamente
    tab = request.GET.get('tab', 'today')
    
    if tab == 'today' or tab == '':
        return redirect('citas_dia')
    elif tab == 'list':
        return redirect('citas_disponibles')
    elif tab == 'taken':
        return redirect('citas_tomadas')
    elif tab == 'completed':
        return redirect('citas_completadas')
    elif tab == 'calendar':
        return redirect('calendario_citas')
    else:
        # Por defecto, redirigir a citas del día
        return redirect('citas_dia')

# Vista AJAX para obtener citas del día actualizadas (para actualización automática)
@login_required
def obtener_citas_dia_ajax(request):
    """Vista AJAX que devuelve las citas del día en formato JSON para actualización automática"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'error': 'Cuenta desactivada'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'error': 'Perfil no encontrado'}, status=404)
    
    # Obtener citas del día
    citas_hoy = Cita.objects.filter(
        fecha_hora__date=timezone.now().date()
    ).select_related('tipo_servicio', 'dentista', 'cliente').prefetch_related('odontogramas').order_by('fecha_hora')
    
    # Obtener información de fichas
    from historial_clinico.models import Odontograma
    odontogramas = Odontograma.objects.filter(cita__isnull=False).select_related('cita')
    citas_con_ficha = set(odontogramas.values_list('cita_id', flat=True))
    
    # Preparar datos de citas en formato JSON
    citas_data = []
    for cita in citas_hoy:
        cita.tiene_ficha = cita.id in citas_con_ficha
        odontograma = None
        if cita.tiene_ficha:
            odontograma = odontogramas.filter(cita_id=cita.id).first()
        
        # Convertir fecha_hora a zona horaria de Chile
        from django.utils import timezone
        import pytz
        try:
            chile_tz = pytz.timezone('America/Santiago')
            if timezone.is_naive(cita.fecha_hora):
                fecha_hora_chile = timezone.make_aware(cita.fecha_hora, timezone.utc).astimezone(chile_tz)
            else:
                fecha_hora_chile = cita.fecha_hora.astimezone(chile_tz)
        except Exception:
            fecha_hora_chile = cita.fecha_hora
        
        citas_data.append({
            'id': cita.id,
            'estado': cita.estado,
            'estado_display': cita.get_estado_display(),
            'fecha': fecha_hora_chile.strftime('%d/%m/%Y'),
            'fecha_hora': fecha_hora_chile.strftime('%Y-%m-%d %H:%M'),
            'hora': fecha_hora_chile.strftime('%H:%M'),
            'paciente_nombre': cita.paciente_nombre or (cita.cliente.nombre_completo if cita.cliente else 'Sin asignar'),
            'dentista_id': cita.dentista.id if cita.dentista else None,
            'dentista_nombre': cita.dentista.nombre_completo if cita.dentista else 'Sin asignar',
            'tipo_servicio': cita.tipo_servicio.nombre if cita.tipo_servicio else (cita.tipo_consulta or 'Sin servicio'),
            'tiene_ficha': cita.tiene_ficha,
            'odontograma_id': odontograma.id if odontograma else None,
            'hora_llegada': cita.hora_llegada.strftime('%Y-%m-%d %H:%M') if cita.hora_llegada else None,
        })
    
    return JsonResponse({
        'success': True,
        'citas': citas_data,
        'timestamp': timezone.now().isoformat()
    })

# Agregar hora disponible (solo administrativos)
# Esta vista solo procesa POST desde el modal, nunca renderiza una página
@login_required
def agregar_hora(request):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para agregar horas.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    # Si es GET, redirigir al panel (no debería accederse directamente)
    if request.method == 'GET':
        return redirect('panel_trabajador')
    
    # Procesar POST desde el modal
    if request.method == 'POST':
        fecha_hora_str = request.POST.get('fecha_hora', '')
        tipo_consulta = request.POST.get('tipo_consulta', '').strip()
        notas = request.POST.get('notas', '')
        dentista_id = request.POST.get('dentista_id', '')
        tipo_servicio_id = request.POST.get('tipo_servicio', '').strip()
        
        errores = []
        
        try:
            # Validar campos requeridos
            if not fecha_hora_str:
                errores.append('Debe seleccionar una fecha y hora.')
            
            if not dentista_id:
                errores.append('Debe seleccionar un dentista.')
            
            # El modal puede enviar tipo_consulta (nombre del servicio) o tipo_servicio (ID)
            # Si viene tipo_consulta, buscar el servicio por nombre
            tipo_servicio = None
            if tipo_servicio_id:
                try:
                    tipo_servicio = TipoServicio.objects.get(id=tipo_servicio_id, activo=True)
                except TipoServicio.DoesNotExist:
                    errores.append('El servicio seleccionado no existe o está inactivo.')
            elif tipo_consulta:
                # Buscar servicio por nombre
                try:
                    tipo_servicio = TipoServicio.objects.get(nombre=tipo_consulta, activo=True)
                    tipo_servicio_id = str(tipo_servicio.id)
                except TipoServicio.DoesNotExist:
                    # Si no se encuentra, se creará la cita con tipo_consulta como texto libre
                    pass
            
            # Si hay errores, devolver JSON para mostrar alerta en el modal
            if errores:
                return JsonResponse({'success': False, 'error': errores[0]}, status=400)
            
            # Convertir fecha y hacerla timezone-aware
            from django.utils import timezone
            fecha_hora_naive = datetime.fromisoformat(fecha_hora_str)
            # Hacer el datetime aware usando la timezone del sistema
            fecha_hora = timezone.make_aware(fecha_hora_naive)
            
            # Validar que la fecha no sea en el pasado
            ahora = timezone.now()
            if fecha_hora < ahora:
                return JsonResponse({'success': False, 'error': 'No se pueden crear citas en fechas pasadas. Por favor, seleccione una fecha y hora futura.'}, status=400)
            
            # Obtener el dentista seleccionado
            try:
                dentista = Perfil.objects.get(id=dentista_id, rol='dentista', activo=True)
            except Perfil.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Dentista seleccionado no válido o inactivo.'}, status=400)
            
            # Validar que la fecha/hora esté dentro del horario de trabajo del dentista
            dia_semana = fecha_hora.weekday()  # 0=Lunes, 6=Domingo
            hora_cita = fecha_hora.time()
            
            # Obtener horarios activos del dentista para ese día
            horarios_dia = HorarioDentista.objects.filter(
                dentista=dentista,
                dia_semana=dia_semana,
                activo=True
            )
            
            if not horarios_dia.exists():
                dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
                return JsonResponse({'success': False, 'error': f'El dentista {dentista.nombre_completo} no trabaja los {dias_nombres[dia_semana]}.'}, status=400)
            
            # Verificar que la hora esté dentro de algún bloque de horario
            hora_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita < horario.hora_fin:
                    hora_valida = True
                    break
            
            if not hora_valida:
                horarios_str = ', '.join([f"{h.hora_inicio.strftime('%H:%M')}-{h.hora_fin.strftime('%H:%M')}" for h in horarios_dia])
                return JsonResponse({'success': False, 'error': f'La hora seleccionada no está dentro del horario de trabajo del dentista. Horarios disponibles: {horarios_str}'}, status=400)
            
            # Validar duración de la cita
            duracion_minutos = 30  # Duración por defecto
            if tipo_servicio and tipo_servicio.duracion_estimada:
                duracion_minutos = tipo_servicio.duracion_estimada
            
            # Calcular hora de fin de la cita
            from datetime import timedelta
            fecha_hora_fin = fecha_hora + timedelta(minutes=duracion_minutos)
            hora_fin_cita = fecha_hora_fin.time()
            
            # Verificar que la duración completa quepa en el bloque de horario
            duracion_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita and hora_fin_cita <= horario.hora_fin:
                    duracion_valida = True
                    break
            
            if not duracion_valida:
                return JsonResponse({'success': False, 'error': f'La duración del servicio ({duracion_minutos} minutos) no cabe en el horario seleccionado. Por favor, seleccione una hora más temprana.'}, status=400)
            
            # Verificar que no se solape con otra cita del mismo dentista
            citas_existentes = Cita.objects.filter(
                dentista=dentista,
                fecha_hora__date=fecha_hora.date(),
                estado__in=['disponible', 'reservada', 'confirmada']
            ).exclude(id=None)  # Excluir la cita actual si se está editando
            
            for cita_existente in citas_existentes:
                fecha_hora_existente_fin = cita_existente.fecha_hora
                if cita_existente.tipo_servicio and cita_existente.tipo_servicio.duracion_estimada:
                    fecha_hora_existente_fin += timedelta(minutes=cita_existente.tipo_servicio.duracion_estimada)
                else:
                    fecha_hora_existente_fin += timedelta(minutes=30)  # Duración por defecto
                
                # Verificar solapamiento
                if (fecha_hora < fecha_hora_existente_fin and fecha_hora_fin > cita_existente.fecha_hora):
                    return JsonResponse({'success': False, 'error': f'La cita se solapa con otra cita existente del dentista a las {cita_existente.fecha_hora.strftime("%H:%M")}.'}, status=400)
            
            # Verificar que no exista ya una cita en esa fecha/hora exacta
            if Cita.objects.filter(fecha_hora=fecha_hora).exists():
                return JsonResponse({'success': False, 'error': 'Ya existe una cita en esa fecha y hora exacta.'}, status=400)
            
            # Obtener precio del servicio si existe
            precio_cobrado = None
            if tipo_servicio:
                precio_cobrado = tipo_servicio.precio_base
                
                # Si hay precio_cobrado en el formulario, usarlo (puede ser un ajuste)
                precio_cobrado_str = request.POST.get('precio_cobrado', '').strip()
                if precio_cobrado_str:
                    try:
                        # Limpiar formato si tiene puntos
                        precio_cobrado_limpio = precio_cobrado_str.replace('.', '').replace(',', '.')
                        precio_ajustado = float(precio_cobrado_limpio)
                        if precio_ajustado > 0:
                            precio_cobrado = precio_ajustado
                    except ValueError:
                        pass  # Usar precio base si el ajuste no es válido
            
            # Manejar asignación de paciente (cliente existente o nuevo)
            cliente_obj = None
            paciente_nombre = None
            paciente_email = None
            paciente_telefono = None
            estado_cita = 'disponible'
            
            # Solo procesar cliente si el checkbox está marcado
            asignar_cliente = request.POST.get('asignar_cliente', '') == 'on'
            cliente_id = request.POST.get('cliente_id', '').strip()
            
            # Log para debugging (solo en desarrollo)
            from django.conf import settings
            if settings.DEBUG:
                logger.info(f"Procesando cita - asignar_cliente: {asignar_cliente}, cliente_id: '{cliente_id}'")
            
            if asignar_cliente and cliente_id:
                if cliente_id == 'nuevo':
                    # Crear nuevo cliente - solo si se proporcionaron los datos necesarios
                    paciente_nombre = request.POST.get('paciente_nombre', '').strip()
                    paciente_email = request.POST.get('paciente_email', '').strip()
                    paciente_telefono_raw = request.POST.get('paciente_telefono', '').strip()
                    
                    if paciente_nombre and paciente_email:
                        # Normalizar teléfono si se proporcionó
                        paciente_telefono = None
                        if paciente_telefono_raw:
                            paciente_telefono = normalizar_telefono_chileno(paciente_telefono_raw)
                            if not paciente_telefono:
                                return JsonResponse({
                                    'success': False, 
                                    'error': f'El número de teléfono "{paciente_telefono_raw}" no es válido. Por favor, ingrese un número de celular chileno de 8 dígitos (ejemplo: 20589344).'
                                }, status=400)
                        
                        try:
                            # Si no se proporcionó teléfono, usar uno por defecto válido
                            telefono_para_guardar = paciente_telefono if paciente_telefono else '+56900000000'
                            
                            # Crear o obtener cliente
                            cliente_obj, created = Cliente.objects.get_or_create(
                                email=paciente_email,
                                defaults={
                                    'nombre_completo': paciente_nombre,
                                    'telefono': telefono_para_guardar,
                                    'activo': True
                                }
                            )
                            # Si el cliente ya existía, actualizar información si es necesario
                            if not created:
                                cliente_obj.nombre_completo = paciente_nombre
                                if paciente_telefono:
                                    cliente_obj.telefono = paciente_telefono
                                cliente_obj.activo = True
                                try:
                                    cliente_obj.save()
                                except Exception as save_error:
                                    error_msg = str(save_error)
                                    if 'telefono' in error_msg.lower() or 'phone' in error_msg.lower() or 'regex' in error_msg.lower():
                                        return JsonResponse({
                                            'success': False, 
                                            'error': f'El número de teléfono no es válido. Por favor, ingrese un número de celular chileno de 8 dígitos (ejemplo: 20589344).'
                                        }, status=400)
                                    raise
                            
                            estado_cita = 'reservada'
                        except Cliente.MultipleObjectsReturned:
                            # Si hay múltiples clientes con el mismo email (no debería pasar, pero por seguridad)
                            cliente_obj = Cliente.objects.filter(email=paciente_email, activo=True).first()
                            if cliente_obj:
                                cliente_obj.nombre_completo = paciente_nombre
                                if paciente_telefono:
                                    cliente_obj.telefono = paciente_telefono
                                cliente_obj.activo = True
                                cliente_obj.save()
                                estado_cita = 'reservada'
                            else:
                                return JsonResponse({
                                    'success': False, 
                                    'error': 'Error al procesar el cliente. Por favor, intente nuevamente.'
                                }, status=400)
                        except Exception as e:
                            error_msg = str(e)
                            import traceback
                            logger.error(f"Error al crear/actualizar cliente: {e}")
                            logger.error(traceback.format_exc())
                            
                            if 'email' in error_msg.lower() or 'unique' in error_msg.lower() or 'duplicate' in error_msg.lower():
                                return JsonResponse({
                                    'success': False, 
                                    'error': f'Ya existe un cliente con el email "{paciente_email}". Por favor, seleccione el cliente existente o use otro email.'
                                }, status=400)
                            elif 'telefono' in error_msg.lower() or 'phone' in error_msg.lower() or 'regex' in error_msg.lower() or 'validator' in error_msg.lower():
                                return JsonResponse({
                                    'success': False, 
                                    'error': f'El número de teléfono no es válido. Por favor, ingrese un número de celular chileno de 8 dígitos (ejemplo: 20589344).'
                                }, status=400)
                            else:
                                # Re-lanzar el error para que se capture en el bloque general
                                raise
                    else:
                        # Si se marcó el checkbox pero no se proporcionaron los datos necesarios
                        return JsonResponse({
                            'success': False, 
                            'error': 'Para crear un nuevo cliente, debe proporcionar al menos el nombre y el email.'
                        }, status=400)
                elif cliente_id and cliente_id != 'nuevo':
                    # Usar cliente existente
                    try:
                        # Convertir cliente_id a entero
                        cliente_id_int = int(cliente_id)
                        cliente_obj = Cliente.objects.get(id=cliente_id_int, activo=True)
                        paciente_nombre = cliente_obj.nombre_completo
                        paciente_email = cliente_obj.email
                        paciente_telefono = cliente_obj.telefono
                        estado_cita = 'reservada'
                    except ValueError:
                        # Si cliente_id no es un número válido
                        logger.error(f"cliente_id no es un número válido: {cliente_id}")
                        return JsonResponse({
                            'success': False, 
                            'error': 'El ID del cliente no es válido. Por favor, seleccione un cliente de la lista.'
                        }, status=400)
                    except Cliente.DoesNotExist:
                        logger.error(f"Cliente con ID {cliente_id} no encontrado o inactivo")
                        return JsonResponse({
                            'success': False, 
                            'error': 'El cliente seleccionado no existe o está inactivo. Por favor, seleccione otro cliente.'
                        }, status=400)
                    except Exception as e:
                        logger.error(f"Error al obtener cliente {cliente_id}: {e}")
                        import traceback
                        logger.error(traceback.format_exc())
                        return JsonResponse({
                            'success': False, 
                            'error': 'Error al procesar el cliente seleccionado. Por favor, intente nuevamente.'
                        }, status=400)
            
            # Validar que si se asignó un cliente, el objeto cliente existe
            if asignar_cliente and estado_cita == 'reservada' and not cliente_obj:
                return JsonResponse({
                    'success': False, 
                    'error': 'Error: Se intentó asignar un cliente pero no se pudo obtener. Por favor, verifique que el cliente esté activo y exista en el sistema.'
                }, status=400)
            
            # Crear la cita
            try:
                cita = Cita.objects.create(
                    fecha_hora=fecha_hora,
                    tipo_consulta=tipo_consulta,
                    tipo_servicio=tipo_servicio,
                    precio_cobrado=precio_cobrado,
                    notas=notas,
                    dentista=dentista,
                    creada_por=perfil,
                    cliente=cliente_obj,  # Puede ser None si no se asignó cliente
                    paciente_nombre=paciente_nombre,
                    paciente_email=paciente_email,
                    paciente_telefono=paciente_telefono,
                    estado=estado_cita
                )
            except Exception as e:
                error_msg = str(e)
                import traceback
                logger.error(f"Error al crear la cita: {e}")
                logger.error(traceback.format_exc())
                
                # Si el error está relacionado con el cliente o foreign key
                if 'cliente' in error_msg.lower() or 'foreign key' in error_msg.lower() or 'constraint' in error_msg.lower():
                    # Verificar si el cliente existe
                    if cliente_obj:
                        cliente_existe = Cliente.objects.filter(id=cliente_obj.id, activo=True).exists()
                        if not cliente_existe:
                            return JsonResponse({
                                'success': False, 
                                'error': 'El cliente seleccionado no existe o está inactivo. Por favor, seleccione otro cliente o cree la cita sin asignar cliente.'
                            }, status=400)
                    
                    return JsonResponse({
                        'success': False, 
                        'error': 'Error al asignar el cliente a la cita. Por favor, verifique que el cliente esté activo o cree la cita sin asignar cliente.'
                    }, status=400)
                raise
            
            # Registrar en auditoría
            cliente_info = cliente_obj.nombre_completo if cliente_obj else paciente_nombre or "Sin cliente"
            servicio_info = tipo_servicio.nombre if tipo_servicio else tipo_consulta or "Sin servicio"
            registrar_auditoria(
                usuario=perfil,
                accion='crear',
                modulo='citas',
                descripcion=f'Cita creada para {cliente_info} con {dentista.nombre_completo} - {servicio_info}',
                detalles=f'Fecha: {fecha_hora.strftime("%d/%m/%Y %H:%M")}, Estado: {estado_cita}, Precio: ${precio_cobrado if precio_cobrado else "N/A"}',
                objeto_id=cita.id,
                tipo_objeto='Cita',
                request=request
            )
            
            # Enviar notificaciones (WhatsApp Y SMS) si la cita está reservada y tiene teléfono
            # Usar el teléfono del cliente si existe, sino el de la cita
            telefono_para_notificacion = None
            if cita.cliente and cita.cliente.telefono:
                telefono_para_notificacion = cita.cliente.telefono
            elif paciente_telefono:
                telefono_para_notificacion = paciente_telefono
            
            if estado_cita == 'reservada' and telefono_para_notificacion:
                try:
                    from citas.mensajeria_service import enviar_notificaciones_cita
                    resultado = enviar_notificaciones_cita(cita, telefono_override=telefono_para_notificacion)
                    
                    # Construir mensaje de confirmación
                    canales_enviados = []
                    if resultado.get('whatsapp', {}).get('enviado'):
                        canales_enviados.append('WhatsApp')
                    if resultado.get('sms', {}).get('enviado'):
                        canales_enviados.append('SMS')
                    if resultado.get('email', {}).get('enviado'):
                        canales_enviados.append('Correo')
                    
                    if canales_enviados:
                        logger.info(f"Notificaciones enviadas por {', '.join(canales_enviados)} para cita {cita.id} a {telefono_para_notificacion}")
                    else:
                        logger.warning(f"No se pudieron enviar notificaciones para cita {cita.id}. Errores: WhatsApp={resultado.get('whatsapp', {}).get('error')}, SMS={resultado.get('sms', {}).get('error')}, Email={resultado.get('email', {}).get('error')}")
                except Exception as e:
                    # No fallar la creación de la cita si las notificaciones fallan
                    logger.error(f"Error al enviar notificaciones para cita {cita.id}: {e}")
                    import traceback
                    logger.error(traceback.format_exc())
            
            # Si todo salió bien, devolver JSON para que el modal muestre el mensaje
            if estado_cita == 'reservada':
                mensaje = f'Cita creada y asignada a {paciente_nombre} para {dentista.nombre_completo}.'
                # Las notificaciones ya se enviaron arriba (línea 692), no duplicar
                if telefono_para_notificacion or cita.paciente_email:
                    mensaje += ' Se han enviado notificaciones al paciente.'
            else:
                mensaje = f'Cita creada correctamente para {dentista.nombre_completo}.'
            
            return JsonResponse({'success': True, 'message': mensaje})
            
        except ValueError as e:
            error_msg = str(e)
            if 'time data' in error_msg.lower() or 'date' in error_msg.lower():
                return JsonResponse({'success': False, 'error': 'El formato de fecha y hora no es válido. Por favor, seleccione una fecha y hora correcta.'}, status=400)
            return JsonResponse({'success': False, 'error': f'Error en el formato de los datos: {error_msg}'}, status=400)
        except Exception as e:
            # Log del error completo para debugging
            import traceback
            logger.error(f"Error al crear cita: {e}")
            logger.error(traceback.format_exc())
            # Mensaje genérico pero útil para el usuario
            error_msg = str(e)
            error_trace = traceback.format_exc()
            
            # Detectar errores específicos
            if 'dentista' in error_msg.lower():
                return JsonResponse({'success': False, 'error': 'Error al procesar la información del dentista. Por favor, verifique que el dentista esté activo y tenga horarios configurados.'}, status=400)
            elif 'horario' in error_msg.lower() or 'schedule' in error_msg.lower():
                return JsonResponse({'success': False, 'error': 'Error al validar el horario. Por favor, verifique que la fecha y hora seleccionadas sean válidas.'}, status=400)
            elif 'cliente' in error_msg.lower() or 'paciente' in error_msg.lower() or 'email' in error_msg.lower() or 'unique' in error_msg.lower():
                # Error relacionado con cliente
                if 'email' in error_msg.lower() or 'unique' in error_msg.lower():
                    return JsonResponse({'success': False, 'error': 'Ya existe un cliente con ese email. Por favor, seleccione el cliente existente o use otro email.'}, status=400)
                elif 'telefono' in error_msg.lower() or 'phone' in error_msg.lower() or 'regex' in error_msg.lower():
                    return JsonResponse({'success': False, 'error': 'El número de teléfono no es válido. Por favor, ingrese un número de celular chileno de 8 dígitos (ejemplo: 20589344).'}, status=400)
                else:
                    return JsonResponse({'success': False, 'error': 'Error al procesar la información del cliente. Por favor, verifique los datos ingresados (nombre, email y teléfono).'}, status=400)
            elif 'telefono' in error_msg.lower() or 'phone' in error_msg.lower() or 'regex' in error_msg.lower() or 'validator' in error_msg.lower():
                return JsonResponse({'success': False, 'error': 'El número de teléfono no es válido. Por favor, ingrese un número de celular chileno de 8 dígitos (ejemplo: 20589344).'}, status=400)
            else:
                # Mensaje más detallado para debugging (solo en desarrollo)
                import sys
                if 'django' in sys.modules and hasattr(sys.modules['django'], 'conf'):
                    from django.conf import settings
                    if settings.DEBUG:
                        return JsonResponse({'success': False, 'error': f'Error al crear la cita: {error_msg}. Detalles: {error_trace[:200]}'}, status=400)
                return JsonResponse({'success': False, 'error': f'Error al crear la cita: {error_msg}. Por favor, verifique todos los campos e intente nuevamente.'}, status=400)

# Editar cita creada (solo administrativos)
# - Citas disponibles: se puede editar todo
# - Citas reservadas/confirmadas: se puede editar fecha/hora, dentista, servicio, precio (NO cliente)
# - Citas completadas: NO se pueden editar (usar ajustar_precio_cita)
@login_required
def editar_cita(request, cita_id):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para editar citas.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    cita = get_object_or_404(Cita, id=cita_id)
    
    # Restricción: No se pueden editar citas completadas
    if cita.estado == 'completada':
        messages.error(request, 'Las citas completadas no se pueden editar. Use "Ajustar Precio" para modificar el precio o notas.')
        return redirect('panel_trabajador')
    
    # Permitir editar cliente si la cita está disponible o si se quiere asignar un cliente a una cita existente
    puede_editar_cliente = (cita.estado == 'disponible')

    if request.method == 'POST':
        fecha_hora_str = request.POST.get('fecha_hora')
        tipo_consulta = request.POST.get('tipo_consulta', '')
        tipo_servicio_id = request.POST.get('tipo_servicio', '')
        precio_cobrado = request.POST.get('precio_cobrado', '')
        notas = request.POST.get('notas', '')
        dentista_id = request.POST.get('dentista', '')
        cliente_id = request.POST.get('cliente_id', '')
        
        try:
            fecha_hora = datetime.fromisoformat(fecha_hora_str)
            cita.fecha_hora = fecha_hora
            cita.tipo_consulta = tipo_consulta
            
            # Para citas reservadas/confirmadas: permitir cambiar dentista
            if dentista_id and cita.estado in ['reservada', 'confirmada']:
                try:
                    dentista = Perfil.objects.get(id=dentista_id, es_dentista=True, activo=True)
                    cita.dentista = dentista
                except Perfil.DoesNotExist:
                    pass
            
            # Actualizar tipo de servicio
            if tipo_servicio_id:
                try:
                    tipo_servicio = TipoServicio.objects.get(id=tipo_servicio_id, activo=True)
                    cita.tipo_servicio = tipo_servicio
                    # Si no se especifica precio manualmente, usar el precio base del servicio
                    if not precio_cobrado:
                        cita.precio_cobrado = tipo_servicio.precio_base
                    else:
                        try:
                            cita.precio_cobrado = float(precio_cobrado)
                        except ValueError:
                            cita.precio_cobrado = tipo_servicio.precio_base
                except TipoServicio.DoesNotExist:
                    messages.warning(request, 'El servicio seleccionado no existe o está inactivo.')
                    cita.tipo_servicio = None
            else:
                # Si se cambia el precio manualmente sin servicio
                if precio_cobrado:
                    try:
                        cita.precio_cobrado = float(precio_cobrado)
                    except ValueError:
                        pass
                elif cita.estado == 'disponible':
                    # Solo para citas disponibles se puede quitar servicio
                    cita.tipo_servicio = None
                    cita.precio_cobrado = None
            
            # Asignar cliente si se proporciona y la cita está disponible
            if cliente_id and cita.estado == 'disponible':
                try:
                    cliente = Cliente.objects.get(id=cliente_id, activo=True)
                    cita.cliente = cliente
                    cita.paciente_nombre = cliente.nombre_completo
                    cita.paciente_email = cliente.email
                    cita.paciente_telefono = cliente.telefono
                    # Cambiar estado a reservada si se asigna un cliente
                    cita.estado = 'reservada'
                except Cliente.DoesNotExist:
                    messages.warning(request, 'El cliente seleccionado no existe o está inactivo.')
            elif cliente_id and cita.estado in ['reservada', 'confirmada']:
                # Permitir cambiar cliente incluso en citas reservadas si se proporciona explícitamente
                try:
                    cliente = Cliente.objects.get(id=cliente_id, activo=True)
                    cita.cliente = cliente
                    cita.paciente_nombre = cliente.nombre_completo
                    cita.paciente_email = cliente.email
                    cita.paciente_telefono = cliente.telefono
                except Cliente.DoesNotExist:
                    messages.warning(request, 'El cliente seleccionado no existe o está inactivo.')
            
            cita.notas = notas
            cita.save()
            
            # Registrar en auditoría
            cliente_info = cita.cliente.nombre_completo if cita.cliente else cita.paciente_nombre or "Sin cliente"
            servicio_info = cita.tipo_servicio.nombre if cita.tipo_servicio else cita.tipo_consulta or "Sin servicio"
            registrar_auditoria(
                usuario=perfil,
                accion='editar',
                modulo='citas',
                descripcion=f'Cita editada: {cliente_info} - {servicio_info}',
                detalles=f'Fecha: {cita.fecha_hora.strftime("%d/%m/%Y %H:%M")}, Estado: {cita.estado}, Dentista: {cita.dentista.nombre_completo if cita.dentista else "N/A"}',
                objeto_id=cita.id,
                tipo_objeto='Cita',
                request=request
            )
            
            mensaje = 'Cita actualizada correctamente.'
            if cliente_id and cita.cliente:
                mensaje += f' Cliente asignado: {cita.cliente.nombre_completo}.'
            messages.success(request, mensaje)
            
            # Handle AJAX requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': True, 'message': mensaje})
            
            return redirect('panel_trabajador')
        except Exception as e:
            messages.error(request, f'Error al actualizar la cita: {e}')
            
            # Handle AJAX requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'message': f'Error al actualizar la cita: {e}'}, status=400)

    # Obtener servicios activos para el selector
    servicios_activos = TipoServicio.objects.filter(activo=True).order_by('categoria', 'nombre')
    
    # Obtener dentistas para el selector (solo si es reservada/confirmada)
    dentistas = None
    if cita.estado in ['reservada', 'confirmada']:
        dentistas = Perfil.objects.filter(es_dentista=True, activo=True).order_by('nombre', 'apellido')
    
    # Obtener clientes activos para el selector
    clientes = Cliente.objects.filter(activo=True).order_by('nombre_completo')
    
    context = {
        'perfil': perfil,
        'cita': cita,
        'servicios': servicios_activos,
        'dentistas': dentistas,
        'clientes': clientes,
        'puede_editar_cliente': puede_editar_cliente,
    }
    return render(request, 'citas/citas/editar_cita.html', context)


# Acciones sobre citas

@login_required
def cancelar_cita_admin(request, cita_id):
    try:
        perfil = Perfil.objects.get(user=request.user)
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'Perfil no encontrado.'}, status=404)
        return redirect('login')
    
    cita = get_object_or_404(Cita, id=cita_id)
    
    # Verificar permisos: administrativos pueden cancelar cualquier cita, dentistas solo las suyas
    if perfil.es_dentista() and cita.dentista != perfil:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'No tienes permisos para cancelar esta cita.'}, status=403)
        messages.error(request, 'No tienes permisos para cancelar esta cita.')
        return redirect('mis_citas_dentista')
    elif not perfil.es_administrativo() and not perfil.es_dentista():
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'No tienes permisos para cancelar citas.'}, status=403)
        messages.error(request, 'No tienes permisos para cancelar citas.')
        return redirect('panel_trabajador')
    
    # Verificar que la cita pueda ser cancelada (solo reservadas)
    if cita.estado != 'reservada':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': f'No se puede cancelar una cita en estado "{cita.get_estado_display()}". Solo se pueden cancelar citas reservadas.'}, status=400)
        messages.error(request, f'No se puede cancelar una cita en estado "{cita.get_estado_display()}".')
        if perfil.es_dentista():
            return redirect('mis_citas_dentista')
        else:
            return redirect('panel_trabajador')
    
    if cita.cancelar():
        # Enviar notificaciones de cancelación por WhatsApp, SMS y correo electrónico
        try:
            from citas.mensajeria_service import enviar_notificaciones_cancelacion_cita
            import logging
            logger = logging.getLogger(__name__)
            
            resultado = enviar_notificaciones_cancelacion_cita(cita)
            
            # Log del resultado
            canales_exitosos = []
            if resultado.get('whatsapp', {}).get('enviado'):
                canales_exitosos.append('WhatsApp')
            if resultado.get('sms', {}).get('enviado'):
                canales_exitosos.append('SMS')
            if resultado.get('email', {}).get('enviado'):
                canales_exitosos.append('Correo')
            
            if canales_exitosos:
                logger.info(f"Notificaciones de cancelación enviadas por: {', '.join(canales_exitosos)} para cita {cita.id}")
            else:
                logger.warning(f"No se pudieron enviar notificaciones de cancelación para cita {cita.id}. Errores: WhatsApp={resultado.get('whatsapp', {}).get('error')}, SMS={resultado.get('sms', {}).get('error')}, Email={resultado.get('email', {}).get('error')}")
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error al enviar notificaciones de cancelación para cita {cita.id}: {e}", exc_info=True)
        
        mensaje = f'Cita del {cita.fecha_hora.strftime("%d/%m/%Y a las %H:%M")} cancelada exitosamente.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': True, 'message': mensaje})
        messages.success(request, mensaje)
        # Redirigir según el rol
        if perfil.es_dentista():
            return redirect('mis_citas_dentista')
        else:
            return redirect('panel_trabajador')
    else:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'No se pudo cancelar la cita. Verifique que la cita esté en estado reservada.'}, status=400)
        messages.error(request, 'No se pudo cancelar la cita')
        if perfil.es_dentista():
            return redirect('mis_citas_dentista')
        else:
            return redirect('panel_trabajador')

@login_required
def marcar_no_show(request, cita_id):
    """
    Vista para marcar una cita como "No Show" (el paciente no se presentó)
    SOLO DISPONIBLE PARA ADMINISTRATIVOS
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción.'}, status=403)
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'Perfil no encontrado.'}, status=404)
        messages.error(request, 'Perfil no encontrado.')
        return redirect('login')
    
    cita = get_object_or_404(Cita, id=cita_id)
    
    # Solo se puede marcar como no_show si está en estado reservada o confirmada
    if cita.estado not in ['reservada', 'confirmada']:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': f'No se puede marcar como "No Show" una cita en estado "{cita.get_estado_display()}".'}, status=400)
        messages.error(request, f'No se puede marcar como "No Show" una cita en estado "{cita.get_estado_display()}".')
        return redirect('panel_trabajador')
    
    # Obtener motivo de no asistencia (OBLIGATORIO)
    motivo_no_asistencia = request.POST.get('motivo_no_asistencia', '').strip()
    
    # Validar que el motivo sea obligatorio
    if not motivo_no_asistencia:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'El motivo de no asistencia es obligatorio. Por favor, indique la razón por la cual el paciente no asistió.'}, status=400)
        messages.error(request, 'El motivo de no asistencia es obligatorio. Por favor, indique la razón por la cual el paciente no asistió.')
        return redirect('panel_trabajador')
    
    # Marcar como no_show y guardar motivo
    cita.estado = 'no_show'
    cita.motivo_no_asistencia = motivo_no_asistencia
    cita.save()
    
    mensaje = f'Cita marcada como "No Show" para {cita.paciente_nombre or "Sin nombre"} ({cita.fecha_hora.strftime("%d/%m/%Y %H:%M")}). Motivo registrado en el historial.'
    
    # Handle AJAX requests
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        from django.http import JsonResponse
        return JsonResponse({'success': True, 'message': mensaje})
    
    messages.success(request, mensaje)
    return redirect('panel_trabajador')

@login_required
def marcar_listo_para_atender(request, cita_id):
    """
    Vista para que el dentista marque que está listo para atender al paciente
    Cambia el estado de "en_espera" a "listo_para_atender"
    SOLO DISPONIBLE PARA DENTISTAS (el dentista asignado a la cita)
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': 'Solo los dentistas pueden marcar citas como listas para atender.'}, status=403)
            messages.error(request, 'Solo los dentistas pueden realizar esta acción.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'Perfil no encontrado.'}, status=404)
        messages.error(request, 'Perfil no encontrado.')
        return redirect('login')
    
    cita = get_object_or_404(Cita, id=cita_id)
    
    # Verificar que el dentista esté asignado a esta cita
    if not cita.dentista or cita.dentista != perfil:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'No tienes permisos para modificar esta cita.'}, status=403)
        messages.error(request, 'No tienes permisos para modificar esta cita.')
        return redirect('panel_trabajador')
    
    # Solo se puede marcar como listo si está en estado "en_espera"
    if cita.estado != 'en_espera':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': f'La cita debe estar en estado "En Espera" para marcarla como lista. Estado actual: "{cita.get_estado_display()}".'}, status=400)
        messages.error(request, f'La cita debe estar en estado "En Espera". Estado actual: "{cita.get_estado_display()}".')
        return redirect('panel_trabajador')
    
    # Cambiar estado
    cita.estado = 'listo_para_atender'
    cita.save()
    
    mensaje = f'Cita marcada como "Listo para Atender". La recepcionista será notificada para pasar al paciente.'
    
    # Handle AJAX requests
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        from django.http import JsonResponse
        return JsonResponse({'success': True, 'message': mensaje})
    
    messages.success(request, mensaje)
    return redirect('panel_trabajador')

@login_required
def iniciar_atencion(request, cita_id):
    """
    Vista para que el dentista marque que inició la atención del paciente
    Cambia el estado de "listo_para_atender" a "en_progreso"
    SOLO DISPONIBLE PARA DENTISTAS (el dentista asignado a la cita)
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': 'Solo los dentistas pueden iniciar la atención.'}, status=403)
            messages.error(request, 'Solo los dentistas pueden realizar esta acción.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'Perfil no encontrado.'}, status=404)
        messages.error(request, 'Perfil no encontrado.')
        return redirect('login')
    
    cita = get_object_or_404(Cita, id=cita_id)
    
    # Verificar que el dentista esté asignado a esta cita
    if not cita.dentista or cita.dentista != perfil:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'No tienes permisos para modificar esta cita.'}, status=403)
        messages.error(request, 'No tienes permisos para modificar esta cita.')
        return redirect('panel_trabajador')
    
    # Solo se puede iniciar atención si está en estado "listo_para_atender"
    if cita.estado != 'listo_para_atender':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': f'La cita debe estar en estado "Listo para Atender". Estado actual: "{cita.get_estado_display()}".'}, status=400)
        messages.error(request, f'La cita debe estar en estado "Listo para Atender". Estado actual: "{cita.get_estado_display()}".')
        return redirect('panel_trabajador')
    
    # Cambiar estado
    cita.estado = 'en_progreso'
    cita.save()
    
    mensaje = f'Atención iniciada para {cita.paciente_nombre or "el paciente"}.'
    
    # Handle AJAX requests
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        from django.http import JsonResponse
        return JsonResponse({'success': True, 'message': mensaje})
    
    messages.success(request, mensaje)
    return redirect('panel_trabajador')

@login_required
def finalizar_atencion(request, cita_id):
    """
    Vista para que el dentista marque que finalizó la atención (después de crear la ficha)
    Cambia el estado de "en_progreso" a "finalizada"
    SOLO DISPONIBLE PARA DENTISTAS (el dentista asignado a la cita)
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': 'Solo los dentistas pueden finalizar la atención.'}, status=403)
            messages.error(request, 'Solo los dentistas pueden realizar esta acción.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'Perfil no encontrado.'}, status=404)
        messages.error(request, 'Perfil no encontrado.')
        return redirect('login')
    
    cita = get_object_or_404(Cita, id=cita_id)
    
    # Verificar que el dentista esté asignado a esta cita
    if not cita.dentista or cita.dentista != perfil:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'No tienes permisos para modificar esta cita.'}, status=403)
        messages.error(request, 'No tienes permisos para modificar esta cita.')
        return redirect('panel_trabajador')
    
    # Solo se puede finalizar atención si está en estado "en_progreso"
    if cita.estado != 'en_progreso':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': f'La cita debe estar en estado "En Progreso". Estado actual: "{cita.get_estado_display()}".'}, status=400)
        messages.error(request, f'La cita debe estar en estado "En Progreso". Estado actual: "{cita.get_estado_display()}".')
        return redirect('panel_trabajador')
    
    # Cambiar estado
    cita.estado = 'finalizada'
    cita.save()
    
    mensaje = f'Atención finalizada. El paciente puede dirigirse a recepción para pagar.'
    
    # Handle AJAX requests
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        from django.http import JsonResponse
        return JsonResponse({'success': True, 'message': mensaje})
    
    messages.success(request, mensaje)
    return redirect('panel_trabajador')

@login_required
def completar_cita_recepcion(request, cita_id):
    """
    Vista para que la recepcionista marque la cita como completada (después de recibir el pago)
    Cambia el estado de "finalizada" a "completada" y registra el precio cobrado
    SOLO DISPONIBLE PARA ADMINISTRATIVOS
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': 'Solo el personal administrativo puede completar citas.'}, status=403)
            messages.error(request, 'Solo el personal administrativo puede realizar esta acción.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'Perfil no encontrado.'}, status=404)
        messages.error(request, 'Perfil no encontrado.')
        return redirect('login')
    
    cita = get_object_or_404(Cita, id=cita_id)
    
    # Solo se puede completar si está en estado "finalizada"
    if cita.estado != 'finalizada':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': f'La cita debe estar en estado "Finalizada" para completarla. Estado actual: "{cita.get_estado_display()}".'}, status=400)
        messages.error(request, f'La cita debe estar en estado "Finalizada". Estado actual: "{cita.get_estado_display()}".')
        return redirect('panel_trabajador')
    
    # Obtener precio cobrado
    if request.method == 'POST':
        import json
        if request.headers.get('Content-Type') == 'application/json':
            data = json.loads(request.body)
            precio_cobrado = data.get('precio_cobrado')
        else:
            precio_cobrado = request.POST.get('precio_cobrado')
        
        if precio_cobrado:
            try:
                precio_cobrado = float(precio_cobrado)
                if precio_cobrado <= 0:
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        from django.http import JsonResponse
                        return JsonResponse({'success': False, 'error': 'El precio debe ser mayor a 0.'}, status=400)
                    messages.error(request, 'El precio debe ser mayor a 0.')
                    return redirect('panel_trabajador')
                cita.precio_cobrado = precio_cobrado
            except (ValueError, TypeError):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    from django.http import JsonResponse
                    return JsonResponse({'success': False, 'error': 'El precio ingresado no es válido.'}, status=400)
                messages.error(request, 'El precio ingresado no es válido.')
                return redirect('panel_trabajador')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': 'Debe ingresar el precio cobrado.'}, status=400)
            messages.error(request, 'Debe ingresar el precio cobrado.')
            return redirect('panel_trabajador')
    
    # Cambiar estado
    cita.estado = 'completada'
    cita.save()
    
    # Registrar en auditoría
    registrar_auditoria(
        usuario=perfil,
        accion='cambio_estado',
        modulo='citas',
        descripcion=f'Cita completada - Precio cobrado: ${cita.precio_cobrado:,.0f}',
        detalles=f'Cita ID: {cita.id}, Estado anterior: Finalizada, Estado nuevo: Completada, Precio cobrado: ${cita.precio_cobrado:,.0f}',
        objeto_id=cita.id,
        tipo_objeto='Cita',
        request=request
    )
    
    mensaje = f'Cita completada exitosamente para {cita.paciente_nombre or "el paciente"}. Precio cobrado: ${cita.precio_cobrado:,.0f}'
    
    # Handle AJAX requests
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        from django.http import JsonResponse
        return JsonResponse({'success': True, 'message': mensaje})
    
    messages.success(request, mensaje)
    return redirect('panel_trabajador')

@login_required
def marcar_llegada(request, cita_id):
    """
    Vista para marcar que el paciente llegó físicamente a la clínica
    Cambia el estado a "en_espera" y registra la hora de llegada
    SOLO DISPONIBLE PARA ADMINISTRATIVOS
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción.'}, status=403)
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'Perfil no encontrado.'}, status=404)
        messages.error(request, 'Perfil no encontrado.')
        return redirect('login')
    
    cita = get_object_or_404(Cita, id=cita_id)
    
    # Se puede marcar llegada en cualquier estado excepto los finales (completada, no_show, cancelada)
    # Esto permite manejar casos especiales: llegadas tempranas, tardías, o cambios de estado
    # Pero NO se puede marcar llegada a una cita disponible (nadie la tomó)
    estados_no_permitidos = ['completada', 'no_show', 'cancelada', 'disponible']
    if cita.estado in estados_no_permitidos:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': f'No se puede marcar llegada para una cita en estado "{cita.get_estado_display()}".'}, status=400)
        messages.error(request, f'No se puede marcar llegada para una cita en estado "{cita.get_estado_display()}".')
        return redirect('panel_trabajador')
    
    # Marcar como en espera y registrar hora de llegada
    cita.estado = 'en_espera'
    cita.hora_llegada = timezone.now()
    cita.save()
    
    mensaje = f'Paciente {cita.paciente_nombre or "Sin nombre"} marcado como "En Espera". Hora de llegada: {cita.hora_llegada.strftime("%H:%M")}.'
    
    # Handle AJAX requests
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        from django.http import JsonResponse
        return JsonResponse({'success': True, 'message': mensaje})
    
    messages.success(request, mensaje)
    return redirect('panel_trabajador')

@login_required
def marcar_no_llego(request, cita_id):
    """
    Vista para marcar que el paciente no llegó a la cita
    Cambia el estado a "no_show" (igual que marcar_no_show pero con nombre diferente para claridad)
    SOLO DISPONIBLE PARA ADMINISTRATIVOS
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción.'}, status=403)
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'Perfil no encontrado.'}, status=404)
        messages.error(request, 'Perfil no encontrado.')
        return redirect('login')
    
    cita = get_object_or_404(Cita, id=cita_id)
    
    # Solo se puede marcar como "No Llegó" si la cita está reservada o confirmada
    # No tiene sentido marcar "No Llegó" a una cita disponible (nadie la tomó)
    if cita.estado not in ['reservada', 'confirmada']:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': f'No se puede marcar como "No Llegó" una cita en estado "{cita.get_estado_display()}". Solo se puede marcar en citas reservadas o confirmadas.'}, status=400)
        messages.error(request, f'No se puede marcar como "No Llegó" una cita en estado "{cita.get_estado_display()}". Solo se puede marcar en citas reservadas o confirmadas.')
        return redirect('panel_trabajador')
    
    # Obtener motivo de no asistencia (OBLIGATORIO)
    motivo_no_asistencia = request.POST.get('motivo_no_asistencia', '').strip()
    
    # Validar que el motivo sea obligatorio
    if not motivo_no_asistencia:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'El motivo de no asistencia es obligatorio. Por favor, indique la razón por la cual el paciente no asistió.'}, status=400)
        messages.error(request, 'El motivo de no asistencia es obligatorio. Por favor, indique la razón por la cual el paciente no asistió.')
        return redirect('panel_trabajador')
    
    # Marcar como no_show y guardar motivo
    cita.estado = 'no_show'
    cita.motivo_no_asistencia = motivo_no_asistencia
    cita.save()
    
    mensaje = f'Cita marcada como "No Llegó" para {cita.paciente_nombre or "Sin nombre"} ({cita.fecha_hora.strftime("%d/%m/%Y %H:%M")}). Motivo registrado en el historial del cliente.'
    
    # Handle AJAX requests
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        from django.http import JsonResponse
        return JsonResponse({'success': True, 'message': mensaje})
    
    messages.success(request, mensaje)
    return redirect('panel_trabajador')

@login_required
def reagendar_cita(request, cita_id):
    """
    Vista para reagendar una cita (cambiar fecha/hora)
    SOLO DISPONIBLE PARA ADMINISTRATIVOS
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': 'No tienes permisos para reagendar citas.'}, status=403)
            messages.error(request, 'No tienes permisos para reagendar citas.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'Perfil no encontrado.'}, status=404)
        messages.error(request, 'Perfil no encontrado.')
        return redirect('login')
    
    cita = get_object_or_404(Cita.objects.select_related('dentista', 'tipo_servicio'), id=cita_id)
    
    # Solo se puede reagendar si está en estado reservada, confirmada o disponible
    if cita.estado not in ['reservada', 'confirmada', 'disponible']:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': f'No se puede reagendar una cita en estado "{cita.get_estado_display()}".'}, status=400)
        messages.error(request, f'No se puede reagendar una cita en estado "{cita.get_estado_display()}".')
        return redirect('panel_trabajador')
    
    if request.method == 'POST':
        nueva_fecha_hora_str = request.POST.get('nueva_fecha_hora', '').strip()
        nuevo_dentista_id = request.POST.get('nuevo_dentista', '').strip()
        
        if not nueva_fecha_hora_str:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': 'Debe seleccionar una nueva fecha y hora.'}, status=400)
            messages.error(request, 'Debe seleccionar una nueva fecha y hora.')
            return redirect('panel_trabajador')
        
        try:
            from datetime import datetime
            fecha_hora_naive = datetime.fromisoformat(nueva_fecha_hora_str)
            nueva_fecha_hora = timezone.make_aware(fecha_hora_naive)
        except ValueError:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': 'Formato de fecha/hora inválido.'}, status=400)
            messages.error(request, 'Formato de fecha/hora inválido.')
            return redirect('panel_trabajador')
        
        # Validar que la nueva fecha no sea en el pasado
        ahora = timezone.now()
        if nueva_fecha_hora < ahora:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': 'No se puede reagendar a una fecha/hora pasada.'}, status=400)
            messages.error(request, 'No se puede reagendar a una fecha/hora pasada.')
            return redirect('panel_trabajador')
        
        # Determinar el dentista (mantener el actual o cambiar)
        dentista = cita.dentista
        if nuevo_dentista_id:
            try:
                nuevo_dentista = Perfil.objects.get(id=nuevo_dentista_id, rol='dentista', activo=True)
                dentista = nuevo_dentista
            except Perfil.DoesNotExist:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    from django.http import JsonResponse
                    return JsonResponse({'success': False, 'error': 'Dentista seleccionado no válido.'}, status=400)
                messages.error(request, 'Dentista seleccionado no válido.')
                return redirect('panel_trabajador')
        
        # Validar horario del dentista
        if dentista:
            dia_semana = nueva_fecha_hora.weekday()
            hora_cita = nueva_fecha_hora.time()
            
            horarios_dia = HorarioDentista.objects.filter(
                dentista=dentista,
                dia_semana=dia_semana,
                activo=True
            )
            
            if not horarios_dia.exists():
                dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    from django.http import JsonResponse
                    return JsonResponse({'success': False, 'error': f'El dentista no trabaja los {dias_nombres[dia_semana]}.'}, status=400)
                messages.error(request, f'El dentista no trabaja los {dias_nombres[dia_semana]}.')
                return redirect('panel_trabajador')
            
            hora_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita < horario.hora_fin:
                    hora_valida = True
                    break
            
            if not hora_valida:
                horarios_str = ', '.join([f"{h.hora_inicio.strftime('%H:%M')}-{h.hora_fin.strftime('%H:%M')}" for h in horarios_dia])
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    from django.http import JsonResponse
                    return JsonResponse({'success': False, 'error': f'La hora seleccionada no está dentro del horario de trabajo del dentista. Horarios disponibles: {horarios_str}'}, status=400)
                messages.error(request, f'La hora seleccionada no está dentro del horario de trabajo del dentista. Horarios disponibles: {horarios_str}')
                return redirect('panel_trabajador')
        
        # Verificar que no se solape con otra cita del mismo dentista
        citas_existentes = Cita.objects.filter(
            dentista=dentista,
            fecha_hora__date=nueva_fecha_hora.date(),
            estado__in=['disponible', 'reservada', 'confirmada']
        ).exclude(id=cita.id)
        
        from datetime import timedelta
        duracion_minutos = 30
        if cita.tipo_servicio and cita.tipo_servicio.duracion_estimada:
            duracion_minutos = cita.tipo_servicio.duracion_estimada
        
        fecha_hora_fin = nueva_fecha_hora + timedelta(minutes=duracion_minutos)
        
        for cita_existente in citas_existentes:
            fecha_hora_existente_fin = cita_existente.fecha_hora
            if cita_existente.tipo_servicio and cita_existente.tipo_servicio.duracion_estimada:
                fecha_hora_existente_fin += timedelta(minutes=cita_existente.tipo_servicio.duracion_estimada)
            else:
                fecha_hora_existente_fin += timedelta(minutes=30)
            
            if (nueva_fecha_hora < fecha_hora_existente_fin and fecha_hora_fin > cita_existente.fecha_hora):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    from django.http import JsonResponse
                    return JsonResponse({'success': False, 'error': f'La nueva fecha/hora se solapa con otra cita existente a las {cita_existente.fecha_hora.strftime("%H:%M")}.'}, status=400)
                messages.error(request, f'La nueva fecha/hora se solapa con otra cita existente a las {cita_existente.fecha_hora.strftime("%H:%M")}.')
                return redirect('panel_trabajador')
        
        # Verificar que no exista ya una cita en esa fecha/hora exacta
        if Cita.objects.filter(fecha_hora=nueva_fecha_hora).exclude(id=cita.id).exists():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': 'Ya existe una cita en esa fecha y hora exacta.'}, status=400)
            messages.error(request, 'Ya existe una cita en esa fecha y hora exacta.')
            return redirect('panel_trabajador')
        
        # Guardar fecha/hora anterior para el mensaje
        fecha_hora_anterior = cita.fecha_hora
        
        # Guardar dentista anterior para el mensaje
        dentista_anterior = cita.dentista
        
        # Actualizar la cita
        cita.fecha_hora = nueva_fecha_hora
        if dentista:
            cita.dentista = dentista
        cita.save()
        
        mensaje = f'Cita reagendada de {fecha_hora_anterior.strftime("%d/%m/%Y %H:%M")} a {nueva_fecha_hora.strftime("%d/%m/%Y %H:%M")}.'
        if dentista and dentista != dentista_anterior:
            mensaje += f' Asignada a {dentista.nombre_completo}.'
        
        # Handle AJAX requests
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': True, 'message': mensaje})
        
        messages.success(request, mensaje)
        return redirect('panel_trabajador')
    
    # Si es GET, devolver información de la cita para el modal
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        from django.http import JsonResponse
        return JsonResponse({
            'success': True,
            'cita': {
                'id': cita.id,
                'fecha_hora_actual': cita.fecha_hora.strftime('%Y-%m-%dT%H:%M'),
                'paciente_nombre': cita.paciente_nombre or 'Sin asignar',
                'dentista_id': cita.dentista.id if cita.dentista else None,
                'dentista_nombre': cita.dentista.nombre_completo if cita.dentista else 'Sin asignar',
            }
        })
    
    return redirect('panel_trabajador')

@login_required
def confirmar_cita(request, cita_id):
    """
    Vista para confirmar una cita (cambiar de 'reservada' a 'confirmada')
    SOLO DISPONIBLE PARA ADMINISTRATIVOS (recepcionistas)
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para confirmar citas. Solo el personal administrativo puede realizar esta acción.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'Perfil no encontrado.')
        return redirect('login')
    
    cita = get_object_or_404(Cita, id=cita_id)
    
    # Verificar que la cita esté en estado "reservada"
    if cita.estado != 'reservada':
        messages.error(request, f'Esta cita está en estado "{cita.get_estado_display()}" y no puede ser confirmada.')
        return redirect('panel_trabajador')
    
    # Cambiar a estado confirmada
    cita.estado = 'confirmada'
    cita.save()
    
    messages.success(request, f'✅ Cita del {cita.fecha_hora.strftime("%d/%m/%Y a las %H:%M")} confirmada exitosamente.')
    return redirect('panel_trabajador')


@login_required
def completar_cita(request, cita_id):
    """
    Vista para marcar una cita como completada
    SOLO DISPONIBLE PARA ADMINISTRATIVOS (recepcionistas)
    Los dentistas NO pueden completar sus propias citas
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para completar citas. Solo el personal administrativo puede realizar esta acción.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'Perfil no encontrado.')
        return redirect('login')
    
    # Cargar la cita con el tipo_servicio relacionado para acceder al precio_base
    cita = get_object_or_404(Cita.objects.select_related('tipo_servicio', 'cliente', 'dentista'), id=cita_id)
    
    # Verificar que la cita esté en estado "confirmada" o "reservada"
    if cita.estado not in ['confirmada', 'reservada']:
        messages.error(request, f'Esta cita está en estado "{cita.get_estado_display()}" y no puede ser completada.')
        return redirect('panel_trabajador')
    
    # VALIDACIÓN CRÍTICA 1: Verificar que la cita no sea futura
    from django.utils import timezone
    ahora = timezone.now()
    if cita.fecha_hora > ahora:
        tiempo_faltante = cita.fecha_hora - ahora
        horas_faltantes = int(tiempo_faltante.total_seconds() / 3600)
        messages.error(
            request,
            f'⚠️ No se puede completar una cita futura. '
            f'Esta cita está programada para el {cita.fecha_hora.strftime("%d/%m/%Y a las %H:%M")} '
            f'(faltan aproximadamente {horas_faltantes} horas). '
            f'Solo se pueden completar citas que ya han ocurrido.'
        )
        return redirect('panel_trabajador')
    
    # VALIDACIÓN CRÍTICA 2: Verificar si la cita tiene ficha odontológica creada
    from historial_clinico.models import Odontograma
    tiene_ficha = Odontograma.objects.filter(cita=cita).exists()
    odontograma = None
    ficha_completa = False
    
    if tiene_ficha:
        odontograma = Odontograma.objects.filter(cita=cita).first()
        # Verificar que la ficha tenga los campos mínimos requeridos
        if odontograma:
            ficha_completa = (
                odontograma.paciente_nombre and
                odontograma.motivo_consulta and
                odontograma.dentista
            )
    
    # Si es POST, validar y actualizar
    if request.method == 'POST':
        # Validar que tenga ficha antes de completar
        if not tiene_ficha:
            messages.error(
                request, 
                '⚠️ No se puede completar la cita sin crear la ficha odontológica. '
                'Por favor, asegúrate de que el dentista haya creado la ficha antes de marcar la cita como completada.'
            )
            return redirect('completar_cita', cita_id=cita_id)
        
        # Validar que la ficha esté completa
        if tiene_ficha and not ficha_completa:
            messages.error(
                request,
                '⚠️ La ficha odontológica existe pero está incompleta. '
                'Por favor, asegúrate de que la ficha tenga al menos: nombre del paciente, motivo de consulta y dentista asignado.'
            )
            return redirect('completar_cita', cita_id=cita_id)
        precio_cobrado = request.POST.get('precio_cobrado', '')
        precio_servicio_base = request.POST.get('precio_servicio_base', '')
        
        # Si hay precio_servicio_base (campo oculto), usarlo si no se proporcionó precio_cobrado
        if not precio_cobrado and precio_servicio_base:
            try:
                precio = float(precio_servicio_base)
                if precio >= 0:
                    cita.precio_cobrado = precio
            except ValueError:
                pass
        
        # Si se proporcionó precio_cobrado, usarlo (puede ser un ajuste)
        if precio_cobrado:
            try:
                precio = float(precio_cobrado)
                if precio >= 0:
                    cita.precio_cobrado = precio
                else:
                    messages.error(request, 'El precio debe ser un valor positivo.')
                    return redirect('completar_cita', cita_id=cita_id)
            except ValueError:
                # Si no es válido, usar el precio del servicio
                if cita.tipo_servicio and cita.tipo_servicio.precio_base:
                    cita.precio_cobrado = cita.tipo_servicio.precio_base
                else:
                    messages.error(request, 'El precio ingresado no es válido. Por favor, ingrese un número válido.')
                    return redirect('completar_cita', cita_id=cita_id)
        
        # VALIDACIÓN CRÍTICA 3: Validar que el precio esté asignado
        # Si aún no hay precio pero hay tipo_servicio, usar el precio_base
        if not cita.precio_cobrado and cita.tipo_servicio and cita.tipo_servicio.precio_base:
            cita.precio_cobrado = cita.tipo_servicio.precio_base
        
        # Validación final: el precio es obligatorio
        if not cita.precio_cobrado or cita.precio_cobrado <= 0:
            messages.error(
                request, 
                '⚠️ Debe asignar un precio válido (mayor a 0) para completar la cita. '
                'Si el servicio no tiene precio base, debe ingresarlo manualmente.'
            )
            return redirect('completar_cita', cita_id=cita_id)
        
        # Obtener y guardar los nuevos campos
        metodo_pago = request.POST.get('metodo_pago', '')
        if metodo_pago in ['efectivo', 'transferencia', 'tarjeta']:
            cita.metodo_pago = metodo_pago
        
        motivo_ajuste = request.POST.get('motivo_ajuste_precio', '').strip()
        if motivo_ajuste:
            cita.motivo_ajuste_precio = motivo_ajuste
        
        notas_finalizacion = request.POST.get('notas_finalizacion', '').strip()
        if notas_finalizacion:
            cita.notas_finalizacion = notas_finalizacion
        
        # Guardar información de quién completó y cuándo
        cita.completada_por = perfil
        cita.fecha_completada = timezone.now()
        
        # Todas las validaciones pasaron, marcar como completada
        if cita.completar():
            # Guardar los campos adicionales
            cita.save()
            
            mensaje = f'✅ Cita del {cita.fecha_hora.strftime("%d/%m/%Y a las %H:%M")} marcada como completada exitosamente.'
            if cita.precio_cobrado:
                mensaje += f' Precio a cobrar: ${cita.precio_cobrado:,.0f}'
            if cita.metodo_pago:
                mensaje += f' | Método de pago: {cita.get_metodo_pago_display()}'
            messages.success(request, mensaje)
            return redirect('panel_trabajador')
        else:
            messages.error(request, 'No se pudo completar la cita. Intenta nuevamente.')
            return redirect('panel_trabajador')
    
    # Si es GET, mostrar formulario de confirmación con precio
    # Calcular validaciones para mostrar en el template
    validaciones = {
        'fecha_valida': cita.fecha_hora <= ahora,
        'tiene_ficha': tiene_ficha,
        'ficha_completa': ficha_completa,
        'tiene_precio': False,  # Se calculará después
        'puede_completar': False,  # Se calculará después
    }
    
    # Verificar si tiene precio (actual o del servicio)
    tiene_precio_actual = cita.precio_cobrado and cita.precio_cobrado > 0
    tiene_precio_servicio = (cita.tipo_servicio and 
                            cita.tipo_servicio.precio_base and 
                            cita.tipo_servicio.precio_base > 0)
    validaciones['tiene_precio'] = tiene_precio_actual or tiene_precio_servicio
    
    # Determinar si puede completar (todas las validaciones deben pasar)
    validaciones['puede_completar'] = (
        validaciones['fecha_valida'] and
        validaciones['tiene_ficha'] and
        validaciones['ficha_completa'] and
        validaciones['tiene_precio']
    )
    
    # Asegurar que el tipo_servicio se cargue correctamente
    # SIEMPRE recargar el tipo_servicio directamente desde la base de datos para obtener el precio actualizado
    from citas.models import TipoServicio
    precio_servicio = None
    tipo_servicio_encontrado = None
    
    # Primero intentar cargar por tipo_servicio_id
    if cita.tipo_servicio_id:
        try:
            tipo_servicio_obj = TipoServicio.objects.get(id=cita.tipo_servicio_id)
            cita.tipo_servicio = tipo_servicio_obj
            tipo_servicio_encontrado = tipo_servicio_obj
            
            if tipo_servicio_obj.precio_base is not None:
                try:
                    precio_servicio = float(tipo_servicio_obj.precio_base)
                except (ValueError, TypeError):
                    precio_servicio = tipo_servicio_obj.precio_base
        except TipoServicio.DoesNotExist:
            cita.tipo_servicio = None
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error al cargar tipo_servicio {cita.tipo_servicio_id}: {e}")
    
    # Si no se encontró por ID pero hay tipo_consulta, intentar buscar por nombre
    if not tipo_servicio_encontrado and cita.tipo_consulta:
        try:
            # Buscar servicio por nombre (case-insensitive, sin espacios)
            nombre_buscar = cita.tipo_consulta.strip()
            tipo_servicio_obj = TipoServicio.objects.filter(
                nombre__iexact=nombre_buscar,
                activo=True
            ).first()
            
            if tipo_servicio_obj:
                cita.tipo_servicio = tipo_servicio_obj
                tipo_servicio_encontrado = tipo_servicio_obj
                
                if tipo_servicio_obj.precio_base is not None:
                    try:
                        precio_servicio = float(tipo_servicio_obj.precio_base)
                    except (ValueError, TypeError):
                        precio_servicio = tipo_servicio_obj.precio_base
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Error al buscar servicio por nombre '{cita.tipo_consulta}': {e}")
    
    # Si aún no se encontró, intentar búsqueda parcial
    if not tipo_servicio_encontrado and cita.tipo_consulta:
        try:
            nombre_buscar = cita.tipo_consulta.strip()
            # Buscar servicios que contengan el nombre (case-insensitive)
            tipo_servicio_obj = TipoServicio.objects.filter(
                nombre__icontains=nombre_buscar,
                activo=True
            ).first()
            
            if tipo_servicio_obj:
                cita.tipo_servicio = tipo_servicio_obj
                tipo_servicio_encontrado = tipo_servicio_obj
                
                if tipo_servicio_obj.precio_base is not None:
                    try:
                        precio_servicio = float(tipo_servicio_obj.precio_base)
                    except (ValueError, TypeError):
                        precio_servicio = tipo_servicio_obj.precio_base
        except Exception:
            pass
    
    context = {
        'perfil': perfil,
        'cita': cita,
        'tiene_ficha': tiene_ficha,
        'ficha_completa': ficha_completa,
        'odontograma': odontograma,
        'precio_servicio': precio_servicio,  # Pasar el precio directamente al template
        'validaciones': validaciones,  # Pasar validaciones al template
        'ahora': ahora,  # Para mostrar en el template si es necesario
    }
    return render(request, 'citas/citas/completar_cita.html', context)


# Ajustar precio de cita completada (solo administrativos)
# Solo permite ajustar precio y notas, nada más
@login_required
def ajustar_precio_cita(request, cita_id):
    """
    Vista para ajustar precio y notas de una cita completada
    SOLO DISPONIBLE PARA ADMINISTRATIVOS (recepcionistas)
    Solo permite modificar precio y notas, nada más
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para ajustar precios de citas. Solo el personal administrativo puede realizar esta acción.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'Perfil no encontrado.')
        return redirect('login')
    
    cita = get_object_or_404(Cita, id=cita_id)
    
    # Restricción: Solo citas completadas pueden ajustar precio
    if cita.estado != 'completada':
        messages.error(request, 'Solo se puede ajustar el precio de citas completadas. Para editar otros campos, use "Editar Cita".')
        return redirect('panel_trabajador')
    
    if request.method == 'POST':
        precio_cobrado = request.POST.get('precio_cobrado', '')
        notas = request.POST.get('notas', '')
        
        # Actualizar precio si se proporciona
        if precio_cobrado:
            try:
                precio = float(precio_cobrado)
                if precio >= 0:
                    cita.precio_cobrado = precio
                else:
                    messages.error(request, 'El precio debe ser un valor positivo.')
            except ValueError:
                messages.error(request, 'El precio ingresado no es válido.')
        else:
            # Si se envía vacío, mantener el precio actual o usar el del servicio
            if not cita.precio_cobrado and cita.tipo_servicio:
                cita.precio_cobrado = cita.tipo_servicio.precio_base
        
        # Actualizar notas
        cita.notas = notas
        cita.save()
        
        mensaje = f'✅ Precio y notas de la cita del {cita.fecha_hora.strftime("%d/%m/%Y a las %H:%M")} actualizados correctamente.'
        if cita.precio_cobrado:
            mensaje += f' Precio a cobrar: ${cita.precio_cobrado:,.0f}'
        messages.success(request, mensaje)
        return redirect('panel_trabajador')
    
    # Si es GET, mostrar formulario de ajuste
    # Asegurar que el tipo_servicio se cargue correctamente
    if cita.tipo_servicio_id:
        try:
            from citas.models import TipoServicio
            tipo_servicio_obj = TipoServicio.objects.get(id=cita.tipo_servicio_id)
            cita.tipo_servicio = tipo_servicio_obj
        except TipoServicio.DoesNotExist:
            pass
    
    context = {
        'perfil': perfil,
        'cita': cita,
    }
    return render(request, 'citas/citas/ajustar_precio_cita.html', context)


# Eliminar cita (solo administrativos)
@login_required
def eliminar_cita(request, cita_id):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': 'No tienes permisos para eliminar citas.'}, status=403)
            messages.error(request, 'No tienes permisos para eliminar citas.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': 'Perfil no encontrado.'}, status=404)
        return redirect('login')

    cita = get_object_or_404(Cita, id=cita_id)
    
    # Solo se pueden eliminar citas disponibles
    if cita.estado != 'disponible':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': f'No se puede eliminar una cita en estado "{cita.get_estado_display()}". Solo se pueden eliminar citas disponibles.'}, status=400)
        messages.error(request, f'No se puede eliminar una cita en estado "{cita.get_estado_display()}".')
        return redirect('panel_trabajador')
    
    fecha_texto = cita.fecha_hora.strftime('%d/%m/%Y a las %H:%M')
    cliente_info = cita.cliente.nombre_completo if cita.cliente else cita.paciente_nombre or "Sin cliente"
    servicio_info = cita.tipo_servicio.nombre if cita.tipo_servicio else cita.tipo_consulta or "Sin servicio"
    
    try:
        cita.delete()
        
        # Registrar en auditoría
        registrar_auditoria(
            usuario=perfil,
            accion='eliminar',
            modulo='citas',
            descripcion=f'Cita eliminada: {cliente_info} - {servicio_info}',
            detalles=f'Fecha original: {fecha_texto}, Estado: {cita.estado}',
            objeto_id=cita_id,
            tipo_objeto='Cita',
            request=request
        )
        
        mensaje = f'Cita del {fecha_texto} eliminada exitosamente.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': True, 'message': mensaje})
        messages.success(request, mensaje)
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'error': f'No se pudo eliminar la cita: {str(e)}'}, status=500)
        messages.error(request, f'No se pudo eliminar la cita: {e}')

    # Volver al panel del trabajador para ver los cambios
    return redirect('panel_trabajador')

# ==========================================
# VISTAS SEPARADAS PARA GESTIÓN DE CITAS CON NAVBAR LATERAL
# ==========================================

@login_required
def citas_dia(request):
    """Vista para citas del día con navbar lateral"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a este panel.')
        return redirect('login')
    
    # Verificar permisos pero no redirigir a panel_trabajador para evitar bucles
    # Si no es administrativo, mostrar funcionalidad limitada
    es_admin = perfil.es_administrativo()
    if not es_admin:
        # Si no es administrativo, mostrar mensaje pero permitir acceso básico
        # para evitar bucles de redirección
        messages.warning(request, 'Acceso limitado: Solo usuarios administrativos pueden gestionar citas.')
    
    # Filtro de búsqueda
    search_query = request.GET.get('search', '').strip()
    
    # Citas del día
    citas_hoy = Cita.objects.filter(
        fecha_hora__date=timezone.now().date()
    ).exclude(estado='cancelada').select_related('tipo_servicio', 'dentista', 'cliente').prefetch_related('odontogramas')
    
    # Aplicar filtro de búsqueda si existe
    if search_query:
        citas_hoy = citas_hoy.filter(
            Q(cliente__nombre_completo__icontains=search_query) |
            Q(cliente__email__icontains=search_query) |
            Q(cliente__telefono__icontains=search_query) |
            Q(paciente_nombre__icontains=search_query) |
            Q(paciente_email__icontains=search_query) |
            Q(tipo_servicio__nombre__icontains=search_query) |
            Q(tipo_consulta__icontains=search_query) |
            Q(dentista__nombre_completo__icontains=search_query) |
            Q(notas__icontains=search_query)
        )
    
    citas_hoy = citas_hoy.order_by('fecha_hora')
    
    # Limpiar referencias inválidas
    from pacientes.models import Cliente
    from django.db import connection
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT c.id 
            FROM citas_cita c
            LEFT JOIN pacientes_cliente p ON c.cliente_id = p.id
            WHERE c.cliente_id IS NOT NULL 
            AND p.id IS NULL
        """)
        citas_con_cliente_invalido = [row[0] for row in cursor.fetchall()]
    
    if citas_con_cliente_invalido:
        Cita.objects.filter(id__in=citas_con_cliente_invalido).update(cliente=None)
    
    # Obtener información de fichas
    from historial_clinico.models import Odontograma
    odontogramas = Odontograma.objects.filter(cita__isnull=False).select_related('cita')
    citas_con_ficha = set(odontogramas.values_list('cita_id', flat=True))
    
    citas_hoy_list = list(citas_hoy)
    for cita in citas_hoy_list:
        cita.tiene_ficha = cita.id in citas_con_ficha
        if cita.tiene_ficha:
            cita.odontograma = odontogramas.filter(cita_id=cita.id).first()
    
    # Obtener citas pasadas que requieren atención (para el sidebar)
    citas_pasadas = Cita.objects.filter(
        fecha_hora__lt=timezone.now()
    ).exclude(
        estado__in=['completada', 'cancelada', 'no_show']
    ).filter(
        estado__in=['disponible', 'reservada', 'confirmada']
    ).select_related('tipo_servicio', 'dentista', 'cliente').order_by('-fecha_hora')[:10]
    
    # Paginación - 6 registros por página
    page = request.GET.get('page', 1)
    paginator = Paginator(citas_hoy_list, 6)
    try:
        citas_pag = paginator.page(page)
    except PageNotAnInteger:
        citas_pag = paginator.page(1)
    except EmptyPage:
        citas_pag = paginator.page(paginator.num_pages)
    
    # Estadísticas
    estadisticas = {
        'citas_hoy': Cita.objects.filter(fecha_hora__date=timezone.now().date()).count(),
        'disponibles': Cita.objects.filter(estado='disponible').count(),
        'realizadas': Cita.objects.filter(estado='completada').count(),
    }
    
    # Obtener datos necesarios
    dentistas = Perfil.objects.filter(rol='dentista', activo=True).select_related('user')
    servicios_activos = TipoServicio.objects.filter(activo=True).order_by('categoria', 'nombre')
    clientes = Cliente.objects.filter(activo=True).order_by('nombre_completo')
    
    context = {
        'perfil': perfil,
        'citas': citas_pag,
        'estadisticas': estadisticas,
        'dentistas': dentistas,
        'servicios_activos': servicios_activos,
        'clientes': clientes,
        'es_admin': es_admin,
        'seccion_activa': 'dia',
        'search_query': search_query,
        'citas_pasadas': citas_pasadas,
    }
    return render(request, 'citas/citas/gestor_citas_base.html', context)

@login_required
def citas_disponibles(request):
    """Vista para citas disponibles con navbar lateral"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a este panel.')
        return redirect('login')
    
    # Verificar permisos pero no redirigir a panel_trabajador para evitar bucles
    es_admin = perfil.es_administrativo()
    if not es_admin:
        messages.warning(request, 'Acceso limitado: Solo usuarios administrativos pueden gestionar citas.')
    
    # Filtro de búsqueda
    search_query = request.GET.get('search', '').strip()
    
    # Citas disponibles
    citas_list = Cita.objects.filter(estado='disponible').select_related('tipo_servicio', 'dentista', 'cliente')
    
    # Aplicar filtro de búsqueda si existe
    if search_query:
        citas_list = citas_list.filter(
            Q(cliente__nombre_completo__icontains=search_query) |
            Q(cliente__email__icontains=search_query) |
            Q(tipo_servicio__nombre__icontains=search_query) |
            Q(tipo_consulta__icontains=search_query) |
            Q(dentista__nombre_completo__icontains=search_query) |
            Q(fecha_hora__date__icontains=search_query)
        )
    
    citas_list = citas_list.order_by('fecha_hora')
    
    # Obtener citas pasadas que requieren atención (para el sidebar)
    citas_pasadas = Cita.objects.filter(
        fecha_hora__lt=timezone.now()
    ).exclude(
        estado__in=['completada', 'cancelada', 'no_show']
    ).filter(
        estado__in=['disponible', 'reservada', 'confirmada']
    ).select_related('tipo_servicio', 'dentista', 'cliente').order_by('-fecha_hora')[:10]
    
    # Paginación - 6 registros por página
    page = request.GET.get('page', 1)
    paginator = Paginator(citas_list, 6)
    try:
        citas_pag = paginator.page(page)
    except PageNotAnInteger:
        citas_pag = paginator.page(1)
    except EmptyPage:
        citas_pag = paginator.page(paginator.num_pages)
    
    # Estadísticas
    estadisticas = {
        'citas_hoy': Cita.objects.filter(fecha_hora__date=timezone.now().date()).count(),
        'disponibles': Cita.objects.filter(estado='disponible').count(),
        'realizadas': Cita.objects.filter(estado='completada').count(),
    }
    
    # Obtener datos necesarios
    dentistas = Perfil.objects.filter(rol='dentista', activo=True).select_related('user')
    servicios_activos = TipoServicio.objects.filter(activo=True).order_by('categoria', 'nombre')
    from pacientes.models import Cliente
    clientes = Cliente.objects.filter(activo=True).order_by('nombre_completo')
    
    context = {
        'perfil': perfil,
        'citas': citas_pag,
        'estadisticas': estadisticas,
        'dentistas': dentistas,
        'servicios_activos': servicios_activos,
        'clientes': clientes,
        'es_admin': es_admin,
        'seccion_activa': 'disponibles',
        'search_query': search_query,
        'citas_pasadas': citas_pasadas,
    }
    return render(request, 'citas/citas/gestor_citas_base.html', context)

# Listado de citas tomadas (reservadas o confirmadas) - solo administrativos
@login_required
def citas_tomadas(request):
    try:
        perfil = Perfil.objects.get(user=request.user)
        # Verificar permisos pero no redirigir a panel_trabajador para evitar bucles
        es_admin = perfil.es_administrativo()
        if not es_admin:
            messages.warning(request, 'Acceso limitado: Solo usuarios administrativos pueden gestionar citas.')
    except Perfil.DoesNotExist:
        return redirect('login')

    # Filtro de búsqueda
    search_query = request.GET.get('search', '').strip()
    
    citas_list = Cita.objects.filter(estado__in=['reservada', 'confirmada']).select_related('tipo_servicio', 'dentista', 'cliente')
    
    # Aplicar filtro de búsqueda si existe
    if search_query:
        citas_list = citas_list.filter(
            Q(cliente__nombre_completo__icontains=search_query) |
            Q(cliente__email__icontains=search_query) |
            Q(cliente__telefono__icontains=search_query) |
            Q(paciente_nombre__icontains=search_query) |
            Q(paciente_email__icontains=search_query) |
            Q(tipo_servicio__nombre__icontains=search_query) |
            Q(tipo_consulta__icontains=search_query) |
            Q(dentista__nombre_completo__icontains=search_query) |
            Q(fecha_hora__date__icontains=search_query) |
            Q(notas__icontains=search_query)
        )
    
    citas_list = citas_list.order_by('fecha_hora')
    
    # Obtener información de fichas
    from historial_clinico.models import Odontograma
    odontogramas = Odontograma.objects.filter(cita__isnull=False).select_related('cita')
    citas_con_ficha = set(odontogramas.values_list('cita_id', flat=True))
    
    citas_list_processed = list(citas_list)
    for cita in citas_list_processed:
        cita.tiene_ficha = cita.id in citas_con_ficha
        if cita.tiene_ficha:
            cita.odontograma = odontogramas.filter(cita_id=cita.id).first()
    
    # Paginación - 6 registros por página
    paginator = Paginator(citas_list_processed, 6)
    page = request.GET.get('page', 1)
    
    try:
        citas = paginator.page(page)
    except PageNotAnInteger:
        citas = paginator.page(1)
    except EmptyPage:
        citas = paginator.page(paginator.num_pages)

    # Obtener citas pasadas que requieren atención (para el sidebar)
    citas_pasadas = Cita.objects.filter(
        fecha_hora__lt=timezone.now()
    ).exclude(
        estado__in=['completada', 'cancelada', 'no_show']
    ).filter(
        estado__in=['disponible', 'reservada', 'confirmada']
    ).select_related('tipo_servicio', 'dentista', 'cliente').order_by('-fecha_hora')[:10]
    
    # Estadísticas
    estadisticas = {
        'citas_hoy': Cita.objects.filter(fecha_hora__date=timezone.now().date()).count(),
        'disponibles': Cita.objects.filter(estado='disponible').count(),
        'realizadas': Cita.objects.filter(estado='completada').count(),
    }
    
    # Obtener datos necesarios
    dentistas = Perfil.objects.filter(rol='dentista', activo=True).select_related('user')
    servicios_activos = TipoServicio.objects.filter(activo=True).order_by('categoria', 'nombre')
    from pacientes.models import Cliente
    clientes = Cliente.objects.filter(activo=True).order_by('nombre_completo')
    
    context = {
        'perfil': perfil,
        'citas': citas,
        'estadisticas': estadisticas,
        'dentistas': dentistas,
        'servicios_activos': servicios_activos,
        'clientes': clientes,
        'es_admin': es_admin,
        'seccion_activa': 'tomadas',
        'search_query': search_query,
        'citas_pasadas': citas_pasadas,
    }
    return render(request, 'citas/citas/gestor_citas_base.html', context)

# Listado de citas completadas - solo administrativos
@login_required
def citas_completadas(request):
    try:
        perfil = Perfil.objects.get(user=request.user)
        # Verificar permisos pero no redirigir a panel_trabajador para evitar bucles
        es_admin = perfil.es_administrativo()
        if not es_admin:
            messages.warning(request, 'Acceso limitado: Solo usuarios administrativos pueden gestionar citas.')
    except Perfil.DoesNotExist:
        return redirect('login')

    # Filtro de búsqueda
    search_query = request.GET.get('search', '').strip()
    
    # Citas completadas incluyen 'completada', 'no_show' y 'cancelada' (todas son parte del historial)
    citas_list = Cita.objects.filter(estado__in=['completada', 'no_show', 'cancelada']).select_related('tipo_servicio', 'dentista', 'cliente')
    
    # Aplicar filtro de búsqueda si existe
    if search_query:
        citas_list = citas_list.filter(
            Q(cliente__nombre_completo__icontains=search_query) |
            Q(cliente__email__icontains=search_query) |
            Q(cliente__telefono__icontains=search_query) |
            Q(paciente_nombre__icontains=search_query) |
            Q(paciente_email__icontains=search_query) |
            Q(tipo_servicio__nombre__icontains=search_query) |
            Q(tipo_consulta__icontains=search_query) |
            Q(dentista__nombre_completo__icontains=search_query) |
            Q(fecha_hora__date__icontains=search_query) |
            Q(notas__icontains=search_query)
        )
    
    citas_list = citas_list.order_by('-fecha_hora')
    
    # Obtener información de fichas
    from historial_clinico.models import Odontograma
    odontogramas = Odontograma.objects.filter(cita__isnull=False).select_related('cita')
    citas_con_ficha = set(odontogramas.values_list('cita_id', flat=True))
    
    citas_list_processed = list(citas_list)
    for cita in citas_list_processed:
        cita.tiene_ficha = cita.id in citas_con_ficha
        if cita.tiene_ficha:
            cita.odontograma = odontogramas.filter(cita_id=cita.id).first()
    
    # Paginación - 6 registros por página
    paginator = Paginator(citas_list_processed, 6)
    page = request.GET.get('page', 1)
    
    try:
        citas = paginator.page(page)
    except PageNotAnInteger:
        citas = paginator.page(1)
    except EmptyPage:
        citas = paginator.page(paginator.num_pages)

    # Estadísticas
    estadisticas = {
        'citas_hoy': Cita.objects.filter(fecha_hora__date=timezone.now().date()).count(),
        'disponibles': Cita.objects.filter(estado='disponible').count(),
        'realizadas': Cita.objects.filter(estado='completada').count(),
    }
    
    # Obtener datos necesarios
    dentistas = Perfil.objects.filter(rol='dentista', activo=True).select_related('user')
    servicios_activos = TipoServicio.objects.filter(activo=True).order_by('categoria', 'nombre')
    from pacientes.models import Cliente
    clientes = Cliente.objects.filter(activo=True).order_by('nombre_completo')

    context = {
        'perfil': perfil,
        'citas': citas,
        'estadisticas': estadisticas,
        'dentistas': dentistas,
        'servicios_activos': servicios_activos,
        'clientes': clientes,
        'es_admin': es_admin,
        'seccion_activa': 'completadas',
        'search_query': search_query,
    }
    return render(request, 'citas/citas/gestor_citas_base.html', context)

@login_required
def calendario_citas(request):
    """Vista para calendario general con navbar lateral"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a este panel.')
        return redirect('login')
    
    if not perfil.es_administrativo():
        messages.error(request, 'No tienes permisos para ver este listado.')
        return redirect('panel_trabajador')
    
    # Filtro de búsqueda (opcional para el calendario)
    search_query = request.GET.get('search', '').strip()
    
    # Obtener todas las citas para el calendario (con filtro si existe)
    citas_calendario = Cita.objects.all().select_related('tipo_servicio', 'dentista', 'cliente').order_by('fecha_hora')
    
    # Aplicar filtro de búsqueda si existe
    if search_query:
        citas_calendario = citas_calendario.filter(
            Q(cliente__nombre_completo__icontains=search_query) |
            Q(cliente__email__icontains=search_query) |
            Q(cliente__telefono__icontains=search_query) |
            Q(paciente_nombre__icontains=search_query) |
            Q(paciente_email__icontains=search_query) |
            Q(tipo_servicio__nombre__icontains=search_query) |
            Q(tipo_consulta__icontains=search_query) |
            Q(dentista__nombre_completo__icontains=search_query) |
            Q(notas__icontains=search_query)
        )
    
    # Estadísticas
    estadisticas = {
        'citas_hoy': Cita.objects.filter(fecha_hora__date=timezone.now().date()).count(),
        'disponibles': Cita.objects.filter(estado='disponible').count(),
        'realizadas': Cita.objects.filter(estado='completada').count(),
    }
    
    # Obtener datos necesarios
    dentistas = Perfil.objects.filter(rol='dentista', activo=True).select_related('user')
    servicios_activos = TipoServicio.objects.filter(activo=True).order_by('categoria', 'nombre')
    from pacientes.models import Cliente
    clientes = Cliente.objects.filter(activo=True).order_by('nombre_completo')
    
    context = {
        'perfil': perfil,
        'estadisticas': estadisticas,
        'dentistas': dentistas,
        'servicios_activos': servicios_activos,
        'clientes': clientes,
        'es_admin': True,
        'seccion_activa': 'calendario',
        'search_query': search_query,
        'citas_calendario': citas_calendario,
    }
    return render(request, 'citas/citas/gestor_citas_base.html', context)

# Registrar nuevo trabajador con protección de seguridad
@never_cache
@csrf_protect
def registro_trabajador(request):
    # Rate limiting para registro (máximo 3 registros por IP en 1 hora)
    ip_address = request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0] if request.META.get('HTTP_X_FORWARDED_FOR') else request.META.get('REMOTE_ADDR', '')
    cache_key_reg = f'register_attempts_{ip_address}'
    reg_attempts = cache.get(cache_key_reg, 0)
    
    if reg_attempts >= 3:
        messages.error(request, '⚠️ Demasiados intentos de registro. Por favor, espera 1 hora antes de intentar nuevamente.')
        logger.warning(f'Registro bloqueado por rate limiting - IP: {ip_address}')
        form = RegistroTrabajadorForm()
        return render(request, 'citas/auth/registro_trabajador.html', {'form': form})
    
    if request.method == 'POST':
        form = RegistroTrabajadorForm(request.POST)
        
        if form.is_valid():
            # Formulario válido, intentar guardar
            try:
                user = form.save()
                
                # Limpiar contador de intentos
                cache.delete(cache_key_reg)
                
                # Log de registro exitoso
                logger.info(f'Registro exitoso - Usuario: {user.username}, IP: {ip_address}')
                
                messages.success(request, '✅ Trabajador registrado correctamente. Ya puedes iniciar sesión.')
                return redirect('login')
            except Exception as e:
                # Incrementar contador de intentos fallidos
                cache.set(cache_key_reg, reg_attempts + 1, 3600)  # 1 hora
                
                logger.error(f'Error al registrar trabajador - IP: {ip_address}, Error: {str(e)}')
                
                # Mensajes de error genéricos (no revelar detalles técnicos)
                if 'username' in str(e).lower() and 'already exists' in str(e).lower():
                    messages.error(request, '❌ Este usuario ya existe. Por favor, elige otro nombre de usuario.')
                elif 'email' in str(e).lower() and 'already exists' in str(e).lower():
                    messages.error(request, '❌ Este email ya está registrado. Por favor, usa otro email.')
                else:
                    messages.error(request, '❌ Error al registrar trabajador. Por favor, verifica todos los campos e intenta nuevamente.')
        else:
            # Incrementar contador de intentos fallidos solo si hay errores de validación
            if form.errors:
                cache.set(cache_key_reg, reg_attempts + 1, 3600)  # 1 hora
                logger.warning(f'Intento de registro con errores - IP: {ip_address}')
            
            # Los errores del formulario se mostrarán automáticamente en el template
            # No necesitamos agregar validaciones manuales aquí, el formulario ya las tiene
    else:
        form = RegistroTrabajadorForm()
    
    return render(request, 'citas/auth/registro_trabajador.html', {'form': form})

# Editar perfil
@login_required
def editar_perfil(request):
    try:
        perfil = Perfil.objects.get(user=request.user)
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes un perfil de trabajador válido.')
        return redirect('login')
    
    if request.method == 'POST':
        form = PerfilForm(request.POST, instance=perfil)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado correctamente.')
            return redirect('panel_trabajador')
    else:
        form = PerfilForm(instance=perfil)
    
    return render(request, 'citas/perfil/editar_perfil.html', {'form': form, 'perfil': perfil})

# Dashboard con estadísticas
@login_required
def dashboard(request):
    """Vista principal del Dashboard - Página de inicio del sistema"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    # Fechas de referencia
    ahora = timezone.now()
    hoy = ahora.date()
    inicio_mes = hoy.replace(day=1)
    fin_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    semana_actual = hoy - timedelta(days=7)
    mes_actual = hoy - timedelta(days=30)
    
    if perfil.es_administrativo():
        # ===== ESTADÍSTICAS GENERALES =====
        total_citas = Cita.objects.count()
        total_clientes = Cliente.objects.filter(activo=True).count()
        total_personal = Perfil.objects.filter(activo=True).count()
        total_insumos = Insumo.objects.count()
        
        # Citas
        citas_hoy = Cita.objects.filter(fecha_hora__date=hoy).count()
        citas_semana = Cita.objects.filter(fecha_hora__date__gte=semana_actual).count()
        citas_mes = Cita.objects.filter(fecha_hora__date__gte=inicio_mes, fecha_hora__date__lte=fin_mes).count()
        citas_disponibles = Cita.objects.filter(estado='disponible').count()
        citas_reservadas = Cita.objects.filter(estado='reservada').count()
        citas_confirmadas = Cita.objects.filter(estado='confirmada').count()
        citas_completadas = Cita.objects.filter(estado='completada').count()
        
        # Citas próximas (próximas 7 días)
        proximos_7_dias = hoy + timedelta(days=7)
        citas_proximas = Cita.objects.filter(
            fecha_hora__date__gte=hoy,
            fecha_hora__date__lte=proximos_7_dias,
            estado__in=['reservada', 'confirmada']
        ).select_related('dentista', 'cliente').order_by('fecha_hora')[:10]
        
        # Citas de hoy con detalles
        citas_hoy_detalle = Cita.objects.filter(
            fecha_hora__date=hoy
        ).select_related('dentista', 'cliente').order_by('fecha_hora')
        
        # Finanzas
        total_ingresos = IngresoManual.objects.aggregate(Sum('monto'))['monto__sum'] or 0
        total_egresos = EgresoManual.objects.aggregate(Sum('monto'))['monto__sum'] or 0
        ingresos_mes = IngresoManual.objects.filter(
            fecha__gte=inicio_mes,
            fecha__lte=fin_mes
        ).aggregate(Sum('monto'))['monto__sum'] or 0
        egresos_mes = EgresoManual.objects.filter(
            fecha__gte=inicio_mes,
            fecha__lte=fin_mes
        ).aggregate(Sum('monto'))['monto__sum'] or 0
        balance_total = total_ingresos - total_egresos
        balance_mes = ingresos_mes - egresos_mes
        
        # Ingresos de citas completadas
        ingresos_citas_mes = Cita.objects.filter(
            estado='completada',
            fecha_hora__date__gte=inicio_mes,
            fecha_hora__date__lte=fin_mes,
            precio_cobrado__isnull=False
        ).aggregate(Sum('precio_cobrado'))['precio_cobrado__sum'] or 0
        
        # Insumos con stock bajo
        insumos_bajo_stock = Insumo.objects.filter(
            cantidad_actual__lte=F('cantidad_minima')
        ).count()
        
        # Clientes nuevos este mes
        clientes_nuevos_mes = Cliente.objects.filter(
            fecha_registro__date__gte=inicio_mes,
            fecha_registro__date__lte=fin_mes
        ).count()
        
        # Citas por estado (últimos 7 días)
        citas_por_estado = Cita.objects.filter(
            fecha_hora__date__gte=semana_actual
        ).values('estado').annotate(total=Count('estado'))
        
        estadisticas = {
            'total_citas': total_citas,
            'total_clientes': total_clientes,
            'total_personal': total_personal,
            'total_insumos': total_insumos,
            'citas_hoy': citas_hoy,
            'citas_semana': citas_semana,
            'citas_mes': citas_mes,
            'citas_disponibles': citas_disponibles,
            'citas_reservadas': citas_reservadas,
            'citas_confirmadas': citas_confirmadas,
            'citas_completadas': citas_completadas,
            'insumos_bajo_stock': insumos_bajo_stock,
            'clientes_nuevos_mes': clientes_nuevos_mes,
        }
        
    else:
        # Estadísticas para dentistas
        citas_hoy = Cita.objects.filter(fecha_hora__date=hoy, dentista=perfil).count()
        citas_semana = Cita.objects.filter(fecha_hora__date__gte=semana_actual, dentista=perfil).count()
        citas_mes = Cita.objects.filter(fecha_hora__date__gte=inicio_mes, fecha_hora__date__lte=fin_mes, dentista=perfil).count()
        
        # Citas próximas del dentista
        proximos_7_dias = hoy + timedelta(days=7)
        citas_proximas = Cita.objects.filter(
            dentista=perfil,
            fecha_hora__date__gte=hoy,
            fecha_hora__date__lte=proximos_7_dias,
            estado__in=['reservada', 'confirmada']
        ).select_related('cliente').order_by('fecha_hora')[:10]
        
        # Citas de hoy del dentista
        citas_hoy_detalle = Cita.objects.filter(
            fecha_hora__date=hoy,
            dentista=perfil
        ).select_related('cliente').order_by('fecha_hora')
        
        estadisticas = {
            'citas_hoy': citas_hoy,
            'citas_semana': citas_semana,
            'citas_mes': citas_mes,
            'citas_pendientes': Cita.objects.filter(dentista=perfil, estado='reservada').count(),
            'citas_confirmadas': Cita.objects.filter(dentista=perfil, estado='confirmada').count(),
            'citas_completadas': Cita.objects.filter(dentista=perfil, estado='completada').count(),
        }
        
        # Citas por estado para el dentista (últimos 7 días)
        citas_por_estado = Cita.objects.filter(
            fecha_hora__date__gte=semana_actual,
            dentista=perfil
        ).values('estado').annotate(total=Count('estado'))
        
        # Valores por defecto para dentistas
        total_ingresos = 0
        total_egresos = 0
        ingresos_mes = 0
        egresos_mes = 0
        balance_total = 0
        balance_mes = 0
        ingresos_citas_mes = 0
        insumos_bajo_stock = 0
        clientes_nuevos_mes = 0
    
    context = {
        'perfil': perfil,
        'estadisticas': estadisticas,
        'citas_por_estado': citas_por_estado,
        'citas_proximas': citas_proximas,
        'citas_hoy_detalle': citas_hoy_detalle,
        'es_admin': perfil.es_administrativo(),
        # Finanzas (solo para admin)
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'ingresos_mes': ingresos_mes,
        'egresos_mes': egresos_mes,
        'balance_total': balance_total,
        'balance_mes': balance_mes,
        'ingresos_citas_mes': ingresos_citas_mes,
        'insumos_bajo_stock': insumos_bajo_stock,
        'clientes_nuevos_mes': clientes_nuevos_mes,
        'hoy': hoy,
    }
    
    return render(request, 'citas/dashboard/dashboard.html', context)

# Dashboard específico para dentistas
@login_required
def dashboard_dentista(request):
    """Vista de inicio específica para dentistas con accesos rápidos"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
        
        # Solo permitir acceso a dentistas
        if not perfil.es_dentista():
            messages.error(request, 'No tienes permisos para acceder a esta página.')
            return redirect('dashboard')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    # Fechas de referencia
    ahora = timezone.now()
    hoy = ahora.date()
    inicio_mes = hoy.replace(day=1)
    fin_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    semana_actual = hoy - timedelta(days=7)
    proximos_7_dias = hoy + timedelta(days=7)
    
    # Estadísticas del dentista
    citas_hoy = Cita.objects.filter(fecha_hora__date=hoy, dentista=perfil).count()
    citas_semana = Cita.objects.filter(
        fecha_hora__date__gte=semana_actual,
        dentista=perfil
    ).count()
    citas_mes = Cita.objects.filter(
        fecha_hora__date__gte=inicio_mes,
        fecha_hora__date__lte=fin_mes,
        dentista=perfil
    ).count()
    
    citas_pendientes = Cita.objects.filter(dentista=perfil, estado='reservada').count()
    citas_confirmadas = Cita.objects.filter(dentista=perfil, estado='confirmada').count()
    citas_completadas = Cita.objects.filter(dentista=perfil, estado='completada').count()
    
    # Citas de hoy con detalles
    citas_hoy_detalle = Cita.objects.filter(
        fecha_hora__date=hoy,
        dentista=perfil
    ).select_related('cliente').order_by('fecha_hora')[:10]
    
    # Citas próximas (próximas 7 días)
    citas_proximas = Cita.objects.filter(
        dentista=perfil,
        fecha_hora__date__gte=hoy,
        fecha_hora__date__lte=proximos_7_dias,
        estado__in=['reservada', 'confirmada']
    ).select_related('cliente').order_by('fecha_hora')[:10]
    
    # Contar pacientes únicos del dentista
    pacientes_count = Cita.objects.filter(dentista=perfil).values('cliente').distinct().count()
    
    # Planes de tratamiento activos
    planes_activos = PlanTratamiento.objects.filter(
        dentista=perfil,
        estado='en_proceso'
    ).count()
    
    estadisticas = {
        'citas_hoy': citas_hoy,
        'citas_semana': citas_semana,
        'citas_mes': citas_mes,
        'citas_pendientes': citas_pendientes,
        'citas_confirmadas': citas_confirmadas,
        'citas_completadas': citas_completadas,
        'pacientes_count': pacientes_count,
        'planes_activos': planes_activos,
    }
    
    context = {
        'perfil': perfil,
        'estadisticas': estadisticas,
        'citas_proximas': citas_proximas,
        'citas_hoy_detalle': citas_hoy_detalle,
        'hoy': hoy,
    }
    
    return render(request, 'citas/dashboard/dashboard_dentista.html', context)

# Vista de diagnóstico - Todas las citas (solo administrativos)
@login_required
def todas_las_citas(request):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para ver todas las citas.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    # Obtener todas las citas sin filtros, incluyendo información de odontogramas vinculados
    todas_citas_list = Cita.objects.all().select_related('dentista', 'cliente').prefetch_related('odontogramas').order_by('fecha_hora')
    
    # Paginación - 10 registros por página
    paginator = Paginator(todas_citas_list, 10)
    page = request.GET.get('page', 1)
    
    try:
        todas_citas = paginator.page(page)
    except PageNotAnInteger:
        todas_citas = paginator.page(1)
    except EmptyPage:
        todas_citas = paginator.page(paginator.num_pages)
    
    # Agrupar por estado para diagnóstico (usar la lista completa para estadísticas)
    citas_por_estado = {}
    for cita in todas_citas_list:
        estado = cita.estado
        if estado not in citas_por_estado:
            citas_por_estado[estado] = []
        citas_por_estado[estado].append(cita)
    
    context = {
        'perfil': perfil,
        'todas_citas': todas_citas,
        'citas_por_estado': citas_por_estado,
        'total_citas': todas_citas_list.count(),
        'estados_disponibles': Cita.ESTADO_CHOICES
    }
    
    return render(request, 'citas/citas/todas_las_citas.html', context)

# Gestor de Clientes - solo administrativos
@login_required
def gestor_clientes(request):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para gestionar clientes.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        return redirect('login')

    # Filtros de búsqueda
    search = request.GET.get('search', '')
    estado = request.GET.get('estado', '')
    
    # Obtener clientes desde el modelo Cliente (sistema de gestión)
    clientes_query = Cliente.objects.all()
    
    # Aplicar filtros de búsqueda
    if search:
        from django.db.models import Q
        clientes_query = clientes_query.filter(
            Q(nombre_completo__icontains=search) |
            Q(email__icontains=search) |
            Q(telefono__icontains=search) |
            Q(rut__icontains=search)  # Búsqueda por RUT
        )
    
    # Aplicar filtro de estado
    if estado == 'activo':
        clientes_query = clientes_query.filter(activo=True)
    elif estado == 'inactivo':
        clientes_query = clientes_query.filter(activo=False)
    
    # Anotar cada cliente con el número de citas, odontogramas y radiografías
    from django.db.models import Count, Max
    try:
        # Intentar anotar con todas las relaciones
        clientes_query = clientes_query.annotate(
            total_citas=Count('citas', distinct=True)
        )
        # Intentar agregar odontogramas y radiografias si las relaciones existen
        try:
            from historial_clinico.models import Odontograma, Radiografia
            clientes_query = clientes_query.annotate(
                total_odontogramas=Count('odontogramas', distinct=True),
                total_radiografias=Count('radiografias', distinct=True),
                ultima_cita_fecha=Max('citas__fecha_hora')
            )
        except:
            # Si falla, solo usar total_citas
            pass
    except Exception as e:
        # Si hay error con las anotaciones, usar consulta sin anotaciones
        pass
    
    # Obtener emails de clientes ya existentes para no duplicar
    emails_existentes = set(clientes_query.values_list('email', flat=True))
    
    # Obtener clientes registrados desde cliente_web que no tienen Cliente en gestion_clinica
    clientes_web = []
    try:
        from django.contrib.auth.models import User
        from cuentas.models import PerfilCliente
        
        # Obtener todos los PerfilCliente
        perfiles_web = PerfilCliente.objects.all()
        
        # Filtrar por búsqueda si existe
        if search:
            from django.db.models import Q
            perfiles_web = perfiles_web.filter(
                Q(nombre_completo__icontains=search) |
                Q(email__icontains=search) |
                Q(telefono__icontains=search) |
                Q(rut__icontains=search)
            )
        
        # Solo incluir los que NO tienen Cliente asociado (por email)
        for perfil in perfiles_web:
            if perfil.email and perfil.email not in emails_existentes:
                # Crear un objeto similar a Cliente para la vista
                cliente_web = type('ClienteWeb', (), {
                    'id': f'web_{perfil.id}',
                    'nombre_completo': perfil.nombre_completo,
                    'email': perfil.email,
                    'telefono': perfil.telefono,
                    'rut': perfil.rut or '',
                    'fecha_nacimiento': perfil.fecha_nacimiento,
                    'alergias': perfil.alergias or '',
                    'fecha_registro': perfil.user.date_joined if perfil.user else None,
                    'activo': perfil.user.is_active if perfil.user else True,
                    'tiene_alergias': perfil.tiene_alergias,
                    'edad': perfil.edad,
                    'total_citas': 0,  # Se puede calcular después si es necesario
                    'total_odontogramas': 0,
                    'total_radiografias': 0,
                    'es_de_web': True,  # Marca para identificar que viene de cliente_web
                    'user_id': perfil.user.id if perfil.user else None,
                    'username': perfil.user.username if perfil.user else None,
                })()
                clientes_web.append(cliente_web)
    except ImportError:
        # Si no se puede importar PerfilCliente, continuar sin errores
        pass
    except Exception as e:
        # Si hay error, registrar pero continuar
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Error al obtener clientes de cliente_web: {e}")
    
    # Convertir QuerySet a lista para poder combinar
    clientes_list = list(clientes_query)
    
    # Agregar clientes de web
    clientes_list.extend(clientes_web)
    
    # Ordenar por nombre
    clientes_list.sort(key=lambda x: x.nombre_completo.lower())
    
    # Aplicar filtro de estado a clientes_web también
    if estado == 'activo':
        clientes_list = [c for c in clientes_list if getattr(c, 'activo', True)]
    elif estado == 'inactivo':
        clientes_list = [c for c in clientes_list if not getattr(c, 'activo', True)]
    
    # Paginación - 10 registros por página
    paginator = Paginator(clientes_list, 10)
    page = request.GET.get('page', 1)
    
    try:
        clientes = paginator.page(page)
    except PageNotAnInteger:
        clientes = paginator.page(1)
    except EmptyPage:
        clientes = paginator.page(paginator.num_pages)
    
    # Estadísticas (incluyendo clientes de web)
    total_clientes = Cliente.objects.count() + len(clientes_web)
    clientes_con_citas_count = Cliente.objects.filter(citas__isnull=False).distinct().count()
    
    # Calcular clientes nuevos (últimos 30 días)
    fecha_limite = timezone.now() - timedelta(days=30)
    clientes_nuevos = Cliente.objects.filter(fecha_registro__gte=fecha_limite).count()
    
    # Calcular clientes con alergias
    clientes_con_alergias = Cliente.objects.exclude(alergias__isnull=True).exclude(alergias='').exclude(alergias__iexact='ninguna').count()
    
    # Datos adicionales para panel derecho
    # Clientes recientes (últimos 5 registrados)
    clientes_recientes = Cliente.objects.order_by('-fecha_registro')[:5]
    
    # Citas próximas de clientes (próximos 7 días)
    fecha_limite = timezone.now() + timedelta(days=7)
    citas_proximas = Cita.objects.filter(
        fecha_hora__gte=timezone.now(),
        fecha_hora__lte=fecha_limite,
        estado__in=['reservada', 'confirmada']
    ).select_related('cliente', 'tipo_servicio', 'dentista').order_by('fecha_hora')[:5]
    
    estadisticas = {
        'total_clientes': total_clientes,
        'clientes_con_citas': clientes_con_citas_count,
        'clientes_gestion': Cliente.objects.count(),
        'clientes_web': len(clientes_web),
        'clientes_nuevos': clientes_nuevos,
        'clientes_con_alergias': clientes_con_alergias,
    }
    
    context = {
        'perfil': perfil,
        'clientes': clientes,
        'estadisticas': estadisticas,
        'search': search,
        'estado': estado,
        'es_admin': True,
        'clientes_recientes': clientes_recientes,
        'citas_proximas': citas_proximas,
    }
    
    return render(request, 'citas/clientes/gestor_clientes.html', context)

# Exportar clientes a Excel
@login_required
def exportar_excel_clientes(request):
    """Exporta la lista de clientes a un archivo Excel con diseño mejorado"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para exportar clientes.')
            return redirect('gestor_clientes')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    # Intentar usar openpyxl para mejor diseño
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Crear el libro de trabajo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        # Estilos
        header_fill = PatternFill(start_color="3b82f6", end_color="2563eb", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        title_font = Font(bold=True, size=16, color="1e293b")
        subtitle_font = Font(size=10, color="64748b")
        
        # Título y información
        ws.merge_cells('A1:I1')
        ws['A1'] = "LISTA DE CLIENTES - CLÍNICA DENTAL"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells('A2:I2')
        ws['A2'] = f"Fecha de exportación: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Fila vacía
        ws.row_dimensions[3].height = 5
        
        # Encabezados
        headers = [
            'Nombre Completo', 'RUT', 'Email', 'Teléfono', 
            'Fecha Nacimiento', 'Edad', 'Alergias', 'Fecha Registro', 
            'Estado', 'Total Citas'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border_style
        
        # Obtener todos los clientes
        clientes = Cliente.objects.all().order_by('nombre_completo')
        
        # Obtener clientes de web
        clientes_web = []
        try:
            from cuentas.models import PerfilCliente
            perfiles_web = PerfilCliente.objects.all()
            emails_existentes = set(Cliente.objects.values_list('email', flat=True))
            
            for perfil in perfiles_web:
                if perfil.email and perfil.email not in emails_existentes:
                    cliente_web = type('ClienteWeb', (), {
                        'nombre_completo': perfil.nombre_completo,
                        'rut': perfil.rut or '',
                        'email': perfil.email,
                        'telefono': perfil.telefono or '',
                        'fecha_nacimiento': perfil.fecha_nacimiento,
                        'alergias': perfil.alergias or '',
                        'fecha_registro': perfil.user.date_joined if perfil.user else None,
                        'activo': perfil.user.is_active if perfil.user else True,
                        'total_citas': 0,
                        'edad': getattr(perfil, 'edad', None),
                    })()
                    clientes_web.append(cliente_web)
        except ImportError:
            pass
        
        # Escribir datos
        row_num = 5
        for cliente in list(clientes) + clientes_web:
            # Calcular edad si hay fecha de nacimiento
            edad = ''
            if hasattr(cliente, 'fecha_nacimiento') and cliente.fecha_nacimiento:
                try:
                    from datetime import date
                    hoy = date.today()
                    fecha_nac = cliente.fecha_nacimiento
                    if isinstance(fecha_nac, datetime):
                        fecha_nac = fecha_nac.date()
                    edad = hoy.year - fecha_nac.year - ((hoy.month, hoy.day) < (fecha_nac.month, fecha_nac.day))
                except:
                    edad = getattr(cliente, 'edad', '')
            
            # Formatear fechas
            fecha_nac = ''
            if hasattr(cliente, 'fecha_nacimiento') and cliente.fecha_nacimiento:
                if isinstance(cliente.fecha_nacimiento, datetime):
                    fecha_nac = cliente.fecha_nacimiento.strftime('%d/%m/%Y')
                else:
                    fecha_nac = cliente.fecha_nacimiento.strftime('%d/%m/%Y')
            
            fecha_reg = ''
            if hasattr(cliente, 'fecha_registro') and cliente.fecha_registro:
                if isinstance(cliente.fecha_registro, datetime):
                    fecha_reg = cliente.fecha_registro.strftime('%d/%m/%Y %H:%M')
                else:
                    fecha_reg = str(cliente.fecha_registro)
            
            # Escribir fila
            data = [
                cliente.nombre_completo or '',
                getattr(cliente, 'rut', '') or '',
                cliente.email or '',
                getattr(cliente, 'telefono', '') or '',
                fecha_nac,
                str(edad) if edad else '',
                getattr(cliente, 'alergias', '') or 'Ninguna',
                fecha_reg,
                'Activo' if getattr(cliente, 'activo', True) else 'Inactivo',
                getattr(cliente, 'total_citas', 0) or 0,
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border_style
                if col_num == 1:  # Nombre completo
                    cell.font = Font(bold=True)
                elif col_num == 9:  # Estado
                    if value == 'Activo':
                        cell.fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
            
            row_num += 1
        
        # Ajustar ancho de columnas
        column_widths = [30, 15, 30, 15, 15, 8, 25, 18, 12, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Congelar paneles (fijar encabezados)
        ws.freeze_panes = 'A5'
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"clientes_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar el libro de trabajo
        wb.save(response)
        
        return response
        
    except ImportError:
        # Fallback a CSV mejorado si openpyxl no está disponible
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"clientes_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Escribir BOM UTF-8
        response.write('\ufeff')
        
        import csv
        writer = csv.writer(response, delimiter=';')  # Usar punto y coma para mejor compatibilidad
        
        # Título
        writer.writerow(['LISTA DE CLIENTES - CLÍNICA DENTAL'])
        writer.writerow([f'Fecha de exportación: {timezone.now().strftime("%d/%m/%Y %H:%M:%S")}'])
        writer.writerow([])  # Fila vacía
        
        # Encabezados
        writer.writerow(['Nombre Completo', 'RUT', 'Email', 'Teléfono', 'Fecha Nacimiento', 'Edad', 'Alergias', 'Fecha Registro', 'Estado', 'Total Citas'])
        
        # Obtener todos los clientes
        clientes = Cliente.objects.all().order_by('nombre_completo')
        
        # Obtener clientes de web
        clientes_web = []
        try:
            from cuentas.models import PerfilCliente
            perfiles_web = PerfilCliente.objects.all()
            emails_existentes = set(Cliente.objects.values_list('email', flat=True))
            
            for perfil in perfiles_web:
                if perfil.email and perfil.email not in emails_existentes:
                    cliente_web = type('ClienteWeb', (), {
                        'nombre_completo': perfil.nombre_completo,
                        'rut': perfil.rut or '',
                        'email': perfil.email,
                        'telefono': perfil.telefono or '',
                        'fecha_nacimiento': perfil.fecha_nacimiento,
                        'alergias': perfil.alergias or '',
                        'fecha_registro': perfil.user.date_joined if perfil.user else None,
                        'activo': perfil.user.is_active if perfil.user else True,
                        'total_citas': 0,
                        'edad': getattr(perfil, 'edad', None),
                    })()
                    clientes_web.append(cliente_web)
        except ImportError:
            pass
        
        # Escribir datos
        for cliente in list(clientes) + clientes_web:
            # Calcular edad
            edad = ''
            if hasattr(cliente, 'fecha_nacimiento') and cliente.fecha_nacimiento:
                try:
                    from datetime import date
                    hoy = date.today()
                    fecha_nac = cliente.fecha_nacimiento
                    if isinstance(fecha_nac, datetime):
                        fecha_nac = fecha_nac.date()
                    edad = hoy.year - fecha_nac.year - ((hoy.month, hoy.day) < (fecha_nac.month, fecha_nac.day))
                except:
                    edad = getattr(cliente, 'edad', '')
            
            # Formatear fechas
            fecha_nac = ''
            if hasattr(cliente, 'fecha_nacimiento') and cliente.fecha_nacimiento:
                if isinstance(cliente.fecha_nacimiento, datetime):
                    fecha_nac = cliente.fecha_nacimiento.strftime('%d/%m/%Y')
                else:
                    fecha_nac = cliente.fecha_nacimiento.strftime('%d/%m/%Y')
            
            fecha_reg = ''
            if hasattr(cliente, 'fecha_registro') and cliente.fecha_registro:
                if isinstance(cliente.fecha_registro, datetime):
                    fecha_reg = cliente.fecha_registro.strftime('%d/%m/%Y %H:%M')
                else:
                    fecha_reg = str(cliente.fecha_registro)
            
            writer.writerow([
                cliente.nombre_completo or '',
                getattr(cliente, 'rut', '') or '',
                cliente.email or '',
                getattr(cliente, 'telefono', '') or '',
                fecha_nac,
                str(edad) if edad else '',
                getattr(cliente, 'alergias', '') or 'Ninguna',
                fecha_reg,
                'Activo' if getattr(cliente, 'activo', True) else 'Inactivo',
                getattr(cliente, 'total_citas', 0) or 0,
            ])
        
        return response

# Gestor de Insumos - solo administrativos
@login_required
def gestor_insumos(request):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para gestionar insumos.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        return redirect('login')

    # Filtros de búsqueda
    search = request.GET.get('search', '')
    categoria = request.GET.get('categoria', '')
    estado = request.GET.get('estado', '')
    
    insumos = Insumo.objects.all()
    
    if search:
        insumos = insumos.filter(
            Q(nombre__icontains=search) |
            Q(descripcion__icontains=search) |
            Q(proveedor_principal__nombre__icontains=search) |
            Q(proveedor_texto__icontains=search) |
            Q(ubicacion__icontains=search)
        )
    
    if categoria:
        insumos = insumos.filter(categoria=categoria)
    
    if estado:
        insumos = insumos.filter(estado=estado)
    
    insumos_list = insumos.order_by('nombre')
    
    # Paginación - 10 registros por página
    paginator = Paginator(insumos_list, 10)
    page = request.GET.get('page', 1)
    
    try:
        insumos = paginator.page(page)
    except PageNotAnInteger:
        insumos = paginator.page(1)
    except EmptyPage:
        insumos = paginator.page(paginator.num_pages)
    
    # Estadísticas de insumos
    total_insumos = Insumo.objects.count()
    insumos_stock_bajo = Insumo.objects.filter(cantidad_actual__lte=F('cantidad_minima')).count()
    insumos_proximo_vencimiento = Insumo.objects.filter(
        fecha_vencimiento__lte=timezone.now().date() + timedelta(days=30),
        fecha_vencimiento__gte=timezone.now().date()
    ).count()
    insumos_agotados = Insumo.objects.filter(estado='agotado').count()
    
    estadisticas = {
        'total_insumos': total_insumos,
        'insumos_stock_bajo': insumos_stock_bajo,
        'insumos_proximo_vencimiento': insumos_proximo_vencimiento,
        'insumos_agotados': insumos_agotados,
    }
    
    # Obtener proveedores activos para el formulario
    proveedores = Proveedor.objects.filter(activo=True).order_by('nombre')
    
    context = {
        'perfil': perfil,
        'insumos': insumos,
        'estadisticas': estadisticas,
        'categorias': Insumo.CATEGORIA_CHOICES,
        'estados': Insumo.ESTADO_CHOICES,
        'proveedores': proveedores,
        'search': search,
        'categoria': categoria,
        'estado': estado,
        'es_admin': True
    }
    
    return render(request, 'citas/insumos/gestor_insumos.html', context)

# Agregar nuevo insumo
@login_required
def agregar_insumo(request):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para agregar insumos.')
            return redirect('gestor_inventario_unificado')
    except Perfil.DoesNotExist:
        return redirect('login')

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        categoria = request.POST.get('categoria', '')
        descripcion = request.POST.get('descripcion', '').strip()
        cantidad_actual_str = request.POST.get('cantidad_actual', '0')
        cantidad_minima_str = request.POST.get('cantidad_minima', '1')
        unidad_medida = request.POST.get('unidad_medida', 'unidad')
        precio_unitario_str = request.POST.get('precio_unitario_raw', '') or request.POST.get('precio_unitario', '')
        # Limpiar puntos del formato chileno
        if precio_unitario_str:
            precio_unitario_str = precio_unitario_str.replace('.', '')
        proveedor_id = request.POST.get('proveedor_principal', '').strip()
        fecha_vencimiento = request.POST.get('fecha_vencimiento', '')
        ubicacion = request.POST.get('ubicacion', '').strip()
        notas = request.POST.get('notas', '').strip()
        
        # Verificar si es petición AJAX
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        # Validaciones
        errores = []
        
        if not nombre:
            error_msg = 'El nombre del insumo es obligatorio. Por favor, ingrese un nombre para el insumo.'
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
            errores.append(error_msg)
        
        if len(nombre) < 3:
            error_msg = 'El nombre del insumo debe tener al menos 3 caracteres.'
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
            errores.append(error_msg)
        
        if not categoria:
            error_msg = 'La categoría es obligatoria. Por favor, seleccione una categoría para el insumo.'
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
            errores.append(error_msg)
        
        try:
            cantidad_actual = int(cantidad_actual_str)
            if cantidad_actual < 0:
                error_msg = 'La cantidad actual no puede ser negativa. Por favor, ingrese un valor mayor o igual a 0.'
                if is_ajax:
                    return JsonResponse({'success': False, 'message': error_msg}, status=400)
                errores.append(error_msg)
        except (ValueError, TypeError):
            error_msg = 'La cantidad actual debe ser un número entero válido.'
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
            errores.append(error_msg)
        
        try:
            cantidad_minima = int(cantidad_minima_str)
            if cantidad_minima < 1:
                error_msg = 'La cantidad mínima debe ser al menos 1. Por favor, ingrese un valor mayor a 0.'
                if is_ajax:
                    return JsonResponse({'success': False, 'message': error_msg}, status=400)
                errores.append(error_msg)
        except (ValueError, TypeError):
            error_msg = 'La cantidad mínima debe ser un número entero válido.'
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
            errores.append(error_msg)
        
        if cantidad_actual < cantidad_minima:
            error_msg = f'La cantidad actual ({cantidad_actual}) no puede ser menor que la cantidad mínima ({cantidad_minima}). Por favor, ajuste los valores.'
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
            errores.append(error_msg)
        
        precio_unitario = None
        if precio_unitario_str:
            try:
                precio_unitario = round(float(precio_unitario_str))  # Redondear a entero para pesos chilenos
                if precio_unitario < 0:
                    error_msg = 'El precio unitario no puede ser negativo. Por favor, ingrese un valor positivo o deje el campo vacío.'
                    if is_ajax:
                        return JsonResponse({'success': False, 'message': error_msg}, status=400)
                    errores.append(error_msg)
            except (ValueError, TypeError):
                error_msg = 'El precio unitario debe ser un número válido. Por favor, ingrese un valor numérico o deje el campo vacío.'
                if is_ajax:
                    return JsonResponse({'success': False, 'message': error_msg}, status=400)
                errores.append(error_msg)
        
        # Validar fecha de vencimiento
        fecha_vencimiento_obj = None
        if fecha_vencimiento:
            try:
                from datetime import datetime
                fecha_vencimiento_obj = datetime.strptime(fecha_vencimiento, '%Y-%m-%d').date()
                if fecha_vencimiento_obj < timezone.now().date():
                    error_msg = 'La fecha de vencimiento no puede ser anterior a la fecha actual. Por favor, seleccione una fecha válida.'
                    if is_ajax:
                        return JsonResponse({'success': False, 'message': error_msg}, status=400)
                    errores.append(error_msg)
            except (ValueError, TypeError):
                error_msg = 'La fecha de vencimiento debe tener un formato válido (YYYY-MM-DD). Por favor, verifique el formato.'
                if is_ajax:
                    return JsonResponse({'success': False, 'message': error_msg}, status=400)
                errores.append(error_msg)
        
        # Verificar si ya existe un insumo con el mismo nombre
        if Insumo.objects.filter(nombre__iexact=nombre).exists():
            error_msg = f'Ya existe un insumo con el nombre "{nombre}". Por favor, elija un nombre diferente.'
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
            errores.append(error_msg)
        
        # Validar y obtener proveedor si se proporciona
        proveedor_principal = None
        if proveedor_id:
            try:
                proveedor_principal = Proveedor.objects.get(id=proveedor_id, activo=True)
            except (Proveedor.DoesNotExist, ValueError):
                error_msg = 'El proveedor seleccionado no existe o está inactivo.'
                if is_ajax:
                    return JsonResponse({'success': False, 'message': error_msg}, status=400)
                errores.append(error_msg)
        
        if errores:
            for error in errores:
                messages.error(request, error)
            if is_ajax:
                return JsonResponse({'success': False, 'message': errores[0]}, status=400)
        else:
            try:
                # Procesar imagen si se proporciona
                imagen = None
                if 'imagen' in request.FILES:
                    imagen = request.FILES['imagen']
                    # Validar tamaño (máximo 5MB)
                    if imagen.size > 5 * 1024 * 1024:
                        error_msg = 'La imagen es demasiado grande. El tamaño máximo permitido es 5MB.'
                        if is_ajax:
                            return JsonResponse({'success': False, 'message': error_msg}, status=400)
                        messages.error(request, error_msg)
                        context = {
                            'perfil': perfil,
                            'categorias': Insumo.CATEGORIA_CHOICES,
                            'proveedores': Proveedor.objects.filter(activo=True).order_by('nombre'),
                            'es_admin': True,
                            'form_data': request.POST
                        }
                        return render(request, 'citas/insumos/agregar_insumo.html', context)
                
                insumo = Insumo.objects.create(
                    nombre=nombre,
                    categoria=categoria,
                    descripcion=descripcion,
                    imagen=imagen,
                    cantidad_actual=cantidad_actual,
                    cantidad_minima=cantidad_minima,
                    unidad_medida=unidad_medida,
                    precio_unitario=precio_unitario,
                    proveedor_principal=proveedor_principal,
                    fecha_vencimiento=fecha_vencimiento_obj,
                    ubicacion=ubicacion,
                    notas=notas,
                    creado_por=perfil
                )
                
                # Crear movimiento inicial
                MovimientoInsumo.objects.create(
                    insumo=insumo,
                    tipo='entrada',
                    cantidad=cantidad_actual,
                    cantidad_anterior=0,
                    cantidad_nueva=cantidad_actual,
                    motivo='Stock inicial',
                    realizado_por=perfil
                )
                
                success_msg = f'✅ Insumo "{nombre}" agregado correctamente con {cantidad_actual} {unidad_medida} de stock inicial.'
                
                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'message': success_msg,
                        'insumo': {
                            'id': insumo.id,
                            'nombre': insumo.nombre,
                            'categoria': insumo.get_categoria_display(),
                            'cantidad_actual': insumo.cantidad_actual,
                        }
                    })
                
                messages.success(request, success_msg)
                return redirect(reverse('gestor_inventario_unificado') + '?seccion=insumos')
            except Exception as e:
                error_msg = f'❌ Error inesperado al agregar el insumo. Por favor, intente nuevamente. Si el problema persiste, contacte al administrador del sistema. Detalles: {str(e)}'
                if is_ajax:
                    return JsonResponse({'success': False, 'message': error_msg}, status=500)
                messages.error(request, error_msg)
    
    # Obtener proveedores activos para el formulario
    proveedores = Proveedor.objects.filter(activo=True).order_by('nombre')
    
    context = {
        'perfil': perfil,
        'categorias': Insumo.CATEGORIA_CHOICES,
        'proveedores': proveedores,
        'es_admin': True,
        'form_data': request.POST if request.method == 'POST' else {}
    }
    
    return render(request, 'citas/insumos/agregar_insumo.html', context)

# Editar insumo
@login_required
def editar_insumo(request, insumo_id):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para editar insumos.')
            return redirect('gestor_inventario_unificado')
    except Perfil.DoesNotExist:
        return redirect('login')

    insumo = get_object_or_404(Insumo, id=insumo_id)
    
    if request.method == 'POST':
        insumo.nombre = request.POST.get('nombre')
        insumo.categoria = request.POST.get('categoria')
        insumo.descripcion = request.POST.get('descripcion', '')
        insumo.cantidad_actual = int(request.POST.get('cantidad_actual', 0))
        insumo.cantidad_minima = int(request.POST.get('cantidad_minima', 1))
        insumo.unidad_medida = request.POST.get('unidad_medida', 'unidad')
        precio_unitario_str = request.POST.get('precio_unitario')
        if precio_unitario_str:
            try:
                insumo.precio_unitario = round(float(precio_unitario_str))  # Redondear a entero para pesos chilenos
            except (ValueError, TypeError):
                insumo.precio_unitario = None
        else:
            insumo.precio_unitario = None
        # Actualizar proveedor principal
        proveedor_id = request.POST.get('proveedor_principal', '').strip()
        if proveedor_id:
            try:
                insumo.proveedor_principal = Proveedor.objects.get(id=proveedor_id, activo=True)
            except (Proveedor.DoesNotExist, ValueError):
                insumo.proveedor_principal = None
        else:
            insumo.proveedor_principal = None
        
        fecha_vencimiento_str = request.POST.get('fecha_vencimiento', '')
        if fecha_vencimiento_str:
            try:
                from datetime import datetime
                insumo.fecha_vencimiento = datetime.strptime(fecha_vencimiento_str, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                insumo.fecha_vencimiento = None
        else:
            insumo.fecha_vencimiento = None
        insumo.ubicacion = request.POST.get('ubicacion', '')
        insumo.notas = request.POST.get('notas', '')
        
        # Procesar imagen si se proporciona una nueva
        if 'imagen' in request.FILES:
            nueva_imagen = request.FILES['imagen']
            # Validar tamaño (máximo 5MB)
            if nueva_imagen.size > 5 * 1024 * 1024:
                messages.error(request, 'La imagen es demasiado grande. El tamaño máximo permitido es 5MB.')
                context = {
                    'perfil': perfil,
                    'insumo': insumo,
                    'categorias': Insumo.CATEGORIA_CHOICES,
                    'proveedores': Proveedor.objects.filter(activo=True).order_by('nombre'),
                    'es_admin': True
                }
                return render(request, 'citas/insumos/editar_insumo.html', context)
            insumo.imagen = nueva_imagen
        
        try:
            insumo.save()
            messages.success(request, f'✅ Insumo "{insumo.nombre}" actualizado correctamente.')
            return redirect(reverse('gestor_inventario_unificado') + '?seccion=insumos')
        except Exception as e:
            messages.error(request, f'Error al actualizar insumo: {e}')
    
    # Obtener proveedores activos para el formulario
    proveedores = Proveedor.objects.filter(activo=True).order_by('nombre')
    
    context = {
        'perfil': perfil,
        'insumo': insumo,
        'categorias': Insumo.CATEGORIA_CHOICES,
        'proveedores': proveedores,
        'es_admin': True
    }
    
    return render(request, 'citas/insumos/editar_insumo.html', context)

# Movimiento de stock
@login_required
def movimiento_insumo(request, insumo_id):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para realizar movimientos de stock.')
            return redirect('gestor_inventario_unificado')
    except Perfil.DoesNotExist:
        return redirect('login')

    insumo = get_object_or_404(Insumo, id=insumo_id)
    
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        cantidad = int(request.POST.get('cantidad', 0))
        motivo = request.POST.get('motivo', '')
        observaciones = request.POST.get('observaciones', '')
        
        if cantidad <= 0:
            messages.error(request, 'La cantidad debe ser mayor a 0.')
            return redirect(reverse('gestor_inventario_unificado') + '?seccion=insumos')
        
        cantidad_anterior = insumo.cantidad_actual
        
        try:
            if tipo == 'entrada':
                insumo.cantidad_actual += cantidad
            elif tipo == 'salida':
                if cantidad > insumo.cantidad_actual:
                    messages.error(request, 'No hay suficiente stock disponible.')
                    return redirect(reverse('gestor_inventario_unificado') + '?seccion=insumos')
                insumo.cantidad_actual -= cantidad
            elif tipo == 'ajuste':
                insumo.cantidad_actual = cantidad
            
            insumo.save()
            
            # Crear movimiento
            MovimientoInsumo.objects.create(
                insumo=insumo,
                tipo=tipo,
                cantidad=cantidad,
                cantidad_anterior=cantidad_anterior,
                cantidad_nueva=insumo.cantidad_actual,
                motivo=motivo,
                observaciones=observaciones,
                realizado_por=perfil
            )
            
            messages.success(request, f'✅ Movimiento de stock realizado correctamente.')
            return redirect(reverse('gestor_inventario_unificado') + '?seccion=insumos')
        except Exception as e:
            messages.error(request, f'Error al realizar movimiento: {e}')
    
    context = {
        'perfil': perfil,
        'insumo': insumo,
        'tipos_movimiento': MovimientoInsumo.TIPO_CHOICES,
        'es_admin': True
    }
    
    return render(request, 'citas/insumos/movimiento_insumo.html', context)

# Eliminar insumo
@login_required
def eliminar_insumo(request, insumo_id):
    """Vista para eliminar un insumo"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para eliminar insumos.')
            return redirect(reverse('gestor_inventario_unificado') + '?seccion=insumos')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    if request.method == 'POST':
        try:
            insumo = get_object_or_404(Insumo, id=insumo_id)
            nombre = insumo.nombre
            
            # Verificar si tiene movimientos asociados
            movimientos_count = insumo.movimientos.count()
            if movimientos_count > 0:
                messages.warning(request, f'⚠️ El insumo "{nombre}" tiene {movimientos_count} movimiento(s) registrado(s). Se eliminarán también los movimientos asociados.')
            
            insumo.delete()
            messages.success(request, f'✅ Insumo "{nombre}" eliminado correctamente.')
        except Exception as e:
            messages.error(request, f'❌ Error inesperado al eliminar el insumo. Por favor, intente nuevamente. Si el problema persiste, contacte al administrador del sistema. Detalles: {str(e)}')
    
    return redirect(reverse('gestor_inventario_unificado') + '?seccion=insumos')

# Historial de movimientos
@login_required
def historial_movimientos(request):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para ver el historial de movimientos.')
            return redirect('gestor_inventario_unificado')
    except Perfil.DoesNotExist:
        return redirect('login')

    movimientos = MovimientoInsumo.objects.all().order_by('-fecha_movimiento')
    
    # Filtros
    insumo_id = request.GET.get('insumo')
    tipo = request.GET.get('tipo')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if insumo_id:
        movimientos = movimientos.filter(insumo_id=insumo_id)
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)
    if fecha_desde:
        movimientos = movimientos.filter(fecha_movimiento__date__gte=fecha_desde)
    if fecha_hasta:
        movimientos = movimientos.filter(fecha_movimiento__date__lte=fecha_hasta)
    
    context = {
        'perfil': perfil,
        'movimientos': movimientos,
        'insumos': Insumo.objects.all().order_by('nombre'),
        'tipos_movimiento': MovimientoInsumo.TIPO_CHOICES,
        'insumo_actual': insumo_id,
        'tipo_actual': tipo,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'es_admin': True
    }
    
    return render(request, 'citas/insumos/historial_movimientos.html', context)

# Gestión de Personal - solo administrativos
@login_required
def gestor_personal(request):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para gestionar personal.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        return redirect('login')

    # Filtros de búsqueda
    search = request.GET.get('search', '')
    rol = request.GET.get('rol', '')
    estado = request.GET.get('estado', '')
    
    personal = Perfil.objects.all()
    
    if search:
        personal = personal.filter(
            Q(nombre_completo__icontains=search) |
            Q(email__icontains=search) |
            Q(telefono__icontains=search) |
            Q(user__username__icontains=search) |
            Q(especialidad__icontains=search)
        )
    
    if rol:
        personal = personal.filter(rol=rol)
    
    if estado == 'activo':
        personal = personal.filter(activo=True)
    elif estado == 'inactivo':
        personal = personal.filter(activo=False)
    
    personal = personal.order_by('nombre_completo')
    
    # Paginación - 10 registros por página
    paginator = Paginator(personal, 10)
    page = request.GET.get('page', 1)
    
    try:
        personal = paginator.page(page)
    except PageNotAnInteger:
        personal = paginator.page(1)
    except EmptyPage:
        personal = paginator.page(paginator.num_pages)
    
    # Estadísticas de personal
    total_personal = Perfil.objects.count()
    dentistas_count = Perfil.objects.filter(rol='dentista', activo=True).count()
    administrativos_count = Perfil.objects.filter(rol='administrativo', activo=True).count()
    personal_inactivo = Perfil.objects.filter(activo=False).count()
    personal_sin_acceso = Perfil.objects.filter(requiere_acceso_sistema=False).count()
    
    # Estadísticas adicionales
    from datetime import datetime, timedelta
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    nuevos_este_mes = Perfil.objects.filter(fecha_registro__gte=inicio_mes).count()
    
    # Calcular porcentajes
    dentistas_porcentaje = round((dentistas_count / total_personal * 100) if total_personal > 0 else 0, 1)
    administrativos_porcentaje = round((administrativos_count / total_personal * 100) if total_personal > 0 else 0, 1)

    # Datos adicionales para panel derecho
    # Personal reciente (últimos 5 registrados)
    personal_reciente = Perfil.objects.order_by('-fecha_registro')[:5]
    
    # Dentistas con horarios configurados hoy
    dentistas_con_horario_hoy = Perfil.objects.filter(
        rol='dentista',
        activo=True,
        horarios__dia_semana=timezone.now().weekday(),
        horarios__activo=True
    ).distinct()[:5]

    estadisticas = {
        'total_personal': total_personal,
        'dentistas': dentistas_count,
        'administrativos': administrativos_count,
        'inactivos': personal_inactivo,
        'nuevos_este_mes': nuevos_este_mes,
        'dentistas_porcentaje': dentistas_porcentaje,
        'administrativos_porcentaje': administrativos_porcentaje,
        'sin_acceso': personal_sin_acceso,
    }
    
    context = {
        'perfil': perfil,
        'personal': personal,
        'estadisticas': estadisticas,
        'roles': Perfil.ROLE_CHOICES,
        'search': search,
        'rol': rol,
        'estado': estado,
        'es_admin': True,
        'personal_reciente': personal_reciente,
        'dentistas_con_horario_hoy': dentistas_con_horario_hoy,
    }
    
    return render(request, 'citas/personal/gestor_personal.html', context)

# Agregar nuevo personal
@login_required
def agregar_personal(request):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para agregar personal.')
            return redirect('gestor_personal')
    except Perfil.DoesNotExist:
        return redirect('login')

    if request.method == 'POST':
        logging.info('=== RECIBIENDO POST PARA AGREGAR PERSONAL ===')
        logging.info(f'POST data: {dict(request.POST)}')
        logging.info(f'FILES data: {dict(request.FILES)}')
        
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        nombre_completo = request.POST.get('nombre_completo', '').strip()
        email = request.POST.get('email', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        rol = request.POST.get('rol', '')
        especialidad = request.POST.get('especialidad', '').strip()
        numero_colegio = request.POST.get('numero_colegio', '').strip()
        requiere_acceso = request.POST.get('requiere_acceso_sistema') == 'on'
        foto = request.FILES.get('foto')
        
        logging.info(f'Datos procesados: nombre={nombre_completo}, email={email}, telefono={telefono}, rol={rol}, requiere_acceso={requiere_acceso}')
        
        # Validaciones
        errores = []
        
        # Solo validar usuario y contraseña si requiere acceso al sistema
        if requiere_acceso:
            if not username:
                errores.append('El nombre de usuario es obligatorio para personal con acceso al sistema.')
            elif User.objects.filter(username=username).exists():
                errores.append('Ya existe un usuario con este nombre de usuario.')
            
            if not password:
                errores.append('La contraseña es obligatoria para personal con acceso al sistema.')
            elif len(password) < 8:
                errores.append('La contraseña debe tener al menos 8 caracteres.')
            
            if User.objects.filter(email=email).exists():
                errores.append('Ya existe un usuario con este email.')
        
        if not nombre_completo:
            errores.append('El nombre completo es obligatorio.')
        
        if not email:
            errores.append('El email es obligatorio.')
        
        if not telefono:
            errores.append('El teléfono es obligatorio.')
        
        if not rol:
            errores.append('El rol es obligatorio.')
        
        # Validaciones específicas para dentistas
        if rol == 'dentista':
            if not especialidad:
                errores.append('La especialidad es obligatoria para dentistas.')
            if not numero_colegio:
                errores.append('El número de colegio es obligatorio para dentistas.')
        
        if errores:
            logging.error(f'Errores de validación encontrados: {errores}')
            for error in errores:
                messages.error(request, error)
                logging.error(f'Mensaje de error agregado: {error}')
            logging.info('Redirigiendo a gestor_personal con errores')
            return redirect('gestor_personal')
        else:
            try:
                logging.info(f'Iniciando creación de personal: {nombre_completo}, email: {email}, rol: {rol}')
                user = None
                # Solo crear usuario si requiere acceso al sistema
                if requiere_acceso:
                    logging.info(f'Creando usuario para personal con acceso: {username}')
                    user = User.objects.create_user(
                        username=username,
                        email=email,
                        password=password,
                        first_name=nombre_completo.split()[0] if nombre_completo.split() else '',
                        last_name=' '.join(nombre_completo.split()[1:]) if len(nombre_completo.split()) > 1 else ''
                    )
                    logging.info(f'Usuario creado: {user.id}')
                else:
                    logging.info('Personal sin acceso al sistema, no se crea usuario')
                
                # Crear perfil con validación explícita
                logging.info('Creando objeto Perfil...')
                perfil_nuevo = Perfil(
                    user=user,
                    nombre_completo=nombre_completo.strip(),
                    telefono=telefono.strip(),
                    email=email.strip(),
                    rol=rol,
                    especialidad=especialidad.strip() if rol == 'dentista' else '',
                    numero_colegio=numero_colegio.strip() if rol == 'dentista' else '',
                    requiere_acceso_sistema=requiere_acceso,
                    foto=foto if foto else None
                )
                
                # Si el rol es 'general' (administrador), asegurar todos los permisos
                if rol == 'general':
                    perfil_nuevo.puede_gestionar_citas = True
                    perfil_nuevo.puede_gestionar_clientes = True
                    perfil_nuevo.puede_gestionar_insumos = True
                    perfil_nuevo.puede_gestionar_personal = True
                    perfil_nuevo.puede_ver_reportes = True
                    perfil_nuevo.puede_crear_odontogramas = True
                
                # Validar el modelo antes de guardar
                logging.info('Validando perfil...')
                perfil_nuevo.full_clean()
                logging.info('Guardando perfil en base de datos...')
                perfil_nuevo.save()
                logging.info(f'Perfil guardado con ID: {perfil_nuevo.id}')
                
                # Verificar que se guardó correctamente
                perfil_guardado = Perfil.objects.filter(id=perfil_nuevo.id).first()
                if not perfil_guardado:
                    raise Exception('El perfil no se guardó correctamente en la base de datos.')
                
                # Verificar en la tabla
                total_perfiles = Perfil.objects.count()
                logging.info(f'Total de perfiles en la base de datos: {total_perfiles}')
                
                tipo_personal = "con acceso al sistema" if requiere_acceso else "sin acceso al sistema"
                messages.success(request, f'Personal "{nombre_completo}" agregado correctamente ({tipo_personal}).')
                logging.info(f'Personal "{nombre_completo}" agregado exitosamente. ID: {perfil_guardado.id}')
                return redirect('gestor_personal')
            except Exception as e:
                import traceback
                error_detail = str(e)
                # Si es un error de validación, mostrar el mensaje específico
                if hasattr(e, 'message_dict'):
                    error_messages = []
                    for field, messages_list in e.message_dict.items():
                        error_messages.extend(messages_list)
                    error_detail = '; '.join(error_messages)
                elif hasattr(e, 'messages'):
                    error_detail = '; '.join([str(msg) for msg in e.messages])
                
                # Log detallado del error
                error_traceback = traceback.format_exc()
                logging.error(f'Error al agregar personal: {error_detail}\n{error_traceback}')
                print(f'ERROR AL AGREGAR PERSONAL: {error_detail}')
                print(f'TRACEBACK: {error_traceback}')
                
                messages.error(request, f'Error al agregar personal: {error_detail}')
                return redirect('gestor_personal')
    
    # Si es GET, redirigir al gestor de personal
    return redirect('gestor_personal')

# Editar personal
@login_required
def editar_personal(request, personal_id):
    try:
        perfil_admin = Perfil.objects.get(user=request.user)
        if not perfil_admin.es_administrativo():
            messages.error(request, 'No tienes permisos para editar personal.')
            return redirect('gestor_personal')
    except Perfil.DoesNotExist:
        return redirect('login')

    personal = get_object_or_404(Perfil, id=personal_id)
    
    if request.method == 'POST':
        # Obtener datos del formulario - asegurar que nombre_completo se obtiene correctamente
        nombre_completo = request.POST.get('nombre_completo', '').strip()
        email = request.POST.get('email', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        rol = request.POST.get('rol', '').strip()
        especialidad = request.POST.get('especialidad', '').strip()
        numero_colegio = request.POST.get('numero_colegio', '').strip()
        activo = request.POST.get('activo') == 'on'
        requiere_acceso = request.POST.get('requiere_acceso_sistema') == 'on'
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        foto = request.FILES.get('foto')
        eliminar_foto = request.POST.get('eliminar_foto') == 'on'
        
        # Debug: Log de los datos recibidos
        import logging
        logging.info(f'=== EDITAR PERSONAL - Datos recibidos ===')
        logging.info(f'nombre_completo: "{nombre_completo}"')
        logging.info(f'email: "{email}"')
        logging.info(f'telefono: "{telefono}"')
        logging.info(f'rol: "{rol}"')
        logging.info(f'POST completo: {dict(request.POST)}')
        
        # Validaciones
        errores = []
        
        if not nombre_completo:
            errores.append('El nombre completo es obligatorio.')
        
        if not email:
            errores.append('El email es obligatorio.')
        
        # Validar email único considerando si tiene usuario o no
        if personal.user:
            if User.objects.filter(email=email).exclude(id=personal.user.id).exists():
                errores.append('Ya existe otro usuario con este email.')
        else:
            if requiere_acceso and User.objects.filter(email=email).exists():
                errores.append('Ya existe un usuario con este email.')
        
        if not telefono:
            errores.append('El teléfono es obligatorio.')
        
        if not rol:
            errores.append('El rol es obligatorio.')
        
        # Validaciones para personal con acceso al sistema
        if requiere_acceso and not personal.user:
            # Se está creando nuevo usuario
            if not username:
                errores.append('El nombre de usuario es obligatorio para personal con acceso al sistema.')
            elif User.objects.filter(username=username).exists():
                errores.append('Ya existe un usuario con este nombre de usuario.')
            
            if not password:
                errores.append('La contraseña es obligatoria al crear acceso al sistema.')
            elif len(password) < 8:
                errores.append('La contraseña debe tener al menos 8 caracteres.')
        
        # Validaciones específicas para dentistas
        if rol == 'dentista':
            if not especialidad:
                errores.append('La especialidad es obligatoria para dentistas.')
            if not numero_colegio:
                errores.append('El número de colegio es obligatorio para dentistas.')
        
        if errores:
            for error in errores:
                messages.error(request, error)
        else:
            try:
                # Gestionar usuario Django
                if requiere_acceso:
                    if personal.user:
                        # Actualizar usuario existente
                        personal.user.email = email
                        personal.user.first_name = nombre_completo.split()[0] if nombre_completo.split() else ''
                        personal.user.last_name = ' '.join(nombre_completo.split()[1:]) if len(nombre_completo.split()) > 1 else ''
                        personal.user.is_active = activo
                        if username:
                            personal.user.username = username
                        if password:
                            personal.user.set_password(password)
                        personal.user.save()
                    else:
                        # Crear nuevo usuario
                        user = User.objects.create_user(
                            username=username,
                            email=email,
                            password=password,
                            first_name=nombre_completo.split()[0] if nombre_completo.split() else '',
                            last_name=' '.join(nombre_completo.split()[1:]) if len(nombre_completo.split()) > 1 else ''
                        )
                        personal.user = user
                else:
                    # No requiere acceso - eliminar usuario si existe
                    if personal.user:
                        user_to_delete = personal.user
                        personal.user = None
                        user_to_delete.delete()
                
                # Actualizar perfil - IMPORTANTE: Actualizar nombre_completo primero
                personal.nombre_completo = nombre_completo.strip()
                personal.email = email.strip()
                personal.telefono = telefono.strip()
                personal.rol = rol
                personal.especialidad = especialidad.strip() if rol == 'dentista' else ''
                personal.numero_colegio = numero_colegio.strip() if rol == 'dentista' else ''
                personal.activo = activo
                personal.requiere_acceso_sistema = requiere_acceso
                
                # Si el rol es 'general' (administrador), asegurar todos los permisos
                if rol == 'general':
                    personal.puede_gestionar_citas = True
                    personal.puede_gestionar_clientes = True
                    personal.puede_gestionar_insumos = True
                    personal.puede_gestionar_personal = True
                    personal.puede_ver_reportes = True
                    personal.puede_crear_odontogramas = True
                
                # Gestionar foto
                if eliminar_foto and personal.foto:
                    personal.foto.delete()
                    personal.foto = None
                elif foto:
                    # Eliminar foto anterior si existe
                    if personal.foto:
                        personal.foto.delete()
                    personal.foto = foto
                
                # Guardar el perfil (el método save() también establecerá permisos si es rol 'general')
                personal.save()
                
                # Registrar en auditoría
                detalles_personal = f'Rol: {rol}, Email: {email}, Activo: {"Sí" if activo else "No"}'
                if requiere_acceso:
                    detalles_personal += ', Acceso al sistema: Sí'
                registrar_auditoria(
                    usuario=perfil_admin,
                    accion='actualizar',
                    modulo='personal',
                    descripcion=f'Personal editado: {nombre_completo}',
                    detalles=detalles_personal,
                    objeto_id=personal.id,
                    tipo_objeto='Perfil',
                    request=request
                )
                
                messages.success(request, f'Personal "{nombre_completo}" actualizado correctamente.')
                return redirect('gestor_personal')
            except Exception as e:
                messages.error(request, f'Error al actualizar personal: {str(e)}')
    
    context = {
        'perfil_admin': perfil_admin,
        'personal': personal,
        'roles': Perfil.ROLE_CHOICES,
        'es_admin': True
    }
    
    return render(request, 'citas/personal/editar_personal.html', context)

# Eliminar personal
@login_required
def eliminar_personal(request, personal_id):
    try:
        perfil_admin = Perfil.objects.get(user=request.user)
        if not perfil_admin.es_administrativo():
            messages.error(request, 'No tienes permisos para eliminar personal.')
            return redirect('gestor_personal')
    except Perfil.DoesNotExist:
        return redirect('login')

    if request.method == 'POST':
        personal = get_object_or_404(Perfil, id=personal_id)
        nombre_completo = personal.nombre_completo
        
        try:
            # Eliminar foto si existe
            if personal.foto:
                personal.foto.delete()
            
            # Eliminar usuario asociado si existe
            if personal.user:
                user = personal.user
                user.delete()
            
            # Registrar en auditoría ANTES de eliminar
            detalles_eliminacion = f'Rol: {personal.rol}, Email: {personal.email}'
            if personal.user:
                detalles_eliminacion += ', Usuario web eliminado'
            registrar_auditoria(
                usuario=perfil_admin,
                accion='eliminar',
                modulo='personal',
                descripcion=f'Personal eliminado: {nombre_completo}',
                detalles=detalles_eliminacion,
                objeto_id=personal_id,
                tipo_objeto='Perfil',
                request=request
            )
            
            # Eliminar perfil
            personal.delete()
            
            messages.success(request, f'Personal "{nombre_completo}" eliminado correctamente del sistema.')
        except Exception as e:
            messages.error(request, f'Error al eliminar personal: {str(e)}')
    
    return redirect('gestor_personal')

# Toggle estado personal (activar/desactivar)
@login_required
def toggle_estado_personal(request, personal_id):
    """
    Vista para activar/desactivar un miembro del personal
    También desactiva/activa su cuenta de usuario web si existe
    """
    try:
        perfil_admin = Perfil.objects.get(user=request.user)
        if not perfil_admin.es_administrativo():
            return JsonResponse({
                'success': False,
                'error': 'No tienes permisos para realizar esta acción'
            }, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'No tienes permisos'
        }, status=403)
    
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido'
        }, status=405)
    
    try:
        from django.contrib.auth.models import User
        
        personal = Perfil.objects.get(id=personal_id)
        nuevo_estado = not personal.activo
        personal.activo = nuevo_estado
        personal.save()
        
        # Buscar si existe un usuario web asociado
        if personal.user:
            # Desactivar/activar también el usuario de Django
            personal.user.is_active = nuevo_estado
            personal.user.save()
            
            mensaje = f'Personal "{personal.nombre_completo}" {"activado" if nuevo_estado else "desactivado"} exitosamente. '
            mensaje += f'Usuario web también {"activado" if nuevo_estado else "desactivado"} (no podrá hacer login).'
        else:
            # El personal no tiene usuario web
            mensaje = f'Personal "{personal.nombre_completo}" {"activado" if nuevo_estado else "desactivado"} exitosamente.'
        
        return JsonResponse({
            'success': True,
            'activo': personal.activo,
            'message': mensaje
        })
    except Perfil.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Personal no encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al cambiar el estado: {str(e)}'
        }, status=500)

# Calendario personal para dentistas
@login_required
def calendario_personal(request):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'Solo los dentistas pueden acceder al calendario personal.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        return redirect('login')

    # Obtener parámetros de fecha
    fecha_actual = request.GET.get('fecha', timezone.now().strftime('%Y-%m-%d'))
    try:
        fecha_obj = datetime.strptime(fecha_actual, '%Y-%m-%d').date()
    except ValueError:
        fecha_obj = timezone.now().date()
        fecha_actual = fecha_obj.strftime('%Y-%m-%d')
    
    # Obtener citas del dentista para la fecha seleccionada
    citas_dia = Cita.objects.filter(
        dentista=perfil,
        fecha_hora__date=fecha_obj
    ).order_by('fecha_hora')
    
    # Obtener citas de la semana
    inicio_semana = fecha_obj - timedelta(days=fecha_obj.weekday())
    fin_semana = inicio_semana + timedelta(days=6)
    
    citas_semana = Cita.objects.filter(
        dentista=perfil,
        fecha_hora__date__range=[inicio_semana, fin_semana]
    ).order_by('fecha_hora')
    
    # Estadísticas del dentista
    citas_hoy = citas_dia.count()
    citas_semana_count = citas_semana.count()
    citas_pendientes = Cita.objects.filter(
        dentista=perfil,
        estado__in=['reservada', 'confirmada'],
        fecha_hora__gte=timezone.now()
    ).count()
    citas_completadas_mes = Cita.objects.filter(
        dentista=perfil,
        estado='completada',
        fecha_hora__date__gte=timezone.now().date() - timedelta(days=30)
    ).count()
    
    estadisticas = {
        'citas_hoy': citas_hoy,
        'citas_semana': citas_semana_count,
        'citas_pendientes': citas_pendientes,
        'citas_completadas_mes': citas_completadas_mes,
    }
    
    # Crear calendario semanal
    calendario_semana = []
    for i in range(7):
        fecha_dia = inicio_semana + timedelta(days=i)
        citas_dia_semana = citas_semana.filter(fecha_hora__date=fecha_dia)
        calendario_semana.append({
            'fecha': fecha_dia,
            'citas': citas_dia_semana,
            'es_hoy': fecha_dia == fecha_obj,
            'es_pasado': fecha_dia < timezone.now().date()
        })
    
    context = {
        'perfil': perfil,
        'fecha_actual': fecha_actual,
        'fecha_obj': fecha_obj,
        'citas_dia': citas_dia,
        'citas_semana': citas_semana,
        'calendario_semana': calendario_semana,
        'estadisticas': estadisticas,
        'es_dentista': True
    }
    
    return render(request, 'citas/personal/calendario_personal.html', context)

# Vista para mostrar perfil del dentista
@login_required
def obtener_perfil_json(request):
    """Vista AJAX para obtener información del perfil en JSON"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        
        # Obtener estadísticas según el rol
        if perfil.es_dentista():
            citas_totales = Cita.objects.filter(dentista=perfil).count()
            citas_completadas = Cita.objects.filter(dentista=perfil, estado='completada').count()
            citas_pendientes = Cita.objects.filter(dentista=perfil, estado='reservada').count()
            citas_confirmadas = Cita.objects.filter(dentista=perfil, estado='confirmada').count()
            clientes_gestionados = 0
        elif perfil.es_administrativo():
            # Para administrativos: citas creadas por ellos y clientes creados
            citas_totales = Cita.objects.filter(creada_por=perfil).count()
            citas_completadas = Cita.objects.filter(creada_por=perfil, estado='completada').count()
            citas_pendientes = Cita.objects.filter(creada_por=perfil, estado__in=['reservada', 'confirmada']).count()
            citas_confirmadas = Cita.objects.filter(creada_por=perfil, estado='confirmada').count()
            # Clientes creados por este administrativo (si hay campo creado_por)
            try:
                clientes_gestionados = Cliente.objects.filter(activo=True).count()  # Total de clientes activos
            except:
                clientes_gestionados = 0
        else:
            citas_totales = 0
            citas_completadas = 0
            citas_pendientes = 0
            citas_confirmadas = 0
            clientes_gestionados = 0
        
        data = {
            'success': True,
            'perfil': {
                'nombre_completo': perfil.nombre_completo,
                'rol': perfil.get_rol_display(),
                'especialidad': perfil.especialidad or '',
                'email': perfil.user.email,
                'telefono': perfil.telefono or 'No especificado',
                'activo': perfil.activo,
                'numero_colegio': perfil.numero_colegio or '',
                'fecha_registro': perfil.user.date_joined.strftime('%B %Y'),
                'ultimo_acceso': perfil.user.last_login.strftime('%d/%m/%Y %H:%M') if perfil.user.last_login else 'Nunca',
                'username': perfil.user.username,
            },
            'estadisticas': {
                'citas_totales': citas_totales,
                'citas_completadas': citas_completadas,
                'citas_pendientes': citas_pendientes,
                'citas_confirmadas': citas_confirmadas,
                'clientes_gestionados': clientes_gestionados,
            },
            'es_dentista': perfil.es_dentista(),
            'es_administrativo': perfil.es_administrativo(),
        }
        
        return JsonResponse(data)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Perfil no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def mi_perfil(request):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'No tienes permisos para acceder a esta página.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'Perfil no encontrado.')
        return redirect('login')
    
    # Obtener estadísticas del dentista
    citas_totales = Cita.objects.filter(dentista=perfil).count()
    citas_completadas = Cita.objects.filter(dentista=perfil, estado='completada').count()
    citas_pendientes = Cita.objects.filter(dentista=perfil, estado='reservada').count()
    
    context = {
        'perfil': perfil,
        'citas_totales': citas_totales,
        'citas_completadas': citas_completadas,
        'citas_pendientes': citas_pendientes,
    }
    
    return render(request, 'citas/perfil/mi_perfil.html', context)

# Vista para asignar dentista a una cita
@login_required
def asignar_dentista_cita(request, cita_id):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para asignar dentistas.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        return redirect('login')

    cita = get_object_or_404(Cita, id=cita_id)
    
    if request.method == 'POST':
        dentista_id = request.POST.get('dentista') or request.POST.get('dentista_id')
        if dentista_id:
            try:
                dentista = Perfil.objects.get(id=dentista_id, rol='dentista', activo=True)
                # Asignar dentista usando ORM normal
                cita.dentista = dentista
                cita.save()
                messages.success(request, f'Dentista {dentista.nombre_completo} asignado correctamente.')
                return redirect('panel_trabajador')
            except Perfil.DoesNotExist:
                messages.error(request, 'Dentista no encontrado.')
        else:
            messages.error(request, 'Debe seleccionar un dentista.')
    
    # Obtener dentistas disponibles
    dentistas = Perfil.objects.filter(rol='dentista', activo=True).order_by('nombre_completo')
    
    context = {
        'perfil': perfil,
        'cita': cita,
        'dentistas': dentistas,
        'es_admin': True
    }
    
    return render(request, 'citas/citas/asignar_dentista_cita.html', context)

# Vista para que los dentistas vean sus citas asignadas
@login_required
def mis_citas_dentista(request):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'No tienes permisos para acceder a esta función.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    # Filtros
    estado = request.GET.get('estado', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    tab = request.GET.get('tab', 'activas')  # 'activas' o 'completadas'
    
    # Obtener solo citas asignadas al dentista actual (no puede tomar citas disponibles)
    # Incluir relación con cliente para mostrar nombre completo
    from pacientes.models import Cliente
    from historial_clinico.models import Odontograma
    
    # Separar citas en dos grupos: activas (reservadas, confirmadas, en_espera, listo_para_atender, en_progreso, finalizada) y completadas
    citas_activas = Cita.objects.filter(
        dentista=perfil,
        estado__in=['reservada', 'confirmada', 'en_espera', 'listo_para_atender', 'en_progreso', 'finalizada']
    ).select_related('cliente', 'dentista', 'tipo_servicio')
    
    citas_completadas = Cita.objects.filter(
        dentista=perfil,
        estado='completada'
    ).select_related('cliente', 'dentista', 'tipo_servicio')
    
    # Aplicar filtros de fecha si existen
    if fecha_desde:
        citas_activas = citas_activas.filter(fecha_hora__date__gte=fecha_desde)
        citas_completadas = citas_completadas.filter(fecha_hora__date__gte=fecha_desde)
    if fecha_hasta:
        citas_activas = citas_activas.filter(fecha_hora__date__lte=fecha_hasta)
        citas_completadas = citas_completadas.filter(fecha_hora__date__lte=fecha_hasta)
    
    citas_activas = citas_activas.order_by('-fecha_hora')
    citas_completadas = citas_completadas.order_by('-fecha_hora')
    
    # Obtener IDs de citas que tienen fichas asociadas
    odontogramas = Odontograma.objects.filter(dentista=perfil).exclude(cita__isnull=True)
    citas_con_ficha = set(odontogramas.values_list('cita_id', flat=True))
    
    # Procesar citas activas
    citas_activas_list = list(citas_activas)
    for cita in citas_activas_list:
        # Vincular cliente si no está vinculado pero tiene email
        if not cita.cliente and cita.paciente_email:
            try:
                cliente = Cliente.objects.get(email=cita.paciente_email, activo=True)
                cita.cliente = cliente
            except (Cliente.DoesNotExist, Cliente.MultipleObjectsReturned):
                cliente = Cliente.objects.filter(email=cita.paciente_email, activo=True).first()
                if cliente:
                    cita.cliente = cliente
        # Marcar si tiene ficha
        cita.tiene_ficha = cita.id in citas_con_ficha
        if cita.tiene_ficha:
            cita.odontograma = odontogramas.filter(cita_id=cita.id).first()
    
    # Procesar citas completadas
    citas_completadas_list = list(citas_completadas)
    for cita in citas_completadas_list:
        # Vincular cliente si no está vinculado pero tiene email
        if not cita.cliente and cita.paciente_email:
            try:
                cliente = Cliente.objects.get(email=cita.paciente_email, activo=True)
                cita.cliente = cliente
            except (Cliente.DoesNotExist, Cliente.MultipleObjectsReturned):
                cliente = Cliente.objects.filter(email=cita.paciente_email, activo=True).first()
                if cliente:
                    cita.cliente = cliente
        # Marcar si tiene ficha
        cita.tiene_ficha = cita.id in citas_con_ficha
        if cita.tiene_ficha:
            cita.odontograma = odontogramas.filter(cita_id=cita.id).first()
    
    # Calcular estadísticas solo para citas del dentista
    citas_dentista = Cita.objects.filter(dentista=perfil)
    total_citas = citas_dentista.count()
    citas_hoy = citas_dentista.filter(fecha_hora__date=timezone.now().date()).count()
    citas_reservadas = citas_dentista.filter(estado='reservada').count()
    citas_completadas_count = citas_dentista.filter(estado='completada').count()
    
    context = {
        'citas_activas': citas_activas_list,
        'citas_completadas': citas_completadas_list,
        'estados': Cita.ESTADO_CHOICES,
        'estado_actual': estado,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'tab': tab,
        'perfil': perfil,
        'total_citas': total_citas,
        'citas_hoy': citas_hoy,
        'citas_reservadas': citas_reservadas,
        'citas_completadas_count': citas_completadas_count
    }
    
    return render(request, 'citas/citas/mis_citas_dentista.html', context)

# Vista para que el dentista tome una cita disponible
@login_required
def tomar_cita_dentista(request, cita_id):
    # Los dentistas ya no pueden tomar citas directamente
    # Solo los administrativos pueden asignar citas
    messages.error(request, 'Los dentistas no pueden tomar citas directamente. Contacta al administrador para que te asigne citas.')
    return redirect('mis_citas_dentista')

# ELIMINADO: Los dentistas ya no pueden completar sus propias citas
# Solo el personal administrativo puede marcar citas como completadas

# Exportar lista de insumos a PDF
@login_required
def exportar_insumos_pdf(request):
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para exportar insumos.')
            return redirect('gestor_inventario_unificado')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    # Obtener todos los insumos
    insumos = Insumo.objects.all().order_by('categoria', 'nombre')
    
    # Crear el buffer para el PDF
    buffer = BytesIO()
    
    # Crear el documento PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=18
    )
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Estilo personalizado para el título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#3b82f6')
    )
    
    # Estilo para subtítulos
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        alignment=TA_LEFT,
        textColor=colors.HexColor('#1e293b')
    )
    
    # Estilo para información de la clínica
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#64748b')
    )
    
    # Contenido del PDF
    story = []
    
    # Título principal
    title = Paragraph("INVENTARIO DE INSUMOS", title_style)
    story.append(title)
    
    # Información de la clínica y fecha
    clinic_info = Paragraph(
        f"<b>Clínica Dental</b><br/>"
        f"Reporte generado el: {datetime.now().strftime('%d/%m/%Y a las %H:%M')}<br/>"
        f"Generado por: {perfil.nombre_completo}",
        info_style
    )
    story.append(clinic_info)
    story.append(Spacer(1, 20))
    
    # Estadísticas generales
    total_insumos = insumos.count()
    stock_bajo = insumos.filter(cantidad_actual__lte=F('cantidad_minima')).count()
    agotados = insumos.filter(estado='agotado').count()
    
    stats_text = f"""
    <b>ESTADÍSTICAS GENERALES:</b><br/>
    • Total de insumos: {total_insumos}<br/>
    • Stock bajo: {stock_bajo}<br/>
    • Agotados: {agotados}
    """
    stats_para = Paragraph(stats_text, subtitle_style)
    story.append(stats_para)
    story.append(Spacer(1, 20))
    
    # Agrupar insumos por categoría
    categorias = {}
    for insumo in insumos:
        categoria = insumo.get_categoria_display()
        if categoria not in categorias:
            categorias[categoria] = []
        categorias[categoria].append(insumo)
    
    # Crear tabla para cada categoría
    for categoria, insumos_categoria in categorias.items():
        # Subtítulo de categoría
        categoria_title = Paragraph(f"<b>{categoria.upper()}</b>", subtitle_style)
        story.append(categoria_title)
        story.append(Spacer(1, 10))
        
        # Crear tabla de insumos
        table_data = [
            ['Nombre', 'Stock Actual', 'Stock Mínimo', 'Estado', 'Proveedor', 'Ubicación']
        ]
        
        for insumo in insumos_categoria:
            # Determinar estado visual
            if insumo.estado == 'agotado':
                estado = "AGOTADO"
            elif insumo.stock_bajo:
                estado = "STOCK BAJO"
            elif insumo.proximo_vencimiento:
                estado = "PRÓXIMO A VENCER"
            else:
                estado = "DISPONIBLE"
            
            table_data.append([
                insumo.nombre,
                f"{insumo.cantidad_actual} {insumo.unidad_medida}",
                f"{insumo.cantidad_minima} {insumo.unidad_medida}",
                estado,
                insumo.proveedor or "N/A",
                insumo.ubicacion or "N/A"
            ])
        
        # Crear tabla
        table = Table(table_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1.5*inch, 1*inch])
        
        # Estilo de la tabla
        table.setStyle(TableStyle([
            # Encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Filas de datos
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Colores de estado
            ('TEXTCOLOR', (3, 1), (3, -1), colors.black),
        ]))
        
        # Aplicar colores de estado
        for i, insumo in enumerate(insumos_categoria, 1):
            if insumo.estado == 'agotado':
                table.setStyle(TableStyle([
                    ('TEXTCOLOR', (3, i), (3, i), colors.red),
                ]))
            elif insumo.stock_bajo:
                table.setStyle(TableStyle([
                    ('TEXTCOLOR', (3, i), (3, i), colors.orange),
                ]))
            elif insumo.proximo_vencimiento:
                table.setStyle(TableStyle([
                    ('TEXTCOLOR', (3, i), (3, i), colors.red),
                ]))
            else:
                table.setStyle(TableStyle([
                    ('TEXTCOLOR', (3, i), (3, i), colors.green),
                ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
    
    # Pie de página
    footer_text = f"""
    <i>Este reporte fue generado automáticamente por el sistema de gestión de la clínica dental.<br/>
    Para más información, consulte el sistema en línea.</i>
    """
    footer_para = Paragraph(footer_text, info_style)
    story.append(Spacer(1, 30))
    story.append(footer_para)
    
    # Construir el PDF
    doc.build(story)
    
    # Obtener el contenido del buffer
    pdf_content = buffer.getvalue()
    buffer.close()
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="inventario_insumos_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
    response.write(pdf_content)
    
    return response

# Vista personalizada de logout
def custom_logout(request):
    """Vista personalizada para cerrar sesión"""
    logout(request)
    messages.success(request, 'Has cerrado sesión correctamente.')
    return redirect('login')

# ========== GESTIÓN DE PACIENTES POR DENTISTA ==========

# Vista para que los dentistas gestionen sus pacientes
@login_required
def gestionar_pacientes(request):
    """Vista principal para que los dentistas gestionen sus pacientes"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'Solo los dentistas pueden gestionar pacientes.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    # Filtros de búsqueda
    search = request.GET.get('search', '')
    estado = request.GET.get('estado', '')
    
    # Obtener pacientes asignados al dentista
    pacientes = perfil.get_pacientes_asignados()
    
    # Aplicar filtros
    if search:
        pacientes = [
            paciente for paciente in pacientes
            if (search.lower() in paciente['nombre_completo'].lower() or
                search.lower() in paciente['email'].lower() or
                search.lower() in paciente['telefono'].lower())
        ]
    
    if estado == 'activo':
        pacientes = [paciente for paciente in pacientes if paciente['activo']]
    elif estado == 'inactivo':
        pacientes = [paciente for paciente in pacientes if not paciente['activo']]
    
    # Ordenar por nombre
    pacientes.sort(key=lambda x: x['nombre_completo'])
    
    # Obtener estadísticas
    estadisticas = perfil.get_estadisticas_pacientes()
    
    # Obtener citas recientes de los pacientes
    citas_recientes = perfil.get_citas_pacientes().order_by('-fecha_hora')[:10]
    
    context = {
        'perfil': perfil,
        'pacientes': pacientes,
        'estadisticas': estadisticas,
        'citas_recientes': citas_recientes,
        'search': search,
        'estado': estado,
        'es_dentista': True
    }
    
    return render(request, 'citas/clientes/gestionar_pacientes.html', context)

# Endpoint JSON para obtener estadísticas de un paciente (usado por el modal)
@login_required
def estadisticas_paciente_json(request, paciente_id):
    """Endpoint JSON para obtener estadísticas de un paciente específico"""
    from django.http import JsonResponse
    
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            return JsonResponse({'success': False, 'message': 'Solo los dentistas pueden ver estadísticas de pacientes.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'No tienes permisos para acceder a esta función.'}, status=403)

    # Obtener el paciente desde las citas asignadas al dentista
    pacientes = perfil.get_pacientes_asignados()
    paciente = None
    
    for p in pacientes:
        if p['id'] == int(paciente_id):
            paciente = p
            break
    
    if not paciente:
        return JsonResponse({'success': False, 'message': 'No tienes permisos para ver este paciente.'}, status=403)
    
    # Obtener citas del paciente
    citas_paciente = Cita.objects.filter(
        paciente_email=paciente['email'],
        dentista=perfil
    ).order_by('-fecha_hora')
    
    # Estadísticas del paciente
    ultima_cita = citas_paciente.first()
    proxima_cita = citas_paciente.filter(
        fecha_hora__gte=timezone.now(),
        estado__in=['reservada', 'confirmada']
    ).order_by('fecha_hora').first()
    
    data = {
        'success': True,
        'estadisticas': {
            'total_citas': citas_paciente.count(),
            'citas_completadas': citas_paciente.filter(estado='completada').count(),
            'citas_pendientes': citas_paciente.filter(estado='reservada').count(),
            'citas_canceladas': citas_paciente.filter(estado='cancelada').count(),
        },
        'ultima_cita': None,
        'proxima_cita': None,
    }
    
    if ultima_cita:
        data['ultima_cita'] = {
            'fecha': ultima_cita.fecha_hora.strftime('%d/%m/%Y'),
            'hora': ultima_cita.fecha_hora.strftime('%H:%M'),
            'servicio': str(ultima_cita.tipo_servicio.nombre if ultima_cita.tipo_servicio else ultima_cita.tipo_consulta or 'Consulta general'),
            'estado': ultima_cita.get_estado_display(),
        }
    
    if proxima_cita:
        data['proxima_cita'] = {
            'fecha': proxima_cita.fecha_hora.strftime('%d/%m/%Y'),
            'hora': proxima_cita.fecha_hora.strftime('%H:%M'),
            'servicio': str(proxima_cita.tipo_servicio.nombre if proxima_cita.tipo_servicio else proxima_cita.tipo_consulta or 'Consulta general'),
            'estado': proxima_cita.get_estado_display(),
        }
    
    return JsonResponse(data)

# Vista para que los dentistas vean sus estadísticas de pacientes
@login_required
def estadisticas_pacientes(request):
    """Vista para mostrar estadísticas detalladas de los pacientes del dentista"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'Solo los dentistas pueden ver estadísticas de pacientes.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    # Obtener estadísticas básicas
    estadisticas = perfil.get_estadisticas_pacientes()
    
    # Estadísticas adicionales
    pacientes = perfil.get_pacientes_asignados()
    citas_pacientes = perfil.get_citas_pacientes()
    
    # Citas por mes (últimos 6 meses)
    from datetime import datetime, timedelta
    from django.db.models import Count
    
    hoy = timezone.now().date()
    citas_por_mes = []
    
    for i in range(6):
        fecha_inicio = (hoy.replace(day=1) - timedelta(days=30*i)).replace(day=1)
        fecha_fin = (fecha_inicio + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        citas_mes = citas_pacientes.filter(
            fecha_hora__date__range=[fecha_inicio, fecha_fin]
        ).count()
        
        citas_por_mes.append({
            'mes': fecha_inicio.strftime('%B %Y'),
            'citas': citas_mes
        })
    
    citas_por_mes.reverse()  # Mostrar del más antiguo al más reciente
    
    # Pacientes más activos (top 5)
    pacientes_activos = pacientes.annotate(
        total_citas=Count('citas')
    ).order_by('-total_citas')[:5]
    
    # Tipos de consulta más comunes
    tipos_consulta = citas_pacientes.exclude(
        tipo_consulta__isnull=True
    ).exclude(
        tipo_consulta=''
    ).values('tipo_consulta').annotate(
        total=Count('tipo_consulta')
    ).order_by('-total')[:10]
    
    context = {
        'perfil': perfil,
        'estadisticas': estadisticas,
        'citas_por_mes': citas_por_mes,
        'pacientes_activos': pacientes_activos,
        'tipos_consulta': tipos_consulta,
        'es_dentista': True
    }
    
    return render(request, 'citas/clientes/estadisticas_pacientes.html', context)

# Vista para asignar dentista a un cliente (solo administrativos)
@login_required
def asignar_dentista_cliente(request, cliente_id):
    """Vista para que los administrativos asignen un dentista a un cliente"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'Solo los administrativos pueden asignar dentistas a clientes.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    # Buscar el cliente en el modelo Cliente
    try:
        cliente = Cliente.objects.get(id=cliente_id)
    except Cliente.DoesNotExist:
        messages.error(request, 'Cliente no encontrado.')
        return redirect('gestor_clientes')
    
    if request.method == 'POST':
        dentista_id = request.POST.get('dentista')
        if dentista_id:
            try:
                dentista = Perfil.objects.get(id=dentista_id, rol='dentista', activo=True)
                # Asignar dentista al cliente
                cliente.dentista_asignado = dentista
                cliente.save()
                
                # También actualizar las citas del cliente
                Cita.objects.filter(cliente=cliente).update(dentista=dentista)
                Cita.objects.filter(paciente_email=cliente.email).update(dentista=dentista)
                
                messages.success(request, f'Dentista {dentista.nombre_completo} asignado correctamente al cliente {cliente.nombre_completo}.')
                return redirect('gestor_clientes')
            except Perfil.DoesNotExist:
                messages.error(request, 'Dentista no encontrado.')
        else:
            # Remover dentista asignado del cliente
            cliente.dentista_asignado = None
            cliente.save()
            
            # También remover de las citas del cliente
            Cita.objects.filter(cliente=cliente).update(dentista=None)
            Cita.objects.filter(paciente_email=cliente.email).update(dentista=None)
            
            messages.success(request, f'Dentista removido del cliente {cliente.nombre_completo}.')
            return redirect('gestor_clientes')
    
    # Obtener dentistas disponibles
    dentistas = Perfil.objects.filter(rol='dentista', activo=True).order_by('nombre_completo')
    
    context = {
        'perfil': perfil,
        'cliente': cliente,
        'dentistas': dentistas,
        'es_admin': True
    }
    
    return render(request, 'citas/clientes/asignar_dentista_cliente.html', context)

# ========== GESTIÓN DE ODONTOGRAMAS ==========

# Vista principal para listar odontogramas del dentista
@login_required
def listar_odontogramas(request):
    """Vista para listar todos los odontogramas del dentista"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'Solo los dentistas pueden acceder a los odontogramas.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    # Filtros de búsqueda
    search = request.GET.get('search', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Obtener odontogramas del dentista
    odontogramas = Odontograma.objects.filter(dentista=perfil).select_related('cliente', 'cita')
    
    # Aplicar filtros
    if search:
        odontogramas = odontogramas.filter(
            Q(paciente_nombre__icontains=search) |
            Q(paciente_email__icontains=search) |
            Q(motivo_consulta__icontains=search)
        )
    
    if fecha_desde:
        odontogramas = odontogramas.filter(fecha_creacion__date__gte=fecha_desde)
    if fecha_hasta:
        odontogramas = odontogramas.filter(fecha_creacion__date__lte=fecha_hasta)
    
    odontogramas = odontogramas.order_by('-fecha_creacion')
    
    # Agrupar odontogramas por cliente/paciente
    # Usar email como clave única para agrupar
    pacientes_dict = {}
    
    for odontograma in odontogramas:
        # Determinar clave única: cliente.id si existe, sino email
        if odontograma.cliente:
            clave = f"cliente_{odontograma.cliente.id}"
            nombre_completo = odontograma.cliente.nombre_completo
            email = odontograma.cliente.email
            telefono = odontograma.cliente.telefono or ''
            paciente_id = odontograma.cliente.id
        else:
            clave = f"email_{odontograma.paciente_email}"
            nombre_completo = odontograma.paciente_nombre
            email = odontograma.paciente_email
            telefono = odontograma.paciente_telefono or ''
            paciente_id = None
        
        if clave not in pacientes_dict:
            pacientes_dict[clave] = {
                'id': paciente_id,
                'nombre_completo': nombre_completo,
                'email': email,
                'telefono': telefono,
                'odontogramas': [],
                'total_odontogramas': 0,
                'ultimo_odontograma': None,
            }
        
        pacientes_dict[clave]['odontogramas'].append(odontograma)
        pacientes_dict[clave]['total_odontogramas'] += 1
        
        # Actualizar último odontograma si es más reciente
        if not pacientes_dict[clave]['ultimo_odontograma'] or \
           odontograma.fecha_creacion > pacientes_dict[clave]['ultimo_odontograma'].fecha_creacion:
            pacientes_dict[clave]['ultimo_odontograma'] = odontograma
    
    # Convertir a lista y ordenar por nombre
    pacientes_con_odontogramas = list(pacientes_dict.values())
    pacientes_con_odontogramas.sort(key=lambda x: x['nombre_completo'])
    
    # Estadísticas
    total_odontogramas = odontogramas.count()
    odontogramas_mes = odontogramas.filter(
        fecha_creacion__date__gte=timezone.now().date() - timedelta(days=30)
    ).count()
    
    estadisticas = {
        'total_odontogramas': total_odontogramas,
        'odontogramas_mes': odontogramas_mes,
        'total_pacientes': len(pacientes_con_odontogramas),
    }
    
    context = {
        'perfil': perfil,
        'odontogramas': odontogramas,  # Mantener para compatibilidad
        'pacientes_con_odontogramas': pacientes_con_odontogramas,
        'estadisticas': estadisticas,
        'search': search,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'es_dentista': True
    }
    
    return render(request, 'citas/odontogramas/listar_odontogramas.html', context)

# Vista para crear un nuevo odontograma
@login_required
def crear_odontograma(request):
    """Vista para crear un nuevo odontograma"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'Solo los dentistas pueden crear odontogramas.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    # Obtener paciente_id del GET para redirección después de crear
    paciente_id_redirect = request.GET.get('paciente_id', '')
    
    if request.method == 'POST':
        # Obtener cita asociada (OBLIGATORIA)
        cita_id = request.POST.get('cita_id', '').strip()
        if not cita_id:
            messages.error(request, 'Debes seleccionar una cita para crear la ficha odontológica. Las fichas deben estar vinculadas a una cita específica.')
            if paciente_id_redirect:
                return redirect(f"{reverse('crear_odontograma')}?paciente_id={paciente_id_redirect}")
            return redirect('crear_odontograma')
        
        try:
            cita_obj = Cita.objects.select_related('cliente').get(id=cita_id, dentista=perfil)
        except Cita.DoesNotExist:
            messages.error(request, 'La cita seleccionada no existe o no pertenece a este dentista.')
            return redirect('crear_odontograma')
        
        # Obtener datos del formulario (registro manual)
        paciente_nombre = request.POST.get('paciente_nombre', '').strip()
        paciente_email = request.POST.get('paciente_email', '').strip()
        paciente_telefono = request.POST.get('paciente_telefono', '').strip()
        paciente_fecha_nacimiento = request.POST.get('paciente_fecha_nacimiento', '')
        
        # Intentar vincular con cliente existente si existe (opcional, solo para referencia)
        cliente_obj = None
        if paciente_email:
            try:
                cliente_obj = Cliente.objects.get(email=paciente_email, activo=True)
            except (Cliente.DoesNotExist, Cliente.MultipleObjectsReturned):
                cliente_obj = Cliente.objects.filter(email=paciente_email, activo=True).first()
        
        motivo_consulta = request.POST.get('motivo_consulta', '').strip()
        antecedentes_medicos = request.POST.get('antecedentes_medicos', '').strip()
        alergias = request.POST.get('alergias', '').strip()
        medicamentos_actuales = request.POST.get('medicamentos_actuales', '').strip()
        higiene_oral = request.POST.get('higiene_oral', 'buena')
        estado_general = request.POST.get('estado_general', 'buena')
        observaciones = request.POST.get('observaciones', '').strip()
        plan_tratamiento = request.POST.get('plan_tratamiento', '').strip()
        
        # Si no se proporcionó fecha de nacimiento o alergias, intentar obtenerlas del cliente
        if not paciente_fecha_nacimiento and cliente_obj and cliente_obj.fecha_nacimiento:
            paciente_fecha_nacimiento = cliente_obj.fecha_nacimiento.strftime('%Y-%m-%d')
        
        if not alergias and cliente_obj and cliente_obj.alergias:
            alergias = cliente_obj.alergias
        
        # Validaciones
        errores = []
        
        if not paciente_nombre:
            errores.append('El nombre completo del paciente es obligatorio.')
        
        if not paciente_email:
            errores.append('El email del paciente es obligatorio.')
        
        if not paciente_telefono:
            errores.append('El teléfono del paciente es obligatorio.')
        
        if not paciente_fecha_nacimiento:
            errores.append('La fecha de nacimiento del paciente es obligatoria.')
        
        if not motivo_consulta:
            errores.append('El motivo de consulta es obligatorio.')
        
        # Validar fecha de nacimiento
        fecha_nacimiento_obj = None
        if paciente_fecha_nacimiento:
            try:
                fecha_nacimiento_obj = datetime.strptime(paciente_fecha_nacimiento, '%Y-%m-%d').date()
            except ValueError:
                errores.append('La fecha de nacimiento debe tener un formato válido.')
        
        if errores:
            for error in errores:
                messages.error(request, error)
        else:
            try:
                # Crear odontograma
                odontograma = Odontograma.objects.create(
                    cliente=cliente_obj,  # Cliente del sistema (opcional)
                    cita=cita_obj,  # Cita asociada (OBLIGATORIA)
                    paciente_nombre=paciente_nombre,
                    paciente_email=paciente_email,
                    paciente_telefono=paciente_telefono,
                    paciente_fecha_nacimiento=fecha_nacimiento_obj,
                    dentista=perfil,
                    motivo_consulta=motivo_consulta,
                    antecedentes_medicos=antecedentes_medicos,
                    alergias=alergias,
                    medicamentos_actuales=medicamentos_actuales,
                    higiene_oral=higiene_oral,
                    estado_general=estado_general,
                    observaciones=observaciones,
                    plan_tratamiento=plan_tratamiento
                )
                
                # Procesar datos del odontograma interactivo
                odontograma_data = request.POST.get('odontograma_data', '')
                odontograma_data_extended = request.POST.get('odontograma_data_extended', '')
                
                if odontograma_data:
                    try:
                        import json
                        dientes_data = json.loads(odontograma_data)
                        extended_data = {}
                        
                        # Cargar datos extendidos si existen
                        if odontograma_data_extended:
                            try:
                                extended_data = json.loads(odontograma_data_extended)
                            except:
                                extended_data = {}
                        
                        for numero_diente, caras_data in dientes_data.items():
                            # Determinar el estado general del diente
                            estados = list(caras_data.values())
                            if 'ausente' in estados:
                                estado_general_diente = 'ausente'
                            elif 'caries' in estados:
                                estado_general_diente = 'caries'
                            elif 'obturado' in estados:
                                estado_general_diente = 'obturado'
                            elif 'corona' in estados:
                                estado_general_diente = 'corona'
                            elif 'endodoncia' in estados:
                                estado_general_diente = 'endodoncia'
                            elif 'protesis' in estados:
                                estado_general_diente = 'protesis'
                            elif 'implante' in estados:
                                estado_general_diente = 'implante'
                            elif 'fractura' in estados:
                                estado_general_diente = 'fractura'
                            elif 'sellante' in estados:
                                estado_general_diente = 'sellante'
                            else:
                                estado_general_diente = 'sano'
                            
                            # Preparar observaciones con datos extendidos
                            observaciones_data = {
                                'caras': caras_data
                            }
                            
                            # Agregar metadata (diagnóstico y procedimiento) si existe
                            if numero_diente in extended_data:
                                observaciones_data['metadata'] = extended_data[numero_diente]
                            
                            # Crear estado del diente
                            EstadoDiente.objects.create(
                                odontograma=odontograma,
                                numero_diente=int(numero_diente),
                                estado=estado_general_diente,
                                observaciones=f"Datos del odontograma interactivo: {json.dumps(observaciones_data)}"
                            )
                    except (json.JSONDecodeError, ValueError) as e:
                        messages.warning(request, f'Error al procesar datos del odontograma: {str(e)}')
                
                # Procesar insumos utilizados
                insumos_ids = request.POST.getlist('insumo_id[]')
                insumos_cantidades = request.POST.getlist('insumo_cantidad[]')
                
                insumos_procesados = 0
                insumos_errores = []
                
                for i, insumo_id in enumerate(insumos_ids):
                    # Saltar si el ID está vacío o es solo espacios
                    if not insumo_id or not insumo_id.strip():
                        continue
                        
                    if i < len(insumos_cantidades):
                        cantidad_str = insumos_cantidades[i].strip() if insumos_cantidades[i] else ''
                        
                        # Saltar si la cantidad está vacía
                        if not cantidad_str:
                            continue
                            
                        try:
                            cantidad = int(cantidad_str)
                            if cantidad > 0:
                                # Obtener el insumo
                                from inventario.models import Insumo, MovimientoInsumo
                                from historial_clinico.models import InsumoOdontograma
                                insumo = Insumo.objects.get(id=insumo_id)
                                
                                # Verificar que haya suficiente stock
                                if insumo.cantidad_actual >= cantidad:
                                    # Crear registro de insumo utilizado
                                    InsumoOdontograma.objects.create(
                                        odontograma=odontograma,
                                        insumo=insumo,
                                        cantidad_utilizada=cantidad
                                    )
                                    
                                    # Actualizar cantidad del insumo
                                    cantidad_anterior = insumo.cantidad_actual
                                    insumo.cantidad_actual -= cantidad
                                    
                                    # Actualizar estado si se agotó
                                    if insumo.cantidad_actual == 0:
                                        insumo.estado = 'agotado'
                                    
                                    insumo.save()
                                    
                                    # Crear movimiento de insumo
                                    MovimientoInsumo.objects.create(
                                        insumo=insumo,
                                        tipo='salida',
                                        cantidad=cantidad,
                                        cantidad_anterior=cantidad_anterior,
                                        cantidad_nueva=insumo.cantidad_actual,
                                        motivo=f'Uso en odontograma - Paciente: {paciente_nombre}',
                                        observaciones=f'Odontograma ID: {odontograma.id}',
                                        realizado_por=perfil
                                    )
                                    
                                    insumos_procesados += 1
                                else:
                                    insumos_errores.append(f'{insumo.nombre}: Stock insuficiente (disponible: {insumo.cantidad_actual}, solicitado: {cantidad})')
                        except Insumo.DoesNotExist:
                            insumos_errores.append(f'Insumo con ID {insumo_id} no encontrado')
                        except (ValueError, TypeError):
                            insumos_errores.append(f'Cantidad inválida para el insumo')
                
                # Si la cita estaba en "en_progreso" o "listo_para_atender", cambiar automáticamente a "finalizada"
                if cita_obj.estado in ['en_progreso', 'listo_para_atender']:
                    cita_obj.estado = 'finalizada'
                    cita_obj.save()
                
                # Mensajes de resultado
                if insumos_procesados > 0:
                    if cita_obj.estado == 'finalizada':
                        messages.success(request, f'Odontograma creado correctamente para {paciente_nombre}. Se registraron {insumos_procesados} insumo(s) utilizado(s). La cita ha sido marcada como "Finalizada".')
                    else:
                        messages.success(request, f'Odontograma creado correctamente para {paciente_nombre}. Se registraron {insumos_procesados} insumo(s) utilizado(s).')
                else:
                    if cita_obj.estado == 'finalizada':
                        messages.success(request, f'Odontograma creado correctamente para {paciente_nombre}. La cita ha sido marcada como "Finalizada".')
                    else:
                        messages.success(request, f'Odontograma creado correctamente para {paciente_nombre}.')
                
                if insumos_errores:
                    for error in insumos_errores:
                        messages.warning(request, f'Insumo: {error}')
                
                # Redirigir según el origen
                paciente_id_post = request.POST.get('paciente_id_redirect', paciente_id_redirect)
                if paciente_id_post:
                    # Si viene desde mis_pacientes, redirigir ahí
                    return redirect('mis_pacientes_seccion', paciente_id=paciente_id_post, seccion='odontogramas')
                else:
                    # Si no, redirigir al listado de odontogramas
                    return redirect('listar_odontogramas')
            except Exception as e:
                messages.error(request, f'Error al crear odontograma: {str(e)}')
    
    # Obtener insumos disponibles
    from inventario.models import Insumo
    insumos_disponibles = Insumo.objects.filter(
        estado='disponible',
        cantidad_actual__gt=0
    ).order_by('nombre')
    
    # Inicializar variables para datos pre-seleccionados
    form_data = {}
    cita_pre_seleccionada = None
    cliente_pre_seleccionado = None
    
    if request.method == 'GET':
        cita_id = request.GET.get('cita_id', '')
        paciente_id = request.GET.get('paciente_id', '')
        
        # Si viene paciente_id, buscar el cliente y sus citas
        if paciente_id:
            try:
                cliente_pre_seleccionado = Cliente.objects.get(id=paciente_id, activo=True)
                # Buscar la cita más reciente del paciente con este dentista
                cita_pre_seleccionada = Cita.objects.filter(
                    cliente=cliente_pre_seleccionado,
                    dentista=perfil,
                    estado__in=['reservada', 'completada']
                ).order_by('-fecha_hora').first()
                
                # Si no hay cita, buscar por email
                if not cita_pre_seleccionada and cliente_pre_seleccionado.email:
                    cita_pre_seleccionada = Cita.objects.filter(
                        paciente_email=cliente_pre_seleccionado.email,
                        dentista=perfil,
                        estado__in=['reservada', 'completada']
                    ).order_by('-fecha_hora').first()
            except Cliente.DoesNotExist:
                messages.warning(request, 'El paciente seleccionado no existe.')
        
        # Si viene cita_id, usarlo directamente
        if cita_id and not cita_pre_seleccionada:
            try:
                cita_pre_seleccionada = Cita.objects.select_related('cliente', 'dentista').get(id=cita_id, dentista=perfil)
            except Cita.DoesNotExist:
                messages.warning(request, 'La cita seleccionada no existe o no pertenece a este dentista.')
        
        # Si no hay cita ni paciente, redirigir al listado de odontogramas
        if not cita_pre_seleccionada and not cliente_pre_seleccionado:
            messages.info(request, 'Por favor, selecciona una cita desde el listado para crear la ficha. Las fichas deben estar vinculadas a una cita específica.')
            return redirect('listar_odontogramas')
        
        # Si hay cita pre-seleccionada, auto-rellenar datos
        if cita_pre_seleccionada:
            # Intentar obtener el cliente (prioridad: cliente vinculado > búsqueda por email)
            cliente_obj = cita_pre_seleccionada.cliente
            
            # Si no tiene cliente vinculado pero tiene email, buscar por email
            if not cliente_obj and cita_pre_seleccionada.paciente_email:
                try:
                    cliente_obj = Cliente.objects.get(email=cita_pre_seleccionada.paciente_email, activo=True)
                except (Cliente.DoesNotExist, Cliente.MultipleObjectsReturned):
                    cliente_obj = Cliente.objects.filter(email=cita_pre_seleccionada.paciente_email, activo=True).first()
            
            # Auto-rellenar datos del paciente desde la cita
            # Prioridad: Cliente encontrado (con nombre_completo) > Datos de la cita
            if cliente_obj:
                # Si encontramos cliente, usar SIEMPRE sus datos completos (nombre_completo, no username)
                paciente_nombre_final = cliente_obj.nombre_completo
                paciente_email_final = cliente_obj.email
                paciente_telefono_final = cliente_obj.telefono or ''
                paciente_fecha_nacimiento_final = cliente_obj.fecha_nacimiento.strftime('%Y-%m-%d') if cliente_obj.fecha_nacimiento else ''
                alergias_final = cliente_obj.alergias or ''
            else:
                # Si no hay cliente, usar datos de la cita (paciente de recepción)
                # IMPORTANTE: Si paciente_nombre parece ser un username (sin espacios, corto), 
                # se debe completar manualmente
                paciente_nombre_final = cita_pre_seleccionada.paciente_nombre or ''
                paciente_email_final = cita_pre_seleccionada.paciente_email or ''
                paciente_telefono_final = cita_pre_seleccionada.paciente_telefono or ''
                # La cita no tiene fecha de nacimiento, se debe completar manualmente
                paciente_fecha_nacimiento_final = ''
                alergias_final = ''
            
            form_data = {
                'paciente_nombre': paciente_nombre_final,
                'paciente_email': paciente_email_final,
                'paciente_telefono': paciente_telefono_final,
                'paciente_fecha_nacimiento': paciente_fecha_nacimiento_final,
                'alergias': alergias_final,
                'cita_id': str(cita_pre_seleccionada.id),
            }
        # Si hay cliente pre-seleccionado pero no cita, auto-rellenar datos del cliente
        elif cliente_pre_seleccionado:
            form_data = {
                'paciente_nombre': cliente_pre_seleccionado.nombre_completo,
                'paciente_email': cliente_pre_seleccionado.email,
                'paciente_telefono': cliente_pre_seleccionado.telefono or '',
                'paciente_fecha_nacimiento': cliente_pre_seleccionado.fecha_nacimiento.strftime('%Y-%m-%d') if cliente_pre_seleccionado.fecha_nacimiento else '',
                'alergias': cliente_pre_seleccionado.alergias or '',
            }
    elif request.method == 'POST':
        form_data = request.POST
    
    # Obtener citas del dentista (reservadas o completadas) para asociar con la ficha
    # Si hay cliente pre-seleccionado, filtrar sus citas primero
    if cliente_pre_seleccionado:
        citas_disponibles = Cita.objects.filter(
            dentista=perfil,
            estado__in=['reservada', 'completada'],
            cliente=cliente_pre_seleccionado
        ).select_related('cliente').order_by('-fecha_hora')[:50]
        # Si no hay citas del cliente, incluir todas las citas del dentista
        if not citas_disponibles.exists():
            citas_disponibles = Cita.objects.filter(
                dentista=perfil,
                estado__in=['reservada', 'completada']
            ).select_related('cliente').order_by('-fecha_hora')[:50]
    else:
        citas_disponibles = Cita.objects.filter(
            dentista=perfil,
            estado__in=['reservada', 'completada']
        ).select_related('cliente').order_by('-fecha_hora')[:50]  # Últimas 50 citas
    
    context = {
        'perfil': perfil,
        'condiciones': Odontograma.CONDICION_CHOICES,
        'es_dentista': True,
        'form_data': form_data,
        'insumos_disponibles': insumos_disponibles,
        'cita_pre_seleccionada': cita_pre_seleccionada,
        'citas_disponibles': citas_disponibles
    }
    
    return render(request, 'citas/odontogramas/crear_odontograma.html', context)

# Vista para ver detalles de un odontograma
@login_required
def detalle_odontograma(request, odontograma_id):
    """Vista para ver los detalles de un odontograma específico"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not (perfil.es_dentista() or perfil.es_administrativo()):
            messages.error(request, 'Solo los dentistas y administrativos pueden ver odontogramas.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    # Los administrativos pueden ver cualquier odontograma, los dentistas solo los suyos
    if perfil.es_administrativo():
        odontograma = get_object_or_404(Odontograma, id=odontograma_id)
    else:
        odontograma = get_object_or_404(Odontograma, id=odontograma_id, dentista=perfil)
    
    # Obtener estados de los dientes
    estados_dientes = odontograma.dientes.all().order_by('numero_diente')
    
    # Crear diccionario de dientes para facilitar el acceso
    dientes_dict = {diente.numero_diente: diente for diente in estados_dientes}
    
    # Números de dientes según numeración FDI (adultos)
    numeros_dientes_adultos = [
        18, 17, 16, 15, 14, 13, 12, 11, 21, 22, 23, 24, 25, 26, 27, 28,  # Superior
        48, 47, 46, 45, 44, 43, 42, 41, 31, 32, 33, 34, 35, 36, 37, 38   # Inferior
    ]
    
    # Determinar URL de retorno
    # Prioridad: parámetro return > nueva vista Mis Pacientes (si es dentista) > referer > listar odontogramas
    referer = request.META.get('HTTP_REFERER', '')
    return_url = request.GET.get('return', '')
    
    if return_url:
        # Si se pasó un parámetro return, usarlo directamente
        url_retorno = return_url
    elif perfil.es_dentista():
        # Si es dentista, intentar redirigir a Mis Pacientes
        paciente_id = None
        if odontograma.cliente:
            paciente_id = odontograma.cliente.id
        elif odontograma.paciente_email:
            # Buscar cliente por email
            try:
                cliente = Cliente.objects.get(email=odontograma.paciente_email, activo=True)
                paciente_id = cliente.id
            except Cliente.DoesNotExist:
                # Si no hay cliente, usar hash del email
                paciente_id = hash(odontograma.paciente_email) % 1000000
        
        if paciente_id:
            url_retorno = reverse('mis_pacientes_seccion', args=[paciente_id, 'odontogramas'])
        elif referer and referer.startswith(request.build_absolute_uri('/')[:-1]):
            url_retorno = referer
        else:
            url_retorno = reverse('listar_odontogramas')
    elif odontograma.cliente:
        # Si el odontograma tiene un cliente asociado, volver al perfil del cliente
        url_retorno = reverse('perfil_cliente', args=[odontograma.cliente.id])
    elif referer and referer.startswith(request.build_absolute_uri('/')[:-1]):
        # Si el referer es de nuestro sitio, usarlo
        url_retorno = referer
    else:
        # Por defecto, volver a listar odontogramas
        url_retorno = reverse('listar_odontogramas')
    
    context = {
        'perfil': perfil,
        'odontograma': odontograma,
        'estados_dientes': estados_dientes,
        'dientes_dict': dientes_dict,
        'numeros_dientes': numeros_dientes_adultos,
        'estados_diente': Odontograma.ESTADO_DIENTE_CHOICES,
        'es_dentista': perfil.es_dentista(),
        'es_admin': perfil.es_administrativo(),
        'url_retorno': url_retorno
    }
    
    return render(request, 'citas/odontogramas/detalle_odontograma.html', context)

# Vista para editar un odontograma
@login_required
def editar_odontograma(request, odontograma_id):
    """Vista para editar un odontograma existente"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'Solo los dentistas pueden editar odontogramas.')
            # Redirigir de vuelta a la vista de detalle del odontograma para mostrar el mensaje
            return redirect('detalle_odontograma', odontograma_id=odontograma_id)
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    # Los administrativos pueden ver el odontograma pero no editarlo
    if perfil.es_administrativo():
        odontograma = get_object_or_404(Odontograma, id=odontograma_id)
    else:
        odontograma = get_object_or_404(Odontograma, id=odontograma_id, dentista=perfil)
    
    if request.method == 'POST':
        # Actualizar cita asociada (opcional)
        cita_id = request.POST.get('cita_id', '').strip()
        if cita_id:
            try:
                cita_obj = Cita.objects.get(id=cita_id, dentista=perfil)
                odontograma.cita = cita_obj
            except Cita.DoesNotExist:
                messages.warning(request, 'La cita seleccionada no existe o no pertenece a este dentista.')
        else:
            # Si no se selecciona ninguna cita, dejar en None
            odontograma.cita = None
        
        # Actualizar datos del odontograma
        odontograma.paciente_nombre = request.POST.get('paciente_nombre', '').strip()
        odontograma.paciente_email = request.POST.get('paciente_email', '').strip()
        odontograma.paciente_telefono = request.POST.get('paciente_telefono', '').strip()
        odontograma.motivo_consulta = request.POST.get('motivo_consulta', '').strip()
        odontograma.antecedentes_medicos = request.POST.get('antecedentes_medicos', '').strip()
        odontograma.alergias = request.POST.get('alergias', '').strip()
        odontograma.medicamentos_actuales = request.POST.get('medicamentos_actuales', '').strip()
        odontograma.higiene_oral = request.POST.get('higiene_oral', 'buena')
        odontograma.estado_general = request.POST.get('estado_general', 'buena')
        odontograma.observaciones = request.POST.get('observaciones', '').strip()
        odontograma.plan_tratamiento = request.POST.get('plan_tratamiento', '').strip()
        
        # Importar datetime una vez para usar en ambas secciones
        from datetime import datetime
        
        # Actualizar fecha de nacimiento
        paciente_fecha_nacimiento = request.POST.get('paciente_fecha_nacimiento', '')
        if paciente_fecha_nacimiento:
            try:
                odontograma.paciente_fecha_nacimiento = datetime.strptime(paciente_fecha_nacimiento, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, 'La fecha de nacimiento debe tener un formato válido.')
                return redirect('editar_odontograma', odontograma_id=odontograma.id)
        
        try:
            odontograma.save()
            # Procesar datos interactivos si vienen del formulario de edición
            odontograma_data = request.POST.get('odontograma_data', '')
            odontograma_data_extended = request.POST.get('odontograma_data_extended', '')
            
            if odontograma_data:
                try:
                    import json
                    dientes_data = json.loads(odontograma_data)
                    extended_data = {}
                    
                    # Cargar datos extendidos si existen
                    if odontograma_data_extended:
                        try:
                            extended_data = json.loads(odontograma_data_extended)
                        except:
                            extended_data = {}
                    
                    from django.db import transaction
                    with transaction.atomic():
                        for numero_diente, caras_data in dientes_data.items():
                            estados = list(caras_data.values())
                            if 'ausente' in estados:
                                estado_general_diente = 'perdido'
                            elif 'caries' in estados:
                                estado_general_diente = 'cariado'
                            elif 'obturado' in estados:
                                estado_general_diente = 'obturado'
                            elif 'corona' in estados:
                                estado_general_diente = 'corona'
                            elif 'endodoncia' in estados:
                                estado_general_diente = 'endodoncia'
                            elif 'protesis' in estados:
                                estado_general_diente = 'protesis'
                            elif 'implante' in estados:
                                estado_general_diente = 'implante'
                            elif 'fractura' in estados:
                                estado_general_diente = 'extraccion' if 'extraccion' in estados else 'cariado'
                            elif 'sellante' in estados:
                                estado_general_diente = 'obturado'
                            else:
                                estado_general_diente = 'sano'
                            
                            # Preparar observaciones con datos extendidos
                            observaciones_data = {
                                'caras': caras_data
                            }
                            
                            # Agregar metadata (diagnóstico y procedimiento) si existe
                            if numero_diente in extended_data:
                                observaciones_data['metadata'] = extended_data[numero_diente]
                            
                            # Upsert EstadoDiente
                            estado_obj, _ = EstadoDiente.objects.get_or_create(
                                odontograma=odontograma,
                                numero_diente=int(numero_diente),
                                defaults={'estado': estado_general_diente}
                            )
                            estado_obj.estado = estado_general_diente
                            estado_obj.observaciones = f"Datos del odontograma interactivo: {json.dumps(observaciones_data)}"
                            estado_obj.save()
                except Exception as e:
                    messages.warning(request, f'Error al procesar datos del odontograma: {str(e)}')

            messages.success(request, 'Odontograma actualizado correctamente.')
            # Obtener paciente_id para redirigir a Mis Pacientes
            paciente_id = None
            if odontograma.cliente:
                paciente_id = odontograma.cliente.id
            elif odontograma.paciente_email:
                # Buscar cliente por email
                try:
                    cliente = Cliente.objects.get(email=odontograma.paciente_email, activo=True)
                    paciente_id = cliente.id
                except Cliente.DoesNotExist:
                    # Si no hay cliente, usar hash del email
                    paciente_id = hash(odontograma.paciente_email) % 1000000
            
            if paciente_id:
                return redirect('mis_pacientes_seccion', paciente_id=paciente_id, seccion='odontogramas')
            else:
                return redirect('listar_odontogramas')
        except Exception as e:
            messages.error(request, f'Error al actualizar odontograma: {str(e)}')
    
    # Obtener citas del dentista (reservadas o completadas) para asociar con la ficha
    citas_disponibles = Cita.objects.filter(
        dentista=perfil,
        estado__in=['reservada', 'completada']
    ).order_by('-fecha_hora')[:50]  # Últimas 50 citas
    
    # Obtener paciente_id para redirección
    paciente_id = None
    if odontograma.cliente:
        paciente_id = odontograma.cliente.id
    elif odontograma.paciente_email:
        # Buscar cliente por email
        try:
            cliente = Cliente.objects.get(email=odontograma.paciente_email, activo=True)
            paciente_id = cliente.id
        except Cliente.DoesNotExist:
            # Si no hay cliente, usar hash del email
            paciente_id = hash(odontograma.paciente_email) % 1000000
    
    context = {
        'perfil': perfil,
        'odontograma': odontograma,
        'condiciones': Odontograma.CONDICION_CHOICES,
        'es_dentista': True,
        'citas_disponibles': citas_disponibles,
        'paciente_id': paciente_id
    }
    
    return render(request, 'citas/odontogramas/editar_odontograma.html', context)

# Vista para actualizar el estado de un diente
@login_required
def actualizar_diente(request, odontograma_id, numero_diente):
    """Vista para actualizar el estado de un diente específico"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'Solo los dentistas pueden actualizar estados de dientes.')
            # Redirigir de vuelta a la vista de detalle del odontograma
            return redirect('detalle_odontograma', odontograma_id=odontograma_id)
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    odontograma = get_object_or_404(Odontograma, id=odontograma_id, dentista=perfil)
    
    if request.method == 'POST':
        estado = request.POST.get('estado', 'sano')
        observaciones = request.POST.get('observaciones', '').strip()
        fecha_tratamiento = request.POST.get('fecha_tratamiento', '')
        costo_tratamiento = request.POST.get('costo_tratamiento', '')
        
        # Validar fecha de tratamiento
        fecha_tratamiento_obj = None
        if fecha_tratamiento:
            try:
                from datetime import datetime
                fecha_tratamiento_obj = datetime.strptime(fecha_tratamiento, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, 'La fecha de tratamiento debe tener un formato válido.')
                return redirect('detalle_odontograma', odontograma_id=odontograma.id)
        
        # Validar costo
        costo_obj = None
        if costo_tratamiento:
            try:
                costo_obj = float(costo_tratamiento)
                if costo_obj < 0:
                    messages.error(request, 'El costo no puede ser negativo.')
                    return redirect('detalle_odontograma', odontograma_id=odontograma.id)
            except ValueError:
                messages.error(request, 'El costo debe ser un número válido.')
                return redirect('detalle_odontograma', odontograma_id=odontograma.id)
        
        try:
            # Obtener o crear el estado del diente
            estado_diente, created = EstadoDiente.objects.get_or_create(
                odontograma=odontograma,
                numero_diente=numero_diente,
                defaults={
                    'estado': estado,
                    'observaciones': observaciones,
                    'fecha_tratamiento': fecha_tratamiento_obj,
                    'costo_tratamiento': costo_obj
                }
            )
            
            if not created:
                estado_diente.estado = estado
                estado_diente.observaciones = observaciones
                estado_diente.fecha_tratamiento = fecha_tratamiento_obj
                estado_diente.costo_tratamiento = costo_obj
                estado_diente.save()
            
            messages.success(request, f'Estado del diente {numero_diente} actualizado correctamente.')
            return redirect('detalle_odontograma', odontograma_id=odontograma.id)
        except Exception as e:
            messages.error(request, f'Error al actualizar el diente: {str(e)}')
    
    # Obtener el estado actual del diente
    try:
        estado_diente = EstadoDiente.objects.get(odontograma=odontograma, numero_diente=numero_diente)
    except EstadoDiente.DoesNotExist:
        estado_diente = None
    
    context = {
        'perfil': perfil,
        'odontograma': odontograma,
        'numero_diente': numero_diente,
        'estado_diente': estado_diente,
        'estados_diente': Odontograma.ESTADO_DIENTE_CHOICES,
        'es_dentista': True
    }
    
    return render(request, 'citas/odontogramas/actualizar_diente.html', context)

# Vista para eliminar un odontograma
@login_required
def eliminar_odontograma(request, odontograma_id):
    """Vista para eliminar un odontograma"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'Solo los dentistas pueden eliminar odontogramas.')
            # Redirigir de vuelta a la vista de detalle del odontograma
            return redirect('detalle_odontograma', odontograma_id=odontograma_id)
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    odontograma = get_object_or_404(Odontograma, id=odontograma_id, dentista=perfil)
    
    if request.method == 'POST':
        paciente_nombre = odontograma.paciente_nombre
        try:
            odontograma.delete()
            mensaje = f'Ficha odontológica de {paciente_nombre} eliminada exitosamente.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': True, 'message': mensaje})
            messages.success(request, mensaje)
            return redirect('listar_odontogramas')
        except Exception as e:
            error_msg = f'Error al eliminar odontograma: {str(e)}'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'error': error_msg}, status=500)
            messages.error(request, error_msg)
    
    context = {
        'perfil': perfil,
        'odontograma': odontograma,
        'es_dentista': True
    }
    
    return render(request, 'citas/odontogramas/eliminar_odontograma.html', context)

# Vista para exportar odontograma a PDF
@login_required
def exportar_odontograma_pdf(request, odontograma_id):
    """Vista para exportar un odontograma a PDF"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        # Permitir tanto a dentistas como a administrativos
        if not (perfil.es_dentista() or perfil.es_administrativo()):
            messages.error(request, 'No tienes permisos para exportar odontogramas.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    # Los administrativos pueden ver cualquier odontograma, los dentistas solo los suyos
    if perfil.es_administrativo():
        odontograma = get_object_or_404(Odontograma, id=odontograma_id)
    else:
        odontograma = get_object_or_404(Odontograma, id=odontograma_id, dentista=perfil)
    
    # Obtener estados de los dientes
    estados_dientes = odontograma.dientes.all().order_by('numero_diente')
    
    # Crear diccionario de dientes para facilitar el acceso
    dientes_dict = {diente.numero_diente: diente for diente in estados_dientes}
    
    # Función para extraer datos del odontograma interactivo desde observaciones
    def get_tooth_interactive_data(estado_diente):
        """Extrae los datos de las caras del odontograma interactivo desde observaciones"""
        if not estado_diente or not estado_diente.observaciones:
            return None
        if estado_diente.observaciones.startswith('Datos del odontograma interactivo: '):
            try:
                import json
                json_str = estado_diente.observaciones.replace('Datos del odontograma interactivo: ', '')
                data = json.loads(json_str)
                # Si tiene estructura con 'caras', devolver las caras directamente
                if isinstance(data, dict) and 'caras' in data:
                    return data['caras']
                # Si es directamente un diccionario de caras, devolverlo
                elif isinstance(data, dict):
                    return data
                return None
            except:
                return None
        return None
    
    # Crear el buffer para el PDF
    buffer = BytesIO()
    
    # Crear el documento PDF con márgenes mejorados
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=25,
        leftMargin=25,
        topMargin=40,
        bottomMargin=30
    )
    
    # Estilos mejorados
    styles = getSampleStyleSheet()
    
    # Estilo personalizado para el título principal
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e293b'),
        fontName='Helvetica-Bold'
    )
    
    # Estilo para subtítulos con color turquesa
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=11,
        spaceAfter=8,
        spaceBefore=12,
        alignment=TA_LEFT,
        textColor=colors.HexColor('#1e293b'),
        fontName='Helvetica-Bold',
        borderColor=colors.HexColor('#14b8a6'),
        borderWidth=1,
        borderPadding=6,
        backColor=colors.HexColor('#f0fdfa')
    )
    
    # Estilo para información de la clínica
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#64748b')
    )
    
    # Estilo para texto normal
    normal_style = ParagraphStyle(
        'NormalCompact',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=4,
        alignment=TA_LEFT,
        leading=12
    )
    
    # Estilo para texto de secciones
    section_text_style = ParagraphStyle(
        'SectionText',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=6,
        alignment=TA_LEFT,
        leading=13,
        textColor=colors.HexColor('#374151')
    )
    
    # Contenido del PDF
    story = []
    
    # Título principal con diseño mejorado
    title = Paragraph("<b>FICHA ODONTOLÓGICA</b>", title_style)
    story.append(title)
    
    # Información de la clínica y fecha
    clinic_info = Paragraph(
        f"<b>Clínica Dental</b> | Fecha de Emisión: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        info_style
    )
    story.append(clinic_info)
    story.append(Spacer(1, 12))
    
    # Nota introductoria breve para el paciente con color turquesa
    nota_intro = Paragraph(
        "<b>NOTA:</b><br/>"
        "Este documento contiene el registro de su salud dental. "
        "El odontograma muestra la condición de cada diente. "
        "Consulte la leyenda al final para entender los símbolos y colores.",
        ParagraphStyle(
            'NotaIntro',
            parent=styles['Normal'],
            fontSize=8,
            spaceAfter=10,
            alignment=TA_LEFT,
            textColor=colors.HexColor('#374151'),
            leading=12,
            leftIndent=0,
            rightIndent=0,
            backColor=colors.HexColor('#f0fdfa'),
            borderPadding=8,
            borderColor=colors.HexColor('#14b8a6'),
            borderWidth=1
        )
    )
    story.append(nota_intro)
    story.append(Spacer(1, 12))
    
    # Sección: Información del Paciente
    paciente_title = Paragraph("<b>INFORMACIÓN DEL PACIENTE</b>", subtitle_style)
    story.append(paciente_title)
    
    # Tabla de información del paciente (solo datos esenciales)
    paciente_data = [
        ['Nombre:', odontograma.paciente_nombre],
        ['Email:', odontograma.paciente_email],
        ['Teléfono:', odontograma.paciente_telefono or 'No especificado']
    ]
    if odontograma.paciente_fecha_nacimiento:
        paciente_data.append(['Fecha de Nacimiento:', odontograma.paciente_fecha_nacimiento.strftime('%d/%m/%Y')])
    paciente_table = Table(paciente_data, colWidths=[2*inch, 4.5*inch])
    paciente_table_style = TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0fdfa')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1e293b')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#374151')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ccfbf1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ])
    paciente_table.setStyle(paciente_table_style)
    story.append(paciente_table)
    story.append(Spacer(1, 12))
    
    # Sección: Información Clínica
    clinica_title = Paragraph("<b>INFORMACIÓN CLÍNICA</b>", subtitle_style)
    story.append(clinica_title)
    
    # Motivo de consulta (solo si existe)
    if odontograma.motivo_consulta:
        motivo_text = Paragraph(f"<b>Motivo de Consulta:</b> {odontograma.motivo_consulta}", section_text_style)
        story.append(motivo_text)
        story.append(Spacer(1, 4))
    
    # Estado e higiene en tabla (solo si hay datos)
    estado_data = []
    if odontograma.higiene_oral:
        estado_data.append(['Higiene Oral:', odontograma.get_higiene_oral_display()])
    if odontograma.estado_general:
        estado_data.append(['Estado General:', odontograma.get_estado_general_display()])
    if estado_data:
        estado_table = Table(estado_data, colWidths=[2*inch, 4.5*inch])
        estado_table_style = TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0fdfa')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1e293b')),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ccfbf1')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ])
        estado_table.setStyle(estado_table_style)
        story.append(estado_table)
        story.append(Spacer(1, 6))
    
    # Antecedentes médicos (solo si existe)
    if odontograma.antecedentes_medicos:
        antecedentes_text = Paragraph(f"<b>Antecedentes Médicos:</b> {odontograma.antecedentes_medicos}", section_text_style)
        story.append(antecedentes_text)
        story.append(Spacer(1, 4))
    
    # Alergias (solo si existe)
    if odontograma.alergias:
        alergias_text = Paragraph(f"<b>Alergias:</b> {odontograma.alergias}", section_text_style)
        story.append(alergias_text)
        story.append(Spacer(1, 4))
    
    # Medicamentos actuales (solo si existe)
    if odontograma.medicamentos_actuales:
        medicamentos_text = Paragraph(f"<b>Medicamentos Actuales:</b> {odontograma.medicamentos_actuales}", section_text_style)
        story.append(medicamentos_text)
        story.append(Spacer(1, 6))
    
    # Sección: Odontograma Visual
    odontograma_title = Paragraph("<b>ODONTOGRAMA DENTAL</b>", subtitle_style)
    story.append(odontograma_title)
    
    # Explicación clara del odontograma
    explicacion_odontograma = Paragraph(
        "<b>¿Cómo leer este odontograma?</b><br/>"
        "La boca se divide en 4 cuadrantes (superior derecho, superior izquierdo, inferior izquierdo, inferior derecho). "
        "Cada diente tiene un número único según la numeración internacional (FDI). "
        "Los símbolos y colores indican el estado de cada diente. Consulte la leyenda al final para entender cada símbolo.",
        ParagraphStyle(
            'ExplicacionOdontograma',
            parent=styles['Normal'],
            fontSize=9,
            spaceAfter=8,
            alignment=TA_LEFT,
            textColor=colors.HexColor('#374151'),
            leading=13,
            leftIndent=0,
            backColor=colors.HexColor('#f0fdfa'),
            borderPadding=8,
            borderColor=colors.HexColor('#ccfbf1'),
            borderWidth=1
        )
    )
    story.append(explicacion_odontograma)
    story.append(Spacer(1, 6))
    
    # Función para obtener color según estado
    def get_tooth_color(estado):
        # Asegurarse de que estado sea un string
        if not isinstance(estado, str):
            if isinstance(estado, dict):
                return colors.HexColor('#f3f4f6')  # Gris claro por defecto
            estado = str(estado) if estado else 'sano'
        
        color_map = {
            'sano': colors.HexColor('#10b981'),      # Verde
            'cariado': colors.HexColor('#dc2626'),    # Rojo
            'caries': colors.HexColor('#dc2626'),     # Rojo
            'obturado': colors.HexColor('#f59e0b'),  # Amarillo
            'corona': colors.HexColor('#fbbf24'),    # Dorado
            'perdido': colors.HexColor('#6b7280'),   # Gris
            'ausente': colors.HexColor('#6b7280'),    # Gris
            'endodoncia': colors.HexColor('#0ea5e9'), # Azul
            'protesis': colors.HexColor('#8b5cf6'),  # Púrpura
            'implante': colors.HexColor('#06b6d4'),  # Cian
            'sellante': colors.HexColor('#84cc16'),   # Verde claro
            'fractura': colors.HexColor('#ef4444'),  # Rojo oscuro
            'extraccion': colors.HexColor('#991b1b'), # Rojo muy oscuro
        }
        return color_map.get(estado.lower() if estado else 'sano', colors.HexColor('#f3f4f6'))  # Gris claro por defecto
    
    # Función para obtener el estado principal del diente (considerando datos interactivos)
    def get_tooth_main_state(numero_diente):
        estado_diente = dientes_dict.get(numero_diente)
        if not estado_diente:
            return 'sano', None
        
        # Intentar obtener datos interactivos
        interactive_data = get_tooth_interactive_data(estado_diente)
        
        if interactive_data:
            # Si hay datos interactivos, determinar el estado principal
            estados = list(interactive_data.values())
            if 'ausente' in estados or 'perdido' in estados:
                return 'ausente', interactive_data
            elif 'caries' in estados or 'cariado' in estados:
                return 'caries', interactive_data
            elif 'obturado' in estados:
                return 'obturado', interactive_data
            elif 'corona' in estados:
                return 'corona', interactive_data
            elif 'endodoncia' in estados:
                return 'endodoncia', interactive_data
            elif 'protesis' in estados:
                return 'protesis', interactive_data
            elif 'implante' in estados:
                return 'implante', interactive_data
            elif 'fractura' in estados:
                return 'fractura', interactive_data
            elif 'sellante' in estados:
                return 'sellante', interactive_data
            else:
                return 'sano', interactive_data
        
        # Si no hay datos interactivos, usar el estado general
        # Asegurarse de que siempre devolvamos un string
        estado = estado_diente.estado
        if isinstance(estado, dict):
            # Si el estado es un diccionario, usar 'sano' por defecto
            return 'sano', None
        return str(estado) if estado else 'sano', None
    
    # Función para obtener símbolo según estado
    def get_tooth_symbol(estado):
        # Asegurarse de que estado sea un string
        if not isinstance(estado, str):
            if isinstance(estado, dict):
                return '?'
            estado = str(estado) if estado else 'sano'
        
        symbol_map = {
            'sano': '✓',
            'caries': '●',
            'cariado': '●',
            'obturado': '■',
            'corona': '◊',
            'perdido': '✕',
            'ausente': '✕',
            'endodoncia': '◐',
            'protesis': '◈',
            'implante': '◉',
            'sellante': '◯',
            'fractura': '◢',
            'extraccion': '✕',
        }
        return symbol_map.get(estado.lower() if estado else 'sano', '?')
    
    # Función para obtener texto del diente (simplificado - solo número y símbolo principal)
    def get_tooth_display(numero_diente, estado_val, interactive_data):
        # Mostrar solo número de diente y símbolo principal - sin duplicar información
        symbol = get_tooth_symbol(estado_val)
        return f"{numero_diente}\n{symbol}"
    
    # Crear estructura del odontograma anatómico
    odontograma_data = []
    
    # Encabezado con nombres de dientes (más claro)
    header_row = ['CUADRANTE', 'Molar 3', 'Molar 2', 'Molar 1', 'Premolar 2', 'Premolar 1', 'Canino', 'Incisivo 2', 'Incisivo 1']
    odontograma_data.append(header_row)
    
    # Cuadrante Superior Derecho (vista del paciente)
    superior_derecho = ['SUPERIOR\nDERECHO']
    for numero in [18, 17, 16, 15, 14, 13, 12, 11]:
        estado_val, interactive_data = get_tooth_main_state(numero)
        display_text = get_tooth_display(numero, estado_val, interactive_data)
        superior_derecho.append(display_text)
    odontograma_data.append(superior_derecho)
    
    # Cuadrante Superior Izquierdo (vista del paciente)
    superior_izquierdo = ['SUPERIOR\nIZQUIERDO']
    for numero in [21, 22, 23, 24, 25, 26, 27, 28]:
        estado_val, interactive_data = get_tooth_main_state(numero)
        display_text = get_tooth_display(numero, estado_val, interactive_data)
        superior_izquierdo.append(display_text)
    odontograma_data.append(superior_izquierdo)
    
    # Separador visual
    odontograma_data.append(['─' * 10, '─' * 10, '─' * 10, '─' * 10, '─' * 10, '─' * 10, '─' * 10, '─' * 10, '─' * 10])
    
    # Cuadrante Inferior Izquierdo (vista del paciente)
    inferior_izquierdo = ['INFERIOR\nIZQUIERDO']
    for numero in [38, 37, 36, 35, 34, 33, 32, 31]:
        estado_val, interactive_data = get_tooth_main_state(numero)
        display_text = get_tooth_display(numero, estado_val, interactive_data)
        inferior_izquierdo.append(display_text)
    odontograma_data.append(inferior_izquierdo)
    
    # Cuadrante Inferior Derecho (vista del paciente)
    inferior_derecho = ['INFERIOR\nDERECHO']
    for numero in [41, 42, 43, 44, 45, 46, 47, 48]:
        estado_val, interactive_data = get_tooth_main_state(numero)
        display_text = get_tooth_display(numero, estado_val, interactive_data)
        inferior_derecho.append(display_text)
    odontograma_data.append(inferior_derecho)
    
    # Crear tabla del odontograma mejorada (con más espacio para legibilidad)
    odontograma_table = Table(odontograma_data, colWidths=[1*inch, 0.75*inch, 0.75*inch, 0.75*inch, 0.75*inch, 0.75*inch, 0.75*inch, 0.75*inch, 0.75*inch])
    
    # Estilo de la tabla del odontograma mejorada con colores turquesa
    table_style = TableStyle([
        # Encabezado con color turquesa
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#14b8a6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Cuadrantes - colores turquesa diferenciados
        ('BACKGROUND', (0, 1), (0, 1), colors.HexColor('#14b8a6')),  # Sup Der
        ('BACKGROUND', (0, 2), (0, 2), colors.HexColor('#0d9488')),  # Sup Izq
        ('BACKGROUND', (0, 4), (0, 4), colors.HexColor('#5eead4')),  # Inf Izq
        ('BACKGROUND', (0, 5), (0, 5), colors.HexColor('#2dd4bf')),  # Inf Der
        
        # Texto de cuadrantes
        ('TEXTCOLOR', (0, 1), (0, 5), colors.white),
        ('FONTNAME', (0, 1), (0, 5), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (0, 5), 8),
        
        # Dientes - texto más grande y legible
        ('FONTNAME', (1, 1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (1, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#14b8a6')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (1, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (1, 1), (-1, -1), 8),
        ('LEFTPADDING', (1, 1), (-1, -1), 6),
        ('RIGHTPADDING', (1, 1), (-1, -1), 6),
    ])
    
    # Aplicar colores específicos a dientes según su estado
    for row_idx in range(1, len(odontograma_data)):
        if row_idx == 3:  # Fila separadora
            continue
        
        for col_idx in range(1, len(odontograma_data[row_idx])):
            # Determinar el número del diente
            if row_idx == 1:  # Superior derecho
                tooth_nums = [18, 17, 16, 15, 14, 13, 12, 11]
            elif row_idx == 2:  # Superior izquierdo
                tooth_nums = [21, 22, 23, 24, 25, 26, 27, 28]
            elif row_idx == 4:  # Inferior izquierdo
                tooth_nums = [38, 37, 36, 35, 34, 33, 32, 31]
            elif row_idx == 5:  # Inferior derecho
                tooth_nums = [41, 42, 43, 44, 45, 46, 47, 48]
            else:
                continue
            
            if col_idx - 1 < len(tooth_nums):
                tooth_num = tooth_nums[col_idx - 1]
                estado_val, interactive_data = get_tooth_main_state(tooth_num)
                color = get_tooth_color(estado_val)
                table_style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), color)
                # Usar texto blanco solo si el color es oscuro
                if estado_val in ['ausente', 'perdido', 'caries', 'cariado', 'fractura', 'extraccion']:
                    table_style.add('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.white)
                else:
                    table_style.add('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#1e293b'))
    
    # Estilo para la fila separadora con color turquesa
    table_style.add('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#ccfbf1'))
    table_style.add('TEXTCOLOR', (0, 3), (-1, 3), colors.HexColor('#64748b'))
    table_style.add('FONTSIZE', (0, 3), (-1, 3), 6)
    
    odontograma_table.setStyle(table_style)
    story.append(odontograma_table)
    story.append(Spacer(1, 12))
    
    # Resumen de estado dental
    estados_count = {}
    for diente in estados_dientes:
        estado_val, interactive_data = get_tooth_main_state(diente.numero_diente)
        estados_count[estado_val] = estados_count.get(estado_val, 0) + 1
    
    if estados_count:
        resumen_title = Paragraph("<b>RESUMEN DEL ESTADO DENTAL</b>", subtitle_style)
        story.append(resumen_title)
        
        resumen_data = [['Estado Dental', 'Cantidad', 'Explicación']]
        estado_descriptions = {
            'sano': 'Diente en perfecto estado, sin problemas',
            'caries': 'Diente con caries que necesita tratamiento',
            'obturado': 'Diente ya tratado con empaste',
            'corona': 'Diente con corona o funda protectora',
            'ausente': 'Diente que falta o fue extraído',
            'endodoncia': 'Diente con tratamiento de conducto (nervio tratado)',
            'protesis': 'Diente con prótesis o funda',
            'implante': 'Diente reemplazado con implante dental',
            'sellante': 'Diente con sellante preventivo',
            'fractura': 'Diente con fractura o grieta',
        }
        
        for estado, cantidad in sorted(estados_count.items(), key=lambda x: x[1], reverse=True):
            desc = estado_descriptions.get(estado, 'Estado dental')
            # Capitalizar primera letra y el resto en minúsculas
            estado_display = estado.capitalize().replace('_', ' ')
            resumen_data.append([estado_display, str(cantidad), desc])
        
        resumen_table = Table(resumen_data, colWidths=[1.8*inch, 1*inch, 3.7*inch])
        resumen_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#14b8a6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ccfbf1')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (1, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdfa')]),
        ])
        
        # Colorear la columna de cantidad según el estado
        row_idx = 1
        for estado, cantidad in sorted(estados_count.items(), key=lambda x: x[1], reverse=True):
            color = get_tooth_color(estado)
            resumen_style.add('BACKGROUND', (1, row_idx), (1, row_idx), color)
            if estado in ['ausente', 'perdido', 'caries', 'cariado', 'fractura', 'extraccion']:
                resumen_style.add('TEXTCOLOR', (1, row_idx), (1, row_idx), colors.white)
            else:
                resumen_style.add('TEXTCOLOR', (1, row_idx), (1, row_idx), colors.HexColor('#1e293b'))
            row_idx += 1
        
        resumen_table.setStyle(resumen_style)
        story.append(resumen_table)
        story.append(Spacer(1, 12))
    
    # Sección: Detalle de Caras de Dientes (si hay datos interactivos)
    dientes_con_caras = []
    for diente in estados_dientes:
        interactive_data = get_tooth_interactive_data(diente)
        if interactive_data and isinstance(interactive_data, dict):
            caras_afectadas = {}
            for cara, condicion in interactive_data.items():
                if isinstance(condicion, str) and condicion != 'sano':
                    caras_afectadas[cara] = condicion
            
            if caras_afectadas:
                dientes_con_caras.append({
                    'numero': diente.numero_diente,
                    'caras': caras_afectadas
                })
    
    if dientes_con_caras:
        detalle_caras_title = Paragraph("<b>DETALLE DE CARAS DE DIENTES</b>", subtitle_style)
        story.append(detalle_caras_title)
        
        # Explicación sobre las caras
        explicacion_caras = Paragraph(
            "Esta sección muestra qué cara específica de cada diente tiene una condición. "
            "Las caras son: Oclusal (O) - superficie de masticación, Vestibular (V) - lado externo, "
            "Lingual (L) - lado interno, Mesial (M) - lado anterior, Distal (D) - lado posterior.",
            ParagraphStyle(
                'ExplicacionCaras',
                parent=styles['Normal'],
                fontSize=8,
                spaceAfter=6,
                alignment=TA_LEFT,
                textColor=colors.HexColor('#64748b'),
                leading=11
            )
        )
        story.append(explicacion_caras)
        story.append(Spacer(1, 4))
        
        # Crear tabla de detalles de caras (solo 3 columnas, sin duplicar)
        detalle_caras_data = [['Diente', 'Cara', 'Condición']]
        
        cara_nombres_completos = {
            'oclusal': 'Oclusal (O)',
            'vestibular': 'Vestibular (V)',
            'lingual': 'Lingual (L)',
            'mesial': 'Mesial (M)',
            'distal': 'Distal (D)'
        }
        
        estado_nombres = {
            'sano': 'Sano',
            'caries': 'Caries',
            'cariado': 'Cariado',
            'obturado': 'Obturado',
            'corona': 'Corona',
            'ausente': 'Ausente',
            'endodoncia': 'Endodoncia',
            'protesis': 'Prótesis',
            'implante': 'Implante',
            'sellante': 'Sellante',
            'fractura': 'Fractura',
            'perdido': 'Perdido',
            'extraccion': 'Extracción'
        }
        
        # Crear filas con información de caras (una fila por cara afectada)
        for diente_info in dientes_con_caras:
            numero_diente = diente_info['numero']
            for cara, condicion in diente_info['caras'].items():
                cara_nombre = cara_nombres_completos.get(cara, cara.capitalize())
                condicion_nombre = estado_nombres.get(condicion, condicion.capitalize())
                detalle_caras_data.append([str(numero_diente), cara_nombre, condicion_nombre])
        
        if len(detalle_caras_data) > 1:  # Si hay al menos una fila de datos
            detalle_caras_table = Table(detalle_caras_data, colWidths=[0.8*inch, 1.8*inch, 2.9*inch])
            detalle_caras_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#14b8a6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Número de diente centrado
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ccfbf1')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('ROWBACKGROUNDS', (1, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdfa')]),
            ])
            
            # Colorear la columna de condición según el estado
            row_idx = 1
            for fila in detalle_caras_data[1:]:
                if len(fila) >= 3 and fila[2]:
                    condicion = fila[2].lower()
                    color = get_tooth_color(condicion)
                    detalle_caras_style.add('BACKGROUND', (2, row_idx), (2, row_idx), color)
                    if condicion in ['ausente', 'perdido', 'caries', 'cariado', 'fractura', 'extraccion']:
                        detalle_caras_style.add('TEXTCOLOR', (2, row_idx), (2, row_idx), colors.white)
                    else:
                        detalle_caras_style.add('TEXTCOLOR', (2, row_idx), (2, row_idx), colors.HexColor('#1e293b'))
                row_idx += 1
            
            detalle_caras_table.setStyle(detalle_caras_style)
            story.append(detalle_caras_table)
            story.append(Spacer(1, 12))
    
    # Leyenda de símbolos resumida y compacta
    leyenda_title = Paragraph("<b>GUÍA DE SÍMBOLOS</b>", ParagraphStyle(
        'LeyendaTitle',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=4,
        alignment=TA_LEFT,
        textColor=colors.HexColor('#1e293b'),
        fontName='Helvetica-Bold'
    ))
    story.append(leyenda_title)
    
    # Leyenda compacta en formato de lista simple
    leyenda_texto = (
        "<b>✓</b> Sano | <b>●</b> Caries | <b>■</b> Obturado | <b>◊</b> Corona | <b>✕</b> Ausente | "
        "<b>◐</b> Endodoncia | <b>◈</b> Prótesis | <b>◉</b> Implante | <b>◯</b> Sellante | <b>◢</b> Fractura"
    )
    
    leyenda_parrafo = Paragraph(
        leyenda_texto,
        ParagraphStyle(
            'LeyendaCompacta',
            parent=styles['Normal'],
            fontSize=7,
            spaceAfter=6,
            alignment=TA_LEFT,
            textColor=colors.HexColor('#374151'),
            leading=10,
            backColor=colors.HexColor('#f0fdfa'),
            borderPadding=6,
            borderColor=colors.HexColor('#ccfbf1'),
            borderWidth=1
        )
    )
    story.append(leyenda_parrafo)
    story.append(Spacer(1, 8))
    
    # Sección: Plan de Tratamiento y Observaciones
    if odontograma.plan_tratamiento or odontograma.observaciones:
        plan_title = Paragraph("<b>PLAN DE TRATAMIENTO Y RECOMENDACIONES</b>", subtitle_style)
        story.append(plan_title)
        
        # Nota breve sobre el plan de tratamiento
        nota_plan = Paragraph(
            "Plan de tratamiento y recomendaciones:",
            ParagraphStyle(
                'NotaPlan',
                parent=styles['Normal'],
                fontSize=8,
                spaceAfter=4,
                alignment=TA_LEFT,
                textColor=colors.HexColor('#64748b'),
                leading=11
            )
        )
        story.append(nota_plan)
        story.append(Spacer(1, 2))
        
        plan_data = []
        if odontograma.plan_tratamiento:
            plan_data.append(['Plan de Tratamiento:', odontograma.plan_tratamiento])
        
        if odontograma.observaciones:
            plan_data.append(['Observaciones:', odontograma.observaciones])
        
        if plan_data:
            plan_table = Table(plan_data, colWidths=[2*inch, 4.5*inch])
            plan_table_style = TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0fdfa')),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1e293b')),
                ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#374151')),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ccfbf1')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ])
            plan_table.setStyle(plan_table_style)
            story.append(plan_table)
            story.append(Spacer(1, 10))
    
    # Información del sistema (footer)
    system_info = Paragraph(
        f"<b>Odontólogo responsable:</b> Dr. {odontograma.dentista.nombre_completo} | "
        f"<b>Fecha de creación:</b> {odontograma.fecha_creacion.strftime('%d/%m/%Y %H:%M')} | "
        f"<b>Última actualización:</b> {odontograma.fecha_actualizacion.strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle(
            'SystemInfo',
            parent=styles['Normal'],
            fontSize=7,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#64748b'),
            spaceBefore=12
        )
    )
    story.append(system_info)
    
    # Construir el PDF
    doc.build(story)
    
    # Obtener el contenido del buffer
    pdf_content = buffer.getvalue()
    buffer.close()
    
    # Crear o actualizar el documento en la base de datos
    try:
        # Buscar cliente por email si no está asociado directamente
        cliente_doc = odontograma.cliente
        if not cliente_doc and odontograma.paciente_email:
            try:
                cliente_doc = Cliente.objects.get(email=odontograma.paciente_email, activo=True)
            except Cliente.DoesNotExist:
                cliente_doc = None
            except Cliente.MultipleObjectsReturned:
                cliente_doc = Cliente.objects.filter(email=odontograma.paciente_email, activo=True).first()
        
        documento, created = DocumentoCliente.objects.get_or_create(
            odontograma=odontograma,
            tipo='odontograma',
            defaults={
                'cliente': cliente_doc,
                'titulo': f'Ficha Odontológica - {odontograma.paciente_nombre}',
                'descripcion': f'Ficha odontológica del paciente {odontograma.paciente_nombre}',
                'generado_por': perfil,
            }
        )
        if not created:
            # Actualizar fecha de generación y cliente si cambió
            documento.fecha_generacion = timezone.now()
            documento.generado_por = perfil
            if not documento.cliente and cliente_doc:
                documento.cliente = cliente_doc
            documento.save()
    except Exception as e:
        logger.error(f"Error al crear documento de odontograma: {str(e)}")
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ficha_odontologica_{odontograma.paciente_nombre.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
    response.write(pdf_content)
    
    return response


# Vista para exportar presupuesto/tratamiento a PDF
@login_required
def exportar_presupuesto_pdf(request, plan_id):
    """Vista para exportar un plan de tratamiento (presupuesto) a PDF"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not (perfil.es_dentista() or perfil.es_administrativo()):
            messages.error(request, 'No tienes permisos para exportar presupuestos.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    # Obtener el plan de tratamiento
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    else:
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    
    # Obtener información de la clínica
    try:
        from configuracion.models import InformacionClinica
        info_clinica = InformacionClinica.obtener()
        nombre_clinica = info_clinica.nombre_clinica
        direccion_clinica = info_clinica.direccion
        telefono_clinica = info_clinica.telefono
        email_clinica = info_clinica.email
    except:
        nombre_clinica = "Clínica Dental"
        direccion_clinica = ""
        telefono_clinica = ""
        email_clinica = ""
    
    # Crear el buffer para el PDF
    buffer = BytesIO()
    
    # Crear el documento PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=40,
        bottomMargin=30
    )
    
    # Estilos base de ReportLab
    styles = getSampleStyleSheet()
    
    # Paleta de colores (turquesa)
    primary_color = colors.HexColor('#14b8a6')  # turquesa principal
    primary_dark = colors.HexColor('#0f766e')
    soft_bg = colors.HexColor('#ecfeff')
    soft_bg_alt = colors.HexColor('#e0f2f1')
    gray_text = colors.HexColor('#64748b')
    dark_text = colors.HexColor('#0f172a')
    
    # Estilo para el título principal
    title_style = ParagraphStyle(
        'PresupuestoTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=12,
        alignment=TA_CENTER,
        textColor=primary_dark,
        fontName='Helvetica-Bold'
    )
    
    # Estilo para subtítulos
    subtitle_style = ParagraphStyle(
        'PresupuestoSubtitle',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=8,
        spaceBefore=12,
        alignment=TA_LEFT,
        textColor=primary_dark,
        fontName='Helvetica-Bold'
    )
    
    # Estilo para información de la clínica
    clinic_info_style = ParagraphStyle(
        'ClinicInfo',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_CENTER,
        textColor=gray_text
    )
    
    # Estilo para texto normal
    normal_style = ParagraphStyle(
        'NormalPresupuesto',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=4,
        alignment=TA_LEFT,
        leading=14
    )
    
    # Contenido del PDF
    story = []
    
    # Encabezado con información de la clínica
    header_text = f"<b>{nombre_clinica}</b><br/>"
    if direccion_clinica:
        header_text += f"{direccion_clinica}<br/>"
    if telefono_clinica:
        header_text += f"Tel: {telefono_clinica} | "
    if email_clinica:
        header_text += f"Email: {email_clinica}"
    header = Paragraph(header_text, clinic_info_style)
    story.append(header)
    story.append(Spacer(1, 12))
    
    # Título principal
    title = Paragraph("<b>PRESUPUESTO DE TRATAMIENTO</b>", title_style)
    story.append(title)
    
    # Información de fecha y número
    fecha_info = Paragraph(
        f"Fecha de Emisión: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Presupuesto N° {plan.id}",
        clinic_info_style
    )
    story.append(fecha_info)
    story.append(Spacer(1, 16))
    
    # Sección: Información del Paciente
    paciente_title = Paragraph("<b>INFORMACIÓN DEL PACIENTE</b>", subtitle_style)
    story.append(paciente_title)
    
    paciente_data = [
        ['Nombre Completo:', plan.cliente.nombre_completo],
        ['RUT:', plan.cliente.rut or 'No especificado'],
        ['Email:', plan.cliente.email or 'No especificado'],
        ['Teléfono:', plan.cliente.telefono or 'No especificado'],
    ]
    
    paciente_table = Table(paciente_data, colWidths=[2 * inch, 4.5 * inch])
    paciente_table_style = TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), soft_bg),
        ('TEXTCOLOR', (0, 0), (0, -1), primary_dark),
        ('TEXTCOLOR', (1, 0), (1, -1), dark_text),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ])
    paciente_table.setStyle(paciente_table_style)
    story.append(paciente_table)
    story.append(Spacer(1, 12))
    
    # Sección: Información del Tratamiento
    tratamiento_title = Paragraph("<b>INFORMACIÓN DEL TRATAMIENTO</b>", subtitle_style)
    story.append(tratamiento_title)
    
    tratamiento_info = [
        ['Nombre del Plan:', plan.nombre],
        ['Dentista:', plan.dentista.nombre_completo],
        ['Estado:', plan.get_estado_display()],
    ]
    
    if plan.fecha_inicio_estimada:
        tratamiento_info.append(['Fecha Inicio Estimada:', plan.fecha_inicio_estimada.strftime('%d/%m/%Y')])
    if plan.fecha_fin_estimada:
        tratamiento_info.append(['Fecha Fin Estimada:', plan.fecha_fin_estimada.strftime('%d/%m/%Y')])
    if plan.citas_estimadas:
        tratamiento_info.append(['Citas Estimadas:', str(plan.citas_estimadas)])
    
    tratamiento_table = Table(tratamiento_info, colWidths=[2 * inch, 4.5 * inch])
    tratamiento_table.setStyle(paciente_table_style)
    story.append(tratamiento_table)
    story.append(Spacer(1, 12))
    
    # Diagnóstico y Objetivo
    if plan.diagnostico:
        diagnostico_text = Paragraph(f"<b>Diagnóstico:</b><br/>{plan.diagnostico}", normal_style)
        story.append(diagnostico_text)
        story.append(Spacer(1, 8))
    
    if plan.objetivo:
        objetivo_text = Paragraph(f"<b>Objetivo del Tratamiento:</b><br/>{plan.objetivo}", normal_style)
        story.append(objetivo_text)
        story.append(Spacer(1, 12))
    
    # Sección: Detalle del Tratamiento (Fases e Items)
    # Nota: la gestión de fases y pagos se ha simplificado en el sistema,
    # por lo que esta sección se omite del PDF para mantener un diseño limpio.
    # Sección: Resumen Financiero
    resumen_title = Paragraph("<b>RESUMEN FINANCIERO</b>", subtitle_style)
    story.append(resumen_title)
    
    resumen_data = [
        ['Presupuesto Total:', f"${plan.presupuesto_total:,.0f}"],
        ['Descuento:', f"${plan.descuento:,.0f}"],
        ['Precio Final:', f"${plan.precio_final:,.0f}"],
        ['Total Pagado:', f"${plan.total_pagado:,.0f}"],
        ['Saldo Pendiente:', f"${plan.saldo_pendiente:,.0f}"],
    ]
    
    resumen_table = Table(resumen_data, colWidths=[2.5 * inch, 4 * inch])
    resumen_table_style = TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), soft_bg_alt),
        ('TEXTCOLOR', (0, 0), (0, -1), primary_dark),
        ('TEXTCOLOR', (1, 0), (1, -1), dark_text),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 2), (1, 2), soft_bg),  # Precio Final
        ('BACKGROUND', (0, 4), (1, 4), colors.HexColor('#fef2f2')),  # Saldo Pendiente
    ])
    resumen_table.setStyle(resumen_table_style)
    story.append(resumen_table)
    story.append(Spacer(1, 12))
    
    # Historial de Pagos (se omite en la versión actual del presupuesto para mantener el enfoque en el resumen financiero)
    
    # Notas
    if plan.notas_paciente:
        notas_title = Paragraph("<b>NOTAS PARA EL PACIENTE</b>", subtitle_style)
        story.append(notas_title)
        notas_text = Paragraph(plan.notas_paciente, normal_style)
        story.append(notas_text)
        story.append(Spacer(1, 12))
    
    # Footer
    footer_text = f"<b>Generado por:</b> {perfil.nombre_completo} | <b>Fecha:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    footer = Paragraph(footer_text, clinic_info_style)
    story.append(footer)
    
    # Construir el PDF
    doc.build(story)
    
    # Obtener el contenido del buffer
    pdf_content = buffer.getvalue()
    buffer.close()
    
    # Crear o actualizar el documento en la base de datos
    documento, created = DocumentoCliente.objects.get_or_create(
        plan_tratamiento=plan,
        tipo='presupuesto',
        defaults={
            'cliente': plan.cliente,
            'titulo': f'Presupuesto - {plan.nombre}',
            'descripcion': f'Presupuesto del tratamiento {plan.nombre}',
            'generado_por': perfil,
        }
    )
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    filename = f"presupuesto_{plan.cliente.nombre_completo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf_content)
    
    return response


# ============================================================================
# GESTIÓN DE DOCUMENTOS
# ============================================================================

@login_required
def gestor_documentos(request):
    """Vista para gestionar documentos de clientes (solo administradores)"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'Solo los administrativos pueden gestionar documentos.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    # Obtener parámetros de búsqueda
    tipo_filtro = request.GET.get('tipo', '')
    cliente_busqueda = request.GET.get('cliente', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Obtener todos los documentos
    documentos = DocumentoCliente.objects.all().select_related('cliente', 'generado_por', 'plan_tratamiento', 'odontograma')
    
    # Aplicar filtros
    if tipo_filtro:
        documentos = documentos.filter(tipo=tipo_filtro)
    
    if cliente_busqueda:
        documentos = documentos.filter(
            Q(cliente__nombre_completo__icontains=cliente_busqueda) |
            Q(cliente__email__icontains=cliente_busqueda) |
            Q(cliente__rut__icontains=cliente_busqueda) |
            Q(odontograma__paciente_nombre__icontains=cliente_busqueda) |
            Q(odontograma__paciente_email__icontains=cliente_busqueda) |
            Q(plan_tratamiento__cliente__nombre_completo__icontains=cliente_busqueda) |
            Q(plan_tratamiento__cliente__email__icontains=cliente_busqueda)
        )
    
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            documentos = documentos.filter(fecha_generacion__date__gte=fecha_desde_obj)
        except:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            documentos = documentos.filter(fecha_generacion__date__lte=fecha_hasta_obj)
        except:
            pass
    
    # Ordenar por fecha de generación (más recientes primero)
    documentos = documentos.order_by('-fecha_generacion')
    
    # Paginación
    paginator = Paginator(documentos, 20)
    page = request.GET.get('page', 1)
    try:
        documentos_pag = paginator.page(page)
    except PageNotAnInteger:
        documentos_pag = paginator.page(1)
    except EmptyPage:
        documentos_pag = paginator.page(paginator.num_pages)
    
    # Estadísticas
    total_documentos = DocumentoCliente.objects.count()
    documentos_por_tipo = {}
    
    # Tipos de documentos a excluir del selector
    tipos_excluidos = ['receta', 'factura', 'nota_medica', 'otro']
    
    # Filtrar tipos de documentos (excluir los especificados)
    tipos_documento_filtrados = [
        (tipo, nombre) for tipo, nombre in DocumentoCliente.TIPO_DOCUMENTO_CHOICES 
        if tipo not in tipos_excluidos
    ]
    
    # Solo contar documentos de los tipos que se mostrarán
    for tipo, nombre in tipos_documento_filtrados:
        documentos_por_tipo[tipo] = DocumentoCliente.objects.filter(tipo=tipo).count()
    
    # Obtener clientes para el filtro
    clientes = Cliente.objects.filter(activo=True).order_by('nombre_completo')[:50]
    
    context = {
        'perfil': perfil,
        'documentos': documentos_pag,
        'tipos_documento': tipos_documento_filtrados,
        'tipo_filtro': tipo_filtro,
        'cliente_busqueda': cliente_busqueda,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'total_documentos': total_documentos,
        'documentos_por_tipo': documentos_por_tipo,
        'clientes': clientes,
    }
    
    return render(request, 'citas/documentos/gestor_documentos.html', context)


@login_required
def descargar_documento(request, documento_id):
    """Vista para descargar un documento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para descargar documentos.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    documento = get_object_or_404(DocumentoCliente, id=documento_id)
    
    # Generar el PDF según el tipo de documento
    if documento.tipo == 'presupuesto' and documento.plan_tratamiento:
        return exportar_presupuesto_pdf(request, documento.plan_tratamiento.id)
    elif documento.tipo == 'odontograma' and documento.odontograma:
        return exportar_odontograma_pdf(request, documento.odontograma.id)
    elif documento.tipo == 'consentimiento':
        # Buscar el consentimiento asociado
        consentimiento = ConsentimientoInformado.objects.filter(
            cliente=documento.cliente,
            cita=documento.cita,
            plan_tratamiento=documento.plan_tratamiento
        ).first()
        if consentimiento:
            return exportar_consentimiento_pdf(request, consentimiento.id)
    elif documento.archivo_pdf:
        # Si tiene archivo PDF guardado, servirlo
        response = HttpResponse(documento.archivo_pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{documento.titulo}.pdf"'
        return response
    
    messages.error(request, 'El documento no está disponible para descarga.')
    return redirect('gestor_documentos')


@login_required
def enviar_documento_correo(request, documento_id):
    """Vista para enviar un documento por correo"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({'error': 'No tienes permisos para enviar documentos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'error': 'No tienes permisos para acceder a esta función.'}, status=403)
    
    documento = get_object_or_404(DocumentoCliente, id=documento_id)
    email_destino = request.POST.get('email', documento.cliente.email)
    
    if not email_destino:
        return JsonResponse({'error': 'No se especificó un email de destino.'}, status=400)
    
    try:
        from django.core.mail import EmailMessage
        from django.conf import settings
        import os
        
        # Generar el PDF según el tipo
        pdf_content = None
        filename = f"{documento.titulo}.pdf"
        
        if documento.tipo == 'presupuesto' and documento.plan_tratamiento:
            # Generar PDF del presupuesto
            buffer = BytesIO()
            # Reutilizar la lógica de exportar_presupuesto_pdf pero sin crear respuesta HTTP
            # Por ahora, redirigimos a generar el PDF
            response = exportar_presupuesto_pdf(request, documento.plan_tratamiento.id)
            pdf_content = response.content
            filename = f"presupuesto_{documento.cliente.nombre_completo.replace(' ', '_')}.pdf"
        elif documento.tipo == 'odontograma' and documento.odontograma:
            # Generar PDF del odontograma
            response = exportar_odontograma_pdf(request, documento.odontograma.id)
            pdf_content = response.content
            filename = f"ficha_odontologica_{documento.cliente.nombre_completo.replace(' ', '_')}.pdf"
        elif documento.tipo == 'consentimiento':
            # Buscar el consentimiento asociado
            consentimiento = ConsentimientoInformado.objects.filter(
                cliente=documento.cliente,
                cita=documento.cita,
                plan_tratamiento=documento.plan_tratamiento
            ).first()
            if consentimiento:
                response = exportar_consentimiento_pdf(request, consentimiento.id)
                pdf_content = response.content
                filename = f"consentimiento_{documento.cliente.nombre_completo.replace(' ', '_')}.pdf"
            else:
                return JsonResponse({'error': 'Consentimiento no encontrado.'}, status=400)
        elif documento.archivo_pdf:
            # Usar archivo PDF existente
            pdf_content = documento.archivo_pdf.read()
            filename = documento.archivo_pdf.name.split('/')[-1]
        else:
            return JsonResponse({'error': 'El documento no está disponible para envío.'}, status=400)
        
        # Obtener información de la clínica
        try:
            from configuracion.models import InformacionClinica
            info_clinica = InformacionClinica.obtener()
            nombre_clinica = info_clinica.nombre_clinica
            email_clinica = info_clinica.email or getattr(settings, 'DEFAULT_FROM_EMAIL', 'miclinicacontacto@gmail.com')
            direccion_clinica = info_clinica.direccion or ''
            telefono_clinica = info_clinica.telefono or ''
        except:
            nombre_clinica = "Clínica Dental"
            email_clinica = getattr(settings, 'DEFAULT_FROM_EMAIL', 'miclinicacontacto@gmail.com')
            direccion_clinica = ''
            telefono_clinica = ''
        
        # Renderizar template HTML
        from django.template.loader import render_to_string
        mensaje_html = render_to_string('citas/emails/documento_enviado.html', {
            'cliente_nombre': documento.cliente.nombre_completo,
            'tipo_documento': documento.get_tipo_display(),
            'titulo': documento.titulo,
            'fecha_generacion': documento.fecha_generacion,
            'nombre_clinica': nombre_clinica,
            'direccion_clinica': direccion_clinica,
            'telefono_clinica': telefono_clinica,
            'email_clinica': email_clinica,
        })
        
        # Crear el email con contenido HTML
        asunto = f"{documento.get_tipo_display()} - {nombre_clinica}"
        email = EmailMessage(
            asunto,
            mensaje_html,
            email_clinica,
            [email_destino],
        )
        email.content_subtype = "html"  # Indicar que es HTML
        
        # Adjuntar PDF
        email.attach(filename, pdf_content, 'application/pdf')
        
        # Enviar email
        email.send()
        
        # Actualizar documento
        documento.enviado_por_correo = True
        documento.fecha_envio = timezone.now()
        documento.email_destinatario = email_destino
        documento.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Documento enviado exitosamente a {email_destino}'
        })
        
    except Exception as e:
        logger.error(f"Error al enviar documento por correo: {str(e)}")
        return JsonResponse({'error': f'Error al enviar el documento: {str(e)}'}, status=500)


@login_required
def enviar_consentimiento_por_correo(request, consentimiento_id):
    """Vista para enviar un consentimiento por correo con enlace de firma"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({'error': 'No tienes permisos para enviar consentimientos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'error': 'No tienes permisos para acceder a esta función.'}, status=403)
    
    consentimiento = get_object_or_404(ConsentimientoInformado, id=consentimiento_id)
    
    if consentimiento.esta_firmado:
        return JsonResponse({'error': 'Este consentimiento ya está firmado.'}, status=400)
    
    email_destino = request.POST.get('email', consentimiento.cliente.email)
    
    if not email_destino:
        return JsonResponse({'error': 'No se especificó un email de destino.'}, status=400)
    
    try:
        from django.core.mail import EmailMessage
        from django.conf import settings
        
        # Generar token de firma si no existe
        if not consentimiento.token_firma or not consentimiento.token_es_valido():
            consentimiento.generar_token_firma()
        
        # Generar PDF del consentimiento
        response = exportar_consentimiento_pdf(request, consentimiento.id)
        pdf_content = response.content
        filename = f"consentimiento_{consentimiento.cliente.nombre_completo.replace(' ', '_')}.pdf"
        
        # Obtener información de la clínica
        try:
            from configuracion.models import InformacionClinica
            info_clinica = InformacionClinica.obtener()
            nombre_clinica = info_clinica.nombre_clinica
            email_clinica = info_clinica.email or getattr(settings, 'DEFAULT_FROM_EMAIL', 'miclinicacontacto@gmail.com')
            direccion_clinica = info_clinica.direccion or ''
            telefono_clinica = info_clinica.telefono or ''
        except:
            nombre_clinica = "Clínica Dental"
            email_clinica = getattr(settings, 'DEFAULT_FROM_EMAIL', 'miclinicacontacto@gmail.com')
            direccion_clinica = ''
            telefono_clinica = ''
        
        # Construir URL de firma (pública con token)
        dominio = request.get_host()
        # URL para firma en sistema de gestión (pública con token)
        url_firma_gestion = f"{request.scheme}://{dominio}/trabajadores/consentimientos/firmar-publico/{consentimiento.token_firma}/"
        
        # URL para firma en cliente_web (requiere login)
        # Asumiendo que cliente_web está en el mismo dominio o subdominio
        url_firma_web = f"{request.scheme}://{dominio.replace('gestion', 'web')}/consentimientos/{consentimiento.id}/firmar/"
        
        # Renderizar template HTML
        from django.template.loader import render_to_string
        mensaje_html = render_to_string('citas/emails/consentimiento_enviado.html', {
            'cliente_nombre': consentimiento.cliente.nombre_completo,
            'titulo': consentimiento.titulo,
            'tipo_procedimiento': consentimiento.get_tipo_procedimiento_display(),
            'fecha_creacion': consentimiento.fecha_creacion,
            'url_firma_gestion': url_firma_gestion,
            'url_firma_web': url_firma_web,
            'nombre_clinica': nombre_clinica,
            'direccion_clinica': direccion_clinica,
            'telefono_clinica': telefono_clinica,
            'email_clinica': email_clinica,
        })
        
        # Crear el email con contenido HTML
        asunto = f"Consentimiento Informado - {nombre_clinica}"
        email = EmailMessage(
            asunto,
            mensaje_html,
            email_clinica,
            [email_destino],
        )
        email.content_subtype = "html"  # Indicar que es HTML
        
        # Adjuntar PDF
        email.attach(filename, pdf_content, 'application/pdf')
        
        # Enviar email
        email.send()
        
        return JsonResponse({
            'success': True,
            'message': f'Consentimiento enviado exitosamente a {email_destino}. Se ha incluido un enlace para firmar.'
        })
        
    except Exception as e:
        logger.error(f"Error al enviar consentimiento por correo: {str(e)}")
        return JsonResponse({'error': f'Error al enviar el consentimiento: {str(e)}'}, status=500)


# ============================================================================
# GESTIÓN DE CONSENTIMIENTOS INFORMADOS
# ============================================================================

@login_required
def gestor_consentimientos(request):
    """Vista principal para gestionar consentimientos informados"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'Solo los administrativos pueden gestionar consentimientos.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    # Obtener parámetros de búsqueda
    estado_filtro = request.GET.get('estado', '')
    cliente_busqueda = request.GET.get('cliente', '')
    tipo_filtro = request.GET.get('tipo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Obtener todos los consentimientos
    consentimientos = ConsentimientoInformado.objects.all().select_related(
        'cliente', 'dentista', 'cita', 'plan_tratamiento', 'plantilla'
    )
    
    # Aplicar filtros
    if estado_filtro:
        consentimientos = consentimientos.filter(estado=estado_filtro)
    
    if tipo_filtro:
        consentimientos = consentimientos.filter(tipo_procedimiento=tipo_filtro)
    
    if cliente_busqueda:
        consentimientos = consentimientos.filter(
            Q(cliente__nombre_completo__icontains=cliente_busqueda) |
            Q(cliente__email__icontains=cliente_busqueda) |
            Q(cliente__rut__icontains=cliente_busqueda)
        )
    
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            consentimientos = consentimientos.filter(fecha_creacion__date__gte=fecha_desde_obj)
        except:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            consentimientos = consentimientos.filter(fecha_creacion__date__lte=fecha_hasta_obj)
        except:
            pass
    
    # Ordenar por fecha de creación (más recientes primero)
    consentimientos = consentimientos.order_by('-fecha_creacion')
    
    # Paginación
    paginator = Paginator(consentimientos, 20)
    page = request.GET.get('page', 1)
    try:
        consentimientos_pag = paginator.page(page)
    except PageNotAnInteger:
        consentimientos_pag = paginator.page(1)
    except EmptyPage:
        consentimientos_pag = paginator.page(paginator.num_pages)
    
    # Estadísticas
    total_consentimientos = ConsentimientoInformado.objects.count()
    consentimientos_por_estado = {}
    for estado, nombre in ConsentimientoInformado.ESTADO_CHOICES:
        consentimientos_por_estado[estado] = ConsentimientoInformado.objects.filter(estado=estado).count()
    
    # Obtener plantillas activas
    plantillas = PlantillaConsentimiento.objects.filter(activo=True).order_by('tipo_procedimiento', 'nombre')
    
    context = {
        'perfil': perfil,
        'consentimientos': consentimientos_pag,
        'estados': ConsentimientoInformado.ESTADO_CHOICES,
        'tipos_procedimiento': PlantillaConsentimiento.TIPO_PROCEDIMIENTO_CHOICES,
        'estado_filtro': estado_filtro,
        'tipo_filtro': tipo_filtro,
        'cliente_busqueda': cliente_busqueda,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'total_consentimientos': total_consentimientos,
        'consentimientos_por_estado': consentimientos_por_estado,
        'plantillas': plantillas,
    }
    
    return render(request, 'citas/consentimientos/gestor_consentimientos.html', context)


@login_required
def crear_consentimiento(request):
    """Vista para crear un nuevo consentimiento informado"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'Solo los administrativos pueden crear consentimientos.')
            return redirect('gestor_consentimientos')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    # Obtener datos necesarios
    clientes = Cliente.objects.filter(activo=True).order_by('nombre_completo')
    plantillas = PlantillaConsentimiento.objects.filter(activo=True).order_by('tipo_procedimiento', 'nombre')
    dentistas = Perfil.objects.filter(rol='dentista', activo=True).order_by('nombre_completo')
    
    if request.method == 'POST':
        try:
            cliente_id = request.POST.get('cliente')
            plantilla_id = request.POST.get('plantilla', '')
            tipo_procedimiento = request.POST.get('tipo_procedimiento')
            titulo = request.POST.get('titulo')
            diagnostico = request.POST.get('diagnostico', '')
            justificacion = request.POST.get('justificacion', '')
            naturaleza_procedimiento = request.POST.get('naturaleza_procedimiento', '')
            objetivos_tratamiento = request.POST.get('objetivos_tratamiento', '')
            contenido = request.POST.get('contenido', '')
            riesgos = request.POST.get('riesgos', '')
            beneficios = request.POST.get('beneficios', '')
            alternativas = request.POST.get('alternativas', '')
            pronostico = request.POST.get('pronostico', '')
            cuidados_postoperatorios = request.POST.get('cuidados_postoperatorios', '')
            dentista_id = request.POST.get('dentista', '')
            rut_dentista = request.POST.get('rut_dentista', '')
            registro_superintendencia = request.POST.get('registro_superintendencia', '')
            explicado_por_id = request.POST.get('explicado_por', '')
            rut_explicado_por = request.POST.get('rut_explicado_por', '')
            cita_id = request.POST.get('cita', '')
            plan_id = request.POST.get('plan_tratamiento', '')
            fecha_vencimiento = request.POST.get('fecha_vencimiento', '')
            
            cliente = get_object_or_404(Cliente, id=cliente_id)
            
            # Si se seleccionó una plantilla, usar su contenido
            plantilla = None
            if plantilla_id:
                plantilla = get_object_or_404(PlantillaConsentimiento, id=plantilla_id)
                if not diagnostico and plantilla.diagnostico_base:
                    diagnostico = plantilla.diagnostico_base.replace('{nombre_paciente}', cliente.nombre_completo)
                if not naturaleza_procedimiento and plantilla.naturaleza_procedimiento:
                    naturaleza_procedimiento = plantilla.naturaleza_procedimiento
                if not objetivos_tratamiento and plantilla.objetivos_tratamiento:
                    objetivos_tratamiento = plantilla.objetivos_tratamiento
                if not contenido and plantilla.contenido:
                    # Reemplazar variables en el contenido
                    contenido = plantilla.contenido.replace('{nombre_paciente}', cliente.nombre_completo)
                    contenido = contenido.replace('{fecha}', datetime.now().strftime('%d/%m/%Y'))
                if plantilla.riesgos and not riesgos:
                    riesgos = plantilla.riesgos
                if plantilla.beneficios and not beneficios:
                    beneficios = plantilla.beneficios
                if plantilla.alternativas and not alternativas:
                    alternativas = plantilla.alternativas
                if plantilla.pronostico and not pronostico:
                    pronostico = plantilla.pronostico
                if plantilla.cuidados_postoperatorios and not cuidados_postoperatorios:
                    cuidados_postoperatorios = plantilla.cuidados_postoperatorios
            
            # Obtener dentista
            dentista = None
            if dentista_id:
                dentista = get_object_or_404(Perfil, id=dentista_id)
            elif perfil.es_dentista():
                dentista = perfil
            
            # Obtener profesional informante
            explicado_por = perfil
            if explicado_por_id:
                explicado_por = get_object_or_404(Perfil, id=explicado_por_id)
            
            # Crear consentimiento
            consentimiento = ConsentimientoInformado.objects.create(
                cliente=cliente,
                plantilla=plantilla,
                tipo_procedimiento=tipo_procedimiento,
                titulo=titulo,
                diagnostico=diagnostico or None,
                justificacion=justificacion or None,
                naturaleza_procedimiento=naturaleza_procedimiento or None,
                objetivos_tratamiento=objetivos_tratamiento or None,
                contenido=contenido or None,
                riesgos=riesgos or None,
                beneficios=beneficios or None,
                alternativas=alternativas or None,
                pronostico=pronostico or None,
                cuidados_postoperatorios=cuidados_postoperatorios or None,
                dentista=dentista,
                rut_dentista=rut_dentista or None,
                registro_superintendencia=registro_superintendencia or None,
                explicado_por=explicado_por,
                rut_explicado_por=rut_explicado_por or None,
                estado='pendiente',
                fecha_vencimiento=datetime.strptime(fecha_vencimiento, '%Y-%m-%d').date() if fecha_vencimiento else None,
            )
            
            # Asociar con cita si se proporciona
            if cita_id:
                try:
                    cita = Cita.objects.get(id=cita_id)
                    consentimiento.cita = cita
                    consentimiento.save()
                except Cita.DoesNotExist:
                    pass
            
            # Asociar con plan de tratamiento si se proporciona
            if plan_id:
                try:
                    plan = PlanTratamiento.objects.get(id=plan_id)
                    consentimiento.plan_tratamiento = plan
                    consentimiento.save()
                except PlanTratamiento.DoesNotExist:
                    pass
            
            # Crear documento asociado
            DocumentoCliente.objects.create(
                cliente=cliente,
                tipo='consentimiento',
                titulo=f'Consentimiento - {titulo}',
                descripcion=f'Consentimiento informado para {tipo_procedimiento}',
                cita=consentimiento.cita,
                plan_tratamiento=consentimiento.plan_tratamiento,
                generado_por=perfil,
            )
            
            messages.success(request, 'Consentimiento creado exitosamente.')
            return redirect('gestor_consentimientos')
            
        except Exception as e:
            logger.error(f"Error al crear consentimiento: {str(e)}")
            messages.error(request, f'Error al crear el consentimiento: {str(e)}')
    
    context = {
        'perfil': perfil,
        'clientes': clientes,
        'plantillas': plantillas,
        'dentistas': dentistas,
        'tipos_procedimiento': PlantillaConsentimiento.TIPO_PROCEDIMIENTO_CHOICES,
    }
    
    return render(request, 'citas/consentimientos/crear_consentimiento.html', context)


@login_required
def crear_consentimiento_desde_plan(request, plan_id):
    """Vista para crear un consentimiento informado desde un plan de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        messages.error(request, 'No tienes permisos para crear consentimientos.')
        return redirect('detalle_plan_tratamiento', plan_id=plan_id)
    
    # Obtener datos necesarios
    plantillas = PlantillaConsentimiento.objects.filter(activo=True).order_by('tipo_procedimiento', 'nombre')
    dentistas = Perfil.objects.filter(rol='dentista', activo=True).order_by('nombre_completo')
    
    if request.method == 'POST':
        try:
            plantilla_id = request.POST.get('plantilla', '')
            tipo_procedimiento = request.POST.get('tipo_procedimiento')
            titulo = request.POST.get('titulo')
            diagnostico = request.POST.get('diagnostico', '')
            justificacion = request.POST.get('justificacion', '')
            naturaleza_procedimiento = request.POST.get('naturaleza_procedimiento', '')
            objetivos_tratamiento = request.POST.get('objetivos_tratamiento', '')
            contenido = request.POST.get('contenido', '')
            riesgos = request.POST.get('riesgos', '')
            beneficios = request.POST.get('beneficios', '')
            alternativas = request.POST.get('alternativas', '')
            pronostico = request.POST.get('pronostico', '')
            cuidados_postoperatorios = request.POST.get('cuidados_postoperatorios', '')
            dentista_id = request.POST.get('dentista', '')
            rut_dentista = request.POST.get('rut_dentista', '')
            registro_superintendencia = request.POST.get('registro_superintendencia', '')
            explicado_por_id = request.POST.get('explicado_por', '')
            rut_explicado_por = request.POST.get('rut_explicado_por', '')
            fecha_vencimiento = request.POST.get('fecha_vencimiento', '')
            
            cliente = plan.cliente
            
            # Si se seleccionó una plantilla, usar su contenido
            plantilla = None
            if plantilla_id:
                plantilla = get_object_or_404(PlantillaConsentimiento, id=plantilla_id)
                if not diagnostico and plantilla.diagnostico_base:
                    diagnostico = plantilla.diagnostico_base.replace('{nombre_paciente}', cliente.nombre_completo)
                if not naturaleza_procedimiento and plantilla.naturaleza_procedimiento:
                    naturaleza_procedimiento = plantilla.naturaleza_procedimiento
                if not objetivos_tratamiento and plantilla.objetivos_tratamiento:
                    objetivos_tratamiento = plantilla.objetivos_tratamiento
                if not contenido and plantilla.contenido:
                    contenido = plantilla.contenido.replace('{nombre_paciente}', cliente.nombre_completo)
                    contenido = contenido.replace('{fecha}', datetime.now().strftime('%d/%m/%Y'))
                if plantilla.riesgos and not riesgos:
                    riesgos = plantilla.riesgos
                if plantilla.beneficios and not beneficios:
                    beneficios = plantilla.beneficios
                if plantilla.alternativas and not alternativas:
                    alternativas = plantilla.alternativas
                if plantilla.pronostico and not pronostico:
                    pronostico = plantilla.pronostico
                if plantilla.cuidados_postoperatorios and not cuidados_postoperatorios:
                    cuidados_postoperatorios = plantilla.cuidados_postoperatorios
            
            # Usar el dentista del plan por defecto
            dentista = plan.dentista
            if dentista_id:
                dentista = get_object_or_404(Perfil, id=dentista_id)
            
            # Obtener profesional informante
            explicado_por = perfil
            if explicado_por_id:
                explicado_por = get_object_or_404(Perfil, id=explicado_por_id)
            
            # Si no hay diagnóstico, usar el del plan
            if not diagnostico and plan.diagnostico:
                diagnostico = plan.diagnostico
            
            # Crear consentimiento
            consentimiento = ConsentimientoInformado.objects.create(
                cliente=cliente,
                plantilla=plantilla,
                plan_tratamiento=plan,
                tipo_procedimiento=tipo_procedimiento,
                titulo=titulo,
                diagnostico=diagnostico or None,
                justificacion=justificacion or None,
                naturaleza_procedimiento=naturaleza_procedimiento or None,
                objetivos_tratamiento=objetivos_tratamiento or None,
                contenido=contenido or None,
                riesgos=riesgos or None,
                beneficios=beneficios or None,
                alternativas=alternativas or None,
                pronostico=pronostico or None,
                cuidados_postoperatorios=cuidados_postoperatorios or None,
                dentista=dentista,
                rut_dentista=rut_dentista or None,
                registro_superintendencia=registro_superintendencia or None,
                explicado_por=explicado_por,
                rut_explicado_por=rut_explicado_por or None,
                estado='pendiente',
                fecha_vencimiento=datetime.strptime(fecha_vencimiento, '%Y-%m-%d').date() if fecha_vencimiento else None,
            )
            
            # Crear documento asociado
            DocumentoCliente.objects.create(
                cliente=cliente,
                tipo='consentimiento',
                titulo=f'Consentimiento - {titulo}',
                descripcion=f'Consentimiento informado para {tipo_procedimiento}',
                plan_tratamiento=plan,
                generado_por=perfil,
            )
            
            messages.success(request, 'Consentimiento creado exitosamente.')
            return redirect('detalle_plan_tratamiento', plan_id=plan_id)
            
        except Exception as e:
            logger.error(f"Error al crear consentimiento desde plan: {str(e)}")
            messages.error(request, f'Error al crear el consentimiento: {str(e)}')
    
    # Prellenar datos del plan
    context = {
        'perfil': perfil,
        'plan': plan,
        'cliente': plan.cliente,
        'plantillas': plantillas,
        'dentistas': dentistas,
        'dentista_plan': plan.dentista,
        'tipos_procedimiento': PlantillaConsentimiento.TIPO_PROCEDIMIENTO_CHOICES,
        'diagnostico_plan': plan.diagnostico,
    }
    
    return render(request, 'citas/consentimientos/crear_consentimiento_desde_plan.html', context)


@login_required
def editar_consentimiento(request, consentimiento_id):
    """Vista para editar un consentimiento informado"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'Solo los administrativos pueden editar consentimientos.')
            return redirect('gestor_consentimientos')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    consentimiento = get_object_or_404(ConsentimientoInformado, id=consentimiento_id)
    
    # No permitir editar si ya está firmado
    if consentimiento.esta_firmado:
        messages.warning(request, 'No se puede editar un consentimiento que ya está firmado.')
        return redirect('detalle_consentimiento', consentimiento_id=consentimiento_id)
    
    if request.method == 'POST':
        try:
            tipo_procedimiento = request.POST.get('tipo_procedimiento')
            titulo = request.POST.get('titulo')
            contenido = request.POST.get('contenido')
            riesgos = request.POST.get('riesgos', '')
            beneficios = request.POST.get('beneficios', '')
            alternativas = request.POST.get('alternativas', '')
            fecha_vencimiento = request.POST.get('fecha_vencimiento', '')
            
            consentimiento.tipo_procedimiento = tipo_procedimiento
            consentimiento.titulo = titulo
            consentimiento.contenido = contenido
            consentimiento.riesgos = riesgos or None
            consentimiento.beneficios = beneficios or None
            consentimiento.alternativas = alternativas or None
            consentimiento.fecha_vencimiento = datetime.strptime(fecha_vencimiento, '%Y-%m-%d').date() if fecha_vencimiento else None
            consentimiento.save()
            
            messages.success(request, 'Consentimiento actualizado exitosamente.')
            # Si el consentimiento tiene un plan de tratamiento, redirigir al plan
            if consentimiento.plan_tratamiento:
                return redirect('detalle_plan_tratamiento', plan_id=consentimiento.plan_tratamiento.id)
            return redirect('gestor_consentimientos')
            
        except Exception as e:
            logger.error(f"Error al editar consentimiento: {str(e)}")
            messages.error(request, f'Error al actualizar el consentimiento: {str(e)}')
    
    context = {
        'perfil': perfil,
        'consentimiento': consentimiento,
        'tipos_procedimiento': PlantillaConsentimiento.TIPO_PROCEDIMIENTO_CHOICES,
    }
    
    # Si el consentimiento tiene un plan de tratamiento, agregarlo al contexto
    if consentimiento.plan_tratamiento:
        context['plan_tratamiento'] = consentimiento.plan_tratamiento
    
    return render(request, 'citas/consentimientos/editar_consentimiento.html', context)


@login_required
def detalle_consentimiento(request, consentimiento_id):
    """Vista para ver el detalle de un consentimiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not (perfil.es_administrativo() or perfil.es_dentista()):
            messages.error(request, 'No tienes permisos para ver consentimientos.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    consentimiento = get_object_or_404(ConsentimientoInformado, id=consentimiento_id)
    
    # Los dentistas solo pueden ver sus propios consentimientos
    if perfil.es_dentista() and consentimiento.dentista != perfil:
        messages.error(request, 'No tienes permisos para ver este consentimiento.')
        return redirect('gestor_consentimientos')
    
    context = {
        'perfil': perfil,
        'consentimiento': consentimiento,
        'puede_editar': perfil.es_administrativo() and not consentimiento.esta_firmado,
        'puede_firmar': perfil.es_administrativo() and consentimiento.estado == 'pendiente',
    }
    
    return render(request, 'citas/consentimientos/detalle_consentimiento.html', context)


def firmar_consentimiento_publico(request, token):
    """Vista pública para firmar un consentimiento usando token (sin autenticación)"""
    consentimiento = get_object_or_404(ConsentimientoInformado, token_firma=token)
    
    # Verificar que el token sea válido
    if not consentimiento.token_es_valido():
        return render(request, 'citas/consentimientos/firmar_consentimiento_publico.html', {
            'error': 'El enlace de firma ha expirado o ya no es válido. Por favor, contacte a la clínica para obtener un nuevo enlace.',
            'consentimiento': None
        })
    
    if consentimiento.esta_firmado:
        return render(request, 'citas/consentimientos/firmar_consentimiento_publico.html', {
            'error': 'Este consentimiento ya ha sido firmado.',
            'consentimiento': consentimiento
        })
    
    if request.method == 'POST':
        try:
            firma_paciente = request.POST.get('firma_paciente', '')
            nombre_firmante = request.POST.get('nombre_firmante', consentimiento.cliente.nombre_completo)
            rut_firmante = request.POST.get('rut_firmante', consentimiento.cliente.rut or '')
            nombre_testigo = request.POST.get('nombre_testigo', '')
            rut_testigo = request.POST.get('rut_testigo', '')
            firma_testigo = request.POST.get('firma_testigo', '')
            declaracion_comprension = request.POST.get('declaracion_comprension') == 'on'
            derecho_revocacion = request.POST.get('derecho_revocacion') == 'on'
            
            if not firma_paciente:
                return render(request, 'citas/consentimientos/firmar_consentimiento_publico.html', {
                    'error': 'La firma del paciente es obligatoria.',
                    'consentimiento': consentimiento
                })
            
            if not declaracion_comprension:
                return render(request, 'citas/consentimientos/firmar_consentimiento_publico.html', {
                    'error': 'Debe confirmar la declaración de comprensión.',
                    'consentimiento': consentimiento
                })
            
            if not derecho_revocacion:
                return render(request, 'citas/consentimientos/firmar_consentimiento_publico.html', {
                    'error': 'Debe confirmar que conoce su derecho de revocación.',
                    'consentimiento': consentimiento
                })
            
            consentimiento.firma_paciente = firma_paciente
            consentimiento.nombre_firmante = nombre_firmante
            consentimiento.rut_firmante = rut_firmante
            consentimiento.nombre_testigo = nombre_testigo if nombre_testigo else None
            consentimiento.rut_testigo = rut_testigo if rut_testigo else None
            consentimiento.firma_testigo = firma_testigo if firma_testigo else None
            consentimiento.declaracion_comprension = declaracion_comprension
            consentimiento.derecho_revocacion = derecho_revocacion
            consentimiento.estado = 'firmado'
            consentimiento.fecha_firma = timezone.now()
            consentimiento.save()
            
            return render(request, 'citas/consentimientos/firmar_consentimiento_publico.html', {
                'success': True,
                'message': 'Consentimiento firmado exitosamente. Gracias por su tiempo.',
                'consentimiento': consentimiento
            })
            
        except Exception as e:
            logger.error(f"Error al firmar consentimiento público: {str(e)}")
            return render(request, 'citas/consentimientos/firmar_consentimiento_publico.html', {
                'error': f'Error al firmar el consentimiento: {str(e)}',
                'consentimiento': consentimiento
            })
    
    # GET: Mostrar formulario de firma
    return render(request, 'citas/consentimientos/firmar_consentimiento_publico.html', {
        'consentimiento': consentimiento,
        'cliente': consentimiento.cliente
    })


@login_required
def firmar_consentimiento(request, consentimiento_id):
    """Vista para firmar un consentimiento informado (presencial o desde sistema)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        perfil = Perfil.objects.get(user=request.user)
        # Permitir tanto administrativos como dentistas para firma presencial
        if not (perfil.es_administrativo() or perfil.es_dentista()):
            return JsonResponse({'error': 'No tienes permisos para firmar consentimientos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'error': 'No tienes permisos para acceder a esta función.'}, status=403)
    
    consentimiento = get_object_or_404(ConsentimientoInformado, id=consentimiento_id)
    
    if consentimiento.esta_firmado:
        return JsonResponse({'error': 'Este consentimiento ya está firmado.'}, status=400)
    
    try:
        firma_paciente = request.POST.get('firma_paciente', '')
        nombre_firmante = request.POST.get('nombre_firmante', consentimiento.cliente.nombre_completo)
        rut_firmante = request.POST.get('rut_firmante', consentimiento.cliente.rut or '')
        nombre_testigo = request.POST.get('nombre_testigo', '')
        rut_testigo = request.POST.get('rut_testigo', '')
        firma_testigo = request.POST.get('firma_testigo', '')
        declaracion_comprension = request.POST.get('declaracion_comprension') == 'on'
        derecho_revocacion = request.POST.get('derecho_revocacion') == 'on'
        
        if not firma_paciente:
            return JsonResponse({'error': 'La firma del paciente es obligatoria.'}, status=400)
        
        if not declaracion_comprension:
            return JsonResponse({'error': 'Debe confirmar la declaración de comprensión.'}, status=400)
        
        if not derecho_revocacion:
            return JsonResponse({'error': 'Debe confirmar que conoce su derecho de revocación.'}, status=400)
        
        consentimiento.firma_paciente = firma_paciente
        consentimiento.nombre_firmante = nombre_firmante
        consentimiento.rut_firmante = rut_firmante
        consentimiento.nombre_testigo = nombre_testigo if nombre_testigo else None
        consentimiento.rut_testigo = rut_testigo if rut_testigo else None
        consentimiento.firma_testigo = firma_testigo if firma_testigo else None
        consentimiento.declaracion_comprension = declaracion_comprension
        consentimiento.derecho_revocacion = derecho_revocacion
        consentimiento.estado = 'firmado'
        consentimiento.fecha_firma = timezone.now()
        consentimiento.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Consentimiento firmado exitosamente.'
        })
        
    except Exception as e:
        logger.error(f"Error al firmar consentimiento: {str(e)}")
        return JsonResponse({'error': f'Error al firmar el consentimiento: {str(e)}'}, status=500)


@login_required
def exportar_consentimiento_pdf(request, consentimiento_id):
    """Vista para exportar un consentimiento informado a PDF"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not (perfil.es_dentista() or perfil.es_administrativo()):
            messages.error(request, 'No tienes permisos para exportar consentimientos.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    consentimiento = get_object_or_404(ConsentimientoInformado, id=consentimiento_id)
    
    # Obtener información de la clínica
    try:
        from configuracion.models import InformacionClinica
        info_clinica = InformacionClinica.obtener()
        nombre_clinica = info_clinica.nombre_clinica
        direccion_clinica = info_clinica.direccion
        telefono_clinica = info_clinica.telefono
        email_clinica = info_clinica.email
    except:
        nombre_clinica = "Clínica Dental"
        direccion_clinica = ""
        telefono_clinica = ""
        email_clinica = ""
    
    # Crear el buffer para el PDF
    buffer = BytesIO()
    
    # Función para convertir firma base64 a imagen de ReportLab
    def convertir_firma_base64_a_imagen(firma_base64, max_width=3*inch, max_height=1*inch):
        """Convierte una firma en formato base64 a una imagen de ReportLab"""
        if not firma_base64:
            return None
        
        try:
            import base64
            from io import BytesIO
            from PIL import Image as PILImage
            
            # Verificar si es una cadena base64
            if not isinstance(firma_base64, str):
                return None
            
            # Si no empieza con data:image, asumir que es solo base64
            if firma_base64.startswith('data:image'):
                # Extraer solo la parte base64
                firma_base64 = firma_base64.split(',')[1] if ',' in firma_base64 else firma_base64
            
            # Decodificar base64
            imagen_bytes = base64.b64decode(firma_base64)
            imagen_pil = PILImage.open(BytesIO(imagen_bytes))
            
            # Convertir a RGB si es necesario (para PNG con transparencia)
            if imagen_pil.mode in ('RGBA', 'LA', 'P'):
                fondo = PILImage.new('RGB', imagen_pil.size, (255, 255, 255))
                if imagen_pil.mode == 'P':
                    imagen_pil = imagen_pil.convert('RGBA')
                fondo.paste(imagen_pil, mask=imagen_pil.split()[-1] if imagen_pil.mode in ('RGBA', 'LA') else None)
                imagen_pil = fondo
            elif imagen_pil.mode != 'RGB':
                imagen_pil = imagen_pil.convert('RGB')
            
            # Redimensionar si es muy grande
            # Asumir que las imágenes del canvas tienen aproximadamente 96 DPI (estándar web)
            ancho_px, alto_px = imagen_pil.size
            # Convertir max_width y max_height de pulgadas a píxeles (asumiendo 96 DPI para web)
            max_width_px = (float(max_width) / inch) * 96
            max_height_px = (float(max_height) / inch) * 96
            
            # Calcular ratio para mantener proporción
            ratio_ancho = max_width_px / ancho_px if ancho_px > max_width_px else 1
            ratio_alto = max_height_px / alto_px if alto_px > max_height_px else 1
            ratio = min(ratio_ancho, ratio_alto)
            
            if ratio < 1:
                nuevo_ancho = int(ancho_px * ratio)
                nuevo_alto = int(alto_px * ratio)
                imagen_pil = imagen_pil.resize((nuevo_ancho, nuevo_alto), PILImage.Resampling.LANCZOS)
            
            # Guardar en BytesIO
            img_buffer = BytesIO()
            imagen_pil.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Crear Image de ReportLab
            # ReportLab usa puntos (72 puntos = 1 pulgada)
            # Convertir píxeles a pulgadas (asumiendo 96 DPI) y luego a puntos
            ancho_final, alto_final = imagen_pil.size
            # Convertir píxeles a pulgadas (96 píxeles = 1 pulgada para imágenes web)
            width_inches = ancho_final / 96.0
            height_inches = alto_final / 96.0
            
            # Limitar al máximo permitido y convertir a puntos
            width_final = min(width_inches * inch, max_width)
            height_final = min(height_inches * inch, max_height)
            
            return Image(img_buffer, width=width_final, height=height_final)
        except Exception as e:
            logger.error(f"Error al convertir firma base64 a imagen: {str(e)}")
            return None
    
    # Función para agregar marca de agua con logo de diente
    def add_watermark(canvas_obj, doc_obj):
        """Agregar marca de agua con logo de diente en todas las páginas"""
        canvas_obj.saveState()
        # Color turquesa transparente para la marca de agua
        canvas_obj.setFillColor(colors.HexColor('#14b8a6'), alpha=0.06)
        canvas_obj.setFont('Helvetica-Bold', 150)
        # Rotar el texto 45 grados
        canvas_obj.rotate(45)
        # Posicionar en el centro de la página (ajustado para A4)
        # A4: 8.27 x 11.69 pulgadas
        canvas_obj.drawCentredString(4.5*inch, -2.5*inch, '🦷')
        canvas_obj.restoreState()
    
    # Crear el documento PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=35,
        leftMargin=35,
        topMargin=50,
        bottomMargin=40,
        onFirstPage=add_watermark,
        onLaterPages=add_watermark
    )
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Estilo para el título principal
    title_style = ParagraphStyle(
        'ConsentimientoTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=12,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e293b'),
        fontName='Helvetica-Bold'
    )
    
    # Estilo para subtítulos con color turquesa
    subtitle_style = ParagraphStyle(
        'ConsentimientoSubtitle',
        parent=styles['Heading2'],
        fontSize=13,
        spaceAfter=10,
        spaceBefore=14,
        alignment=TA_LEFT,
        textColor=colors.HexColor('#14b8a6'),  # Color turquesa
        fontName='Helvetica-Bold',
        borderWidth=0,
        borderPadding=0,
        leftIndent=0,
        rightIndent=0,
    )
    
    # Estilo para subtítulos de secciones (B.1, B.2, etc.)
    section_subtitle_style = ParagraphStyle(
        'SectionSubtitle',
        parent=styles['Heading3'],
        fontSize=11,
        spaceAfter=6,
        spaceBefore=10,
        alignment=TA_LEFT,
        textColor=colors.HexColor('#0d9488'),  # Turquesa más oscuro
        fontName='Helvetica-Bold'
    )
    
    # Estilo para encabezado de clínica
    clinic_header_style = ParagraphStyle(
        'ClinicHeader',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#14b8a6'),  # Color turquesa
        fontName='Helvetica-Bold',
        spaceAfter=8
    )
    
    # Estilo para información de la clínica
    clinic_info_style = ParagraphStyle(
        'ClinicInfo',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#64748b'),
        spaceAfter=12
    )
    
    # Estilo para texto normal
    normal_style = ParagraphStyle(
        'NormalConsentimiento',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6,
        alignment=TA_LEFT,
        leading=14
    )
    
    # Contenido del PDF - Estructura según Ley 20.584
    story = []
    
    # A. ENCABEZADO - Información de la Clínica (mejorado)
    clinic_header = Paragraph(f"<b>{nombre_clinica}</b>", clinic_header_style)
    story.append(clinic_header)
    
    header_info = []
    if direccion_clinica:
        header_info.append(direccion_clinica)
    contact_info = []
    if telefono_clinica:
        contact_info.append(f"Teléfono: {telefono_clinica}")
    if email_clinica:
        contact_info.append(f"Email: {email_clinica}")
    
    if header_info or contact_info:
        header_text = ""
        if header_info:
            header_text += "<br/>".join(header_info)
        if contact_info:
            if header_text:
                header_text += "<br/>"
            header_text += " | ".join(contact_info)
        header_text += f"<br/><b>Fecha de Generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        header = Paragraph(header_text, clinic_info_style)
        story.append(header)
    
    story.append(Spacer(1, 20))
    
    # Título principal con línea decorativa
    title = Paragraph(f"<b>CONSENTIMIENTO INFORMADO</b>", title_style)
    story.append(title)
    story.append(Spacer(1, 8))
    
    # Subtítulo del procedimiento
    if consentimiento.titulo:
        subtitle_proc = Paragraph(f"<i>{consentimiento.titulo}</i>", ParagraphStyle(
            'SubtitleProc',
            parent=styles['Normal'],
            fontSize=12,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#14b8a6'),
            fontName='Helvetica-Oblique',
            spaceAfter=16
        ))
        story.append(subtitle_proc)
    else:
        story.append(Spacer(1, 16))
    
    # A. IDENTIFICACIÓN Y ANTECEDENTES
    identificacion_title = Paragraph("<b>A. IDENTIFICACIÓN Y ANTECEDENTES</b>", subtitle_style)
    story.append(identificacion_title)
    
    # Limpiar RUT para evitar símbolos extraños (como '$' de datos antiguos)
    rut_paciente = (consentimiento.cliente.rut or 'No especificado').replace('$', '')
    
    paciente_data = [
        ['Nombre Completo del Paciente:', consentimiento.cliente.nombre_completo],
        ['RUT:', rut_paciente],
        ['Email:', consentimiento.cliente.email or 'No especificado'],
        ['Teléfono:', consentimiento.cliente.telefono or 'No especificado'],
    ]
    
    paciente_table = Table(paciente_data, colWidths=[2.2*inch, 4.3*inch])
    paciente_table_style = TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecfeff')),  # Fondo turquesa muy claro
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#0d9488')),  # Texto turquesa oscuro
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#374151')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#14b8a6')),  # Borde turquesa
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#14b8a6')),  # Línea inferior turquesa
    ])
    paciente_table.setStyle(paciente_table_style)
    story.append(paciente_table)
    story.append(Spacer(1, 16))
    
    # B. INFORMACIÓN DETALLADA DEL PROCEDIMIENTO
    info_title = Paragraph("<b>B. INFORMACIÓN DETALLADA DEL PROCEDIMIENTO</b>", subtitle_style)
    story.append(info_title)
    story.append(Spacer(1, 8))
    
    # B.1. Diagnóstico y Justificación
    if consentimiento.diagnostico:
        diagnostico_title = Paragraph("<b>B.1. Diagnóstico y Justificación del Tratamiento</b>", section_subtitle_style)
        story.append(diagnostico_title)
        diagnostico_text = Paragraph(consentimiento.diagnostico or 'No especificado', normal_style)
        story.append(diagnostico_text)
        if consentimiento.justificacion:
            justificacion_text = Paragraph(consentimiento.justificacion or 'No especificado', normal_style)
            story.append(justificacion_text)
        story.append(Spacer(1, 12))
    
    # B.2. Naturaleza y Objetivos del Tratamiento
    if consentimiento.naturaleza_procedimiento or consentimiento.objetivos_tratamiento:
        naturaleza_title = Paragraph("<b>B.2. Naturaleza y Objetivos del Tratamiento</b>", section_subtitle_style)
        story.append(naturaleza_title)
        if consentimiento.naturaleza_procedimiento:
            naturaleza_text = Paragraph(f"<b>Naturaleza del Procedimiento:</b><br/>{consentimiento.naturaleza_procedimiento or 'No especificado'}", normal_style)
            story.append(naturaleza_text)
        if consentimiento.objetivos_tratamiento:
            objetivos_text = Paragraph(f"<b>Objetivos del Tratamiento:</b><br/>{consentimiento.objetivos_tratamiento or 'No especificado'}", normal_style)
            story.append(objetivos_text)
        story.append(Spacer(1, 12))
    
    # B.3. Contenido General del Consentimiento
    if consentimiento.contenido:
        contenido_title = Paragraph("<b>B.3. Información General del Procedimiento</b>", section_subtitle_style)
        story.append(contenido_title)
        contenido_text = Paragraph(consentimiento.contenido or 'No especificado', normal_style)
        story.append(contenido_text)
        story.append(Spacer(1, 12))
    
    # B.4. Alternativas de Tratamiento (OBLIGATORIO - Ley 20.584)
    if consentimiento.alternativas:
        alternativas_title = Paragraph("<b>B.4. Alternativas de Tratamiento</b>", section_subtitle_style)
        story.append(alternativas_title)
        alternativas_text = Paragraph(consentimiento.alternativas or 'No especificado', normal_style)
        story.append(alternativas_text)
        story.append(Spacer(1, 12))
    
    # B.5. Riesgos y Complicaciones Relevantes (OBLIGATORIO - Ley 20.584)
    if consentimiento.riesgos:
        riesgos_title = Paragraph("<b>B.5. Riesgos y Complicaciones Relevantes</b>", section_subtitle_style)
        story.append(riesgos_title)
        riesgos_text = Paragraph(consentimiento.riesgos or 'No especificado', normal_style)
        story.append(riesgos_text)
        story.append(Spacer(1, 12))
    
    # B.6. Beneficios Esperados
    if consentimiento.beneficios:
        beneficios_title = Paragraph("<b>B.6. Beneficios Esperados</b>", section_subtitle_style)
        story.append(beneficios_title)
        beneficios_text = Paragraph(consentimiento.beneficios or 'No especificado', normal_style)
        story.append(beneficios_text)
        story.append(Spacer(1, 12))
    
    # B.7. Pronóstico
    if consentimiento.pronostico:
        pronostico_title = Paragraph("<b>B.7. Pronóstico</b>", section_subtitle_style)
        story.append(pronostico_title)
        pronostico_text = Paragraph(consentimiento.pronostico or 'No especificado', normal_style)
        story.append(pronostico_text)
        story.append(Spacer(1, 12))
    
    # B.8. Cuidados Postoperatorios
    if consentimiento.cuidados_postoperatorios:
        cuidados_title = Paragraph("<b>B.8. Cuidados Postoperatorios</b>", section_subtitle_style)
        story.append(cuidados_title)
        cuidados_text = Paragraph(consentimiento.cuidados_postoperatorios or 'No especificado', normal_style)
        story.append(cuidados_text)
        story.append(Spacer(1, 16))
    
    # C. DECLARACIÓN DEL PACIENTE Y FIRMAS (Ley 20.584)
    declaracion_title = Paragraph("<b>C. DECLARACIÓN DEL PACIENTE Y FIRMAS</b>", subtitle_style)
    story.append(declaracion_title)
    story.append(Spacer(1, 10))
    
    # Declaración de Comprensión (Ley 20.584) - con fondo turquesa
    declaracion_box_style = ParagraphStyle(
        'DeclaracionBox',
        parent=normal_style,
        backColor=colors.HexColor('#ecfeff'),  # Fondo turquesa muy claro
        borderColor=colors.HexColor('#14b8a6'),  # Borde turquesa
        borderWidth=1,
        borderPadding=10,
        leftIndent=0,
        rightIndent=0,
    )
    declaracion_text = Paragraph(
        "<b>DECLARACIÓN DE COMPRENSIÓN:</b><br/>"
        "Yo, el paciente o su representante legal, declaro que he sido informado de forma clara, comprensible y oportuna "
        "sobre mi diagnóstico, los riesgos, beneficios y alternativas del procedimiento propuesto, de acuerdo con lo "
        "establecido en la <b>Ley N° 20.584</b> sobre Derechos y Deberes de las Personas en relación con las Acciones vinculadas "
        "a su Atención en Salud.",
        declaracion_box_style
    )
    story.append(declaracion_text)
    story.append(Spacer(1, 10))
    
    # Derecho de Revocación (Ley 20.584) - con fondo turquesa
    revocacion_text = Paragraph(
        "<b>DERECHO DE REVOCACIÓN:</b><br/>"
        "Conozco que tengo el derecho a revocar libremente este consentimiento en cualquier momento previo al inicio del tratamiento.",
        declaracion_box_style
    )
    story.append(revocacion_text)
    story.append(Spacer(1, 16))
    
    # Espacios para Firmas
    firmas_title = Paragraph("<b>FIRMAS</b>", subtitle_style)
    story.append(firmas_title)
    story.append(Spacer(1, 12))
    
    # Estilo para títulos de sección de firmas
    firma_section_title_style = ParagraphStyle(
        'FirmaSectionTitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#14b8a6'),
        fontName='Helvetica-Bold',
        spaceAfter=6,
        spaceBefore=0
    )
    
    # Tabla de firmas - Estructura mejorada y organizada
    firmas_data = []
    row_idx = 0
    title_rows = []  # Filas que son títulos de sección (para aplicar SPAN)
    
    # Paciente
    if consentimiento.esta_firmado:
        # Título de sección que ocupará ambas columnas
        firmas_data.append([Paragraph('<b>PACIENTE</b>', firma_section_title_style), ''])
        title_rows.append(row_idx)
        row_idx += 1
        
        firmas_data.append(['Nombre:', consentimiento.nombre_firmante or consentimiento.cliente.nombre_completo])
        row_idx += 1
        
        if consentimiento.rut_firmante:
            firmas_data.append(['RUT:', consentimiento.rut_firmante])
            row_idx += 1
        
        # Convertir firma base64 a imagen si es necesario
        firma_paciente_img = convertir_firma_base64_a_imagen(consentimiento.firma_paciente, max_width=2.5*inch, max_height=0.8*inch)
        if firma_paciente_img:
            firmas_data.append(['Firma:', firma_paciente_img])
        else:
            firmas_data.append(['Firma:', consentimiento.firma_paciente or '_________________________'])
        row_idx += 1
        
        firmas_data.append(['Fecha y Hora:', consentimiento.fecha_firma.strftime('%d/%m/%Y %H:%M') if consentimiento.fecha_firma else ''])
        row_idx += 1
    else:
        firmas_data.append([Paragraph('<b>PACIENTE</b>', firma_section_title_style), Paragraph('<i>Pendiente de firma</i>', normal_style)])
        title_rows.append(row_idx)
        row_idx += 1
    
    # Separador visual
    firmas_data.append(['', ''])
    row_idx += 1
    
    # Profesional Tratante
    if consentimiento.dentista:
        # Título de sección que ocupará ambas columnas
        firmas_data.append([Paragraph('<b>PROFESIONAL TRATANTE</b>', firma_section_title_style), ''])
        title_rows.append(row_idx)
        row_idx += 1
        
        firmas_data.append(['Nombre:', consentimiento.dentista.nombre_completo])
        row_idx += 1
        
        if consentimiento.rut_dentista:
            firmas_data.append(['RUT:', consentimiento.rut_dentista])
            row_idx += 1
        
        if consentimiento.registro_superintendencia:
            firmas_data.append(['Registro Superintendencia de Salud:', consentimiento.registro_superintendencia])
            row_idx += 1
        
        firmas_data.append(['Firma:', '_________________________'])
        row_idx += 1
        
        # Espacio
        firmas_data.append(['', ''])
        row_idx += 1
    
    # Testigo (opcional pero recomendado)
    if consentimiento.nombre_testigo:
        firmas_data.append([Paragraph('<b>TESTIGO</b>', firma_section_title_style), ''])
        title_rows.append(row_idx)
        row_idx += 1
        
        firmas_data.append(['Nombre:', consentimiento.nombre_testigo])
        row_idx += 1
        
        if consentimiento.rut_testigo:
            firmas_data.append(['RUT:', consentimiento.rut_testigo])
            row_idx += 1
        
        # Convertir firma testigo base64 a imagen si es necesario
        firma_testigo_img = convertir_firma_base64_a_imagen(consentimiento.firma_testigo, max_width=2.5*inch, max_height=0.8*inch)
        if firma_testigo_img:
            firmas_data.append(['Firma:', firma_testigo_img])
        else:
            firmas_data.append(['Firma:', consentimiento.firma_testigo or '_________________________'])
        row_idx += 1
    
    if firmas_data:
        firmas_table = Table(firmas_data, colWidths=[2.4*inch, 4.1*inch])
        
        # Construir lista de estilos
        style_list = [
            # Estilo base para la primera columna (etiquetas)
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecfeff')),  # Fondo turquesa muy claro
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#0d9488')),  # Texto turquesa oscuro
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, -1), 10),
            
            # Estilo para la segunda columna (valores)
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#374151')),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (1, 0), (1, -1), 10),
            
            # Alineación
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#14b8a6')),  # Borde turquesa
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#14b8a6')),  # Línea superior más gruesa
            
            # Padding base
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]
        
        # Aplicar SPAN a los títulos de sección para que ocupen ambas columnas
        for title_row in title_rows:
            style_list.append(('SPAN', (0, title_row), (1, title_row)))  # Ocupar ambas columnas
            style_list.append(('BACKGROUND', (0, title_row), (1, title_row), colors.HexColor('#b2f5ea')))  # Fondo turquesa más destacado
            style_list.append(('ALIGN', (0, title_row), (1, title_row), 'LEFT'))
            style_list.append(('TOPPADDING', (0, title_row), (1, title_row), 12))
            style_list.append(('BOTTOMPADDING', (0, title_row), (1, title_row), 12))
        
        # Aplicar estilos a filas vacías (separadores)
        for i, row in enumerate(firmas_data):
            if len(row) >= 2 and (row[0] == '' or row[0] is None) and (row[1] == '' or row[1] is None):
                style_list.append(('BACKGROUND', (0, i), (1, i), colors.HexColor('#ffffff')))
                style_list.append(('TOPPADDING', (0, i), (1, i), 4))
                style_list.append(('BOTTOMPADDING', (0, i), (1, i), 4))
        
        firmas_table_style = TableStyle(style_list)
        firmas_table.setStyle(firmas_table_style)
        story.append(firmas_table)
        story.append(Spacer(1, 16))
    
    # Footer con información legal (mejorado)
    story.append(Spacer(1, 12))
    footer_text = f"<b>Documento generado el:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    if consentimiento.fecha_vencimiento:
        footer_text += f" | <b>Válido hasta:</b> {consentimiento.fecha_vencimiento.strftime('%d/%m/%Y')}"
    footer_text += "<br/><i>Este documento cumple con los requisitos de la <b>Ley N° 20.584</b> sobre Derechos y Deberes de las Personas en relación con las Acciones vinculadas a su Atención en Salud.</i>"
    
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=clinic_info_style,
        fontSize=8,
        textColor=colors.HexColor('#14b8a6'),  # Color turquesa para el footer
        spaceBefore=8,
        borderColor=colors.HexColor('#14b8a6'),
        borderWidth=1,
        borderPadding=8,
        backColor=colors.HexColor('#f0fdfa'),  # Fondo turquesa muy claro
    )
    footer = Paragraph(footer_text, footer_style)
    story.append(footer)
    
    # Construir el PDF
    try:
        doc.build(story)
    except Exception as e:
        logger.error(f"Error al construir PDF del consentimiento {consentimiento_id}: {str(e)}")
        buffer.close()
        messages.error(request, f'Error al generar el PDF: {str(e)}')
        return redirect('gestor_consentimientos')
    
    # Obtener el contenido del buffer
    pdf_content = buffer.getvalue()
    buffer.close()
    
    # Crear o actualizar el documento en la base de datos (usando consentimiento como referencia única)
    try:
        documento = DocumentoCliente.objects.get(
            tipo='consentimiento',
            cliente=consentimiento.cliente,
            cita=consentimiento.cita,
            plan_tratamiento=consentimiento.plan_tratamiento
        )
        # Actualizar si existe
        documento.titulo = f'Consentimiento - {consentimiento.titulo}'
        documento.descripcion = f'Consentimiento informado para {consentimiento.get_tipo_procedimiento_display()}'
        documento.generado_por = perfil
        documento.fecha_generacion = timezone.now()
        documento.save()
    except DocumentoCliente.DoesNotExist:
        # Crear nuevo si no existe
        documento = DocumentoCliente.objects.create(
            tipo='consentimiento',
            cliente=consentimiento.cliente,
            titulo=f'Consentimiento - {consentimiento.titulo}',
            descripcion=f'Consentimiento informado para {consentimiento.get_tipo_procedimiento_display()}',
            cita=consentimiento.cita,
            plan_tratamiento=consentimiento.plan_tratamiento,
            generado_por=perfil,
        )
    except DocumentoCliente.MultipleObjectsReturned:
        # Si hay múltiples, tomar el más reciente
        documento = DocumentoCliente.objects.filter(
            tipo='consentimiento',
            cliente=consentimiento.cliente,
            cita=consentimiento.cita,
            plan_tratamiento=consentimiento.plan_tratamiento
        ).order_by('-fecha_generacion').first()
        documento.titulo = f'Consentimiento - {consentimiento.titulo}'
        documento.descripcion = f'Consentimiento informado para {consentimiento.get_tipo_procedimiento_display()}'
        documento.generado_por = perfil
        documento.fecha_generacion = timezone.now()
        documento.save()
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    filename = f"consentimiento_{consentimiento.cliente.nombre_completo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf_content)
    
    return response


@login_required
def eliminar_consentimiento(request, consentimiento_id):
    """Elimina un consentimiento informado (solo administrativos, vía AJAX)."""
    from django.http import JsonResponse
    
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({'success': False, 'error': 'No tienes permisos para eliminar consentimientos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos para acceder a esta función.'}, status=403)
    
    consentimiento = get_object_or_404(ConsentimientoInformado, id=consentimiento_id)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    titulo = consentimiento.titulo or 'Consentimiento'
    try:
        consentimiento.delete()
        return JsonResponse({
            'success': True,
            'message': f'El consentimiento \"{titulo}\" fue eliminado correctamente.'
        })
    except Exception as e:
        logger.error(f"Error al eliminar consentimiento {consentimiento_id}: {e}")
        return JsonResponse({'success': False, 'error': 'Error al eliminar el consentimiento.'}, status=500)


@login_required
def firmar_consentimiento_recepcion(request, plan_id, consentimiento_id):
    """Sube un documento firmado físicamente por el paciente para un consentimiento informado (solo administrativos, vía AJAX)."""
    from django.http import JsonResponse
    from django.core.files.storage import default_storage
    import os
    
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({'success': False, 'error': 'No tienes permisos para subir documentos de consentimientos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos para acceder a esta función.'}, status=403)
    
    # Verificar que el plan existe y el consentimiento pertenece al plan
    plan = get_object_or_404(PlanTratamiento, id=plan_id)
    consentimiento = get_object_or_404(ConsentimientoInformado, id=consentimiento_id, plan_tratamiento=plan)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    # Verificar que el consentimiento no esté ya firmado
    if consentimiento.estado == 'firmado':
        return JsonResponse({'success': False, 'error': 'Este consentimiento ya está firmado.'}, status=400)
    
    # Obtener el archivo del formulario
    documento_firmado = request.FILES.get('documento_firmado_fisico')
    
    # Validaciones
    if not documento_firmado:
        return JsonResponse({'success': False, 'error': 'Debes subir el documento firmado físicamente por el paciente.'}, status=400)
    
    # Validar tipo de archivo
    extensiones_permitidas = ['.pdf', '.jpg', '.jpeg', '.png', '.heic', '.heif']
    nombre_archivo = documento_firmado.name.lower()
    extension_valida = any(nombre_archivo.endswith(ext) for ext in extensiones_permitidas)
    
    if not extension_valida:
        return JsonResponse({
            'success': False, 
            'error': 'Formato de archivo no válido. Se aceptan: PDF, JPG, PNG, HEIC.'
        }, status=400)
    
    # Validar tamaño (máx 10MB)
    if documento_firmado.size > 10 * 1024 * 1024:
        return JsonResponse({
            'success': False, 
            'error': 'El archivo es demasiado grande. El tamaño máximo es 10MB.'
        }, status=400)
    
    try:
        # Si ya existe un documento, eliminarlo antes de subir el nuevo
        if consentimiento.documento_firmado_fisico:
            try:
                if os.path.isfile(consentimiento.documento_firmado_fisico.path):
                    os.remove(consentimiento.documento_firmado_fisico.path)
            except Exception as e:
                logger.warning(f"No se pudo eliminar el archivo anterior del consentimiento {consentimiento_id}: {e}")
        
        # Actualizar el consentimiento con el documento
        consentimiento.documento_firmado_fisico = documento_firmado
        consentimiento.subido_por = perfil
        consentimiento.fecha_subida = timezone.now()
        consentimiento.nombre_firmante = plan.cliente.nombre_completo
        consentimiento.rut_firmante = plan.cliente.rut or ''
        consentimiento.firmado_por_recepcion = True
        consentimiento.recepcionista_firmante = perfil
        consentimiento.declaracion_comprension = True  # Se asume que el paciente firmó, por lo tanto comprendió
        consentimiento.derecho_revocacion = True  # Se asume que el paciente firmó, por lo tanto conoce su derecho
        consentimiento.estado = 'firmado'
        consentimiento.fecha_firma = timezone.now()
        consentimiento.save()
        
        return JsonResponse({
            'success': True,
            'message': f'El documento firmado del consentimiento "{consentimiento.titulo}" fue subido exitosamente.'
        })
    except Exception as e:
        logger.error(f"Error al subir documento firmado para consentimiento {consentimiento_id}: {e}")
        return JsonResponse({'success': False, 'error': 'Error al subir el documento. Por favor, intenta nuevamente.'}, status=500)


@login_required
def descargar_documento_firmado_consentimiento(request, consentimiento_id):
    """Descarga el documento firmado físicamente de un consentimiento (solo administrativos)."""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para descargar documentos de consentimientos.')
            return redirect('gestor_consentimientos')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    consentimiento = get_object_or_404(ConsentimientoInformado, id=consentimiento_id)
    
    if not consentimiento.documento_firmado_fisico:
        messages.error(request, 'Este consentimiento no tiene un documento firmado subido.')
        return redirect('detalle_consentimiento', consentimiento_id=consentimiento.id)
    
    try:
        from django.http import FileResponse
        import os
        
        if os.path.exists(consentimiento.documento_firmado_fisico.path):
            response = FileResponse(
                open(consentimiento.documento_firmado_fisico.path, 'rb'),
                content_type='application/pdf' if consentimiento.documento_firmado_fisico.name.endswith('.pdf') else 'image/jpeg'
            )
            nombre_archivo = os.path.basename(consentimiento.documento_firmado_fisico.name)
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            return response
        else:
            messages.error(request, 'El archivo no se encuentra en el servidor.')
            return redirect('detalle_consentimiento', consentimiento_id=consentimiento.id)
    except Exception as e:
        logger.error(f"Error al descargar documento firmado del consentimiento {consentimiento_id}: {e}")
        messages.error(request, 'Error al descargar el documento.')
        return redirect('detalle_consentimiento', consentimiento_id=consentimiento.id)


@login_required
def obtener_plantilla_consentimiento(request, plantilla_id):
    """Vista AJAX para obtener datos de una plantilla de consentimiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({'error': 'No tienes permisos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'error': 'No tienes permisos.'}, status=403)
    
    plantilla = get_object_or_404(PlantillaConsentimiento, id=plantilla_id)
    
    return JsonResponse({
        'success': True,
        'plantilla': {
            'nombre': plantilla.nombre,
            'tipo_procedimiento': plantilla.tipo_procedimiento,
            'diagnostico_base': plantilla.diagnostico_base or '',
            'naturaleza_procedimiento': plantilla.naturaleza_procedimiento or '',
            'objetivos_tratamiento': plantilla.objetivos_tratamiento or '',
            'contenido': plantilla.contenido,
            'riesgos': plantilla.riesgos or '',
            'beneficios': plantilla.beneficios or '',
            'alternativas': plantilla.alternativas or '',
            'pronostico': plantilla.pronostico or '',
            'cuidados_postoperatorios': plantilla.cuidados_postoperatorios or '',
        }
    })


# ============================================================================
# VISTAS DE GESTIÓN DE EVALUACIONES
# ============================================================================

@login_required
def marcar_evaluacion_revisada(request, evaluacion_id):
    """
    Marca una evaluación como revisada
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('login')
    
    try:
        evaluacion = Evaluacion.objects.get(id=evaluacion_id)
        
        if evaluacion.estado == 'pendiente':
            evaluacion.marcar_como_revisada(perfil)
            messages.success(request, f'Evaluación de {evaluacion.cliente.nombre_completo} marcada como revisada.')
        else:
            messages.info(request, 'Esta evaluación ya fue revisada anteriormente.')
    
    except Evaluacion.DoesNotExist:
        messages.error(request, 'Evaluación no encontrada.')
    
    return redirect('gestor_evaluaciones')


@login_required
def archivar_evaluacion(request, evaluacion_id):
    """
    Archiva una evaluación
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('login')
    
    try:
        evaluacion = Evaluacion.objects.get(id=evaluacion_id)
        evaluacion.archivar()
        messages.success(request, f'Evaluación de {evaluacion.cliente.nombre_completo} archivada.')
    
    except Evaluacion.DoesNotExist:
        messages.error(request, 'Evaluación no encontrada.')
    
    return redirect('gestor_evaluaciones')


@login_required
def eliminar_evaluacion(request, evaluacion_id):
    """
    Elimina una evaluación (solo admin)
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para eliminar evaluaciones.')
            return redirect('gestor_evaluaciones')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('login')
    
    try:
        evaluacion = Evaluacion.objects.get(id=evaluacion_id)
        nombre_cliente = evaluacion.cliente.nombre_completo
        evaluacion.delete()
        messages.success(request, f'Evaluación de {nombre_cliente} eliminada permanentemente.')
    
    except Evaluacion.DoesNotExist:
        messages.error(request, 'Evaluación no encontrada.')
    
    return redirect('gestor_evaluaciones')


@login_required
def gestor_finanzas(request):
    """
    Vista dedicada para gestionar las finanzas de la clínica
    Muestra ingresos de citas completadas y estadísticas financieras
    SOLO DISPONIBLE PARA ADMINISTRATIVOS
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
        
        # Solo permitir acceso a administrativos
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para acceder a esta sección. Solo el personal administrativo puede ver las finanzas.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('login')
    
    # Obtener fechas para filtros
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    fin_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    inicio_año = hoy.replace(month=1, day=1)
    
    # Obtener todas las citas completadas (excluir las que tienen precio_cobrado = None y no tienen tipo_servicio, que fueron eliminadas del historial)
    # Nota: Las citas que tienen precio_cobrado = None pero tienen tipo_servicio aún se incluyen porque tienen precio base
    citas_completadas = Cita.objects.filter(
        estado='completada'
    ).exclude(
        precio_cobrado__isnull=True,
        tipo_servicio__isnull=True
    ).select_related('tipo_servicio', 'cliente', 'dentista').order_by('-fecha_hora')
    
    # ===== INGRESOS MANUALES =====
    ingresos_manuales = IngresoManual.objects.all().select_related('creado_por').order_by('-fecha', '-creado_el')
    
    # Estadísticas financieras generales
    # Calcular ingresos considerando precio_cobrado o precio del servicio
    total_ingresos = 0
    for cita in citas_completadas:
        if cita.precio_cobrado:
            total_ingresos += float(cita.precio_cobrado)
        elif cita.tipo_servicio and cita.tipo_servicio.precio_base:
            total_ingresos += float(cita.tipo_servicio.precio_base)
    
    # Sumar ingresos manuales
    for ingreso_manual in ingresos_manuales:
        total_ingresos += float(ingreso_manual.monto)
    
    # Ingresos del mes actual
    citas_completadas_mes = citas_completadas.filter(
        fecha_hora__date__gte=inicio_mes,
        fecha_hora__date__lte=fin_mes
    )
    ingresos_mes = 0
    for cita in citas_completadas_mes:
        if cita.precio_cobrado:
            ingresos_mes += float(cita.precio_cobrado)
        elif cita.tipo_servicio and cita.tipo_servicio.precio_base:
            ingresos_mes += float(cita.tipo_servicio.precio_base)
    
    # Sumar ingresos manuales del mes
    ingresos_manuales_mes = ingresos_manuales.filter(
        fecha__gte=inicio_mes,
        fecha__lte=fin_mes
    )
    for ingreso_manual in ingresos_manuales_mes:
        ingresos_mes += float(ingreso_manual.monto)
    
    # Ingresos del año actual
    citas_completadas_año = citas_completadas.filter(
        fecha_hora__date__gte=inicio_año
    )
    ingresos_año = 0
    for cita in citas_completadas_año:
        if cita.precio_cobrado:
            ingresos_año += float(cita.precio_cobrado)
        elif cita.tipo_servicio and cita.tipo_servicio.precio_base:
            ingresos_año += float(cita.tipo_servicio.precio_base)
    
    # Sumar ingresos manuales del año
    ingresos_manuales_año = ingresos_manuales.filter(
        fecha__gte=inicio_año
    )
    for ingreso_manual in ingresos_manuales_año:
        ingresos_año += float(ingreso_manual.monto)
    
    # Cantidad de citas
    total_citas_completadas = citas_completadas.count()
    citas_mes = citas_completadas_mes.count()
    citas_año = citas_completadas_año.count()
    
    # Ingresos por servicio (calcular manualmente para considerar precio_cobrado o precio del servicio)
    ingresos_por_servicio_dict = {}
    for cita in citas_completadas:
        servicio_nombre = cita.tipo_servicio.nombre if cita.tipo_servicio else 'Sin servicio'
        servicio_categoria = cita.tipo_servicio.categoria if cita.tipo_servicio else 'otros'
        
        if servicio_nombre not in ingresos_por_servicio_dict:
            ingresos_por_servicio_dict[servicio_nombre] = {
                'tipo_servicio__nombre': servicio_nombre,
                'tipo_servicio__categoria': servicio_categoria,
                'total_ingresos': 0,
                'cantidad_citas': 0
            }
        
        ingresos_por_servicio_dict[servicio_nombre]['cantidad_citas'] += 1
        if cita.precio_cobrado:
            ingresos_por_servicio_dict[servicio_nombre]['total_ingresos'] += float(cita.precio_cobrado)
        elif cita.tipo_servicio and cita.tipo_servicio.precio_base:
            ingresos_por_servicio_dict[servicio_nombre]['total_ingresos'] += float(cita.tipo_servicio.precio_base)
    
    ingresos_por_servicio = sorted(ingresos_por_servicio_dict.values(), key=lambda x: x['total_ingresos'], reverse=True)
    
    # Ingresos por mes (últimos 12 meses)
    ingresos_por_mes = []
    from calendar import monthrange
    for i in range(11, -1, -1):
        # Calcular fecha del mes
        mes_fecha = hoy - timedelta(days=30 * i)
        mes_inicio = mes_fecha.replace(day=1)
        # Obtener último día del mes
        ultimo_dia = monthrange(mes_fecha.year, mes_fecha.month)[1]
        mes_fin = mes_fecha.replace(day=ultimo_dia)
        
        citas_mes_actual = citas_completadas.filter(
            fecha_hora__date__gte=mes_inicio,
            fecha_hora__date__lte=mes_fin
        )
        
        ingresos_mes_actual = 0
        for cita in citas_mes_actual:
            if cita.precio_cobrado:
                ingresos_mes_actual += float(cita.precio_cobrado)
            elif cita.tipo_servicio and cita.tipo_servicio.precio_base:
                ingresos_mes_actual += float(cita.tipo_servicio.precio_base)
        
        ingresos_por_mes.append({
            'mes': mes_fecha.strftime('%B %Y'),
            'ingresos': ingresos_mes_actual,
            'citas': citas_mes_actual.count()
        })
    
    # Precio promedio por cita
    if total_citas_completadas > 0:
        precio_promedio = total_ingresos / total_citas_completadas
    else:
        precio_promedio = 0
    
    # Ingresos por dentista (calcular manualmente)
    ingresos_por_dentista_dict = {}
    for cita in citas_completadas.filter(dentista__isnull=False):
        if not cita.dentista:
            continue
        dentista_nombre = cita.dentista.nombre_completo
        
        if dentista_nombre not in ingresos_por_dentista_dict:
            ingresos_por_dentista_dict[dentista_nombre] = {
                'dentista__nombre_completo': dentista_nombre,
                'total_ingresos': 0,
                'cantidad_citas': 0
            }
        
        ingresos_por_dentista_dict[dentista_nombre]['cantidad_citas'] += 1
        if cita.precio_cobrado:
            ingresos_por_dentista_dict[dentista_nombre]['total_ingresos'] += float(cita.precio_cobrado)
        elif cita.tipo_servicio and cita.tipo_servicio.precio_base:
            ingresos_por_dentista_dict[dentista_nombre]['total_ingresos'] += float(cita.tipo_servicio.precio_base)
    
    ingresos_por_dentista = sorted(ingresos_por_dentista_dict.values(), key=lambda x: x['total_ingresos'], reverse=True)
    
    # Combinar todos los ingresos (citas + manuales) para la tabla
    todos_ingresos_list = []
    
    # Agregar citas completadas (solo las que tienen precio o servicio)
    for cita in citas_completadas:
        # Solo incluir citas que tienen precio_cobrado o tipo_servicio (no las eliminadas del historial)
        if cita.precio_cobrado is not None or (cita.tipo_servicio and cita.tipo_servicio.precio_base):
            monto_cita = 0
            if cita.precio_cobrado:
                monto_cita = float(cita.precio_cobrado)
            elif cita.tipo_servicio and cita.tipo_servicio.precio_base:
                monto_cita = float(cita.tipo_servicio.precio_base)
            
            todos_ingresos_list.append({
                'cita': cita,
                'monto': monto_cita,
                'tipo': 'cita',
                'fecha': cita.fecha_hora
            })
    
    # Agregar ingresos manuales
    for ingreso_manual in ingresos_manuales:
        from datetime import datetime
        # Crear datetime aware para que sea compatible con los otros
        fecha_naive = datetime.combine(ingreso_manual.fecha, datetime.min.time())
        fecha_ingreso = timezone.make_aware(fecha_naive)
        todos_ingresos_list.append({
            'ingreso_manual': ingreso_manual,
            'monto': float(ingreso_manual.monto),
            'tipo': 'ingreso_manual',
            'fecha': fecha_ingreso
        })
    
    # Ordenar por fecha (más recientes primero)
    todos_ingresos_list.sort(key=lambda x: x['fecha'], reverse=True)
    todos_ingresos = todos_ingresos_list[:50]  # Mostrar los últimos 50
    
    # Citas recientes (últimas 10) - para compatibilidad
    citas_recientes = citas_completadas[:10]
    
    # ===== EGRESOS (Gastos) =====
    # Obtener movimientos de insumos tipo "entrada" (compras) como egresos
    # Excluir movimientos creados por recepción de solicitudes (para evitar duplicación)
    from inventario.models import MovimientoInsumo
    from proveedores.models import SolicitudInsumo
    movimientos_entrada = MovimientoInsumo.objects.filter(
        tipo='entrada'
    ).exclude(
        motivo__startswith='Recepción de solicitud'
    ).select_related('insumo', 'realizado_por').order_by('-fecha_movimiento')
    
    # Obtener solicitudes de insumos marcadas como egreso automático
    # Solo mostrar solicitudes que NO han sido recibidas (para evitar duplicación con egresos manuales)
    solicitudes_egreso = SolicitudInsumo.objects.filter(
        monto_egreso__isnull=False,
        estado__in=['pendiente', 'enviada']  # Solo pendientes o enviadas, no recibidas
    ).select_related('insumo', 'proveedor', 'solicitado_por').order_by('-fecha_solicitud')
    
    # Obtener egresos manuales
    egresos_manuales = EgresoManual.objects.all().select_related('creado_por').order_by('-fecha', '-creado_el')
    
    # Calcular egresos totales (basado en compras de insumos con precio)
    total_egresos = 0
    for movimiento in movimientos_entrada:
        if movimiento.insumo and movimiento.insumo.precio_unitario:
            total_egresos += float(movimiento.insumo.precio_unitario) * movimiento.cantidad
    
    # Sumar egresos automáticos de solicitudes
    for solicitud in solicitudes_egreso:
        if solicitud.monto_egreso:
            total_egresos += float(solicitud.monto_egreso)
    
    # Sumar egresos manuales
    for egreso_manual in egresos_manuales:
        total_egresos += float(egreso_manual.monto)
    
    # Egresos del mes
    movimientos_mes = movimientos_entrada.filter(
        fecha_movimiento__date__gte=inicio_mes,
        fecha_movimiento__date__lte=fin_mes
    )
    egresos_mes = 0
    for movimiento in movimientos_mes:
        if movimiento.insumo and movimiento.insumo.precio_unitario:
            egresos_mes += float(movimiento.insumo.precio_unitario) * movimiento.cantidad
    
    # Sumar egresos automáticos del mes
    solicitudes_mes = solicitudes_egreso.filter(
        fecha_solicitud__date__gte=inicio_mes,
        fecha_solicitud__date__lte=fin_mes
    )
    for solicitud in solicitudes_mes:
        if solicitud.monto_egreso:
            egresos_mes += float(solicitud.monto_egreso)
    
    # Sumar egresos manuales del mes
    egresos_manuales_mes = egresos_manuales.filter(
        fecha__gte=inicio_mes,
        fecha__lte=fin_mes
    )
    for egreso_manual in egresos_manuales_mes:
        egresos_mes += float(egreso_manual.monto)
    
    # Egresos del año
    movimientos_año = movimientos_entrada.filter(
        fecha_movimiento__date__gte=inicio_año
    )
    egresos_año = 0
    for movimiento in movimientos_año:
        if movimiento.insumo and movimiento.insumo.precio_unitario:
            egresos_año += float(movimiento.insumo.precio_unitario) * movimiento.cantidad
    
    # Sumar egresos automáticos del año
    solicitudes_año = solicitudes_egreso.filter(
        fecha_solicitud__date__gte=inicio_año
    )
    for solicitud in solicitudes_año:
        if solicitud.monto_egreso:
            egresos_año += float(solicitud.monto_egreso)
    
    # Sumar egresos manuales del año
    egresos_manuales_año = egresos_manuales.filter(
        fecha__gte=inicio_año
    )
    for egreso_manual in egresos_manuales_año:
        egresos_año += float(egreso_manual.monto)
    
    # Egresos por mes (últimos 12 meses)
    egresos_por_mes = []
    from calendar import monthrange
    for i in range(11, -1, -1):
        mes_fecha = hoy - timedelta(days=30 * i)
        mes_inicio = mes_fecha.replace(day=1)
        ultimo_dia = monthrange(mes_fecha.year, mes_fecha.month)[1]
        mes_fin = mes_fecha.replace(day=ultimo_dia)
        
        movimientos_mes_actual = movimientos_entrada.filter(
            fecha_movimiento__date__gte=mes_inicio,
            fecha_movimiento__date__lte=mes_fin
        )
        
        egresos_mes_actual = 0
        for movimiento in movimientos_mes_actual:
            if movimiento.insumo and movimiento.insumo.precio_unitario:
                egresos_mes_actual += float(movimiento.insumo.precio_unitario) * movimiento.cantidad
        
        # Sumar egresos automáticos del mes
        solicitudes_mes_actual = solicitudes_egreso.filter(
            fecha_solicitud__date__gte=mes_inicio,
            fecha_solicitud__date__lte=mes_fin
        )
        for solicitud in solicitudes_mes_actual:
            if solicitud.monto_egreso:
                egresos_mes_actual += float(solicitud.monto_egreso)
        
        # Sumar egresos manuales del mes
        egresos_manuales_mes_actual = egresos_manuales.filter(
            fecha__gte=mes_inicio,
            fecha__lte=mes_fin
        )
        for egreso_manual in egresos_manuales_mes_actual:
            egresos_mes_actual += float(egreso_manual.monto)
        
        egresos_por_mes.append({
            'mes': mes_fecha.strftime('%B %Y'),
            'egresos': egresos_mes_actual,
            'movimientos': movimientos_mes_actual.count() + solicitudes_mes_actual.count() + egresos_manuales_mes_actual.count()
        })
    
    # Todos los egresos combinados para la tabla (compras, solicitudes, manuales)
    todos_egresos_list = []
    
    # Agregar compras
    for movimiento in movimientos_entrada:
        total_movimiento = 0
        if movimiento.insumo and movimiento.insumo.precio_unitario:
            total_movimiento = float(movimiento.insumo.precio_unitario) * movimiento.cantidad
        
        todos_egresos_list.append({
            'movimiento': movimiento,
            'total': total_movimiento,
            'tipo': 'compra'
        })
    
    # Agregar solicitudes
    for solicitud in solicitudes_egreso:
        todos_egresos_list.append({
            'solicitud': solicitud,
            'total': float(solicitud.monto_egreso) if solicitud.monto_egreso else 0,
            'tipo': 'solicitud'
        })
    
    # Agregar egresos manuales
    for egreso_manual in egresos_manuales:
        todos_egresos_list.append({
            'egreso_manual': egreso_manual,
            'total': float(egreso_manual.monto),
            'tipo': 'egreso_manual'
        })
    
    # Ordenar por fecha (más recientes primero)
    def get_fecha(item):
        if item['tipo'] == 'compra':
            return item['movimiento'].fecha_movimiento
        elif item['tipo'] == 'solicitud':
            return item['solicitud'].fecha_solicitud
        else:  # egreso_manual
            from datetime import datetime
            # Crear datetime aware para que sea compatible con los otros
            fecha_naive = datetime.combine(item['egreso_manual'].fecha, datetime.min.time())
            # Convertir a aware usando timezone.now() como referencia
            return timezone.make_aware(fecha_naive)
    
    todos_egresos_list.sort(key=get_fecha, reverse=True)
    movimientos_recientes = todos_egresos_list[:50]  # Mostrar los últimos 50
    
    # Balance (Ingresos - Egresos)
    balance_total = total_ingresos - total_egresos
    balance_mes = ingresos_mes - egresos_mes
    balance_año = ingresos_año - egresos_año
    
    context = {
        'perfil': perfil,
        'es_admin': perfil.es_administrativo(),
        # Ingresos
        'total_ingresos': total_ingresos,
        'ingresos_mes': ingresos_mes,
        'ingresos_año': ingresos_año,
        'total_citas_completadas': total_citas_completadas,
        'citas_mes': citas_mes,
        'citas_año': citas_año,
        'precio_promedio': precio_promedio,
        'ingresos_por_servicio': ingresos_por_servicio[:5],  # Solo top 5
        'ingresos_por_mes': ingresos_por_mes,
        'ingresos_por_dentista': ingresos_por_dentista[:5],  # Solo top 5
        'citas_recientes': citas_recientes,
        'todos_ingresos': todos_ingresos,  # Todos los ingresos combinados
        'ingresos_manuales': ingresos_manuales,  # Para referencia
        # Egresos
        'total_egresos': total_egresos,
        'egresos_mes': egresos_mes,
        'egresos_año': egresos_año,
        'egresos_por_mes': egresos_por_mes,
        'movimientos_recientes': movimientos_recientes,  # Todos los egresos combinados
        'egresos_manuales': egresos_manuales,  # Para referencia
        # Balance
        'balance_total': balance_total,
        'balance_mes': balance_mes,
        'balance_año': balance_año,
    }
    
    return render(request, 'citas/finanzas/gestor_finanzas.html', context)


# Vistas para Ingresos y Egresos Manuales
@login_required
def agregar_ingreso_manual(request):
    """Vista para agregar un ingreso manual"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para agregar ingresos manuales.')
            return redirect('gestor_finanzas')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    if request.method == 'POST':
        monto_str = request.POST.get('monto', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        fecha = request.POST.get('fecha', '')
        notas = request.POST.get('notas', '').strip()
        
        # Validaciones
        if not monto_str or not descripcion or not fecha:
            messages.error(request, 'Todos los campos son obligatorios.')
            return redirect('gestor_finanzas')
        
        try:
            monto = round(float(monto_str))  # Redondear a entero para pesos chilenos
            if monto <= 0:
                messages.error(request, 'El monto debe ser mayor a cero.')
                return redirect('gestor_finanzas')
        except ValueError:
            messages.error(request, 'El monto debe ser un número válido.')
            return redirect('gestor_finanzas')
        
        try:
            fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Formato de fecha inválido.')
            return redirect('gestor_finanzas')
        
        # Crear ingreso manual
        ingreso = IngresoManual.objects.create(
            monto=monto,
            descripcion=descripcion,
            fecha=fecha_obj,
            notas=notas if notas else None,
            creado_por=perfil
        )
        
        # Registrar en auditoría
        registrar_auditoria(
            usuario=perfil,
            accion='crear',
            modulo='finanzas',
            descripcion=f'Ingreso manual creado: {descripcion}',
            detalles=f'Monto: ${monto:,}, Fecha: {fecha_obj.strftime("%d/%m/%Y")}',
            objeto_id=ingreso.id,
            tipo_objeto='IngresoManual',
            request=request
        )
        
        messages.success(request, f'Ingreso manual de ${monto:,} agregado correctamente.')
    
    return redirect('gestor_finanzas')


@login_required
def eliminar_ingreso_manual(request, ingreso_id):
    """Vista para eliminar un ingreso manual"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para eliminar ingresos manuales.')
            return redirect('gestor_finanzas')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    ingreso = get_object_or_404(IngresoManual, id=ingreso_id)
    monto = ingreso.monto
    descripcion = ingreso.descripcion
    
    # Registrar en auditoría ANTES de eliminar
    registrar_auditoria(
        usuario=perfil,
        accion='eliminar',
        modulo='finanzas',
        descripcion=f'Ingreso manual eliminado: {descripcion}',
        detalles=f'Monto: ${monto:,}, Fecha: {ingreso.fecha.strftime("%d/%m/%Y")}',
        objeto_id=ingreso_id,
        tipo_objeto='IngresoManual',
        request=request
    )
    
    ingreso.delete()
    
    messages.success(request, f'Ingreso manual de ${monto:,} eliminado correctamente.')
    return redirect('gestor_finanzas')


@login_required
def agregar_egreso_manual(request):
    """Vista para agregar un egreso manual"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para agregar egresos manuales.')
            return redirect('gestor_finanzas')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    if request.method == 'POST':
        monto_str = request.POST.get('monto', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        fecha = request.POST.get('fecha', '')
        notas = request.POST.get('notas', '').strip()
        
        # Validaciones
        if not monto_str or not descripcion or not fecha:
            messages.error(request, 'Todos los campos son obligatorios.')
            return redirect('gestor_finanzas')
        
        try:
            monto = round(float(monto_str))  # Redondear a entero para pesos chilenos
            if monto <= 0:
                messages.error(request, 'El monto debe ser mayor a cero.')
                return redirect('gestor_finanzas')
        except ValueError:
            messages.error(request, 'El monto debe ser un número válido.')
            return redirect('gestor_finanzas')
        
        try:
            fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Formato de fecha inválido.')
            return redirect('gestor_finanzas')
        
        # Crear egreso manual
        egreso = EgresoManual.objects.create(
            monto=monto,
            descripcion=descripcion,
            fecha=fecha_obj,
            notas=notas if notas else None,
            creado_por=perfil
        )
        
        # Registrar en auditoría
        registrar_auditoria(
            usuario=perfil,
            accion='crear',
            modulo='finanzas',
            descripcion=f'Egreso manual creado: {descripcion}',
            detalles=f'Monto: ${monto:,}, Fecha: {fecha_obj.strftime("%d/%m/%Y")}',
            objeto_id=egreso.id,
            tipo_objeto='EgresoManual',
            request=request
        )
        
        messages.success(request, f'Egreso manual de ${monto:,} agregado correctamente.')
    
    return redirect('gestor_finanzas')


@login_required
def eliminar_egreso_manual(request, egreso_id):
    """Vista para eliminar un egreso manual"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para eliminar egresos manuales.')
            return redirect('gestor_finanzas')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    egreso = get_object_or_404(EgresoManual, id=egreso_id)
    monto = egreso.monto
    descripcion = egreso.descripcion
    
    # Registrar en auditoría ANTES de eliminar
    registrar_auditoria(
        usuario=perfil,
        accion='eliminar',
        modulo='finanzas',
        descripcion=f'Egreso manual eliminado: {descripcion}',
        detalles=f'Monto: ${monto:,}, Fecha: {egreso.fecha.strftime("%d/%m/%Y")}',
        objeto_id=egreso_id,
        tipo_objeto='EgresoManual',
        request=request
    )
    
    egreso.delete()
    
    messages.success(request, f'Egreso manual de ${monto:,} eliminado correctamente.')
    return redirect('gestor_finanzas')


@login_required
def eliminar_ingreso_cita(request, cita_id):
    """
    Vista para eliminar un ingreso de una cita completada del historial financiero
    Marca el precio_cobrado como None para que no aparezca en los cálculos
    SOLO DISPONIBLE PARA ADMINISTRATIVOS
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para eliminar ingresos de citas.')
            return redirect('gestor_finanzas')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    cita = get_object_or_404(Cita, id=cita_id)
    
    # Solo permitir eliminar ingresos de citas completadas
    if cita.estado != 'completada':
        messages.error(request, 'Solo se pueden eliminar ingresos de citas completadas.')
        return redirect('gestor_finanzas')
    
    # Guardar el monto para el mensaje
    monto_anterior = cita.precio_cobrado or (cita.tipo_servicio.precio_base if cita.tipo_servicio else 0)
    servicio_nombre = cita.tipo_servicio.nombre if cita.tipo_servicio else 'Sin servicio'
    
    # Marcar precio_cobrado como None para que no aparezca en los cálculos financieros
    cita.precio_cobrado = None
    cita.save()
    
    # Registrar en auditoría
    registrar_auditoria(
        usuario=perfil,
        accion='eliminar',
        modulo='finanzas',
        descripcion=f'Ingreso de cita eliminado del historial financiero',
        detalles=f'Cita ID: {cita.id}, Servicio: {servicio_nombre}, Monto: ${monto_anterior:,.0f}',
        objeto_id=cita.id,
        tipo_objeto='Cita',
        request=request
    )
    
    messages.success(request, f'Ingreso de cita eliminado correctamente del historial financiero.')
    return redirect('gestor_finanzas')


@login_required
def eliminar_egreso_compra(request, movimiento_id):
    """
    Vista para eliminar un egreso de compra (MovimientoInsumo) del historial financiero
    Elimina el movimiento de insumo
    SOLO DISPONIBLE PARA ADMINISTRATIVOS
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para eliminar egresos de compras.')
            return redirect('gestor_finanzas')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    from inventario.models import MovimientoInsumo
    movimiento = get_object_or_404(MovimientoInsumo, id=movimiento_id)
    
    # Solo permitir eliminar movimientos de entrada (compras)
    if movimiento.tipo != 'entrada':
        messages.error(request, 'Solo se pueden eliminar movimientos de entrada (compras).')
        return redirect('gestor_finanzas')
    
    # Guardar información para el mensaje
    insumo_nombre = movimiento.insumo.nombre if movimiento.insumo else "Sin nombre"
    cantidad = movimiento.cantidad
    monto = float(movimiento.insumo.precio_unitario) * cantidad if movimiento.insumo and movimiento.insumo.precio_unitario else 0
    
    # Registrar en auditoría ANTES de eliminar
    registrar_auditoria(
        usuario=perfil,
        accion='eliminar',
        modulo='finanzas',
        descripcion=f'Egreso de compra eliminado: {insumo_nombre}',
        detalles=f'Cantidad: {cantidad}, Monto: ${monto:,.0f}, Fecha: {movimiento.fecha_movimiento.strftime("%d/%m/%Y")}',
        objeto_id=movimiento_id,
        tipo_objeto='MovimientoInsumo',
        request=request
    )
    
    # Eliminar el movimiento
    movimiento.delete()
    
    messages.success(request, f'Compra de {insumo_nombre} eliminada correctamente del historial financiero.')
    return redirect('gestor_finanzas')


@login_required
def eliminar_egreso_solicitud(request, solicitud_id):
    """
    Vista para eliminar un egreso de solicitud de insumo del historial financiero
    Elimina el monto_egreso de la solicitud
    SOLO DISPONIBLE PARA ADMINISTRATIVOS
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para eliminar egresos de solicitudes.')
            return redirect('gestor_finanzas')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    from proveedores.models import SolicitudInsumo
    solicitud = get_object_or_404(SolicitudInsumo, id=solicitud_id)
    
    # Solo permitir eliminar si tiene monto_egreso
    if not solicitud.monto_egreso:
        messages.error(request, 'Esta solicitud no tiene un egreso registrado.')
        return redirect('gestor_finanzas')
    
    # Guardar el monto para el mensaje
    monto_anterior = solicitud.monto_egreso or 0
    insumo_nombre = solicitud.insumo.nombre if solicitud.insumo else "Sin nombre"
    
    # Eliminar el monto de egreso
    solicitud.monto_egreso = None
    solicitud.save()
    
    # Registrar en auditoría
    registrar_auditoria(
        usuario=perfil,
        accion='eliminar',
        modulo='finanzas',
        descripcion=f'Egreso de solicitud eliminado: {insumo_nombre}',
        detalles=f'Solicitud ID: {solicitud.id}, Monto: ${monto_anterior:,.0f}, Proveedor: {solicitud.proveedor.nombre if solicitud.proveedor else "N/A"}',
        objeto_id=solicitud_id,
        tipo_objeto='SolicitudInsumo',
        request=request
    )
    
    messages.success(request, f'Egreso de solicitud eliminado correctamente del historial financiero.')
    return redirect('gestor_finanzas')


# =====================================================
# VISTAS PARA INFORMACIÓN DE LA CLÍNICA
# =====================================================

@login_required
def obtener_informacion_clinica(request):
    """
    Vista AJAX para obtener la información de contacto de la clínica
    """
    try:
        from configuracion.models import InformacionClinica
        info = InformacionClinica.obtener()
        
        data = {
            'success': True,
            'data': {
                'nombre_clinica': info.nombre_clinica,
                'direccion': info.direccion,
                'telefono': info.telefono,
                'telefono_secundario': info.telefono_secundario or '',
                'email': info.email,
                'email_alternativo': info.email_alternativo or '',
                'horario_atencion': info.horario_atencion,
                'sitio_web': info.sitio_web or '',
                'whatsapp': info.whatsapp or '',
                'facebook': info.facebook or '',
                'instagram': info.instagram or '',
                'notas_adicionales': info.notas_adicionales or '',
                'actualizado_el': info.actualizado_el.strftime('%d/%m/%Y %H:%M') if info.actualizado_el else '',
                'actualizado_por': info.actualizado_por.nombre_completo if info.actualizado_por else ''
            }
        }
        return JsonResponse(data)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def editar_informacion_clinica(request):
    """
    Vista para editar la información de contacto de la clínica
    Solo disponible para administrativos
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para editar la información de la clínica.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('login')
    
    from configuracion.models import InformacionClinica
    info = InformacionClinica.obtener()
    
    if request.method == 'POST':
        try:
            # Actualizar los campos
            info.nombre_clinica = request.POST.get('nombre_clinica', info.nombre_clinica)
            info.direccion = request.POST.get('direccion', info.direccion)
            info.telefono = request.POST.get('telefono', info.telefono)
            info.telefono_secundario = request.POST.get('telefono_secundario', '')
            info.email = request.POST.get('email', info.email)
            info.email_alternativo = request.POST.get('email_alternativo', '')
            info.horario_atencion = request.POST.get('horario_atencion', info.horario_atencion)
            info.sitio_web = request.POST.get('sitio_web', '')
            info.whatsapp = request.POST.get('whatsapp', '')
            info.facebook = request.POST.get('facebook', '')
            info.instagram = request.POST.get('instagram', '')
            info.notas_adicionales = request.POST.get('notas_adicionales', '')
            info.actualizado_por = perfil
            
            info.save()
            
            # Si es una petición AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Información actualizada correctamente'
                })
            
            messages.success(request, 'Información de la clínica actualizada correctamente.')
            return redirect('panel_trabajador')
        
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                }, status=400)
            
            messages.error(request, f'Error al actualizar la información: {str(e)}')
    
    context = {
        'perfil': perfil,
        'info': info
    }
    
    return render(request, 'citas/perfil/editar_informacion_clinica.html', context)


# =====================================================
# VISTAS PARA GESTIÓN DE CLIENTES
# =====================================================

@login_required
def validar_username(request):
    """
    Vista AJAX para validar si un username ya existe y tiene un Cliente activo asociado
    Solo bloquea si el username tiene un Cliente activo, no si es un User huérfano
    """
    username = request.GET.get('username', '').strip()
    
    if not username:
        return JsonResponse({'existe': False})
    
    from django.contrib.auth.models import User
    
    # Verificar si el username existe
    try:
        user = User.objects.get(username=username)
        
        # Verificar si este User tiene un Cliente activo asociado
        # Buscar por email del User en Clientes activos
        existe_cliente_activo = Cliente.objects.filter(
            email__iexact=user.email, 
            activo=True
        ).exists()
        
        # Si no hay Cliente activo, el User es huérfano y el username puede reutilizarse
        return JsonResponse({'existe': existe_cliente_activo})
    except User.DoesNotExist:
        # El username no existe, está disponible
        return JsonResponse({'existe': False})


@login_required
def buscar_cliente_por_email(request):
    """
    Vista AJAX para buscar un cliente por email y devolver sus datos
    """
    email = request.GET.get('email', '').strip()
    
    if not email:
        return JsonResponse({'existe': False})
    
    try:
        cliente = Cliente.objects.get(email__iexact=email, activo=True)
        return JsonResponse({
            'existe': True,
            'cliente': {
                'nombre_completo': cliente.nombre_completo,
                'telefono': cliente.telefono or '',
                'fecha_nacimiento': cliente.fecha_nacimiento.strftime('%Y-%m-%d') if cliente.fecha_nacimiento else '',
                'alergias': cliente.alergias or ''
            }
        })
    except Cliente.DoesNotExist:
        return JsonResponse({'existe': False})
    except Cliente.MultipleObjectsReturned:
        cliente = Cliente.objects.filter(email__iexact=email, activo=True).first()
        if cliente:
            return JsonResponse({
                'existe': True,
                'cliente': {
                    'nombre_completo': cliente.nombre_completo,
                    'telefono': cliente.telefono or '',
                    'fecha_nacimiento': cliente.fecha_nacimiento.strftime('%Y-%m-%d') if cliente.fecha_nacimiento else '',
                    'alergias': cliente.alergias or ''
                }
            })
        return JsonResponse({'existe': False})


@login_required
def validar_email(request):
    """
    Vista AJAX para validar si un email ya existe en Cliente activo
    Solo verifica en Clientes activos, no en Users sueltos (pueden quedar huérfanos)
    """
    email = request.GET.get('email', '').strip().lower()
    
    if not email:
        return JsonResponse({'existe': False})
    
    # Validar formato de email básico
    from django.core.validators import validate_email
    from django.core.exceptions import ValidationError
    try:
        validate_email(email)
    except ValidationError:
        return JsonResponse({'existe': False, 'invalido': True})
    
    # Verificar si existe en Cliente ACTIVO (solo clientes activos bloquean)
    # Si un cliente fue eliminado (activo=False), su email puede ser reutilizado
    # No verificamos Users huérfanos (sin Cliente activo) porque pueden quedar después de eliminar clientes
    existe = Cliente.objects.filter(email__iexact=email, activo=True).exists()
    
    return JsonResponse({'existe': existe})


@login_required
def validar_rut(request):
    """
    Vista AJAX para validar si un RUT ya existe en Cliente
    """
    rut = request.GET.get('rut', '').strip()
    
    if not rut:
        return JsonResponse({'existe': False})
    
    # Limpiar el RUT (quitar puntos y guiones)
    rut_limpio = rut.replace('.', '').replace('-', '').upper()
    
    # Validar formato básico (debe tener al menos 7 dígitos)
    if not rut_limpio[:-1].isdigit() or len(rut_limpio) < 8:
        return JsonResponse({'existe': False, 'invalido': True})
    
    # Buscar en Cliente (comparar RUTs normalizados)
    existe = False
    clientes = Cliente.objects.filter(rut__isnull=False).exclude(rut='')
    for cliente in clientes:
        if cliente.rut:
            rut_cliente_limpio = cliente.rut.replace('.', '').replace('-', '').upper()
            if rut_cliente_limpio == rut_limpio:
                existe = True
                break
    
    return JsonResponse({'existe': existe})


@login_required
def validar_telefono(request):
    """
    Vista AJAX para validar si un teléfono ya existe en Cliente
    """
    telefono = request.GET.get('telefono', '').strip()
    
    if not telefono:
        return JsonResponse({'existe': False})
    
    # Normalizar el teléfono (8 dígitos chilenos)
    telefono_normalizado = normalizar_telefono_chileno(telefono)
    
    if not telefono_normalizado:
        return JsonResponse({'existe': False, 'invalido': True})
    
    # Buscar en Cliente
    existe = Cliente.objects.filter(telefono=telefono_normalizado).exists()
    
    return JsonResponse({'existe': existe})


@login_required
def crear_cliente_presencial(request):
    """
    Vista para crear un cliente que se registra presencialmente
    También crea su cuenta en el sistema de citas online con credenciales manuales
    Solo disponible para administrativos
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para crear clientes.')
            return redirect('gestor_clientes')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('login')
    
    if request.method == 'POST':
        try:
            import logging
            logger = logging.getLogger(__name__)
            from django.contrib.auth.models import User
            from django.conf import settings
            BASE_DIR = settings.BASE_DIR
            
            # Obtener datos del formulario
            nombre_completo = request.POST.get('nombre_completo')
            email = request.POST.get('email')
            telefono_raw = request.POST.get('telefono')
            rut = request.POST.get('rut', '').strip()
            fecha_nacimiento_str = request.POST.get('fecha_nacimiento', '').strip()
            alergias = request.POST.get('alergias', '').strip()
            
            # Validar campos obligatorios
            if not rut:
                messages.error(request, 'El RUT es obligatorio.')
                return redirect('gestor_clientes')
            
            if not fecha_nacimiento_str:
                messages.error(request, 'La fecha de nacimiento es obligatoria.')
                return redirect('gestor_clientes')
            
            if not alergias:
                messages.error(request, 'Las alergias son obligatorias. Si no tiene alergias, escriba "Ninguna".')
                return redirect('gestor_clientes')
            notas = request.POST.get('notas', '')
            crear_usuario_online = request.POST.get('crear_usuario_online') == 'on'
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '').strip()
            enviar_email = request.POST.get('enviar_credenciales_email') == 'on'
            
            # Normalizar el número de teléfono chileno
            telefono = normalizar_telefono_chileno(telefono_raw)
            if not telefono:
                messages.error(
                    request, 
                    'El número de teléfono no es válido. '
                    'Ingrese solo los 8 dígitos del celular (ejemplo: 12345678). '
                    'El sistema agregará automáticamente el 9 y el código de país (+56).'
                )
                return redirect('gestor_clientes')
            
            # Usar funciones de validación centralizadas
            email_existe, email_error = validar_email_cliente(email)
            if email_existe:
                messages.error(request, email_error)
                return redirect('gestor_clientes')
            
            rut_existe, rut_error = validar_rut_cliente(rut) if rut else (False, None)
            if rut_existe:
                messages.error(request, rut_error)
                return redirect('gestor_clientes')
            
            telefono_existe, telefono_error = validar_telefono_cliente(telefono)
            if telefono_existe:
                messages.error(request, telefono_error)
                return redirect('gestor_clientes')
            
            # Si se va a crear usuario online, validar ANTES de crear el Cliente
            if crear_usuario_online:
                # Validar que se hayan proporcionado username y password
                if not username or not password:
                    messages.error(request, 'Debes proporcionar nombre de usuario y contraseña para crear el acceso web.')
                    return redirect('gestor_clientes')
                
                # Validar longitud mínima de contraseña
                if len(password) < 8:
                    messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
                    return redirect('gestor_clientes')
                
                # Usar función de validación centralizada para username
                username_disponible, username_error = validar_username_disponible(username)
                if not username_disponible and username_error and "asociado a un cliente activo" in username_error:
                    messages.error(request, username_error)
                    return redirect('gestor_clientes')
                elif username_error and "Se reutilizará" in username_error:
                    # Advertencia pero permitimos continuar
                    logger.warning(f"Reutilizando username existente: {username}")
                
                # Validar que el email no exista en User que tenga un Cliente activo asociado
                # Solo bloquear si el User tiene un Cliente activo, no si es un User huérfano
                # (Ya validamos arriba que no existe Cliente activo con ese email, así que si hay User, es huérfano)
                # Pero por seguridad, verificamos una vez más
                usuarios_existentes = User.objects.filter(email__iexact=email)
                if usuarios_existentes.exists():
                    # Si ya validamos que no hay Cliente activo con ese email, el User es huérfano
                    # y podemos reutilizar el email. No bloqueamos.
                    # Solo mostramos advertencia si realmente hay un Cliente activo (doble verificación)
                    if Cliente.objects.filter(email__iexact=email, activo=True).exists():
                        usuarios_lista = ', '.join([u.username for u in usuarios_existentes])
                        messages.error(
                            request, 
                            f'Ya existe un usuario activo con ese email en el sistema de citas online. '
                            f'Usuarios encontrados: {usuarios_lista}. '
                            f'Por favor, usa otro email o elimina/edita los usuarios existentes.'
                        )
                        return redirect('gestor_clientes')
                    # Si no hay Cliente activo, el User es huérfano y podemos continuar
            
            # Procesar fecha de nacimiento (ahora obligatorio)
            try:
                from datetime import datetime
                fecha_nacimiento = datetime.strptime(fecha_nacimiento_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, 'La fecha de nacimiento no tiene un formato válido.')
                return redirect('gestor_clientes')
            
            # Usar transacción atómica para crear Cliente, User y PerfilCliente
            # Si falla cualquier parte, se revierte todo
            with transaction.atomic():
                # Crear el cliente en la clínica
                # Todos los campos ahora son obligatorios (rut, fecha_nacimiento, alergias)
                cliente = Cliente.objects.create(
                    nombre_completo=nombre_completo,
                    email=email,
                    telefono=telefono,
                    rut=rut,
                    fecha_nacimiento=fecha_nacimiento,
                    alergias=alergias,
                    notas=notas,
                    activo=True
                )
                
                # Si se solicitó, crear también el usuario en el sistema de citas online
                user = None
                if crear_usuario_online:
                    try:
                        # Ya validamos todo antes, ahora solo crear el usuario
                        
                        # Crear el usuario de Django con las credenciales proporcionadas
                        user = User.objects.create_user(
                            username=username,
                            email=email,
                            password=password,
                            first_name=nombre_completo.split()[0] if nombre_completo.split() else '',
                            last_name=' '.join(nombre_completo.split()[1:]) if len(nombre_completo.split()) > 1 else ''
                        )
                        
                        # Establecer la relación explícita entre Cliente y User
                        cliente.user = user
                        cliente.save()
                        logger.info(f"✅ Relación establecida: Cliente {cliente.id} <-> User {user.id}")
                        
                        # Crear el PerfilCliente directamente usando el ORM
                        # Como ambos proyectos comparten la misma base de datos PostgreSQL,
                        # podemos insertar directamente en la tabla cuentas_perfilcliente
                        perfil_cliente_creado = False
                        try:
                            from django.db import connection
                            
                            # Verificar si ya existe un PerfilCliente para este usuario
                            with connection.cursor() as cursor:
                                cursor.execute(
                                    "SELECT id FROM cuentas_perfilcliente WHERE user_id = %s",
                                    [user.id]
                                )
                                existe = cursor.fetchone()
                                
                                if not existe:
                                    # Crear nuevo PerfilCliente
                                    cursor.execute("""
                                        INSERT INTO cuentas_perfilcliente 
                                        (user_id, nombre_completo, telefono, email, telefono_verificado, rut, fecha_nacimiento, alergias)
                                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                                    """, [
                                        user.id,
                                        nombre_completo,
                                        telefono,
                                        email,
                                        False,  # telefono_verificado
                                        rut if rut else None,
                                        fecha_nacimiento if fecha_nacimiento else None,
                                        alergias if alergias else None
                                    ])
                                    perfil_cliente_creado = True
                                    logger.info(f"✅ PerfilCliente creado exitosamente para usuario {username} (ID: {user.id})")
                                else:
                                    # Actualizar PerfilCliente existente
                                    cursor.execute("""
                                        UPDATE cuentas_perfilcliente 
                                        SET nombre_completo = %s,
                                            telefono = %s,
                                            email = %s,
                                            rut = %s,
                                            fecha_nacimiento = %s,
                                            alergias = %s
                                        WHERE user_id = %s
                                    """, [
                                        nombre_completo,
                                        telefono,
                                        email,
                                        rut if rut else None,
                                        fecha_nacimiento if fecha_nacimiento else None,
                                        alergias if alergias else None,
                                        user.id
                                    ])
                                    logger.info(f"✅ PerfilCliente actualizado para usuario {username} (ID: {user.id})")
                        except Exception as e:
                            # Error al crear el perfil
                            logger.error(f"❌ Error al crear/actualizar PerfilCliente para usuario {username}: {e}", exc_info=True)
                            import traceback
                            logger.error(f"Traceback: {traceback.format_exc()}")
                            messages.warning(
                                request,
                                f'⚠️ Usuario web creado, pero hubo un error al crear el perfil en cliente_web. '
                                f'Error: {str(e)}. Verifica los logs para más detalles.'
                            )
                        
                        # Actualizar notas solo para indicar que tiene acceso web (SIN guardar credenciales por seguridad)
                        # Las credenciales se envían por email, no se guardan en texto plano
                        if notas:
                            # Si ya hay notas, agregar información de acceso web
                            if "[ACCESO WEB]" not in notas:
                                notas_completas = f"{notas}\n\n[ACCESO WEB]\nUsuario web creado el {timezone.now().strftime('%d/%m/%Y %H:%M')}"
                                cliente.notas = notas_completas
                        else:
                            # Si no hay notas, crear una nueva
                            cliente.notas = f"[ACCESO WEB]\nUsuario web creado el {timezone.now().strftime('%d/%m/%Y %H:%M')}"
                        cliente.save()
                        logger.info(f"✅ Notas actualizadas (sin credenciales) para cliente {cliente.id}")
                    except Exception as e:
                        logger.error(f"❌ Error al crear usuario web: {e}", exc_info=True)
                        raise  # Re-lanzar para que la transacción se revierta
                
                # Registrar en auditoría (dentro de la transacción)
                detalles_cliente = f'Email: {email}, Teléfono: {telefono}, RUT: {rut if rut else "N/A"}'
                if crear_usuario_online:
                    detalles_cliente += f', Usuario web: {username}'
                registrar_auditoria(
                    usuario=perfil,
                    accion='crear',
                    modulo='clientes',
                    descripcion=f'Cliente creado: {nombre_completo}',
                    detalles=detalles_cliente,
                    objeto_id=cliente.id,
                    tipo_objeto='Cliente',
                    request=request
                )
            
            # Enviar correo si se solicitó (fuera de la transacción, ya que el email puede fallar)
            if enviar_email and crear_usuario_online:
                try:
                            logger.info(f"Intentando enviar correo a: {email}")
                            
                            # Verificar configuración de email
                            email_host_user = getattr(settings, 'EMAIL_HOST_USER', '')
                            email_host_password = getattr(settings, 'EMAIL_HOST_PASSWORD', '')
                            
                            logger.info(f"EMAIL_HOST_USER configurado: {'Sí' if email_host_user else 'No'}")
                            logger.info(f"EMAIL_HOST_PASSWORD configurado: {'Sí' if email_host_password else 'No'}")
                            
                            if not email_host_user or not email_host_password:
                                error_msg = "Configuración de email incompleta. Verifica EMAIL_HOST_USER y EMAIL_HOST_PASSWORD en settings.py o variables de entorno."
                                logger.error(error_msg)
                                raise Exception(error_msg)
                            
                            # Obtener información de la clínica
                            try:
                                from configuracion.models import InformacionClinica
                                info_clinica = InformacionClinica.obtener()
                                nombre_clinica = info_clinica.nombre_clinica or "Clínica Dental San Felipe"
                                direccion_clinica = info_clinica.direccion or ''
                                telefono_clinica = info_clinica.telefono or ''
                                email_clinica = info_clinica.email or settings.DEFAULT_FROM_EMAIL
                                sitio_web_clinica = info_clinica.sitio_web or ''
                            except Exception as e:
                                logger.warning(f"Error al obtener información de clínica: {e}")
                                nombre_clinica = getattr(settings, 'CLINIC_NAME', 'Clínica Dental San Felipe')
                                direccion_clinica = getattr(settings, 'CLINIC_ADDRESS', '')
                                telefono_clinica = getattr(settings, 'CLINIC_PHONE', '')
                                email_clinica = settings.DEFAULT_FROM_EMAIL
                                sitio_web_clinica = getattr(settings, 'CLINIC_WEBSITE', '')
                            
                            # Obtener URL de login del cliente web
                            # Intentar obtener CLIENTE_WEB_URL si está configurado, sino usar SITE_URL
                            cliente_web_url = getattr(settings, 'CLIENTE_WEB_URL', None)
                            if not cliente_web_url:
                                cliente_web_url = getattr(settings, 'SITE_URL', 'http://localhost:8000')
                            # Si la URL termina con /, quitar la barra final
                            if cliente_web_url.endswith('/'):
                                cliente_web_url = cliente_web_url[:-1]
                            # Construir URL de login del cliente web (ruta: /cuentas/login/)
                            url_login = f"{cliente_web_url}/cuentas/login/"
                            
                            logger.info(f"URL de login del cliente web: {url_login}")
                            
                            # Renderizar template HTML
                            from django.template.loader import render_to_string
                            from django.core.mail import EmailMessage
                            
                            mensaje_html = render_to_string('citas/emails/credenciales_acceso.html', {
                                'nombre_completo': nombre_completo,
                                'username': username,
                                'password': password,
                                'nombre_clinica': nombre_clinica,
                                'direccion_clinica': direccion_clinica,
                                'telefono_clinica': telefono_clinica,
                                'email_clinica': email_clinica,
                                'sitio_web_clinica': sitio_web_clinica,
                                'url_login': url_login,
                            })
                            
                            asunto = f'Credenciales de Acceso - {nombre_clinica}'
                            
                            logger.info(f"Enviando correo desde {email_clinica} a {email}")
                            logger.info(f"EMAIL_BACKEND: {getattr(settings, 'EMAIL_BACKEND', 'No configurado')}")
                            logger.info(f"EMAIL_HOST: {getattr(settings, 'EMAIL_HOST', 'No configurado')}")
                            logger.info(f"EMAIL_PORT: {getattr(settings, 'EMAIL_PORT', 'No configurado')}")
                            logger.info(f"EMAIL_USE_TLS: {getattr(settings, 'EMAIL_USE_TLS', 'No configurado')}")
                            
                            # Crear el email con contenido HTML
                            email_msg = EmailMessage(
                                asunto,
                                mensaje_html,
                                email_clinica,
                                [email],
                            )
                            email_msg.content_subtype = "html"  # Indicar que es HTML
                            
                            # Intentar enviar el correo
                            try:
                                email_msg.send(fail_silently=False)
                                logger.info(f"✅ Correo HTML enviado exitosamente a: {email}")
                            except Exception as send_error:
                                logger.error(f"❌ Error al enviar correo: {send_error}")
                                logger.error(f"Tipo de error: {type(send_error).__name__}")
                                import traceback
                                logger.error(f"Traceback completo: {traceback.format_exc()}")
                                raise  # Re-lanzar el error para que se capture en el except externo
                            
                            messages.success(
                                request,
                                f'✅ Cliente {nombre_completo} creado exitosamente. '
                                f'Usuario web: {username}. '
                                f'📧 Correo con credenciales enviado a {email}.'
                            )
                except Exception as e:
                    logger.error(f"Error al enviar correo: {str(e)}", exc_info=True)
                    import traceback
                    error_detalle = traceback.format_exc()
                    logger.error(f"Traceback completo: {error_detalle}")
                    
                    messages.error(
                        request,
                        f'⚠️ Cliente creado exitosamente, pero NO se pudo enviar el correo. '
                        f'Error: {str(e)}. '
                        f'Las credenciales están guardadas en las notas del cliente. '
                        f'Verifica la configuración de EMAIL_HOST_USER y EMAIL_HOST_PASSWORD en settings.py'
                    )
            elif crear_usuario_online:
                # Si se creó usuario pero no se envió email
                messages.success(
                    request,
                    f'✅ Cliente {nombre_completo} creado exitosamente. '
                    f'Usuario web: {username}. '
                    f'Las credenciales están guardadas en las notas.'
                )
            else:
                # Si no se creó usuario web
                messages.success(request, f'Cliente {nombre_completo} creado exitosamente.')
            
            return redirect('gestor_clientes')
            
        except Exception as e:
            messages.error(request, f'Error al crear el cliente: {str(e)}')
            return redirect('gestor_clientes')
    
    return redirect('gestor_clientes')


@login_required
def sincronizar_cliente_web(request):
    """
    Vista para sincronizar un cliente de cliente_web con el sistema de gestión.
    Crea un Cliente en gestion_clinica basado en los datos de PerfilCliente.
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Perfil no encontrado.'}, status=404)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        email = data.get('email', '').strip()
        username = data.get('username', '').strip()
        
        if not email:
            return JsonResponse({'success': False, 'error': 'Email es requerido.'}, status=400)
        
        # Buscar PerfilCliente en cliente_web
        try:
            from django.contrib.auth.models import User
            from cuentas.models import PerfilCliente
            
            if username:
                user = User.objects.get(username=username)
                perfil_cliente = PerfilCliente.objects.get(user=user)
            else:
                perfil_cliente = PerfilCliente.objects.get(email=email)
            
            # Verificar que no exista ya un Cliente con ese email
            if Cliente.objects.filter(email=perfil_cliente.email).exists():
                return JsonResponse({
                    'success': False, 
                    'error': f'Ya existe un cliente con el email {perfil_cliente.email} en el sistema de gestión.'
                }, status=400)
            
            # Crear Cliente en el sistema de gestión
            cliente = Cliente.objects.create(
                nombre_completo=perfil_cliente.nombre_completo,
                email=perfil_cliente.email,
                telefono=perfil_cliente.telefono or '',
                rut=perfil_cliente.rut or '',
                fecha_nacimiento=perfil_cliente.fecha_nacimiento,
                alergias=perfil_cliente.alergias or 'Ninguna',
                activo=True,
                notas=f'Cliente sincronizado desde la web. Usuario: {perfil_cliente.user.username if perfil_cliente.user else "N/A"}'
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Cliente {cliente.nombre_completo} sincronizado exitosamente.',
                'cliente_id': cliente.id
            })
            
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Usuario no encontrado.'}, status=404)
        except PerfilCliente.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Perfil de cliente no encontrado en el sistema web.'}, status=404)
        except ImportError:
            return JsonResponse({'success': False, 'error': 'No se puede acceder al sistema de cliente_web.'}, status=500)
            
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos.'}, status=400)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error al sincronizar cliente web: {e}")
        return JsonResponse({'success': False, 'error': f'Error al sincronizar: {str(e)}'}, status=500)


@login_required
def editar_cliente(request, cliente_id):
    """
    Vista para editar la información de un cliente
    También permite editar las credenciales web si existen
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para editar clientes.')
            return redirect('gestor_clientes')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('login')
    
    cliente = get_object_or_404(Cliente, id=cliente_id)
    
    if request.method == 'POST':
        try:
            from django.contrib.auth.models import User
            import re
            
            # Actualizar información del cliente
            cliente.nombre_completo = request.POST.get('nombre_completo', cliente.nombre_completo)
            
            # Actualizar RUT
            rut = request.POST.get('rut', '').strip()
            if rut:
                # Validar que el RUT no exista en otro cliente
                if Cliente.objects.filter(rut=rut).exclude(id=cliente_id).exists():
                    messages.error(request, 'Ya existe otro cliente con ese RUT.')
                    return redirect('gestor_clientes')
                cliente.rut = rut
            else:
                cliente.rut = None
            
            # Actualizar fecha de nacimiento
            fecha_nacimiento_str = request.POST.get('fecha_nacimiento', '').strip()
            if fecha_nacimiento_str:
                try:
                    from datetime import datetime
                    cliente.fecha_nacimiento = datetime.strptime(fecha_nacimiento_str, '%Y-%m-%d').date()
                except ValueError:
                    messages.error(request, 'La fecha de nacimiento no tiene un formato válido.')
                    return redirect('gestor_clientes')
            else:
                cliente.fecha_nacimiento = None
            
            # Actualizar alergias
            alergias = request.POST.get('alergias', '').strip()
            cliente.alergias = alergias if alergias else None
            
            # Normalizar el número de teléfono chileno si se proporcionó uno nuevo
            telefono_raw = request.POST.get('telefono', cliente.telefono)
            telefono_normalizado = normalizar_telefono_chileno(telefono_raw)
            if not telefono_normalizado:
                messages.error(
                    request, 
                    'El número de teléfono no es válido. '
                    'Ingrese solo los 8 dígitos del celular (ejemplo: 12345678). '
                    'El sistema agregará automáticamente el 9 y el código de país (+56).'
                )
                return redirect('gestor_clientes')
            cliente.telefono = telefono_normalizado
            cliente.notas = request.POST.get('notas', cliente.notas)
            
            # Validar si se cambió el email
            email_anterior = cliente.email
            nuevo_email = request.POST.get('email')
            if nuevo_email != email_anterior:
                # Verificar que el nuevo email no exista
                if Cliente.objects.filter(email=nuevo_email).exclude(id=cliente_id).exists():
                    messages.error(request, 'Ya existe un cliente con ese email.')
                    return redirect('gestor_clientes')
                
                # Actualizar email del cliente
                cliente.email = nuevo_email
                
                # Actualizar radiografías asociadas con el email anterior
                from historial_clinico.models import Radiografia
                # Primero actualizar las que tienen el cliente asociado
                radiografias_con_cliente = Radiografia.objects.filter(
                    paciente_email=email_anterior,
                    cliente=cliente
                )
                radiografias_con_cliente.update(paciente_email=nuevo_email)
                
                # También actualizar radiografías que tienen el email anterior pero no tienen cliente asociado
                # Esto incluye todas las radiografías con ese email, independientemente del cliente
                radiografias_sin_cliente = Radiografia.objects.filter(
                    paciente_email=email_anterior,
                    cliente__isnull=True
                )
                
                # Actualizar todas las radiografías con el email anterior, asociándolas al cliente
                for radiografia in radiografias_sin_cliente:
                    radiografia.paciente_email = nuevo_email
                    radiografia.paciente_nombre = cliente.nombre_completo  # Actualizar también el nombre
                    radiografia.cliente = cliente
                    radiografia.save()
                
                # También actualizar radiografías que tienen el email pero están asociadas a otro cliente
                # o sin cliente, pero que claramente pertenecen a este cliente
                radiografias_por_email = Radiografia.objects.filter(
                    paciente_email=email_anterior
                ).exclude(cliente=cliente)
                
                # Si hay un cliente con el email anterior, actualizar esas radiografías también
                # (por si acaso hay duplicados o inconsistencias)
                for radiografia in radiografias_por_email:
                    # Solo actualizar si no tiene cliente o si el cliente es el mismo que estamos editando
                    if not radiografia.cliente or radiografia.cliente.id == cliente_id:
                        radiografia.paciente_email = nuevo_email
                        radiografia.paciente_nombre = cliente.nombre_completo
                        radiografia.cliente = cliente
                        radiografia.save()
            
            # Manejar credenciales web si existen
            nuevo_username = request.POST.get('username_web', '').strip()
            nueva_password = request.POST.get('password_web', '').strip()
            
            # Buscar si el cliente tiene usuario web
            try:
                user_web = User.objects.get(email=cliente.email)
                cambios_credenciales = []
                
                # Cambiar username si se proporcionó uno nuevo
                if nuevo_username and nuevo_username != user_web.username:
                    # Validar que el nuevo username no exista
                    if User.objects.filter(username=nuevo_username).exclude(id=user_web.id).exists():
                        messages.error(request, f'El nombre de usuario "{nuevo_username}" ya existe.')
                        return redirect('gestor_clientes')
                    
                    user_web.username = nuevo_username
                    cambios_credenciales.append(f'Usuario: {nuevo_username}')
                
                # Cambiar password si se proporcionó una nueva
                if nueva_password:
                    if len(nueva_password) < 8:
                        messages.error(request, 'La nueva contraseña debe tener al menos 8 caracteres.')
                        return redirect('gestor_clientes')
                    
                    user_web.set_password(nueva_password)
                    cambios_credenciales.append('Contraseña actualizada')
                
                # Actualizar email del usuario si cambió
                if nuevo_email != user_web.email:
                    # Validar que el nuevo email no esté en uso por otro usuario
                    if User.objects.filter(email=nuevo_email).exclude(id=user_web.id).exists():
                        messages.error(request, 'Ya existe un usuario web con ese email.')
                        return redirect('gestor_clientes')
                    user_web.email = nuevo_email
                
                if cambios_credenciales:
                    user_web.save()
                    # Actualizar las notas solo para indicar que tiene acceso web (SIN guardar credenciales por seguridad)
                    # Las credenciales se envían por email si se cambian, no se guardan en texto plano
                    notas = cliente.notas or ""
                    # Eliminar la sección antigua de credenciales si existe
                    notas = re.sub(r'\[ACCESO WEB\].*?(?=\n\n|\Z)', '', notas, flags=re.DOTALL).strip()
                    # Agregar solo información de que tiene acceso web (sin credenciales)
                    info_credenciales = f"\n\n[ACCESO WEB]\nUsuario web actualizado el {timezone.now().strftime('%d/%m/%Y %H:%M')}"
                    if notas:
                        cliente.notas = notas + info_credenciales
                    else:
                        cliente.notas = info_credenciales.strip()
                    logger.info(f"✅ Notas actualizadas (sin credenciales) para cliente {cliente.id}")
            except User.DoesNotExist:
                pass  # El cliente no tiene usuario web
            
            cliente.save()
            
            # Registrar en auditoría
            cambios_info = []
            if 'cambios_credenciales' in locals() and cambios_credenciales:
                cambios_info.extend(cambios_credenciales)
            detalles_edicion = f'Email: {cliente.email}, RUT: {cliente.rut if cliente.rut else "N/A"}'
            if cambios_info:
                detalles_edicion += f', Cambios: {", ".join(cambios_info)}'
            registrar_auditoria(
                usuario=perfil,
                accion='editar',
                modulo='clientes',
                descripcion=f'Cliente editado: {cliente.nombre_completo}',
                detalles=detalles_edicion,
                objeto_id=cliente.id,
                tipo_objeto='Cliente',
                request=request
            )
            
            if 'cambios_credenciales' in locals() and cambios_credenciales:
                messages.success(
                    request, 
                    f'Cliente {cliente.nombre_completo} actualizado exitosamente. ' + 
                    'Cambios: ' + ', '.join(cambios_credenciales) + '.'
                )
            else:
                messages.success(request, f'Cliente {cliente.nombre_completo} actualizado exitosamente.')
            
        except Exception as e:
            messages.error(request, f'Error al actualizar el cliente: {str(e)}')
    
    return redirect('gestor_clientes')


@login_required
def obtener_cliente(request, cliente_id):
    """
    Vista AJAX para obtener los datos de un cliente
    Incluye información del usuario web si existe
    """
    try:
        from django.contrib.auth.models import User
        
        cliente = Cliente.objects.get(id=cliente_id)
        
        # Buscar si tiene usuario web
        username_web = None
        tiene_usuario_web = False
        try:
            user = User.objects.get(email=cliente.email)
            username_web = user.username
            tiene_usuario_web = True
        except User.DoesNotExist:
            pass
        
        data = {
            'success': True,
            'cliente': {
                'id': cliente.id,
                'nombre_completo': cliente.nombre_completo,
                'email': cliente.email,
                'telefono': cliente.telefono,
                'rut': cliente.rut or '',
                'fecha_nacimiento': cliente.fecha_nacimiento.strftime('%Y-%m-%d') if cliente.fecha_nacimiento else '',
                'alergias': cliente.alergias or '',
                'notas': cliente.notas or '',
                'activo': cliente.activo,
                'fecha_registro': cliente.fecha_registro.strftime('%d/%m/%Y'),
                'tiene_usuario_web': tiene_usuario_web,
                'username_web': username_web,
            }
        }
        return JsonResponse(data)
    except Cliente.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cliente no encontrado'
        }, status=404)


@login_required
def obtener_citas_cliente(request, cliente_id):
    """
    Vista AJAX para obtener las citas de un cliente
    """
    try:
        cliente = Cliente.objects.get(id=cliente_id)
        citas = Cita.objects.filter(cliente=cliente).order_by('-fecha_hora')
        
        citas_data = []
        for cita in citas:
            citas_data.append({
                'id': cita.id,
                'fecha_hora': cita.fecha_hora.strftime('%d/%m/%Y %H:%M'),
                'estado': cita.get_estado_display(),
                'estado_class': cita.estado,
                'tipo_consulta': cita.tipo_consulta or 'No especificado',
                'dentista': cita.dentista.nombre_completo if cita.dentista else 'Sin asignar',
                'notas': cita.notas or '',
            })
        
        data = {
            'success': True,
            'cliente': {
                'nombre': cliente.nombre_completo,
                'email': cliente.email,
            },
            'citas': citas_data,
            'total_citas': len(citas_data)
        }
        return JsonResponse(data)
        
    except Cliente.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cliente no encontrado'
        }, status=404)


@login_required
def toggle_estado_cliente(request, cliente_id):
    """
    Vista para activar/desactivar un cliente
    También desactiva/activa su cuenta de usuario web si existe
    ÚTIL PARA BANEAR USUARIOS PROBLEMÁTICOS
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({
                'success': False,
                'error': 'No tienes permisos para realizar esta acción'
            }, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'No tienes permisos'
        }, status=403)
    
    try:
        from django.contrib.auth.models import User
        
        cliente = Cliente.objects.get(id=cliente_id)
        nuevo_estado = not cliente.activo
        cliente.activo = nuevo_estado
        cliente.save()
        
        # Buscar si existe un usuario web asociado (por email)
        try:
            user = User.objects.get(email=cliente.email)
            # Desactivar/activar también el usuario de Django
            user.is_active = nuevo_estado
            user.save()
            
            mensaje = f'Cliente {"activado" if nuevo_estado else "desactivado"} exitosamente. '
            mensaje += f'Usuario web también {"activado" if nuevo_estado else "desactivado"} (no podrá hacer login).'
        except User.DoesNotExist:
            # El cliente no tiene usuario web
            mensaje = f'Cliente {"activado" if nuevo_estado else "desactivado"} exitosamente.'
        
        return JsonResponse({
            'success': True,
            'activo': cliente.activo,
            'message': mensaje
        })
    except Cliente.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cliente no encontrado'
        }, status=404)


@login_required
def eliminar_cliente(request, cliente_id):
    """
    Vista para eliminar un cliente del sistema
    También elimina su cuenta de usuario web si existe
    ESTA ACCIÓN ES PERMANENTE
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({
                'success': False,
                'error': 'No tienes permisos para realizar esta acción'
            }, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'No tienes permisos'
        }, status=403)
    
    try:
        from django.contrib.auth.models import User
        
        cliente = Cliente.objects.get(id=cliente_id)
        nombre_cliente = cliente.nombre_completo
        email_cliente = cliente.email
        
        # ANTES de eliminar: Manejar las citas reservadas del cliente
        citas_reservadas = cliente.citas.filter(estado='reservada')
        citas_actualizadas = 0
        
        for cita in citas_reservadas:
            # Asegurarse de que los campos de respaldo estén llenos antes de eliminar el cliente
            if not cita.paciente_nombre:
                cita.paciente_nombre = cliente.nombre_completo
            if not cita.paciente_email:
                cita.paciente_email = cliente.email
            if not cita.paciente_telefono:
                cita.paciente_telefono = cliente.telefono
            
            # Cambiar el estado de "reservada" a "disponible" para que aparezca como disponible en el calendario
            # O podrías cambiarlo a "cancelada" si prefieres
            cita.estado = 'disponible'
            cita.save()
            citas_actualizadas += 1
        
        # Verificar relaciones antes de eliminar (para debugging)
        try:
            citas_count = cliente.citas.count()
            odontogramas_count = cliente.odontogramas.count()
            radiografias_count = cliente.radiografias.count() if hasattr(cliente, 'radiografias') else 0
            mensajes_count = cliente.mensajes.count() if hasattr(cliente, 'mensajes') else 0
            evaluaciones_count = cliente.evaluaciones.count() if hasattr(cliente, 'evaluaciones') else 0
            
            # Log para debugging (opcional)
            if citas_count > 0 or odontogramas_count > 0 or radiografias_count > 0:
                print(f"Cliente {cliente_id} tiene: {citas_count} citas, {odontogramas_count} odontogramas, {radiografias_count} radiografias")
        except Exception as e:
            print(f"Error al verificar relaciones: {str(e)}")
        
        # Buscar y eliminar el usuario web asociado si existe
        # Usar transacción atómica para asegurar consistencia
        usuario_web_eliminado = False
        
        try:
            import logging
            logger = logging.getLogger(__name__)
            
            # PRIMERO: Intentar usar la relación explícita (nueva mejora)
            if cliente.user:
                logger.info(f"Encontrado User asociado directamente: {cliente.user.username} (ID: {cliente.user.id})")
                user_a_eliminar = cliente.user
                
                # Usar transacción atómica para la eliminación de User/PerfilCliente
                with transaction.atomic():
                    # Eliminar PerfilCliente primero
                    try:
                        from django.db import connection
                        with connection.cursor() as cursor:
                            cursor.execute(
                                "DELETE FROM cuentas_perfilcliente WHERE user_id = %s",
                                [user_a_eliminar.id]
                            )
                            logger.info(f"✅ PerfilCliente eliminado para usuario {user_a_eliminar.id}")
                    except Exception as e:
                        logger.warning(f"⚠️ No se pudo eliminar PerfilCliente: {e}")
                    
                    # Eliminar User
                    user_a_eliminar.delete()
                    usuario_web_eliminado = True
                    logger.info(f"✅ Usuario {user_a_eliminar.username} eliminado exitosamente usando relación explícita")
            else:
                # FALLBACK: Buscar por email/username (método anterior para compatibilidad)
                # Ya no extraemos el username de las notas (por seguridad, las credenciales no se guardan ahí)
                username_cliente = None
                
                # Usar transacción atómica para la eliminación de User/PerfilCliente
                with transaction.atomic():
                    # Buscar User por email
                    users_por_email = User.objects.filter(email__iexact=email_cliente)
                    
                    # Buscar User por username si lo tenemos
                    users_por_username = User.objects.none()
                    if username_cliente:
                        users_por_username = User.objects.filter(username=username_cliente)
                    
                    # Combinar ambas búsquedas (sin duplicados)
                    users = (users_por_email | users_por_username).distinct()
                    
                    if users.exists():
                        logger.info(f"Encontrados {users.count()} usuario(s) asociado(s) al cliente {cliente_id}")
                        for user in users:
                            logger.info(f"Eliminando usuario: {user.username} (email: {user.email})")
                            
                            # Eliminar también el perfil de cliente si existe (ANTES de eliminar el User)
                            # Esto es importante porque si eliminamos el User primero, el CASCADE eliminará el PerfilCliente
                            # pero queremos asegurarnos de que se elimine correctamente
                            try:
                                from django.db import connection
                                with connection.cursor() as cursor:
                                    # Verificar si existe PerfilCliente
                                    cursor.execute(
                                        "SELECT id FROM cuentas_perfilcliente WHERE user_id = %s",
                                        [user.id]
                                    )
                                    perfil_existe = cursor.fetchone()
                                    
                                    if perfil_existe:
                                        cursor.execute(
                                            "DELETE FROM cuentas_perfilcliente WHERE user_id = %s",
                                            [user.id]
                                        )
                                        logger.info(f"✅ PerfilCliente eliminado para usuario {user.id}")
                                    else:
                                        logger.info(f"ℹ️ No se encontró PerfilCliente para usuario {user.id}")
                            except Exception as e:
                                logger.error(f"❌ Error al eliminar PerfilCliente: {e}", exc_info=True)
                                # Continuar con la eliminación del User aunque falle el PerfilCliente
                            
                            # Eliminar el usuario de Django
                            # Si el PerfilCliente no se eliminó antes, el CASCADE lo eliminará automáticamente
                            try:
                                user.delete()
                                usuario_web_eliminado = True
                                logger.info(f"✅ Usuario {user.username} (ID: {user.id}) eliminado exitosamente")
                            except Exception as e:
                                logger.error(f"❌ Error al eliminar User {user.username}: {e}", exc_info=True)
                                raise  # Re-lanzar el error para que se capture arriba
                    else:
                        logger.info(f"ℹ️ No se encontraron usuarios asociados al cliente {cliente_id} (email: {email_cliente}, username: {username_cliente})")
                        
                        # Si no encontramos User por email/username, buscar PerfilCliente directamente por email
                        # y eliminar tanto el PerfilCliente como su User asociado
                        try:
                            from django.db import connection
                            with connection.cursor() as cursor:
                                # Buscar PerfilCliente por email
                                cursor.execute(
                                    "SELECT user_id FROM cuentas_perfilcliente WHERE email = %s",
                                    [email_cliente]
                                )
                                perfil_result = cursor.fetchone()
                                
                                if perfil_result:
                                    user_id_perfil = perfil_result[0]
                                    logger.info(f"🔍 Encontrado PerfilCliente con user_id={user_id_perfil} para email {email_cliente}")
                                    
                                    # Eliminar PerfilCliente
                                    cursor.execute(
                                        "DELETE FROM cuentas_perfilcliente WHERE user_id = %s",
                                        [user_id_perfil]
                                    )
                                    logger.info(f"✅ PerfilCliente eliminado (user_id: {user_id_perfil})")
                                    
                                    # Eliminar User
                                    try:
                                        user_orphan = User.objects.get(id=user_id_perfil)
                                        user_orphan.delete()
                                        usuario_web_eliminado = True
                                        logger.info(f"✅ Usuario huérfano {user_orphan.username} eliminado")
                                    except User.DoesNotExist:
                                        logger.warning(f"⚠️ User {user_id_perfil} no existe, solo se eliminó PerfilCliente")
                        except Exception as e2:
                            logger.warning(f"⚠️ No se pudo buscar/eliminar PerfilCliente por email: {e2}")
                    
        except Exception as e:
            # Si hay un error en la transacción, se revierte automáticamente
            # Si hay un error al eliminar el usuario, continuar con la eliminación del cliente
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"❌ Error al eliminar usuario web: {str(e)}", exc_info=True)
            import traceback
            logger.error(f"Traceback completo: {traceback.format_exc()}")
        
        # Registrar en auditoría ANTES de eliminar
        detalles_eliminacion = f'Email: {email_cliente}, RUT: {cliente.rut if cliente.rut else "N/A"}'
        if citas_actualizadas > 0:
            detalles_eliminacion += f', {citas_actualizadas} cita(s) actualizada(s)'
        if usuario_web_eliminado:
            detalles_eliminacion += ', Usuario web eliminado'
        registrar_auditoria(
            usuario=perfil,
            accion='eliminar',
            modulo='clientes',
            descripcion=f'Cliente eliminado: {nombre_cliente}',
            detalles=detalles_eliminacion,
            objeto_id=cliente_id,
            tipo_objeto='Cliente',
            request=request
        )
        
        # Eliminar el cliente
        cliente.delete()
        
        # Mensaje de confirmación
        mensaje = f'Cliente "{nombre_cliente}" eliminado exitosamente.'
        if citas_actualizadas > 0:
            mensaje += f' {citas_actualizadas} cita(s) reservada(s) fueron marcadas como disponibles.'
        if usuario_web_eliminado:
            mensaje += ' Su cuenta de usuario web también fue eliminada.'
        
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'citas_actualizadas': citas_actualizadas
        })
    except Cliente.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cliente no encontrado'
        }, status=404)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al eliminar el cliente: {str(e)}',
            'details': str(e)  # Incluir detalles del error
        }, status=500)


# ========== MIS PACIENTES (Vista Unificada para Dentistas) ==========

@login_required
def mis_pacientes(request, paciente_id=None, seccion=None):
    """Vista principal unificada para gestionar pacientes del dentista"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'Solo los dentistas pueden acceder a esta sección.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    # Obtener pacientes vinculados al dentista
    pacientes_vinculados = perfil.get_pacientes_asignados()
    clientes_ids = [p['id'] for p in pacientes_vinculados if 'id' in p and isinstance(p['id'], int)]
    emails_vinculados = [p['email'] for p in pacientes_vinculados if 'email' in p]
    
    # Filtros de búsqueda
    search = request.GET.get('search', '')
    
    # Construir lista de pacientes con información consolidada
    pacientes_lista = []
    
    # Obtener clientes del sistema vinculados
    clientes_sistema = Cliente.objects.filter(
        id__in=clientes_ids,
        activo=True
    ).distinct()
    
    for cliente in clientes_sistema:
        # Contar radiografías y odontogramas
        radiografias_count = Radiografia.objects.filter(
            Q(dentista=perfil) & (Q(cliente=cliente) | Q(paciente_email=cliente.email))
        ).count()
        
        odontogramas_count = Odontograma.objects.filter(
            Q(dentista=perfil) & (Q(cliente=cliente) | Q(paciente_email=cliente.email))
        ).count()
        
        # Aplicar filtro de búsqueda
        if search:
            if not (search.lower() in cliente.nombre_completo.lower() or
                    search.lower() in cliente.email.lower() or
                    search.lower() in (cliente.telefono or '').lower()):
                continue
        
        pacientes_lista.append({
            'id': cliente.id,
            'nombre_completo': cliente.nombre_completo,
            'email': cliente.email,
            'telefono': cliente.telefono or '',
            'total_radiografias': radiografias_count,
            'total_odontogramas': odontogramas_count,
            'tiene_radiografias': radiografias_count > 0,
            'tiene_odontogramas': odontogramas_count > 0,
        })
    
    # Agregar pacientes vinculados que no están en el sistema pero tienen documentos
    emails_procesados = [p['email'] for p in pacientes_lista]
    
    for paciente_vinculado in pacientes_vinculados:
        email_paciente = paciente_vinculado.get('email', '')
        if not email_paciente or email_paciente in emails_procesados:
            continue
        
        # Verificar si tiene radiografías u odontogramas
        radiografias_count = Radiografia.objects.filter(
            dentista=perfil,
            paciente_email=email_paciente
        ).count()
        
        odontogramas_count = Odontograma.objects.filter(
            dentista=perfil,
            paciente_email=email_paciente
        ).count()
        
        # Solo agregar si tiene documentos o si no hay filtro de búsqueda
        if radiografias_count == 0 and odontogramas_count == 0:
            if search:
                continue
        
        # Aplicar filtro de búsqueda
        if search:
            nombre = paciente_vinculado.get('nombre_completo', '')
            if not (search.lower() in nombre.lower() or search.lower() in email_paciente.lower()):
                continue
        
        pacientes_lista.append({
            'id': paciente_vinculado.get('id'),
            'nombre_completo': paciente_vinculado.get('nombre_completo', email_paciente),
            'email': email_paciente,
            'telefono': paciente_vinculado.get('telefono', ''),
            'total_radiografias': radiografias_count,
            'total_odontogramas': odontogramas_count,
            'tiene_radiografias': radiografias_count > 0,
            'tiene_odontogramas': odontogramas_count > 0,
        })
    
    # Ordenar por nombre
    pacientes_lista.sort(key=lambda x: x['nombre_completo'])
    
    # Obtener información del paciente seleccionado si existe
    paciente_seleccionado = None
    cliente_obj = None
    if paciente_id:
        paciente_seleccionado = next((p for p in pacientes_lista if p['id'] == paciente_id), None)
        if not paciente_seleccionado:
            # Intentar buscar por email si no se encontró por ID
            for paciente_vinculado in pacientes_vinculados:
                if paciente_vinculado.get('id') == paciente_id:
                    paciente_seleccionado = {
                        'id': paciente_vinculado.get('id'),
                        'nombre_completo': paciente_vinculado.get('nombre_completo', paciente_vinculado.get('email', '')),
                        'email': paciente_vinculado.get('email', ''),
                        'telefono': paciente_vinculado.get('telefono', ''),
                        'total_radiografias': 0,
                        'total_odontogramas': 0,
                        'tiene_radiografias': False,
                        'tiene_odontogramas': False,
                    }
                    break
        
        # Obtener el objeto Cliente completo si existe
        if paciente_seleccionado and paciente_seleccionado.get('id'):
            try:
                cliente_obj = Cliente.objects.get(id=paciente_seleccionado['id'], activo=True)
            except Cliente.DoesNotExist:
                cliente_obj = None
    
    # Determinar sección por defecto
    if not seccion:
        seccion = 'resumen'
    
    # Obtener datos adicionales según la sección seleccionada
    radiografias = None
    odontogramas = None
    
    if paciente_seleccionado and seccion == 'radiografias':
        # Obtener radiografías del paciente
        paciente_email = paciente_seleccionado.get('email', '')
        radiografias = Radiografia.objects.filter(
            dentista=perfil,
            paciente_email=paciente_email
        ).order_by('-fecha_carga')
        
    elif paciente_seleccionado and seccion == 'odontogramas':
        # Obtener odontogramas del paciente
        paciente_email = paciente_seleccionado.get('email', '')
        paciente_id_seleccionado = paciente_seleccionado.get('id')
        
        # Buscar odontogramas por cliente o por email
        odontogramas_query = Odontograma.objects.filter(dentista=perfil)
        
        if paciente_id_seleccionado:
            try:
                cliente_obj = Cliente.objects.get(id=paciente_id_seleccionado, activo=True)
                odontogramas = odontogramas_query.filter(
                    Q(cliente=cliente_obj) | Q(paciente_email=cliente_obj.email)
                ).prefetch_related(
                    Prefetch('dientes', queryset=EstadoDiente.objects.order_by('numero_diente'))
                ).order_by('-fecha_creacion')
            except Cliente.DoesNotExist:
                odontogramas = odontogramas_query.filter(
                    paciente_email=paciente_email
                ).prefetch_related(
                    Prefetch('dientes', queryset=EstadoDiente.objects.order_by('numero_diente'))
                ).order_by('-fecha_creacion')
        else:
            odontogramas = odontogramas_query.filter(
                paciente_email=paciente_email
            ).prefetch_related(
                Prefetch('dientes', queryset=EstadoDiente.objects.order_by('numero_diente'))
            ).order_by('-fecha_creacion')
    
    context = {
        'perfil': perfil,
        'pacientes': pacientes_lista,
        'paciente_seleccionado': paciente_seleccionado,
        'cliente': cliente_obj,  # Objeto Cliente completo para mostrar información detallada
        'seccion_actual': seccion,
        'search': search,
        'es_dentista': True,
        'radiografias': radiografias,
        'odontogramas': odontogramas,
    }
    
    return render(request, 'citas/mis_pacientes/mis_pacientes.html', context)


# ========== GESTIÓN DE RADIOGRAFÍAS ==========

@login_required
def radiografias_listar(request):
    """Vista principal para listar pacientes vinculados con sus radiografías"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'Solo los dentistas pueden gestionar radiografías.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    # Filtros de búsqueda
    search = request.GET.get('search', '')
    
    # Obtener pacientes vinculados al dentista
    pacientes_vinculados = perfil.get_pacientes_asignados()
    clientes_ids = [p['id'] for p in pacientes_vinculados if 'id' in p and isinstance(p['id'], int)]
    emails_vinculados = [p['email'] for p in pacientes_vinculados if 'email' in p]
    
    # Obtener solo radiografías de pacientes vinculados
    radiografias = Radiografia.objects.filter(
        dentista=perfil
    ).filter(
        Q(cliente_id__in=clientes_ids) | Q(paciente_email__in=emails_vinculados)
    ).select_related('cliente')

    # Obtener todos los emails únicos de pacientes vinculados que tienen radiografías
    emails_unicos = radiografias.values_list('paciente_email', flat=True).distinct()
    
    # Obtener clientes vinculados que tienen radiografías asociadas
    clientes_con_radiografias = Cliente.objects.filter(
        id__in=clientes_ids,
        radiografias__dentista=perfil,
        activo=True
    ).distinct()
    
    # Crear diccionario de pacientes con radiografías
    pacientes_dict = {}
    
    # Primero procesar clientes vinculados del sistema que tienen radiografías
    for cliente in clientes_con_radiografias:
        # Obtener todas las radiografías de este cliente (por relación cliente o por email)
        radiografias_cliente = radiografias.filter(
            Q(cliente=cliente) | Q(paciente_email=cliente.email)
        )
        
        if not radiografias_cliente.exists():
            continue
        
        # Usar la información actualizada del cliente
        nombre_completo = cliente.nombre_completo
        telefono = cliente.telefono or ''
        email_actual = cliente.email
        paciente_id = cliente.id
        
        # Aplicar filtro de búsqueda
        if search:
            if not (search.lower() in nombre_completo.lower() or
                    search.lower() in email_actual.lower() or
                    search.lower() in (telefono or '').lower()):
                continue
        
        # Usar el email actual del cliente como clave
        pacientes_dict[email_actual] = {
            'id': paciente_id,
            'nombre_completo': nombre_completo,
            'email': email_actual,
            'telefono': telefono,
            'total_radiografias': radiografias_cliente.count(),
            'ultima_radiografia': radiografias_cliente.order_by('-fecha_carga').first(),
        }
    
    # Luego procesar emails vinculados que no tienen cliente asociado en el sistema
    for email in emails_unicos:
        # Solo procesar si el email está en la lista de pacientes vinculados
        if email not in emails_vinculados:
            continue
            
        # Si ya procesamos este email (tiene cliente), saltar
        if email in pacientes_dict:
            continue
        
        # Buscar si existe un cliente vinculado en el sistema con este email
        try:
            cliente = Cliente.objects.get(email=email, id__in=clientes_ids, activo=True)
            nombre_completo = cliente.nombre_completo
            telefono = cliente.telefono or ''
            paciente_id = cliente.id
            email_actual = cliente.email
        except Cliente.DoesNotExist:
            # Si no existe cliente, obtener información de las radiografías
            primera_radiografia = radiografias.filter(paciente_email=email).first()
            if primera_radiografia:
                nombre_completo = primera_radiografia.paciente_nombre
                telefono = ''
                # Buscar en pacientes vinculados para obtener el ID correcto
                paciente_vinculado = next((p for p in pacientes_vinculados if p.get('email') == email), None)
                paciente_id = paciente_vinculado['id'] if paciente_vinculado and 'id' in paciente_vinculado else hash(email) % 1000000
                email_actual = email
            else:
                continue
        
        # Aplicar filtro de búsqueda
        if search:
            if not (search.lower() in nombre_completo.lower() or
                    search.lower() in email_actual.lower()):
                continue
        
        # Contar radiografías de este paciente
        radiografias_paciente = radiografias.filter(paciente_email=email)
        
        pacientes_dict[email_actual] = {
            'id': paciente_id,
            'nombre_completo': nombre_completo,
            'email': email_actual,
            'telefono': telefono,
            'total_radiografias': radiografias_paciente.count(),
            'ultima_radiografia': radiografias_paciente.order_by('-fecha_carga').first(),
        }
    
    # Convertir a lista y ordenar por nombre
    pacientes_con_radiografias = list(pacientes_dict.values())
    pacientes_con_radiografias.sort(key=lambda x: x['nombre_completo'])
    
    # Obtener pacientes vinculados que aún no tienen radiografías (para mostrar opción de agregar)
    pacientes_sin_radiografias = []
    for paciente_vinculado in pacientes_vinculados:
        email_paciente = paciente_vinculado.get('email', '')
        if email_paciente and email_paciente not in [p['email'] for p in pacientes_con_radiografias]:
            pacientes_sin_radiografias.append({
                'id': paciente_vinculado.get('id'),
                'nombre_completo': paciente_vinculado.get('nombre_completo', ''),
                'email': email_paciente,
                'telefono': paciente_vinculado.get('telefono', ''),
            })
    
    # Aplicar filtro de búsqueda a pacientes sin radiografías
    if search:
        pacientes_sin_radiografias = [
            p for p in pacientes_sin_radiografias
            if search.lower() in p['nombre_completo'].lower() or search.lower() in p['email'].lower()
        ]
    
    # Calcular estadísticas
    total_radiografias = radiografias.count()
    radiografias_mes = radiografias.filter(
        fecha_carga__month=timezone.now().month,
        fecha_carga__year=timezone.now().year
    ).count()
    radiografias_semana = radiografias.filter(
        fecha_carga__gte=timezone.now() - timedelta(days=7)
    ).count()
    total_pacientes_vinculados = len(pacientes_vinculados)
    
    context = {
        'perfil': perfil,
        'pacientes': pacientes_con_radiografias,
        'pacientes_sin_radiografias': pacientes_sin_radiografias,
        'search': search,
        'es_dentista': True,
        'estadisticas': {
            'total_radiografias': total_radiografias,
            'total_pacientes_con_radiografias': len(pacientes_con_radiografias),
            'total_pacientes_vinculados': total_pacientes_vinculados,
            'radiografias_mes': radiografias_mes,
            'radiografias_semana': radiografias_semana,
        }
    }
    
    return render(request, 'citas/radiografias/radiografias_listar.html', context)


@login_required
def radiografias_paciente(request, paciente_id):
    """Vista para ver y gestionar radiografías de un paciente vinculado específico"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'Solo los dentistas pueden gestionar radiografías.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    # Obtener pacientes vinculados al dentista
    pacientes_vinculados = perfil.get_pacientes_asignados()
    clientes_ids = [p['id'] for p in pacientes_vinculados if 'id' in p and isinstance(p['id'], int)]
    emails_vinculados = [p['email'] for p in pacientes_vinculados if 'email' in p]
    
    # Obtener todas las radiografías del dentista de pacientes vinculados
    radiografias_todas = Radiografia.objects.filter(
        dentista=perfil
    ).filter(
        Q(cliente_id__in=clientes_ids) | Q(paciente_email__in=emails_vinculados)
    )
    
    # Intentar encontrar el paciente por ID (debe ser paciente vinculado)
    paciente = None
    paciente_email = None
    
    # Primero intentar buscar como cliente vinculado del sistema
    try:
        cliente = Cliente.objects.get(id=paciente_id, id__in=clientes_ids, activo=True)
        paciente_email = cliente.email
        paciente = {
            'id': cliente.id,
            'nombre_completo': cliente.nombre_completo,
            'email': cliente.email,
            'telefono': cliente.telefono or '',
        }
    except Cliente.DoesNotExist:
        # Si no es cliente del sistema, buscar por email en las radiografías de pacientes vinculados
        # usando el hash del email
        for radiografia in radiografias_todas:
            email_hash = hash(radiografia.paciente_email) % 1000000
            if email_hash == int(paciente_id) and radiografia.paciente_email in emails_vinculados:
                paciente_email = radiografia.paciente_email
                # Buscar si hay un cliente vinculado con este email
                try:
                    cliente = Cliente.objects.get(email=paciente_email, id__in=clientes_ids, activo=True)
                    paciente = {
                        'id': cliente.id,
                        'nombre_completo': cliente.nombre_completo,
                        'email': cliente.email,
                        'telefono': cliente.telefono or '',
                    }
                except Cliente.DoesNotExist:
                    # Si no existe cliente, buscar en pacientes vinculados
                    paciente_vinculado = next((p for p in pacientes_vinculados if p.get('email') == paciente_email), None)
                    if paciente_vinculado:
                        primera_radiografia = radiografias_todas.filter(paciente_email=paciente_email).first()
                        if primera_radiografia:
                            paciente = {
                                'id': paciente_vinculado.get('id', int(paciente_id)),
                                'nombre_completo': primera_radiografia.paciente_nombre,
                                'email': paciente_email,
                                'telefono': paciente_vinculado.get('telefono', ''),
                            }
                break
    
    # Verificar que el paciente esté vinculado
    if not paciente or not paciente_email or paciente_email not in emails_vinculados:
        messages.error(request, 'Solo puedes ver radiografías de tus pacientes vinculados.')
        return redirect('radiografias_listar')
    
    # Obtener todas las radiografías del paciente vinculado
    radiografias = Radiografia.objects.filter(
        dentista=perfil,
        paciente_email=paciente_email
    ).order_by('-fecha_carga')
    
    context = {
        'perfil': perfil,
        'paciente': paciente,
        'radiografias': radiografias,
        'es_dentista': True
    }
    
    return render(request, 'citas/radiografias/radiografias_paciente.html', context)


@login_required
def agregar_radiografia(request, paciente_id):
    """Vista para agregar una nueva radiografía - Solo pacientes vinculados"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'Solo los dentistas pueden agregar radiografías.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')

    # Obtener pacientes vinculados al dentista
    pacientes_vinculados = perfil.get_pacientes_asignados()
    clientes_ids = [p['id'] for p in pacientes_vinculados if 'id' in p and isinstance(p['id'], int)]
    emails_vinculados = [p['email'] for p in pacientes_vinculados if 'email' in p]
    
    # Obtener el paciente - debe ser un paciente vinculado
    paciente = None
    paciente_email = None
    
    # Primero intentar buscar como cliente vinculado del sistema
    try:
        cliente = Cliente.objects.get(id=paciente_id, id__in=clientes_ids, activo=True)
        paciente_email = cliente.email
        paciente = {
            'id': cliente.id,
            'nombre_completo': cliente.nombre_completo,
            'email': cliente.email,
            'telefono': cliente.telefono or '',
        }
    except Cliente.DoesNotExist:
        # Si no es cliente del sistema, buscar por email en pacientes vinculados
        # usando el hash del email
        radiografias_todas = Radiografia.objects.filter(dentista=perfil)
        for radiografia in radiografias_todas:
            email_hash = hash(radiografia.paciente_email) % 1000000
            if email_hash == int(paciente_id) and radiografia.paciente_email in emails_vinculados:
                paciente_email = radiografia.paciente_email
                # Buscar si hay un cliente vinculado con este email
                try:
                    cliente = Cliente.objects.get(email=paciente_email, id__in=clientes_ids, activo=True)
                    paciente = {
                        'id': cliente.id,
                        'nombre_completo': cliente.nombre_completo,
                        'email': cliente.email,
                        'telefono': cliente.telefono or '',
                    }
                except Cliente.DoesNotExist:
                    # Si no existe cliente, buscar en pacientes vinculados
                    paciente_vinculado = next((p for p in pacientes_vinculados if p.get('email') == paciente_email), None)
                    if paciente_vinculado:
                        primera_radiografia = radiografias_todas.filter(paciente_email=paciente_email).first()
                        if primera_radiografia:
                            paciente = {
                                'id': paciente_vinculado.get('id', int(paciente_id)),
                                'nombre_completo': primera_radiografia.paciente_nombre,
                                'email': paciente_email,
                                'telefono': paciente_vinculado.get('telefono', ''),
                            }
                break
    
    # Verificar que el paciente esté vinculado al dentista
    if not paciente or not paciente_email:
        messages.error(request, 'No tienes permisos para agregar radiografías a este paciente.')
        return redirect('radiografias_listar')
    
    if paciente_email not in emails_vinculados:
        messages.error(request, 'Solo puedes agregar radiografías a tus pacientes vinculados.')
        return redirect('radiografias_listar')
    
    if request.method == 'POST':
        try:
            tipo = request.POST.get('tipo', 'periapical')
            descripcion = request.POST.get('descripcion', '')
            fecha_tomada = request.POST.get('fecha_tomada', '')
            imagen = request.FILES.get('imagen')
            
            if not imagen:
                messages.error(request, 'Debes seleccionar una imagen.')
                return redirect('mis_pacientes_seccion', paciente_id=paciente_id, seccion='radiografias')
            
            # Buscar si existe un cliente con ese email
            cliente_obj = None
            try:
                cliente_obj = Cliente.objects.get(email=paciente['email'], activo=True)
            except Cliente.DoesNotExist:
                pass
            
            # Crear la radiografía
            radiografia = Radiografia.objects.create(
                cliente=cliente_obj,  # Cliente del sistema (opcional)
                paciente_email=paciente['email'],
                paciente_nombre=paciente['nombre_completo'],
                dentista=perfil,
                tipo=tipo,
                descripcion=descripcion,
                imagen=imagen,
                fecha_tomada=fecha_tomada if fecha_tomada else None
            )
            
            messages.success(request, 'Radiografía agregada correctamente.')
            # Redirigir a la nueva vista de Mis Pacientes con sección de radiografías
            return redirect('mis_pacientes_seccion', paciente_id=paciente_id, seccion='radiografias')
            
        except Exception as e:
            messages.error(request, f'Error al agregar la radiografía: {str(e)}')
    
    context = {
        'perfil': perfil,
        'paciente': paciente,
        'tipos_radiografia': Radiografia.TIPO_RADIOGRAFIA_CHOICES,
        'es_dentista': True
    }
    
    return render(request, 'citas/radiografias/agregar_radiografia.html', context)


@login_required
def editar_radiografia(request, radiografia_id):
    """Vista para editar una radiografía existente"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'Solo los dentistas pueden editar radiografías.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    try:
        radiografia = Radiografia.objects.get(id=radiografia_id, dentista=perfil)
    except Radiografia.DoesNotExist:
        messages.error(request, 'Radiografía no encontrada.')
        return redirect('radiografias_listar')
    
    # Obtener paciente_id para redirección
    paciente_id = None
    if radiografia.cliente:
        paciente_id = radiografia.cliente.id
    else:
        paciente_id = hash(radiografia.paciente_email) % 1000000
    
    # Obtener citas del paciente para el selector
    citas_paciente = []
    if radiografia.cliente:
        citas_paciente = Cita.objects.filter(
            cliente=radiografia.cliente,
            estado__in=['reservada', 'completada']
        ).order_by('-fecha_hora')[:20]
    elif radiografia.paciente_email:
        # Buscar citas por email
        citas_paciente = Cita.objects.filter(
            paciente_email=radiografia.paciente_email,
            estado__in=['reservada', 'completada']
        ).order_by('-fecha_hora')[:20]
    
    if request.method == 'POST':
        try:
            tipo = request.POST.get('tipo', radiografia.tipo)
            descripcion = request.POST.get('descripcion', '')
            fecha_tomada = request.POST.get('fecha_tomada', '')
            cita_id = request.POST.get('cita', '')
            imagen = request.FILES.get('imagen')
            
            # Actualizar campos
            radiografia.tipo = tipo
            radiografia.descripcion = descripcion if descripcion else None
            radiografia.fecha_tomada = fecha_tomada if fecha_tomada else None
            
            # Actualizar cita asociada
            if cita_id:
                try:
                    cita = Cita.objects.get(id=cita_id)
                    radiografia.cita = cita
                except Cita.DoesNotExist:
                    radiografia.cita = None
            else:
                radiografia.cita = None
            
            # Actualizar imagen si se proporciona una nueva
            if imagen:
                radiografia.imagen = imagen
                # Si hay nueva imagen, eliminar anotaciones anteriores
                if radiografia.imagen_anotada:
                    radiografia.imagen_anotada.delete()
                    radiografia.imagen_anotada = None
            
            radiografia.save()
            
            messages.success(request, 'Radiografía actualizada correctamente.')
            # Redirigir a la nueva vista de Mis Pacientes con sección de radiografías
            return redirect('mis_pacientes_seccion', paciente_id=paciente_id, seccion='radiografias')
            
        except Exception as e:
            messages.error(request, f'Error al actualizar la radiografía: {str(e)}')
    
    context = {
        'perfil': perfil,
        'radiografia': radiografia,
        'paciente_id': paciente_id,
        'tipos_radiografia': Radiografia.TIPO_RADIOGRAFIA_CHOICES,
        'citas_paciente': citas_paciente,
    }
    
    return render(request, 'citas/radiografias/editar_radiografia.html', context)


@login_required
def guardar_anotaciones_radiografia(request, radiografia_id):
    """Vista AJAX para guardar anotaciones de una radiografía"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            return JsonResponse({'success': False, 'error': 'No tienes permisos para guardar anotaciones.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    try:
        radiografia = Radiografia.objects.get(id=radiografia_id, dentista=perfil)
    except Radiografia.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Radiografía no encontrada.'}, status=404)
    
    try:
        from django.core.files.base import ContentFile
        import base64
        
        # Obtener imagen base64 del canvas
        imagen_data = request.POST.get('imagen_data', '')
        if not imagen_data:
            return JsonResponse({'success': False, 'error': 'No se proporcionó imagen.'}, status=400)
        
        # Decodificar imagen base64
        if ',' in imagen_data:
            imagen_data = imagen_data.split(',')[1]
        
        image_data = base64.b64decode(imagen_data)
        image_file = ContentFile(image_data, name=f'radiografia_{radiografia_id}_anotada.png')
        
        # Eliminar imagen anterior si existe
        if radiografia.imagen_anotada:
            radiografia.imagen_anotada.delete()
        
        # Guardar nueva imagen con anotaciones
        radiografia.imagen_anotada = image_file
        radiografia.save()
        
        return JsonResponse({'success': True, 'message': 'Anotaciones guardadas correctamente.'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al guardar anotaciones: {str(e)}'}, status=500)


@login_required
def obtener_anotaciones_radiografia(request, radiografia_id):
    """Vista AJAX para obtener la imagen con anotaciones de una radiografía"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    try:
        radiografia = Radiografia.objects.get(id=radiografia_id, dentista=perfil)
        
        if radiografia.imagen_anotada:
            return JsonResponse({
                'success': True,
                'imagen_anotada': radiografia.imagen_anotada.url
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'No hay anotaciones guardadas'
            })
            
    except Radiografia.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Radiografía no encontrada.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error: {str(e)}'}, status=500)


@login_required
def obtener_cita(request, cita_id):
    """
    Vista AJAX para obtener los datos de una cita
    """
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Cuenta desactivada'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Perfil no encontrado'}, status=404)
    
    try:
        cita = Cita.objects.select_related('tipo_servicio', 'dentista', 'cliente').get(id=cita_id)
        
        # Convertir fecha_hora a zona horaria de Chile
        from django.utils import timezone
        import pytz
        try:
            chile_tz = pytz.timezone('America/Santiago')
            if timezone.is_naive(cita.fecha_hora):
                fecha_hora_chile = timezone.make_aware(cita.fecha_hora, timezone.utc).astimezone(chile_tz)
            else:
                fecha_hora_chile = cita.fecha_hora.astimezone(chile_tz)
        except Exception:
            fecha_hora_chile = cita.fecha_hora
        
        # Formatear fecha y hora en zona horaria de Chile
        fecha_hora = fecha_hora_chile.strftime('%d/%m/%Y %H:%M')
        fecha = fecha_hora_chile.strftime('%d/%m/%Y')
        hora = fecha_hora_chile.strftime('%H:%M')
        
        # Obtener paciente
        paciente = 'Sin asignar'
        if cita.cliente:
            paciente = cita.cliente.nombre_completo
        elif cita.paciente_nombre:
            paciente = cita.paciente_nombre
        
        # Obtener tipo de consulta
        tipo_consulta = 'Sin especificar'
        if cita.tipo_servicio:
            tipo_consulta = cita.tipo_servicio.nombre
        elif cita.tipo_consulta:
            tipo_consulta = cita.tipo_consulta
        
        # Obtener dentista
        dentista = 'Sin asignar'
        if cita.dentista:
            dentista = cita.dentista.nombre_completo
        
        # Obtener precio
        precio = None
        if cita.precio_cobrado:
            precio = float(cita.precio_cobrado)
        
        data = {
            'success': True,
            'cita': {
                'id': cita.id,
                'fecha_hora': fecha_hora,
                'fecha': fecha,
                'hora': hora,
                'paciente': paciente,
                'paciente_nombre': paciente,  # Alias para compatibilidad
                'tipo_consulta': tipo_consulta,
                'paciente': paciente,
                'dentista': dentista,
                'estado': cita.estado,
                'estado_display': cita.get_estado_display(),
                'notas': cita.notas or '',
                'precio_cobrado': precio,
                'motivo_no_asistencia': cita.motivo_no_asistencia or '',
            }
        }
        return JsonResponse(data)
    except Cita.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Cita no encontrada'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al obtener datos de la cita: {str(e)}'
        }, status=500)


@login_required
def eliminar_radiografia(request, radiografia_id):
    """Vista para eliminar una radiografía"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'Solo los dentistas pueden eliminar radiografías.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    try:
        radiografia = Radiografia.objects.get(id=radiografia_id, dentista=perfil)
        paciente_email = radiografia.paciente_email
        
        # Obtener el paciente para redirigir
        pacientes = perfil.get_pacientes_asignados()
        paciente_id = None
        for p in pacientes:
            if p['email'] == paciente_email:
                paciente_id = p['id']
                break
        
        radiografia.delete()
        messages.success(request, 'Radiografía eliminada correctamente.')
        
        if paciente_id:
            return redirect('radiografias_paciente', paciente_id=paciente_id)
        else:
            return redirect('radiografias_listar')
            
    except Radiografia.DoesNotExist:
        messages.error(request, 'Radiografía no encontrada.')
        return redirect('radiografias_listar')
    except Exception as e:
        messages.error(request, f'Error al eliminar la radiografía: {str(e)}')
        return redirect('radiografias_listar')


@login_required
def perfil_cliente(request, cliente_id):
    """Vista de perfil completo del cliente para administrativos con todos sus historiales"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'Solo los administrativos pueden ver perfiles de clientes.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    try:
        cliente = Cliente.objects.get(id=cliente_id)
    except Cliente.DoesNotExist:
        messages.error(request, 'Cliente no encontrado.')
        return redirect('gestor_clientes')
    
    # Obtener historiales del cliente
    # Para odontogramas y radiografías, buscar tanto por cliente directo como por email
    # para incluir registros que no tienen cliente asociado pero tienen el mismo email
    from django.db.models import Q
    odontogramas = Odontograma.objects.filter(
        Q(cliente=cliente) | Q(paciente_email=cliente.email)
    ).order_by('-fecha_creacion')
    
    radiografias = Radiografia.objects.filter(
        Q(cliente=cliente) | Q(paciente_email=cliente.email)
    ).order_by('-fecha_carga')
    
    citas = Cita.objects.filter(cliente=cliente).order_by('-fecha_hora')
    
    # Obtener planes de tratamiento del cliente
    planes_tratamiento = PlanTratamiento.objects.filter(cliente=cliente).order_by('-creado_el')
    
    # Agregar información de permisos de edición a cada plan
    for plan in planes_tratamiento:
        plan.puede_editar = plan.puede_ser_editado_por(perfil)
    
    # Estadísticas
    estadisticas = {
        'total_citas': citas.count(),
        'citas_completadas': citas.filter(estado='completada').count(),
        'citas_no_asistidas': citas.filter(estado='no_show').count(),
        'citas_pendientes': citas.filter(estado__in=['reservada', 'confirmada', 'en_espera', 'listo_para_atender', 'en_progreso', 'finalizada']).count(),
        'total_odontogramas': odontogramas.count(),
        'total_radiografias': radiografias.count(),
        'total_planes_tratamiento': planes_tratamiento.count(),
        'ultima_cita': citas.first(),
        'ultimo_odontograma': odontogramas.first(),
        'ultima_radiografia': radiografias.first(),
    }
    
    context = {
        'perfil': perfil,
        'cliente': cliente,
        'odontogramas': odontogramas,
        'radiografias': radiografias,
        'citas': citas[:10],  # Últimas 10 citas
        'planes_tratamiento': planes_tratamiento,
        'estadisticas': estadisticas,
        'es_admin': True
    }
    
    return render(request, 'citas/clientes/perfil_cliente.html', context)


@login_required
def enviar_radiografia_por_correo(request, radiografia_id):
    """Vista para enviar una radiografía al correo del cliente"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'Solo los administrativos pueden enviar radiografías por correo.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    try:
        radiografia = Radiografia.objects.get(id=radiografia_id)
        cliente = radiografia.cliente
        
        # Si no hay cliente asociado, intentar buscar por email
        if not cliente:
            try:
                cliente = Cliente.objects.get(email=radiografia.paciente_email, activo=True)
            except Cliente.DoesNotExist:
                messages.error(request, 'No se encontró el cliente asociado a esta radiografía.')
                return redirect('gestor_clientes')
        
        # Verificar que el cliente tenga email
        if not cliente.email:
            messages.error(request, 'El cliente no tiene un email registrado.')
            return redirect('perfil_cliente', cliente_id=cliente.id)
        
        # Importar EmailMessage
        from django.core.mail import EmailMessage
        from django.conf import settings
        from django.utils import timezone
        
        # Obtener información de la clínica
        try:
            from configuracion.models import InformacionClinica
            info_clinica = InformacionClinica.obtener()
            nombre_clinica = info_clinica.nombre_clinica
            email_clinica = info_clinica.email or settings.DEFAULT_FROM_EMAIL
            direccion_clinica = info_clinica.direccion or ''
            telefono_clinica = info_clinica.telefono or ''
        except:
            nombre_clinica = "Clínica Dental"
            email_clinica = settings.DEFAULT_FROM_EMAIL
            direccion_clinica = ''
            telefono_clinica = ''
        
        # Renderizar template HTML
        from django.template.loader import render_to_string
        mensaje_html = render_to_string('citas/emails/radiografia_enviada.html', {
            'cliente_nombre': cliente.nombre_completo,
            'tipo_radiografia': radiografia.get_tipo_display(),
            'fecha_carga': radiografia.fecha_carga,
            'dentista_nombre': radiografia.dentista.nombre_completo if radiografia.dentista else 'No especificado',
            'descripcion': radiografia.descripcion,
            'nombre_clinica': nombre_clinica,
            'direccion_clinica': direccion_clinica,
            'telefono_clinica': telefono_clinica,
            'email_clinica': email_clinica,
        })
        
        # Preparar el mensaje profesional
        asunto = f'Radiografía Dental - {radiografia.get_tipo_display()}'
        
        # Crear el correo con la imagen adjunta
        email = EmailMessage(
            subject=asunto,
            body=mensaje_html,
            from_email=email_clinica,
            to=[cliente.email],
        )
        email.content_subtype = "html"  # Indicar que es HTML
        
        # Adjuntar la imagen de la radiografía
        try:
            if radiografia.imagen:
                email.attach_file(radiografia.imagen.path)
        except Exception as e:
            messages.warning(request, f'Error al adjuntar la imagen: {str(e)}')
        
        # Enviar el correo
        try:
            email.send(fail_silently=False)
            messages.success(request, f'Radiografía enviada correctamente al correo {cliente.email}.')
        except Exception as e:
            messages.error(request, f'Error al enviar el correo: {str(e)}. Verifique la configuración de email.')
            return redirect('perfil_cliente', cliente_id=cliente.id)
        
        return redirect('perfil_cliente', cliente_id=cliente.id)
        
    except Radiografia.DoesNotExist:
        messages.error(request, 'Radiografía no encontrada.')
        return redirect('gestor_clientes')
    except Exception as e:
        messages.error(request, f'Error al procesar la solicitud: {str(e)}')
        return redirect('gestor_clientes')

# ==========================================
# GESTIÓN DE SERVICIOS DENTALES
# ==========================================

@login_required
def gestor_servicios(request):
    """Vista para gestionar tipos de servicios dentales"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'Solo los administrativos pueden gestionar servicios.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    # Filtros de búsqueda
    search = request.GET.get('search', '')
    categoria = request.GET.get('categoria', '')
    estado = request.GET.get('estado', '')
    
    # Obtener servicios
    servicios = TipoServicio.objects.all()
    
    # Aplicar filtros
    if search:
        servicios = servicios.filter(
            Q(nombre__icontains=search) |
            Q(descripcion__icontains=search)
        )
    
    if categoria:
        servicios = servicios.filter(categoria=categoria)
    
    if estado == 'activo':
        servicios = servicios.filter(activo=True)
    elif estado == 'inactivo':
        servicios = servicios.filter(activo=False)
    
    servicios = servicios.order_by('categoria', 'nombre')
    
    # Estadísticas
    total_servicios = TipoServicio.objects.count()
    servicios_activos = TipoServicio.objects.filter(activo=True).count()
    servicios_inactivos = TipoServicio.objects.filter(activo=False).count()
    servicios_por_categoria = TipoServicio.objects.values('categoria').annotate(
        total=Count('id')
    )
    
    estadisticas = {
        'total_servicios': total_servicios,
        'servicios_activos': servicios_activos,
        'servicios_inactivos': servicios_inactivos,
        'servicios_por_categoria': servicios_por_categoria,
    }
    
    context = {
        'perfil': perfil,
        'servicios': servicios,
        'estadisticas': estadisticas,
        'search': search,
        'categoria': categoria,
        'estado': estado,
        'categorias': TipoServicio.CATEGORIA_CHOICES,
    }
    
    return render(request, 'citas/servicios/gestor_servicios.html', context)

@login_required
def crear_servicio(request):
    """Vista para crear un nuevo servicio"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'Solo los administrativos pueden crear servicios.')
            return redirect('gestor_servicios')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre', '').strip()
            descripcion = request.POST.get('descripcion', '').strip()
            categoria = request.POST.get('categoria', 'otros')
            precio_base = request.POST.get('precio_base', '0')
            requiere_dentista = request.POST.get('requiere_dentista') == 'on'
            duracion_estimada = request.POST.get('duracion_estimada', '')
            activo = request.POST.get('activo') == 'on'
            
            # Validaciones
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            
            if not nombre:
                error_msg = 'El nombre del servicio es obligatorio. Por favor, ingrese un nombre para el servicio.'
                if is_ajax:
                    return JsonResponse({'success': False, 'message': error_msg}, status=400)
                messages.error(request, error_msg)
                return redirect('crear_servicio')
            
            if len(nombre) < 3:
                error_msg = 'El nombre del servicio debe tener al menos 3 caracteres.'
                if is_ajax:
                    return JsonResponse({'success': False, 'message': error_msg}, status=400)
                messages.error(request, error_msg)
                return redirect('crear_servicio')
            
            try:
                precio_base = float(precio_base)
                if precio_base < 0:
                    error_msg = 'El precio base debe ser un valor positivo. Por favor, ingrese un precio válido.'
                    if is_ajax:
                        return JsonResponse({'success': False, 'message': error_msg}, status=400)
                    messages.error(request, error_msg)
                    return redirect('crear_servicio')
                if precio_base == 0:
                    warning_msg = 'El precio base es $0. ¿Está seguro de que desea crear un servicio gratuito?'
                    if is_ajax:
                        # Permitir continuar pero con advertencia
                        pass
                    else:
                        messages.warning(request, warning_msg)
            except (ValueError, TypeError):
                error_msg = 'El precio base debe ser un número válido. Por favor, ingrese un valor numérico (ejemplo: 50000).'
                if is_ajax:
                    return JsonResponse({'success': False, 'message': error_msg}, status=400)
                messages.error(request, error_msg)
                return redirect('crear_servicio')
            
            duracion_estimada_int = None
            if duracion_estimada:
                try:
                    duracion_estimada_int = int(duracion_estimada)
                    if duracion_estimada_int < 0:
                        error_msg = 'La duración estimada no puede ser un número negativo.'
                        if is_ajax:
                            return JsonResponse({'success': False, 'message': error_msg}, status=400)
                        messages.error(request, error_msg)
                        return redirect('crear_servicio')
                    if duracion_estimada_int > 1440:
                        warning_msg = 'La duración estimada es muy alta (más de 24 horas). Verifique que el valor sea correcto.'
                        if is_ajax:
                            # Permitir continuar pero con advertencia
                            pass
                        else:
                            messages.warning(request, warning_msg)
                except (ValueError, TypeError):
                    error_msg = 'La duración estimada debe ser un número entero válido (en minutos).'
                    if is_ajax:
                        return JsonResponse({'success': False, 'message': error_msg}, status=400)
                    messages.error(request, error_msg)
                    return redirect('crear_servicio')
            
            # Verificar que no exista un servicio con el mismo nombre
            if TipoServicio.objects.filter(nombre__iexact=nombre).exists():
                error_msg = f'Ya existe un servicio con el nombre "{nombre}". Por favor, elija un nombre diferente.'
                if is_ajax:
                    return JsonResponse({'success': False, 'message': error_msg}, status=400)
                messages.error(request, error_msg)
                return redirect('crear_servicio')
            
            # Crear el servicio
            servicio = TipoServicio.objects.create(
                nombre=nombre,
                descripcion=descripcion,
                categoria=categoria,
                precio_base=precio_base,
                requiere_dentista=requiere_dentista,
                duracion_estimada=duracion_estimada_int,
                activo=activo,
                creado_por=perfil
            )
            
            # Registrar en auditoría
            registrar_auditoria(
                usuario=perfil,
                accion='crear',
                modulo='servicios',
                descripcion=f'Servicio creado: {nombre}',
                detalles=f'Categoría: {servicio.get_categoria_display()}, Precio: ${precio_base:,.0f}, Activo: {"Sí" if activo else "No"}',
                objeto_id=servicio.id,
                tipo_objeto='TipoServicio',
                request=request
            )
            
            # Si es una petición AJAX, devolver JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'✅ Servicio "{nombre}" creado correctamente con un precio base de ${precio_base:,.0f}.',
                    'servicio': {
                        'id': servicio.id,
                        'nombre': servicio.nombre,
                        'categoria': servicio.get_categoria_display(),
                        'precio_base': float(servicio.precio_base),
                        'activo': servicio.activo,
                    }
                })
            
            messages.success(request, f'✅ Servicio "{nombre}" creado correctamente con un precio base de ${precio_base:,.0f}.')
            return redirect('gestor_servicios')
            
        except Exception as e:
            error_msg = f'❌ Error inesperado al crear el servicio. Por favor, intente nuevamente. Si el problema persiste, contacte al administrador del sistema. Detalles: {str(e)}'
            
            # Si es una petición AJAX, devolver JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': error_msg
                }, status=400)
            
            messages.error(request, error_msg)
            return redirect('crear_servicio')
    
    # Si es GET y AJAX, devolver las categorías
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        categorias = [{'code': code, 'name': name} for code, name in TipoServicio.CATEGORIA_CHOICES]
        return JsonResponse({
            'categorias': categorias
        })
    
    context = {
        'perfil': perfil,
        'categorias': TipoServicio.CATEGORIA_CHOICES,
    }
    
    return render(request, 'citas/servicios/crear_servicio.html', context)

@login_required
def editar_servicio(request, servicio_id):
    """Vista para editar un servicio existente"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'Solo los administrativos pueden editar servicios.')
            return redirect('gestor_servicios')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    servicio = get_object_or_404(TipoServicio, id=servicio_id)
    
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre', '').strip()
            descripcion = request.POST.get('descripcion', '').strip()
            categoria = request.POST.get('categoria', 'otros')
            precio_base = request.POST.get('precio_base', '0')
            requiere_dentista = request.POST.get('requiere_dentista') == 'on'
            duracion_estimada = request.POST.get('duracion_estimada', '')
            activo = request.POST.get('activo') == 'on'
            
            # Validaciones
            if not nombre:
                messages.error(request, 'El nombre del servicio es obligatorio. Por favor, ingrese un nombre para el servicio.')
                return redirect('editar_servicio', servicio_id=servicio_id)
            
            if len(nombre) < 3:
                messages.error(request, 'El nombre del servicio debe tener al menos 3 caracteres.')
                return redirect('editar_servicio', servicio_id=servicio_id)
            
            try:
                precio_base = float(precio_base)
                if precio_base < 0:
                    messages.error(request, 'El precio base debe ser un valor positivo. Por favor, ingrese un precio válido.')
                    return redirect('editar_servicio', servicio_id=servicio_id)
                if precio_base == 0:
                    messages.warning(request, 'El precio base es $0. ¿Está seguro de que desea que este servicio sea gratuito?')
            except (ValueError, TypeError):
                messages.error(request, 'El precio base debe ser un número válido. Por favor, ingrese un valor numérico (ejemplo: 50000).')
                return redirect('editar_servicio', servicio_id=servicio_id)
            
            duracion_estimada_int = None
            if duracion_estimada:
                try:
                    duracion_estimada_int = int(duracion_estimada)
                    if duracion_estimada_int < 0:
                        messages.error(request, 'La duración estimada no puede ser un número negativo.')
                        return redirect('editar_servicio', servicio_id=servicio_id)
                    if duracion_estimada_int > 1440:
                        messages.warning(request, 'La duración estimada es muy alta (más de 24 horas). Verifique que el valor sea correcto.')
                except (ValueError, TypeError):
                    messages.error(request, 'La duración estimada debe ser un número entero válido (en minutos).')
                    return redirect('editar_servicio', servicio_id=servicio_id)
            
            # Verificar que no exista otro servicio con el mismo nombre (excepto el actual)
            if TipoServicio.objects.filter(nombre__iexact=nombre).exclude(id=servicio_id).exists():
                messages.error(request, f'Ya existe otro servicio con el nombre "{nombre}". Por favor, elija un nombre diferente.')
                return redirect('editar_servicio', servicio_id=servicio_id)
            
            # Guardar valores anteriores para auditoría
            nombre_anterior = servicio.nombre
            precio_anterior = servicio.precio_base
            
            # Actualizar el servicio
            servicio.nombre = nombre
            servicio.descripcion = descripcion
            servicio.categoria = categoria
            servicio.precio_base = precio_base
            servicio.requiere_dentista = requiere_dentista
            servicio.duracion_estimada = duracion_estimada_int
            servicio.activo = activo
            servicio.save()
            
            # Registrar en auditoría
            detalles_cambio = []
            if nombre_anterior != nombre:
                detalles_cambio.append(f'Nombre: {nombre_anterior} → {nombre}')
            if precio_anterior != precio_base:
                detalles_cambio.append(f'Precio: ${precio_anterior:,.0f} → ${precio_base:,.0f}')
            if not detalles_cambio:
                detalles_cambio.append('Información actualizada')
            
            registrar_auditoria(
                usuario=perfil,
                accion='editar',
                modulo='servicios',
                descripcion=f'Servicio editado: {nombre}',
                detalles='; '.join(detalles_cambio),
                objeto_id=servicio.id,
                tipo_objeto='TipoServicio',
                request=request
            )
            
            messages.success(request, f'✅ Servicio "{nombre}" actualizado correctamente.')
            return redirect('gestor_servicios')
            
        except Exception as e:
            messages.error(request, f'❌ Error inesperado al actualizar el servicio. Por favor, intente nuevamente. Si el problema persiste, contacte al administrador del sistema. Detalles: {str(e)}')
            return redirect('editar_servicio', servicio_id=servicio_id)
    
    context = {
        'perfil': perfil,
        'servicio': servicio,
        'categorias': TipoServicio.CATEGORIA_CHOICES,
    }
    
    return render(request, 'citas/servicios/editar_servicio.html', context)

@login_required
def eliminar_servicio(request, servicio_id):
    """Vista para eliminar un servicio"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'Solo los administrativos pueden eliminar servicios.')
            return redirect('gestor_servicios')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    servicio = get_object_or_404(TipoServicio, id=servicio_id)
    
    if request.method == 'POST':
        nombre_servicio = servicio.nombre
        
        # Verificar si hay citas asociadas a este servicio
        citas_asociadas = Cita.objects.filter(tipo_servicio=servicio).count()
        if citas_asociadas > 0:
            messages.warning(
                request, 
                f'⚠️ No se puede eliminar el servicio "{nombre_servicio}" porque tiene {citas_asociadas} cita(s) asociada(s). '
                f'Para desactivar el servicio, edítelo y desmarque la opción "Servicio Activo".'
            )
            return redirect('gestor_servicios')
        
        try:
            # Guardar información para auditoría antes de eliminar
            servicio_id = servicio.id
            servicio_categoria = servicio.get_categoria_display()
            servicio_precio = servicio.precio_base
            
            # Registrar en auditoría ANTES de eliminar
            registrar_auditoria(
                usuario=perfil,
                accion='eliminar',
                modulo='servicios',
                descripcion=f'Servicio eliminado: {nombre_servicio}',
                detalles=f'Categoría: {servicio_categoria}, Precio: ${servicio_precio:,.0f}',
                objeto_id=servicio_id,
                tipo_objeto='TipoServicio',
                request=request
            )
            
            servicio.delete()
            messages.success(request, f'✅ Servicio "{nombre_servicio}" eliminado correctamente del sistema.')
        except Exception as e:
            messages.error(request, f'❌ Error al eliminar el servicio. Por favor, intente nuevamente. Si el problema persiste, contacte al administrador del sistema. Detalles: {str(e)}')
        
        return redirect('gestor_servicios')
    
    context = {
        'perfil': perfil,
        'servicio': servicio,
    }
    
    return render(request, 'citas/servicios/eliminar_servicio.html', context)

# ==========================================
# GESTIÓN DE HORARIOS DE DENTISTAS
# ==========================================

@login_required
def gestor_horarios(request):
    """Vista para que el administrador gestione horarios de dentistas"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'Solo los administrativos pueden gestionar horarios.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    # Obtener todos los dentistas
    dentistas = Perfil.objects.filter(rol='dentista', activo=True).order_by('nombre_completo')
    
    # Obtener horarios agrupados por dentista
    horarios_por_dentista = {}
    for dentista in dentistas:
        horarios = HorarioDentista.objects.filter(dentista=dentista, activo=True).order_by('dia_semana', 'hora_inicio')
        horarios_por_dentista[dentista.id] = horarios
    
    context = {
        'perfil': perfil,
        'dentistas': dentistas,
        'horarios_por_dentista': horarios_por_dentista,
        'dias_semana': HorarioDentista.DIA_SEMANA_CHOICES,
    }
    return render(request, 'citas/horarios/gestor_horarios.html', context)

@login_required
def gestionar_horario_dentista(request, dentista_id):
    """Vista para que el administrador gestione el horario de un dentista específico"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'Solo los administrativos pueden gestionar horarios.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    dentista = get_object_or_404(Perfil, id=dentista_id, rol='dentista')
    
    # Obtener horarios del dentista agrupados por día
    horarios = HorarioDentista.objects.filter(dentista=dentista).order_by('dia_semana', 'hora_inicio')
    horarios_por_dia = {}
    for dia_num, dia_nombre in HorarioDentista.DIA_SEMANA_CHOICES:
        horarios_por_dia[dia_num] = horarios.filter(dia_semana=dia_num)
    
    context = {
        'perfil': perfil,
        'dentista': dentista,
        'horarios_por_dia': horarios_por_dia,
        'dias_semana': HorarioDentista.DIA_SEMANA_CHOICES,
    }
    return render(request, 'citas/horarios/gestionar_horario_dentista.html', context)

# Vista AJAX para agregar horario
@login_required
def agregar_horario_ajax(request, dentista_id):
    """Vista AJAX para agregar un nuevo horario con validaciones"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        dentista = Perfil.objects.get(id=dentista_id, rol='dentista')
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Dentista no encontrado'}, status=404)
    
    dia_semana = request.POST.get('dia_semana')
    hora_inicio = request.POST.get('hora_inicio')
    hora_fin = request.POST.get('hora_fin')
    
    # Validaciones básicas
    if not dia_semana or not hora_inicio or not hora_fin:
        return JsonResponse({'success': False, 'error': 'Todos los campos son obligatorios'}, status=400)
    
    try:
        dia_semana = int(dia_semana)
        if dia_semana < 0 or dia_semana > 6:
            return JsonResponse({'success': False, 'error': 'Día de la semana inválido'}, status=400)
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Día de la semana inválido'}, status=400)
    
    # Convertir strings a time
    try:
        from datetime import datetime
        hora_inicio_obj = datetime.strptime(hora_inicio, '%H:%M').time()
        hora_fin_obj = datetime.strptime(hora_fin, '%H:%M').time()
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Formato de hora inválido'}, status=400)
    
    # Validar que hora_fin > hora_inicio
    if hora_fin_obj <= hora_inicio_obj:
        return JsonResponse({'success': False, 'error': 'La hora de fin debe ser mayor que la hora de inicio'}, status=400)
    
    # Validar solapamiento con otros horarios del mismo día
    horarios_existentes = HorarioDentista.objects.filter(
        dentista=dentista,
        dia_semana=dia_semana,
        activo=True
    )
    
    for horario_existente in horarios_existentes:
        # Verificar si hay solapamiento
        if not (hora_fin_obj <= horario_existente.hora_inicio or hora_inicio_obj >= horario_existente.hora_fin):
            return JsonResponse({
                'success': False, 
                'error': f'El horario se solapa con otro existente: {horario_existente.hora_inicio.strftime("%H:%M")}-{horario_existente.hora_fin.strftime("%H:%M")}'
            }, status=400)
    
    # Crear el horario
    try:
        horario = HorarioDentista.objects.create(
            dentista=dentista,
            dia_semana=dia_semana,
            hora_inicio=hora_inicio_obj,
            hora_fin=hora_fin_obj,
            activo=True
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Horario agregado: {horario.get_dia_semana_display()} {hora_inicio}-{hora_fin}',
            'horario': {
                'id': horario.id,
                'dia_semana': horario.dia_semana,
                'dia_nombre': horario.get_dia_semana_display(),
                'hora_inicio': horario.hora_inicio.strftime('%H:%M'),
                'hora_fin': horario.hora_fin.strftime('%H:%M'),
                'activo': horario.activo
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al crear horario: {str(e)}'}, status=500)

# Vista AJAX para editar horario
@login_required
def editar_horario_ajax(request, horario_id):
    """Vista AJAX para editar un horario existente"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        horario = HorarioDentista.objects.get(id=horario_id)
    except HorarioDentista.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Horario no encontrado'}, status=404)
    
    dia_semana = request.POST.get('dia_semana')
    hora_inicio = request.POST.get('hora_inicio')
    hora_fin = request.POST.get('hora_fin')
    
    # Validaciones básicas
    if not dia_semana or not hora_inicio or not hora_fin:
        return JsonResponse({'success': False, 'error': 'Todos los campos son obligatorios'}, status=400)
    
    try:
        dia_semana = int(dia_semana)
        if dia_semana < 0 or dia_semana > 6:
            return JsonResponse({'success': False, 'error': 'Día de la semana inválido'}, status=400)
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Día de la semana inválido'}, status=400)
    
    # Convertir strings a time
    try:
        from datetime import datetime
        hora_inicio_obj = datetime.strptime(hora_inicio, '%H:%M').time()
        hora_fin_obj = datetime.strptime(hora_fin, '%H:%M').time()
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Formato de hora inválido'}, status=400)
    
    # Validar que hora_fin > hora_inicio
    if hora_fin_obj <= hora_inicio_obj:
        return JsonResponse({'success': False, 'error': 'La hora de fin debe ser mayor que la hora de inicio'}, status=400)
    
    # Validar solapamiento con otros horarios del mismo día (excluyendo el actual)
    horarios_existentes = HorarioDentista.objects.filter(
        dentista=horario.dentista,
        dia_semana=dia_semana,
        activo=True
    ).exclude(id=horario_id)
    
    for horario_existente in horarios_existentes:
        # Verificar si hay solapamiento
        if not (hora_fin_obj <= horario_existente.hora_inicio or hora_inicio_obj >= horario_existente.hora_fin):
            return JsonResponse({
                'success': False, 
                'error': f'El horario se solapa con otro existente: {horario_existente.hora_inicio.strftime("%H:%M")}-{horario_existente.hora_fin.strftime("%H:%M")}'
            }, status=400)
    
    # Actualizar el horario
    try:
        horario.dia_semana = dia_semana
        horario.hora_inicio = hora_inicio_obj
        horario.hora_fin = hora_fin_obj
        horario.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Horario actualizado: {horario.get_dia_semana_display()} {hora_inicio}-{hora_fin}',
            'horario': {
                'id': horario.id,
                'dia_semana': horario.dia_semana,
                'dia_nombre': horario.get_dia_semana_display(),
                'hora_inicio': horario.hora_inicio.strftime('%H:%M'),
                'hora_fin': horario.hora_fin.strftime('%H:%M'),
                'activo': horario.activo
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al actualizar horario: {str(e)}'}, status=500)

# Vista AJAX para eliminar horarios
@login_required
def eliminar_horarios_ajax(request, dentista_id):
    """Vista AJAX para eliminar horarios"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        dentista = Perfil.objects.get(id=dentista_id, rol='dentista')
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Dentista no encontrado'}, status=404)
    
    import json
    data = json.loads(request.body)
    horarios_ids = data.get('horarios_ids', [])
    
    if not horarios_ids:
        return JsonResponse({'success': False, 'error': 'No se seleccionaron horarios para eliminar'}, status=400)
    
    try:
        horarios_eliminados = HorarioDentista.objects.filter(id__in=horarios_ids, dentista=dentista)
        cantidad = horarios_eliminados.count()
        horarios_eliminados.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'{cantidad} horario(s) eliminado(s) correctamente'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al eliminar horarios: {str(e)}'}, status=500)

@login_required
def ver_mi_horario(request):
    """Vista para que el dentista vea su horario (solo lectura)"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_dentista():
            messages.error(request, 'Solo los dentistas pueden ver su horario.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    # Obtener horarios del dentista agrupados por día
    horarios = HorarioDentista.objects.filter(dentista=perfil, activo=True).order_by('dia_semana', 'hora_inicio')
    horarios_por_dia = {}
    for dia_num, dia_nombre in HorarioDentista.DIA_SEMANA_CHOICES:
        horarios_por_dia[dia_num] = horarios.filter(dia_semana=dia_num)
    
    context = {
        'perfil': perfil,
        'horarios_por_dia': horarios_por_dia,
        'dias_semana': HorarioDentista.DIA_SEMANA_CHOICES,
    }
    return render(request, 'citas/horarios/ver_mi_horario.html', context)


# ==========================================
# GESTIÓN DE PLANES DE TRATAMIENTO
# ==========================================

@login_required
def listar_planes_tratamiento(request):
    """Lista todos los planes de tratamiento según el rol del usuario"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    # FILTRO SEGÚN ROL
    if perfil.es_administrativo():
        # ADMINISTRADOR: Ve TODOS los planes
        planes = PlanTratamiento.objects.all()
        
        # Filtros adicionales para admin
        dentista_filtro = request.GET.get('dentista_id', '')
        if dentista_filtro:
            planes = planes.filter(dentista_id=dentista_filtro)
        
        # Estadísticas globales
        from django.db.models import Sum
        estadisticas = {
            'total_planes': PlanTratamiento.objects.count(),
            'planes_activos': PlanTratamiento.objects.filter(estado='en_progreso').count(),
            'planes_completados': PlanTratamiento.objects.filter(estado='completado').count(),
            'ingresos_estimados': PlanTratamiento.objects.filter(
                estado__in=['aprobado', 'en_progreso']
            ).aggregate(Sum('precio_final'))['precio_final__sum'] or 0,
        }
        
        # Obtener dentistas para el filtro
        dentistas = Perfil.objects.filter(rol='dentista', activo=True).order_by('nombre_completo')
        
    elif perfil.es_dentista():
        # DENTISTA: Solo ve planes de SUS clientes vinculados
        # Obtener clientes vinculados al dentista
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        
        # Filtrar planes por dentista Y por clientes vinculados
        planes = PlanTratamiento.objects.filter(
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
        
        # Estadísticas solo de sus planes
        estadisticas = {
            'total_planes': planes.count(),
            'planes_activos': planes.filter(estado='en_progreso').count(),
            'planes_completados': planes.filter(estado='completado').count(),
        }
        
        dentistas = None
        dentista_filtro = None
    else:
        messages.error(request, 'No tienes permisos para ver planes de tratamiento.')
        return redirect('panel_trabajador')
    
    # Filtros comunes
    search = request.GET.get('search', '')
    estado_filtro = request.GET.get('estado', '')
    
    if search:
        planes = planes.filter(
            Q(nombre__icontains=search) |
            Q(cliente__nombre_completo__icontains=search) |
            Q(cliente__email__icontains=search) |
            Q(cliente__rut__icontains=search)
        )
    
    if estado_filtro:
        planes = planes.filter(estado=estado_filtro)
    
    # Ordenamiento
    orden = request.GET.get('orden', '-creado_el')
    ordenamientos_validos = {
        '-creado_el': '-creado_el',
        'creado_el': 'creado_el',
        '-precio_final': '-precio_final',
        'precio_final': 'precio_final',
        '-progreso_porcentaje': '-progreso_porcentaje',
        'progreso_porcentaje': 'progreso_porcentaje',
        'nombre': 'nombre',
        '-nombre': '-nombre',
    }
    orden_seleccionado = ordenamientos_validos.get(orden, '-creado_el')
    
    planes = planes.select_related('cliente', 'dentista', 'odontograma_inicial').prefetch_related('citas', 'consentimientos').order_by(orden_seleccionado)
    
    # Paginación
    paginator = Paginator(planes, 20)
    page = request.GET.get('page')
    try:
        planes_paginados = paginator.page(page)
    except PageNotAnInteger:
        planes_paginados = paginator.page(1)
    except EmptyPage:
        planes_paginados = paginator.page(paginator.num_pages)
    
    # Calcular última y próxima cita para cada plan
    from django.utils import timezone
    for plan in planes_paginados:
        citas_ordenadas = plan.citas.all().order_by('fecha_hora')
        if citas_ordenadas.exists():
            # Última cita (la más reciente)
            plan.ultima_cita_obj = citas_ordenadas.last()
            # Próxima cita (la primera futura con estado reservada o confirmada)
            proximas = citas_ordenadas.filter(
                fecha_hora__gte=timezone.now(),
                estado__in=['reservada', 'confirmada']
            ).first()
            plan.proxima_cita_obj = proximas
        else:
            plan.ultima_cita_obj = None
            plan.proxima_cita_obj = None
        
        # Verificar si hay consentimientos pendientes
        plan.tiene_consentimientos_pendientes = plan.consentimientos.exclude(estado='firmado').exists()
    
    context = {
        'perfil': perfil,
        'planes': planes_paginados,
        'estadisticas': estadisticas,
        'es_admin': perfil.es_administrativo(),
        'es_dentista': perfil.es_dentista(),
        'search': search,
        'estado_filtro': estado_filtro,
        'dentistas': dentistas,
        'dentista_filtro': dentista_filtro,
        'orden': orden,
    }
    
    # Usar template diferente para dentistas
    if perfil.es_dentista():
        return render(request, 'citas/planes_tratamiento/mis_tratamientos_dentista.html', context)
    else:
        return render(request, 'citas/planes_tratamiento/listar_planes_tratamiento.html', context)


@login_required
def crear_plan_tratamiento(request):
    """Crea un nuevo plan de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    if not (perfil.es_dentista() or perfil.es_administrativo()):
        messages.error(request, 'Solo dentistas y administrativos pueden crear planes de tratamiento.')
        return redirect('panel_trabajador')
    
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente_id')
        dentista_id = request.POST.get('dentista_id')
        
        # VERIFICACIONES DIFERENTES
        if perfil.es_dentista():
            # DENTISTA: Restricciones estrictas
            pacientes_dentista = perfil.get_pacientes_asignados()
            clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
            
            # 1. Solo puede crear para SUS clientes
            if int(cliente_id) not in clientes_ids:
                messages.error(request, 'Solo puedes crear planes para tus pacientes.')
                return redirect('crear_plan_tratamiento')
            
            # 2. Solo puede asignarse a sí mismo
            if int(dentista_id) != perfil.id:
                messages.error(request, 'Solo puedes crear planes como dentista asignado.')
                return redirect('crear_plan_tratamiento')
        
        # Validar datos
        nombre = request.POST.get('nombre', '').strip()
        if not nombre:
            messages.error(request, 'El nombre del plan es requerido.')
            return redirect('crear_plan_tratamiento')
        
        try:
            cliente = Cliente.objects.get(id=cliente_id, activo=True)
            dentista = Perfil.objects.get(id=dentista_id, rol='dentista', activo=True)
        except (Cliente.DoesNotExist, Perfil.DoesNotExist):
            messages.error(request, 'Cliente o dentista no válido.')
            return redirect('crear_plan_tratamiento')
        
        # Obtener odontograma inicial si se proporciona
        odontograma_id = request.POST.get('odontograma_inicial_id', '')
        odontograma_inicial = None
        if odontograma_id:
            try:
                odontograma_inicial = Odontograma.objects.get(id=odontograma_id)
            except Odontograma.DoesNotExist:
                pass
        
        # Crear el plan
        try:
            # Obtener valores de presupuesto (usar campos _raw si existen, sino procesar el formateado)
            presupuesto_str = request.POST.get('presupuesto_total_raw') or request.POST.get('presupuesto_total', '0')
            descuento_str = request.POST.get('descuento_raw') or request.POST.get('descuento', '0')
            
            # Remover caracteres no numéricos (por si acaso viene formateado)
            presupuesto_str = re.sub(r'[^\d]', '', presupuesto_str) or '0'
            descuento_str = re.sub(r'[^\d]', '', descuento_str) or '0'
            
            presupuesto_total = float(presupuesto_str)
            descuento = float(descuento_str)
            precio_final = presupuesto_total - descuento
            
            citas_estimadas = request.POST.get('citas_estimadas', '').strip()
            citas_estimadas = int(citas_estimadas) if citas_estimadas and citas_estimadas.isdigit() else None
            
            plan = PlanTratamiento.objects.create(
                cliente=cliente,
                dentista=dentista,
                odontograma_inicial=odontograma_inicial,
                nombre=nombre,
                descripcion=request.POST.get('descripcion', ''),
                diagnostico=request.POST.get('diagnostico', ''),
                objetivo=request.POST.get('objetivo', ''),
                presupuesto_total=presupuesto_total,
                descuento=descuento,
                precio_final=precio_final,
                estado=request.POST.get('estado', 'borrador'),
                fecha_inicio_estimada=request.POST.get('fecha_inicio_estimada') or None,
                fecha_fin_estimada=request.POST.get('fecha_fin_estimada') or None,
                citas_estimadas=citas_estimadas,
                notas_internas=request.POST.get('notas_internas', '') if perfil.es_administrativo() else '',
                notas_paciente=request.POST.get('notas_paciente', ''),
                creado_por=perfil,
            )
            
            # Registrar en auditoría
            registrar_auditoria(
                usuario=perfil,
                accion='crear',
                modulo='planes_tratamiento',
                descripcion=f'Plan de tratamiento creado: {nombre}',
                detalles=f'Cliente: {cliente.nombre_completo}, Dentista: {dentista.nombre_completo}, Presupuesto: ${presupuesto_total:,.0f}, Estado: {plan.get_estado_display()}',
                objeto_id=plan.id,
                tipo_objeto='PlanTratamiento',
                request=request
            )
            
            # Procesar fases si se enviaron
            fases_data = request.POST.getlist('fases[]')
            if fases_data:
                try:
                    fases_json = json.loads(fases_data[0]) if fases_data else []
                    for fase_data in fases_json:
                        fase = FaseTratamiento.objects.create(
                            plan=plan,
                            nombre=fase_data.get('nombre', ''),
                            descripcion=fase_data.get('descripcion', ''),
                            orden=fase_data.get('orden', 1),
                            presupuesto=float(fase_data.get('presupuesto', 0)),
                        )
                        
                        # Procesar items de la fase
                        items_data = fase_data.get('items', [])
                        for item_data in items_data:
                            servicio_id = item_data.get('servicio_id')
                            servicio = None
                            if servicio_id:
                                try:
                                    servicio = TipoServicio.objects.get(id=servicio_id)
                                except TipoServicio.DoesNotExist:
                                    pass
                            
                            ItemTratamiento.objects.create(
                                fase=fase,
                                servicio=servicio,
                                descripcion=item_data.get('descripcion', ''),
                                cantidad=int(item_data.get('cantidad', 1)),
                                precio_unitario=float(item_data.get('precio_unitario', 0)),
                            )
                except (json.JSONDecodeError, ValueError) as e:
                    logger.error(f"Error al procesar fases: {e}")
            
            messages.success(request, f'Plan de tratamiento "{plan.nombre}" creado exitosamente.')
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
            
        except ValueError as e:
            messages.error(request, f'Error en los datos numéricos: {str(e)}')
            return redirect('crear_plan_tratamiento')
        except Exception as e:
            messages.error(request, f'Error al crear el plan: {str(e)}')
            logger.error(f"Error al crear plan de tratamiento: {e}")
            return redirect('crear_plan_tratamiento')
    
    # GET: Mostrar formulario
    if perfil.es_administrativo():
        # ADMINISTRADOR: Ve todos los clientes y dentistas
        clientes = Cliente.objects.filter(activo=True).order_by('nombre_completo')
        dentistas = Perfil.objects.filter(rol='dentista', activo=True).order_by('nombre_completo')
        mostrar_selector_dentista = True
    else:
        # DENTISTA: Solo ve SUS clientes y solo puede ser él mismo
        clientes = obtener_clientes_permitidos(perfil)
        dentistas = [perfil]
        mostrar_selector_dentista = False
    
    # Obtener cliente pre-seleccionado si viene de otra vista
    cliente_pre_seleccionado = request.GET.get('cliente_id')
    
    # Obtener odontogramas recientes
    if perfil.es_dentista():
        odontogramas_recientes = Odontograma.objects.filter(
            dentista=perfil
        ).select_related('cliente').order_by('-fecha_creacion')[:10]
    else:
        odontogramas_recientes = Odontograma.objects.all().select_related('cliente').order_by('-fecha_creacion')[:10]
    
    # Obtener servicios disponibles
    servicios = TipoServicio.objects.filter(activo=True).order_by('categoria', 'nombre')
    
    context = {
        'perfil': perfil,
        'clientes': clientes,
        'dentistas': dentistas,
        'mostrar_selector_dentista': mostrar_selector_dentista,
        'odontogramas_recientes': odontogramas_recientes,
        'cliente_pre_seleccionado': cliente_pre_seleccionado,
        'servicios': servicios,
        'es_admin': perfil.es_administrativo(),
        'es_dentista': perfil.es_dentista(),
    }
    
    return render(request, 'citas/planes_tratamiento/crear_plan_tratamiento.html', context)


@login_required
def detalle_plan_tratamiento(request, plan_id):
    """Muestra el detalle de un plan de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('login')
    
    # FILTRO SEGÚN ROL
    if perfil.es_administrativo():
        # ADMINISTRADOR: Puede ver cualquier plan
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        # DENTISTA: Solo puede ver planes de SUS clientes
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        messages.error(request, 'No tienes permisos para ver planes de tratamiento.')
        return redirect('panel_trabajador')
    
    # Obtener citas vinculadas al plan
    citas = plan.citas.all().order_by('fecha_hora').select_related('tipo_servicio')
    
    # Calcular total de precios de citas
    from decimal import Decimal
    total_precios_citas = Decimal('0.00')
    for cita in citas:
        if cita.precio_cobrado:
            total_precios_citas += Decimal(str(cita.precio_cobrado))
    
    # Calcular saldo disponible (precio_final - total_precios_citas)
    saldo_disponible_citas = plan.precio_final - total_precios_citas
    
    # Contar total de citas
    total_citas = citas.count()
    
    # Obtener servicios activos para los modales
    from citas.models import TipoServicio
    servicios_activos = TipoServicio.objects.filter(activo=True).order_by('nombre')
    
    # Obtener consentimientos informados asociados al plan
    consentimientos = plan.consentimientos.all().order_by('-fecha_creacion')
    
    # Obtener plantillas de consentimiento para el botón de crear
    plantillas_consentimiento = PlantillaConsentimiento.objects.filter(activo=True).order_by('tipo_procedimiento', 'nombre')
    
    context = {
        'perfil': perfil,
        'plan': plan,
        'citas': citas,
        'total_precios_citas': total_precios_citas,
        'saldo_disponible_citas': saldo_disponible_citas,
        'total_citas': total_citas,
        'servicios_activos': servicios_activos,
        'consentimientos': consentimientos,
        'plantillas_consentimiento': plantillas_consentimiento,
        'es_admin': perfil.es_administrativo(),
        'es_dentista': perfil.es_dentista(),
        'puede_editar': plan.puede_ser_editado_por(perfil),
        'puede_eliminar': plan.puede_ser_eliminado_por(perfil),
        'puede_cancelar': plan.puede_ser_cancelado_por(perfil),
    }
    
    # Verificar si puede aprobar el tratamiento
    tiene_consentimiento_firmado = plan.consentimientos.filter(estado='firmado').exists()
    presupuesto_aceptado = plan.presupuesto_aceptado
    
    puede_aprobar = False
    if perfil.es_administrativo() or (perfil.es_dentista() and plan.dentista == perfil):
        # Verificar condiciones para aprobar
        puede_aprobar = tiene_consentimiento_firmado and presupuesto_aceptado and plan.estado in ['borrador', 'pendiente_aprobacion']
    
    # Verificar si puede finalizar el tratamiento
    puede_finalizar = False
    if perfil.es_administrativo() or (perfil.es_dentista() and plan.dentista == perfil):
        # Solo se puede finalizar si está aprobado o en progreso
        puede_finalizar = plan.estado in ['aprobado', 'en_progreso']
    
    # Verificar si puede crear citas (solo administrativos pueden crear citas)
    # Los dentistas NO pueden crear citas, eso lo maneja la recepcionista
    puede_crear_citas = False
    if perfil.es_administrativo():
        puede_crear_citas = tiene_consentimiento_firmado and presupuesto_aceptado and plan.estado in ['aprobado', 'en_progreso']
    
    # Calcular total de precios de citas (para validar que no exceda el monto total)
    from decimal import Decimal
    total_precios_citas = Decimal('0.00')
    for cita in citas:
        if cita.precio_cobrado:
            total_precios_citas += Decimal(str(cita.precio_cobrado))
    
    # Calcular saldo disponible para nuevas citas
    saldo_disponible_citas = plan.precio_final - total_precios_citas
    
    context.update({
        'puede_aprobar': puede_aprobar,
        'puede_finalizar': puede_finalizar,
        'puede_crear_citas': puede_crear_citas,
        'tiene_consentimiento_firmado': tiene_consentimiento_firmado,
        'presupuesto_aceptado': presupuesto_aceptado,
        'total_precios_citas': total_precios_citas,
        'saldo_disponible_citas': saldo_disponible_citas,
    })
    
    return render(request, 'citas/planes_tratamiento/detalle_plan_tratamiento.html', context)


@login_required
def aceptar_presupuesto_tratamiento(request, plan_id):
    """Marca el presupuesto de un tratamiento como aceptado"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción.'}, status=403)
    
    # Verificar permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción.'}, status=403)
    
    # Marcar presupuesto como aceptado
    plan.presupuesto_aceptado = True
    plan.fecha_aceptacion_presupuesto = timezone.now()
    plan.save()
    
    return JsonResponse({
        'success': True,
        'message': 'Presupuesto marcado como aceptado correctamente.'
    })


@login_required
def aprobar_tratamiento(request, plan_id):
    """Aprueba un tratamiento validando que tenga consentimiento firmado y presupuesto aceptado"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción.'}, status=403)
    
    # Verificar permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción.'}, status=403)
    
    # Validar que el estado permita la aprobación
    if plan.estado not in ['borrador', 'pendiente_aprobacion']:
        return JsonResponse({
                'success': False,
                'error': f'No se puede aprobar un tratamiento en estado "{plan.get_estado_display()}".'
            }, status=400)
    
    # Validar que tenga consentimiento informado firmado
    consentimientos_firmados = plan.consentimientos.filter(estado='firmado')
    if not consentimientos_firmados.exists():
        return JsonResponse({
                'success': False,
                'error': 'No se puede aprobar el tratamiento. Debe tener al menos un consentimiento informado firmado.'
            }, status=400)
    
    # Validar que el presupuesto esté aceptado
    if not plan.presupuesto_aceptado:
        return JsonResponse({
                'success': False,
                'error': 'No se puede aprobar el tratamiento. El presupuesto debe estar aceptado.'
            }, status=400)
    
    # Aprobar el tratamiento
    plan.estado = 'aprobado'
    plan.fecha_aprobacion = timezone.now()
    plan.save()
    
    return JsonResponse({
        'success': True,
        'message': 'Tratamiento aprobado correctamente. El tratamiento puede comenzar.'
    })


@login_required
def finalizar_tratamiento(request, plan_id):
    """Marca un tratamiento como finalizado"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción.'}, status=403)
    
    # Verificar permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción.'}, status=403)
    
    # Validar que el estado permita la finalización
    if plan.estado not in ['aprobado', 'en_progreso']:
        return JsonResponse({
                'success': False,
                'error': f'No se puede finalizar un tratamiento en estado "{plan.get_estado_display()}". Solo se pueden finalizar tratamientos aprobados o en progreso.'
            }, status=400)
    
    # Validar que no esté ya completado
    if plan.estado == 'completado':
        return JsonResponse({
                'success': False,
                'error': 'Este tratamiento ya está finalizado.'
            }, status=400)
    
    # Obtener motivo de finalización si se proporciona
    motivo_finalizacion = request.POST.get('motivo_finalizacion', '').strip()
    
    # Finalizar el tratamiento
    plan.estado = 'completado'
    plan.fecha_completado = timezone.now()
    if motivo_finalizacion:
        plan.notas_internas = (plan.notas_internas or '') + f'\n\n[Finalizado el {timezone.now().strftime("%d/%m/%Y %H:%M")}] Motivo: {motivo_finalizacion}'
    plan.save()
    
    return JsonResponse({
        'success': True,
        'message': 'Tratamiento finalizado correctamente.'
    })


@login_required
def crear_cita_desde_plan(request, plan_id):
    """Crea una cita asociada a un plan de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Validar que el plan esté en un estado que permita crear citas
    # Solo se pueden crear citas si el plan está aprobado o en progreso
    # Y si tiene consentimiento firmado y presupuesto aceptado
    if request.method == 'POST':
        # Verificar que el plan tenga consentimiento firmado y presupuesto aceptado
        tiene_consentimiento_firmado = plan.consentimientos.filter(estado='firmado').exists()
        presupuesto_aceptado = plan.presupuesto_aceptado
        
        if not tiene_consentimiento_firmado:
            return JsonResponse({
                'success': False, 
                'error': 'No se pueden crear citas. El tratamiento debe tener al menos un consentimiento informado firmado.'
            }, status=400)
        
        if not presupuesto_aceptado:
            return JsonResponse({
                'success': False, 
                'error': 'No se pueden crear citas. El presupuesto del tratamiento debe estar aceptado.'
            }, status=400)
        
        # Verificar que el plan esté en un estado válido para crear citas
        if plan.estado not in ['aprobado', 'en_progreso']:
            return JsonResponse({
                'success': False, 
                'error': f'No se pueden crear citas para un tratamiento en estado "{plan.get_estado_display()}". El tratamiento debe estar aprobado o en progreso.'
            }, status=400)
        
        # Si el plan está aprobado pero no en progreso, cambiarlo automáticamente a en_progreso
        if plan.estado == 'aprobado':
            plan.estado = 'en_progreso'
            plan.save()
    
    if request.method == 'POST':
        from django.utils import timezone
        from datetime import timedelta
        
        fecha_hora_str = request.POST.get('fecha_hora', '')
        tipo_servicio_id = request.POST.get('tipo_servicio', '').strip()
        notas = request.POST.get('notas', '').strip()
        precio_cobrado_str = request.POST.get('precio_cobrado', '').strip()
        
        errores = []
        
        if not fecha_hora_str:
            errores.append('Debe seleccionar una fecha y hora.')
        
        # Validar precio
        precio_cobrado = None
        if precio_cobrado_str:
            try:
                precio_cobrado = Decimal(str(precio_cobrado_str))
                if precio_cobrado < 0:
                    errores.append('El precio no puede ser negativo.')
            except (ValueError, InvalidOperation):
                errores.append('El precio ingresado no es válido.')
        else:
            errores.append('Debe ingresar un precio para la cita.')
        
        if errores:
            return JsonResponse({'success': False, 'error': errores[0]}, status=400)
        
        # Validar que el precio no exceda el saldo disponible del tratamiento
        total_precios_citas = Decimal('0.00')
        for cita in plan.citas.all():
            if cita.precio_cobrado:
                total_precios_citas += Decimal(str(cita.precio_cobrado))
        
        saldo_disponible = plan.precio_final - total_precios_citas
        
        if precio_cobrado > saldo_disponible:
            return JsonResponse({
                'success': False, 
                'error': f'El precio de la cita (${precio_cobrado:,.0f}) excede el saldo disponible del tratamiento (${saldo_disponible:,.0f}). El total de precios de citas no puede exceder el precio final del tratamiento (${plan.precio_final:,.0f}).'
            }, status=400)
        
        try:
            # Convertir fecha y hacerla timezone-aware
            fecha_hora_naive = datetime.fromisoformat(fecha_hora_str)
            fecha_hora = timezone.make_aware(fecha_hora_naive)
            
            # Validar que la fecha no sea en el pasado
            ahora = timezone.now()
            if fecha_hora < ahora:
                return JsonResponse({'success': False, 'error': 'No se pueden crear citas en fechas pasadas.'}, status=400)
            
            # Obtener tipo de servicio si se proporcionó
            tipo_servicio = None
            if tipo_servicio_id:
                try:
                    tipo_servicio = TipoServicio.objects.get(id=tipo_servicio_id, activo=True)
                except TipoServicio.DoesNotExist:
                    pass  # No es obligatorio
            
            # No se usan fases, simplificado
            
            # Validar horario del dentista
            dentista = plan.dentista
            dia_semana = fecha_hora.weekday()
            hora_cita = fecha_hora.time()
            
            horarios_dia = HorarioDentista.objects.filter(
                dentista=dentista,
                dia_semana=dia_semana,
                activo=True
            )
            
            if not horarios_dia.exists():
                dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
                return JsonResponse({'success': False, 'error': f'El dentista no trabaja los {dias_nombres[dia_semana]}.'}, status=400)
            
            hora_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita < horario.hora_fin:
                    hora_valida = True
                    break
            
            if not hora_valida:
                return JsonResponse({'success': False, 'error': 'La hora seleccionada no está dentro del horario de trabajo del dentista.'}, status=400)
            
            # Validar duración y solapamiento
            duracion_minutos = tipo_servicio.duracion_estimada if tipo_servicio and tipo_servicio.duracion_estimada else 30
            fecha_hora_fin = fecha_hora + timedelta(minutes=duracion_minutos)
            hora_fin_cita = fecha_hora_fin.time()
            
            duracion_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita and hora_fin_cita <= horario.hora_fin:
                    duracion_valida = True
                    break
            
            if not duracion_valida:
                return JsonResponse({'success': False, 'error': f'La duración del servicio ({duracion_minutos} minutos) no cabe en el horario seleccionado.'}, status=400)
            
            # Verificar solapamiento con otras citas
            citas_existentes = Cita.objects.filter(
                dentista=dentista,
                fecha_hora__date=fecha_hora.date(),
                estado__in=['disponible', 'reservada', 'confirmada', 'en_progreso']
            )
            
            for cita_existente in citas_existentes:
                fecha_hora_existente_fin = cita_existente.fecha_hora
                if cita_existente.tipo_servicio and cita_existente.tipo_servicio.duracion_estimada:
                    fecha_hora_existente_fin += timedelta(minutes=cita_existente.tipo_servicio.duracion_estimada)
                else:
                    fecha_hora_existente_fin += timedelta(minutes=30)
                
                if (fecha_hora < fecha_hora_existente_fin and fecha_hora_fin > cita_existente.fecha_hora):
                    return JsonResponse({'success': False, 'error': f'La cita se solapa con otra cita existente a las {cita_existente.fecha_hora.strftime("%H:%M")}.'}, status=400)
            
            # Verificar que no exista ya una cita en esa fecha/hora exacta
            if Cita.objects.filter(fecha_hora=fecha_hora).exists():
                return JsonResponse({'success': False, 'error': 'Ya existe una cita en esa fecha y hora exacta.'}, status=400)
            
            # El precio ya fue validado arriba, usar el que viene del formulario
            # Si no se proporcionó precio en el formulario pero hay tipo_servicio, usar el precio base como fallback
            if not precio_cobrado and tipo_servicio and tipo_servicio.precio_base:
                precio_cobrado = Decimal(str(tipo_servicio.precio_base))
            
            # Crear la cita asociada al plan (sin fase)
            cita = Cita.objects.create(
                fecha_hora=fecha_hora,
                tipo_servicio=tipo_servicio,
                tipo_consulta=tipo_servicio.nombre if tipo_servicio else f"Cita - {plan.nombre}",
                precio_cobrado=precio_cobrado,
                notas=notas or f"Cita del plan de tratamiento: {plan.nombre}",
                dentista=dentista,
                cliente=plan.cliente,
                paciente_nombre=plan.cliente.nombre_completo,
                paciente_email=plan.cliente.email,
                paciente_telefono=plan.cliente.telefono,
                estado='reservada',
                creada_por=perfil,
                plan_tratamiento=plan,
                fase_tratamiento=None
            )
            
            # Enviar notificaciones (WhatsApp Y SMS) si tiene teléfono
            # Usar el teléfono del cliente del plan
            telefono_cliente = plan.cliente.telefono if plan.cliente else None
            logger.info(f"Intentando enviar notificaciones para cita {cita.id} del plan {plan.id}. Teléfono cliente: {telefono_cliente}, Teléfono cita: {cita.paciente_telefono}, Teléfono paciente (propiedad): {cita.telefono_paciente}")
            
            if telefono_cliente:
                try:
                    from citas.mensajeria_service import enviar_notificaciones_cita
                    resultado = enviar_notificaciones_cita(cita, telefono_override=telefono_cliente)
                    canales_enviados = []
                    if resultado['whatsapp']['enviado']:
                        canales_enviados.append('WhatsApp')
                    if resultado['sms']['enviado']:
                        canales_enviados.append('SMS')
                    if canales_enviados:
                        logger.info(f"Notificaciones enviadas por {', '.join(canales_enviados)} para cita {cita.id} del plan {plan.id}")
                    else:
                        logger.warning(f"No se pudieron enviar notificaciones para cita {cita.id}. WhatsApp error: {resultado.get('whatsapp', {}).get('error')}, SMS error: {resultado.get('sms', {}).get('error')}")
                except Exception as e:
                    logger.error(f"Error al enviar notificaciones para cita {cita.id}: {e}", exc_info=True)
            else:
                logger.warning(f"No hay teléfono del cliente para enviar notificaciones. Cliente: {plan.cliente}, Teléfono: {telefono_cliente}")
            
            mensaje = f'Cita creada exitosamente para {plan.cliente.nombre_completo} el {fecha_hora.strftime("%d/%m/%Y a las %H:%M")}.'
            if plan.cliente.telefono:
                mensaje += ' Se ha enviado un SMS de confirmación.'
            
            return JsonResponse({
                'success': True,
                'message': mensaje,
                'cita_id': cita.id
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en el formato de fecha: {e}'}, status=400)
        except Exception as e:
            logger.error(f"Error al crear cita desde plan: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return JsonResponse({'success': False, 'error': f'Error al crear la cita: {str(e)}'}, status=500)
    
    # GET: Devolver información del plan y opciones disponibles
    tipos_servicio = TipoServicio.objects.filter(activo=True).order_by('nombre')
    
    return JsonResponse({
        'success': True,
        'plan': {
            'id': plan.id,
            'nombre': plan.nombre,
            'cliente': {
                'id': plan.cliente.id,
                'nombre': plan.cliente.nombre_completo,
                'email': plan.cliente.email,
                'telefono': plan.cliente.telefono
            },
            'dentista': {
                'id': plan.dentista.id,
                'nombre': plan.dentista.nombre_completo
            }
        },
        'tipos_servicio': [{'id': t.id, 'nombre': t.nombre, 'duracion': t.duracion_estimada or 30, 'precio': float(t.precio_base) if t.precio_base else 0} for t in tipos_servicio]
    })


@login_required
def editar_plan_tratamiento(request, plan_id):
    """Edita un plan de tratamiento existente - FUNCIONALIDAD DESHABILITADA"""
    plan = get_object_or_404(PlanTratamiento, id=plan_id)
    messages.warning(
        request, 
        'La funcionalidad de editar planes de tratamiento ha sido deshabilitada. '
        'Si necesita realizar cambios, por favor cancele y elimine el plan actual, '
        'luego cree uno nuevo con la información correcta.'
    )
    return redirect('detalle_plan_tratamiento', plan_id=plan.id)


@login_required
def eliminar_plan_tratamiento(request, plan_id):
    """Elimina definitivamente un plan (SOLO ADMINISTRADOR)"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'Solo los administrativos pueden eliminar planes definitivamente.')
            return redirect('listar_planes_tratamiento')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos.')
        return redirect('login')
    
    plan = get_object_or_404(PlanTratamiento, id=plan_id)
    
    if request.method == 'POST':
        # Guardar información para auditoría antes de eliminar
        plan_id = plan.id
        plan_nombre = plan.nombre
        plan_cliente = plan.cliente.nombre_completo
        motivo_eliminacion = request.POST.get('motivo_eliminacion', '')
        
        plan.eliminado_por = perfil
        plan.fecha_eliminacion = timezone.now()
        plan.motivo_eliminacion = motivo_eliminacion
        plan.save()
        
        # Registrar en auditoría
        registrar_auditoria(
            usuario=perfil,
            accion='eliminar',
            modulo='planes_tratamiento',
            descripcion=f'Plan de tratamiento eliminado: {plan_nombre}',
            detalles=f'Cliente: {plan_cliente}, Motivo: {motivo_eliminacion or "No especificado"}',
            objeto_id=plan_id,
            tipo_objeto='PlanTratamiento',
            request=request
        )
        
        messages.success(request, 'Plan eliminado exitosamente.')
        return redirect('listar_planes_tratamiento')
    
    return render(request, 'citas/planes_tratamiento/eliminar_plan_tratamiento.html', {'plan': plan, 'perfil': perfil})


@login_required
def agregar_fase_tratamiento(request, plan_id):
    """Agrega una nueva fase a un plan de tratamiento existente"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Verificar que el plan puede ser editado
    if not plan.puede_ser_editado_por(perfil):
        return JsonResponse({'success': False, 'error': f'No puedes agregar fases a este plan en estado "{plan.get_estado_display()}".'}, status=400)
    
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre', '').strip()
            descripcion = request.POST.get('descripcion', '').strip()
            presupuesto_str = request.POST.get('presupuesto', '0').strip()
            
            if not nombre:
                return JsonResponse({'success': False, 'error': 'El nombre de la fase es requerido.'}, status=400)
            
            # Obtener el siguiente orden
            ultima_fase = plan.fases.all().order_by('-orden').first()
            nuevo_orden = (ultima_fase.orden + 1) if ultima_fase else 1
            
            # Procesar presupuesto
            presupuesto_str = re.sub(r'[^\d.]', '', presupuesto_str) or '0'
            presupuesto = float(presupuesto_str)
            
            # Crear la fase
            fase = FaseTratamiento.objects.create(
                plan=plan,
                nombre=nombre,
                descripcion=descripcion,
                orden=nuevo_orden,
                presupuesto=presupuesto
            )
            
            # Procesar items si se enviaron
            items_data = request.POST.getlist('items[]')
            if items_data:
                try:
                    items_json = json.loads(items_data[0]) if items_data else []
                    for item_data in items_json:
                        servicio_id = item_data.get('servicio_id')
                        servicio = None
                        if servicio_id:
                            try:
                                servicio = TipoServicio.objects.get(id=servicio_id, activo=True)
                            except TipoServicio.DoesNotExist:
                                pass
                        
                        ItemTratamiento.objects.create(
                            fase=fase,
                            servicio=servicio,
                            descripcion=item_data.get('descripcion', ''),
                            cantidad=int(item_data.get('cantidad', 1)),
                            precio_unitario=float(item_data.get('precio_unitario', 0)),
                        )
                except (json.JSONDecodeError, ValueError) as e:
                    logger.error(f"Error al procesar items de fase: {e}")
            
            return JsonResponse({
                'success': True,
                'message': f'Fase "{fase.nombre}" agregada exitosamente.',
                'fase': {
                    'id': fase.id,
                    'nombre': fase.nombre,
                    'orden': fase.orden,
                    'presupuesto': str(fase.presupuesto),
                }
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en los datos: {str(e)}'}, status=400)
        except Exception as e:
            logger.error(f"Error al agregar fase: {e}")
            return JsonResponse({'success': False, 'error': f'Error al agregar la fase: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def editar_fase_tratamiento(request, plan_id, fase_id):
    """Edita una fase de un plan de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener la fase
    fase = get_object_or_404(FaseTratamiento, id=fase_id, plan=plan)
    
    # Verificar que el plan puede ser editado
    if not plan.puede_ser_editado_por(perfil):
        return JsonResponse({'success': False, 'error': f'No puedes editar fases de este plan en estado "{plan.get_estado_display()}".'}, status=400)
    
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre', '').strip()
            descripcion = request.POST.get('descripcion', '').strip()
            presupuesto_str = request.POST.get('presupuesto', '0').strip()
            orden_str = request.POST.get('orden', '').strip()
            
            if not nombre:
                return JsonResponse({'success': False, 'error': 'El nombre de la fase es requerido.'}, status=400)
            
            # Actualizar campos básicos
            fase.nombre = nombre
            fase.descripcion = descripcion
            
            # Procesar presupuesto
            presupuesto_str = re.sub(r'[^\d.]', '', presupuesto_str) or '0'
            fase.presupuesto = float(presupuesto_str)
            
            # Actualizar orden si se proporcionó
            if orden_str and orden_str.isdigit():
                nuevo_orden = int(orden_str)
                # Verificar que no haya conflicto con otra fase
                fase_existente = plan.fases.filter(orden=nuevo_orden).exclude(id=fase.id).first()
                if fase_existente:
                    # Intercambiar órdenes
                    orden_anterior = fase.orden
                    fase_existente.orden = orden_anterior
                    fase_existente.save()
                fase.orden = nuevo_orden
            
            fase.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Fase "{fase.nombre}" actualizada exitosamente.',
                'fase': {
                    'id': fase.id,
                    'nombre': fase.nombre,
                    'orden': fase.orden,
                    'presupuesto': str(fase.presupuesto),
                }
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en los datos: {str(e)}'}, status=400)
        except Exception as e:
            logger.error(f"Error al editar fase: {e}")
            return JsonResponse({'success': False, 'error': f'Error al editar la fase: {str(e)}'}, status=500)
    
    # GET: Devolver datos de la fase
    items = fase.items.all()
    items_data = []
    for item in items:
        items_data.append({
            'id': item.id,
            'descripcion': item.descripcion,
            'servicio_id': item.servicio.id if item.servicio else None,
            'servicio_nombre': item.servicio.nombre if item.servicio else '',
            'cantidad': item.cantidad,
            'precio_unitario': str(item.precio_unitario),
        })
    
    return JsonResponse({
        'success': True,
        'fase': {
            'id': fase.id,
            'nombre': fase.nombre,
            'descripcion': fase.descripcion,
            'orden': fase.orden,
            'presupuesto': str(fase.presupuesto),
            'completada': fase.completada,
            'items': items_data,
        }
    })


@login_required
def eliminar_fase_tratamiento(request, plan_id, fase_id):
    """Elimina una fase de un plan de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener la fase
    fase = get_object_or_404(FaseTratamiento, id=fase_id, plan=plan)
    
    # Verificar que el plan puede ser editado
    if not plan.puede_ser_editado_por(perfil):
        return JsonResponse({'success': False, 'error': f'No puedes eliminar fases de este plan en estado "{plan.get_estado_display()}".'}, status=400)
    
    # Verificar si la fase tiene citas asociadas
    citas_asociadas = Cita.objects.filter(fase_tratamiento=fase).count()
    if citas_asociadas > 0:
        return JsonResponse({
            'success': False,
            'error': f'No se puede eliminar la fase porque tiene {citas_asociadas} cita(s) asociada(s).'
        }, status=400)
    
    if request.method == 'POST':
        try:
            nombre_fase = fase.nombre
            fase.delete()
            
            # Reordenar las fases restantes
            fases_restantes = plan.fases.all().order_by('orden')
            for idx, fase_restante in enumerate(fases_restantes, start=1):
                fase_restante.orden = idx
                fase_restante.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Fase "{nombre_fase}" eliminada exitosamente.'
            })
            
        except Exception as e:
            logger.error(f"Error al eliminar fase: {e}")
            return JsonResponse({'success': False, 'error': f'Error al eliminar la fase: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def crear_cita_desde_fase(request, plan_id, fase_id):
    """Crea una cita específica para una fase de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener la fase
    fase = get_object_or_404(FaseTratamiento, id=fase_id, plan=plan)
    
    # Validar que el plan esté en un estado que permita crear citas
    if request.method == 'POST':
        # Verificar que el plan tenga consentimiento firmado y presupuesto aceptado
        tiene_consentimiento_firmado = plan.consentimientos.filter(estado='firmado').exists()
        presupuesto_aceptado = plan.presupuesto_aceptado
        
        if not tiene_consentimiento_firmado:
            return JsonResponse({
                'success': False, 
                'error': 'No se pueden crear citas. El tratamiento debe tener al menos un consentimiento informado firmado.'
            }, status=400)
        
        if not presupuesto_aceptado:
            return JsonResponse({
                'success': False, 
                'error': 'No se pueden crear citas. El presupuesto del tratamiento debe estar aceptado.'
            }, status=400)
        
        # Verificar que el plan esté en un estado válido para crear citas
        if plan.estado not in ['aprobado', 'en_progreso']:
            return JsonResponse({
                'success': False, 
                'error': f'No se pueden crear citas para un tratamiento en estado "{plan.get_estado_display()}". El tratamiento debe estar aprobado o en progreso.'
            }, status=400)
        
        # Si el plan está aprobado pero no en progreso, cambiarlo automáticamente a en_progreso
        if plan.estado == 'aprobado':
            plan.estado = 'en_progreso'
            plan.save()
    
    if request.method == 'POST':
        fecha_hora_str = request.POST.get('fecha_hora', '')
        tipo_servicio_id = request.POST.get('tipo_servicio', '').strip()
        notas = request.POST.get('notas', '').strip()
        
        if not fecha_hora_str:
            return JsonResponse({'success': False, 'error': 'Debe seleccionar una fecha y hora.'}, status=400)
        
        try:
            # Convertir fecha y hacerla timezone-aware
            fecha_hora_naive = datetime.fromisoformat(fecha_hora_str)
            fecha_hora = timezone.make_aware(fecha_hora_naive)
            
            # Validar que la fecha no sea en el pasado
            ahora = timezone.now()
            if fecha_hora < ahora:
                return JsonResponse({'success': False, 'error': 'No se pueden crear citas en fechas pasadas.'}, status=400)
            
            # Obtener tipo de servicio si se proporcionó
            tipo_servicio = None
            if tipo_servicio_id:
                try:
                    tipo_servicio = TipoServicio.objects.get(id=tipo_servicio_id, activo=True)
                except TipoServicio.DoesNotExist:
                    pass
            
            # Validar horario del dentista
            dentista = plan.dentista
            dia_semana = fecha_hora.weekday()
            hora_cita = fecha_hora.time()
            
            horarios_dia = HorarioDentista.objects.filter(
                dentista=dentista,
                dia_semana=dia_semana,
                activo=True
            )
            
            if not horarios_dia.exists():
                dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
                return JsonResponse({'success': False, 'error': f'El dentista no trabaja los {dias_nombres[dia_semana]}.'}, status=400)
            
            hora_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita < horario.hora_fin:
                    hora_valida = True
                    break
            
            if not hora_valida:
                return JsonResponse({'success': False, 'error': 'La hora seleccionada no está dentro del horario de trabajo del dentista.'}, status=400)
            
            # Validar duración y solapamiento
            duracion_minutos = tipo_servicio.duracion_estimada if tipo_servicio and tipo_servicio.duracion_estimada else 30
            fecha_hora_fin = fecha_hora + timedelta(minutes=duracion_minutos)
            hora_fin_cita = fecha_hora_fin.time()
            
            duracion_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita and hora_fin_cita <= horario.hora_fin:
                    duracion_valida = True
                    break
            
            if not duracion_valida:
                return JsonResponse({'success': False, 'error': f'La duración del servicio ({duracion_minutos} minutos) no cabe en el horario seleccionado.'}, status=400)
            
            # Verificar solapamiento con otras citas
            citas_existentes = Cita.objects.filter(
                dentista=dentista,
                fecha_hora__date=fecha_hora.date(),
                estado__in=['disponible', 'reservada', 'confirmada', 'en_progreso']
            )
            
            for cita_existente in citas_existentes:
                fecha_hora_existente_fin = cita_existente.fecha_hora
                if cita_existente.tipo_servicio and cita_existente.tipo_servicio.duracion_estimada:
                    fecha_hora_existente_fin += timedelta(minutes=cita_existente.tipo_servicio.duracion_estimada)
                else:
                    fecha_hora_existente_fin += timedelta(minutes=30)
                
                if (fecha_hora < fecha_hora_existente_fin and fecha_hora_fin > cita_existente.fecha_hora):
                    return JsonResponse({'success': False, 'error': f'La cita se solapa con otra cita existente a las {cita_existente.fecha_hora.strftime("%H:%M")}.'}, status=400)
            
            # Obtener precio del servicio
            precio_cobrado = None
            if tipo_servicio:
                precio_cobrado = tipo_servicio.precio_base
            
            # Crear la cita asociada al plan y fase
            cita = Cita.objects.create(
                fecha_hora=fecha_hora,
                tipo_servicio=tipo_servicio,
                tipo_consulta=tipo_servicio.nombre if tipo_servicio else f"Cita - {fase.nombre}",
                precio_cobrado=precio_cobrado,
                notas=notas or f"Cita de la fase: {fase.nombre}",
                dentista=dentista,
                cliente=plan.cliente,
                paciente_nombre=plan.cliente.nombre_completo,
                paciente_email=plan.cliente.email,
                paciente_telefono=plan.cliente.telefono,
                estado='reservada',
                creada_por=perfil,
                plan_tratamiento=plan,
                fase_tratamiento=fase
            )
            
            # Enviar notificaciones (WhatsApp Y SMS) si tiene teléfono
            # Usar el teléfono del cliente del plan
            telefono_cliente = plan.cliente.telefono if plan.cliente else None
            if telefono_cliente:
                try:
                    from citas.mensajeria_service import enviar_notificaciones_cita
                    resultado = enviar_notificaciones_cita(cita, telefono_override=telefono_cliente)
                    canales_enviados = []
                    if resultado['whatsapp']['enviado']:
                        canales_enviados.append('WhatsApp')
                    if resultado['sms']['enviado']:
                        canales_enviados.append('SMS')
                    if canales_enviados:
                        logger.info(f"Notificaciones enviadas por {', '.join(canales_enviados)} para cita {cita.id} del plan {plan.id}")
                except Exception as e:
                    logger.error(f"Error al enviar notificaciones para cita {cita.id}: {e}")
            
            mensaje = f'Cita creada exitosamente para la fase "{fase.nombre}" el {fecha_hora.strftime("%d/%m/%Y a las %H:%M")}.'
            if plan.cliente.telefono:
                mensaje += ' Se ha enviado un SMS de confirmación.'
            
            return JsonResponse({
                'success': True,
                'message': mensaje,
                'cita_id': cita.id
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en el formato de fecha: {e}'}, status=400)
        except Exception as e:
            logger.error(f"Error al crear cita desde fase: {e}")
            return JsonResponse({'success': False, 'error': f'Error al crear la cita: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def cancelar_plan_tratamiento(request, plan_id):
    """Cancela un plan (DENTISTA puede cancelar sus planes)"""
    from django.http import JsonResponse
    
    try:
        perfil = Perfil.objects.get(user=request.user)
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
        messages.error(request, 'No tienes permisos.')
        return redirect('login')
    
    # Obtener plan con permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    else:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
        messages.error(request, 'No tienes permisos.')
        return redirect('panel_trabajador')
    
    if not plan.puede_ser_cancelado_por(perfil):
        error_msg = f'No puedes cancelar este plan en estado "{plan.get_estado_display()}".'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg}, status=400)
        messages.error(request, error_msg)
        return redirect('detalle_plan_tratamiento', plan_id=plan.id)
    
    if request.method == 'POST':
        motivo_cancelacion = request.POST.get('motivo_cancelacion', '').strip()
        
        if not motivo_cancelacion:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'El motivo de cancelación es obligatorio.'}, status=400)
            messages.error(request, 'El motivo de cancelación es obligatorio.')
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
        
        try:
            plan.estado = 'cancelado'
            plan.motivo_cancelacion = motivo_cancelacion
            plan.fecha_cancelacion = timezone.now()
            plan.save()
            
            success_msg = f'El tratamiento "{plan.nombre}" ha sido cancelado exitosamente. Quedará registrado como cancelado en el historial del cliente.'
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': success_msg
                })
            
            messages.success(request, success_msg)
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
        except Exception as e:
            logger.error(f"Error al cancelar plan {plan_id}: {e}")
            error_msg = 'Error al cancelar el plan. Por favor, intenta nuevamente.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg}, status=500)
            messages.error(request, error_msg)
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
    
    # GET: Mostrar formulario (si se accede directamente)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    return render(request, 'citas/planes_tratamiento/cancelar_plan_tratamiento.html', {'plan': plan, 'perfil': perfil})


@login_required
def crear_plan_desde_odontograma(request, odontograma_id):
    """Crea un plan de tratamiento basado en un odontograma"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos.')
        return redirect('login')
    
    if not (perfil.es_dentista() or perfil.es_administrativo()):
        messages.error(request, 'Solo dentistas y administrativos pueden crear planes.')
        return redirect('panel_trabajador')
    
    # Obtener odontograma
    if perfil.es_administrativo():
        odontograma = get_object_or_404(Odontograma, id=odontograma_id)
    else:
        odontograma = get_object_or_404(Odontograma, id=odontograma_id, dentista=perfil)
    
    # Redirigir a crear plan con datos pre-llenados
    return redirect(f"{reverse('crear_plan_tratamiento')}?odontograma_id={odontograma_id}&cliente_id={odontograma.cliente.id if odontograma.cliente else ''}")


@login_required
def registrar_pago_tratamiento(request, plan_id):
    """Registra un pago parcial para un plan de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    if request.method == 'POST':
        try:
            monto_str = request.POST.get('monto', '0').strip()
            fecha_pago_str = request.POST.get('fecha_pago', '')
            metodo_pago = request.POST.get('metodo_pago', '').strip()
            numero_comprobante = request.POST.get('numero_comprobante', '').strip()
            notas = request.POST.get('notas', '').strip()
            cita_id = request.POST.get('cita_id', '').strip()
            
            if not monto_str or float(monto_str) <= 0:
                return JsonResponse({'success': False, 'error': 'El monto debe ser mayor a 0.'}, status=400)
            
            if not fecha_pago_str:
                return JsonResponse({'success': False, 'error': 'La fecha de pago es requerida.'}, status=400)
            
            if not metodo_pago or metodo_pago not in ['efectivo', 'transferencia', 'tarjeta', 'cheque']:
                return JsonResponse({'success': False, 'error': 'El método de pago es requerido.'}, status=400)
            
            # Procesar monto
            monto_str = re.sub(r'[^\d.]', '', monto_str) or '0'
            monto = float(monto_str)
            
            # Validar que el monto no exceda el saldo pendiente
            saldo_pendiente = plan.saldo_pendiente
            if monto > saldo_pendiente:
                return JsonResponse({
                    'success': False,
                    'error': f'El monto excede el saldo pendiente (${saldo_pendiente:,.0f}).'
                }, status=400)
            
            # Procesar fecha
            fecha_pago = datetime.strptime(fecha_pago_str, '%Y-%m-%d').date()
            
            # Obtener cita si se proporcionó
            cita = None
            if cita_id:
                try:
                    cita = Cita.objects.get(id=cita_id, plan_tratamiento=plan)
                except Cita.DoesNotExist:
                    pass
            
            # Crear el pago
            pago = PagoTratamiento.objects.create(
                plan_tratamiento=plan,
                monto=monto,
                fecha_pago=fecha_pago,
                metodo_pago=metodo_pago,
                numero_comprobante=numero_comprobante if numero_comprobante else None,
                notas=notas if notas else None,
                cita=cita,
                registrado_por=perfil
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Pago de ${monto:,.0f} registrado exitosamente.',
                'pago': {
                    'id': pago.id,
                    'monto': str(pago.monto),
                    'fecha_pago': pago.fecha_pago.strftime('%d/%m/%Y'),
                    'metodo_pago': pago.get_metodo_pago_display(),
                },
                'total_pagado': str(plan.total_pagado),
                'saldo_pendiente': str(plan.saldo_pendiente),
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en los datos: {str(e)}'}, status=400)
        except Exception as e:
            logger.error(f"Error al registrar pago: {e}")
            return JsonResponse({'success': False, 'error': f'Error al registrar el pago: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def eliminar_pago_tratamiento(request, plan_id, pago_id):
    """Elimina un pago registrado (solo administrativos)"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({'success': False, 'error': 'Solo los administrativos pueden eliminar pagos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    plan = get_object_or_404(PlanTratamiento, id=plan_id)
    pago = get_object_or_404(PagoTratamiento, id=pago_id, plan_tratamiento=plan)
    
    if request.method == 'POST':
        try:
            monto = pago.monto
            pago.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Pago de ${monto:,.0f} eliminado exitosamente.',
                'total_pagado': str(plan.total_pagado),
                'saldo_pendiente': str(plan.saldo_pendiente),
            })
            
        except Exception as e:
            logger.error(f"Error al eliminar pago: {e}")
            return JsonResponse({'success': False, 'error': f'Error al eliminar el pago: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def enviar_documentos_tratamiento(request, plan_id):
    """Vista para enviar presupuesto y consentimientos de un tratamiento por correo"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not (perfil.es_administrativo() or perfil.es_dentista()):
            return JsonResponse({'success': False, 'error': 'No tienes permisos para enviar documentos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan de tratamiento
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    else:
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    # Verificar que el cliente tenga email
    if not plan.cliente.email:
        return JsonResponse({'success': False, 'error': 'El cliente no tiene un email registrado.'}, status=400)
    
    try:
        from django.core.mail import EmailMultiAlternatives
        from django.conf import settings
        from io import BytesIO
        from datetime import datetime
        
        # Obtener información de la clínica
        try:
            from configuracion.models import InformacionClinica
            info_clinica = InformacionClinica.obtener()
            nombre_clinica = info_clinica.nombre_clinica
            email_clinica = info_clinica.email or getattr(settings, 'EMAIL_FROM', 'noreply@clinica.com')
        except:
            nombre_clinica = "Clínica Dental"
            email_clinica = getattr(settings, 'EMAIL_FROM', 'noreply@clinica.com')
        
        # Generar PDF del presupuesto
        presupuesto_response = exportar_presupuesto_pdf(request, plan.id)
        presupuesto_pdf = presupuesto_response.content
        
        # Obtener consentimientos pendientes o no firmados
        consentimientos = plan.consentimientos.filter(estado__in=['pendiente', 'firmado']).order_by('-fecha_creacion')
        
        # Crear el email
        asunto = f"Documentos del Tratamiento - {plan.nombre} - {nombre_clinica}"
        
        mensaje_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: linear-gradient(135deg, #14b8a6 0%, #0d9488 100%); min-height: 100vh;">
            <div style="max-width: 600px; margin: 0 auto; padding: 40px 20px;">
                <!-- Header con logo -->
                <div style="background: white; border-radius: 12px 12px 0 0; padding: 30px; text-align: center; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    <div style="display: inline-flex; align-items: center; gap: 12px; margin-bottom: 10px;">
                        <div style="width: 50px; height: 50px; background: linear-gradient(135deg, #14b8a6, #0d9488); border-radius: 50%; display: flex; align-items: center; justify-content: center; color: white; font-size: 1.8rem;">
                            <span style="font-size: 1.8rem;">🦷</span>
                        </div>
                        <h1 style="margin: 0; color: #0f766e; font-size: 1.75rem; font-weight: 700;">Clínica San Felipe</h1>
                    </div>
                    <p style="margin: 0; color: #64748b; font-size: 0.9rem;">Victoria, Región de la Araucanía</p>
                </div>
                
                <!-- Contenido del correo -->
                <div style="background: white; padding: 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    <h2 style="color: #1e293b; margin-top: 0; font-size: 1.5rem;">Estimado/a {plan.cliente.nombre_completo},</h2>
                    
                    <p style="color: #475569; line-height: 1.6; font-size: 1rem;">Le enviamos los documentos relacionados con su tratamiento: <strong style="color: #0f766e;">{plan.nombre}</strong></p>
                    
                    <div style="background: linear-gradient(135deg, #f0fdfa, #e0f2f1); padding: 20px; border-radius: 8px; margin: 25px 0; border-left: 4px solid #14b8a6;">
                        <h3 style="margin-top: 0; color: #0f766e; font-size: 1.1rem; margin-bottom: 12px;">Documentos adjuntos:</h3>
                        <ul style="margin: 0; padding-left: 20px; color: #475569;">
                            <li style="margin-bottom: 8px;"><strong style="color: #0f766e;">Presupuesto del Tratamiento</strong></li>
        """
        
        if consentimientos.exists():
            mensaje_html += f"<li style='margin-bottom: 8px;'><strong style='color: #0f766e;'>Consentimientos Informados</strong> ({consentimientos.count()} documento(s))</li>"
        
        mensaje_html += f"""
                        </ul>
                    </div>
                    
                    <p style="color: #475569; line-height: 1.6;">Por favor, revise cuidadosamente estos documentos. Si tiene alguna consulta, no dude en contactarnos.</p>
                    
                    <div style="background: #fef3c7; padding: 18px; border-radius: 8px; margin: 25px 0; border-left: 4px solid #f59e0b;">
                        <p style="margin: 0 0 12px 0; color: #92400e; font-weight: 600; font-size: 1rem;"><strong>Importante:</strong> Para proceder con el tratamiento, necesitamos que:</p>
                        <ul style="margin: 0; padding-left: 20px; color: #92400e;">
                            <li style="margin-bottom: 6px;">Acepte el presupuesto</li>
                            <li style="margin-bottom: 6px;">Firme el consentimiento informado</li>
                        </ul>
                    </div>
                    
                    <p style="color: #475569; line-height: 1.6;">Puede hacerlo desde su panel de cliente en nuestra página web o contactándonos directamente.</p>
                </div>
                
                <!-- Footer -->
                <div style="background: white; border-radius: 0 0 12px 12px; padding: 25px 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border-top: 1px solid #e0f2f1;">
                    <p style="margin: 0; color: #64748b; font-size: 0.9rem; line-height: 1.6;">
                        Saludos cordiales,<br>
                        <strong style="color: #0f766e; font-size: 1rem;">{nombre_clinica}</strong>
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        mensaje_texto = f"""
        Estimado/a {plan.cliente.nombre_completo},
        
        Le enviamos los documentos relacionados con su tratamiento: {plan.nombre}
        
        Documentos adjuntos:
        - Presupuesto del Tratamiento
        - Consentimientos Informados
        
        Por favor, revise cuidadosamente estos documentos. Si tiene alguna consulta, no dude en contactarnos.
        
        Importante: Para proceder con el tratamiento, necesitamos que:
        - Acepte el presupuesto
        - Firme el consentimiento informado
        
        Puede hacerlo desde su panel de cliente en nuestra página web o contactándonos directamente.
        
        Saludos cordiales,
        {nombre_clinica}
        """
        
        # Crear el email
        email = EmailMultiAlternatives(
            asunto,
            mensaje_texto,
            email_clinica,
            [plan.cliente.email],
        )
        email.attach_alternative(mensaje_html, "text/html")
        
        # Adjuntar presupuesto
        filename_presupuesto = f"presupuesto_{plan.cliente.nombre_completo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        email.attach(filename_presupuesto, presupuesto_pdf, 'application/pdf')
        
        # Adjuntar consentimientos
        for consentimiento in consentimientos:
            try:
                consentimiento_response = exportar_consentimiento_pdf(request, consentimiento.id)
                consentimiento_pdf = consentimiento_response.content
                filename_consentimiento = f"consentimiento_{consentimiento.titulo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
                email.attach(filename_consentimiento, consentimiento_pdf, 'application/pdf')
            except Exception as e:
                logger.error(f"Error al generar PDF del consentimiento {consentimiento.id}: {e}")
                # Continuar con los demás documentos
        
        # Enviar email
        email.send()
        
        # Actualizar fecha de envío en documentos relacionados
        from historial_clinico.models import DocumentoCliente
        documentos_actualizados = 0
        for consentimiento in consentimientos:
            # Actualizar el documento relacionado si existe
            try:
                documento = DocumentoCliente.objects.filter(
                    plan_tratamiento=plan,
                    tipo='consentimiento',
                    cliente=plan.cliente
                ).filter(
                    titulo__icontains=consentimiento.titulo
                ).first()
                
                if documento:
                    documento.enviado_por_correo = True
                    documento.fecha_envio = timezone.now()
                    documento.email_destinatario = plan.cliente.email
                    documento.save(update_fields=['enviado_por_correo', 'fecha_envio', 'email_destinatario'])
                    documentos_actualizados += 1
            except Exception as e:
                logger.warning(f"No se pudo actualizar documento para consentimiento {consentimiento.id}: {e}")
        
        # Actualizar también el documento del presupuesto si existe
        try:
            doc_presupuesto = DocumentoCliente.objects.filter(
                plan_tratamiento=plan,
                tipo='presupuesto',
                cliente=plan.cliente
            ).first()
            
            if doc_presupuesto:
                doc_presupuesto.enviado_por_correo = True
                doc_presupuesto.fecha_envio = timezone.now()
                doc_presupuesto.email_destinatario = plan.cliente.email
                doc_presupuesto.save(update_fields=['enviado_por_correo', 'fecha_envio', 'email_destinatario'])
                documentos_actualizados += 1
        except Exception as e:
            logger.warning(f"No se pudo actualizar documento de presupuesto: {e}")
        
        return JsonResponse({
            'success': True,
            'message': f'Documentos enviados exitosamente a {plan.cliente.email}. Se enviaron {1 + consentimientos.count()} documento(s).'
        })
        
    except Exception as e:
        logger.error(f"Error al enviar documentos del tratamiento: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'error': f'Error al enviar los documentos: {str(e)}'}, status=500)

    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener la fase
    fase = get_object_or_404(FaseTratamiento, id=fase_id, plan=plan)
    
    # Verificar que el plan puede ser editado
    if not plan.puede_ser_editado_por(perfil):
        return JsonResponse({'success': False, 'error': f'No puedes eliminar fases de este plan en estado "{plan.get_estado_display()}".'}, status=400)
    
    # Verificar si la fase tiene citas asociadas
    citas_asociadas = Cita.objects.filter(fase_tratamiento=fase).count()
    if citas_asociadas > 0:
        return JsonResponse({
            'success': False,
            'error': f'No se puede eliminar la fase porque tiene {citas_asociadas} cita(s) asociada(s).'
        }, status=400)
    
    if request.method == 'POST':
        try:
            nombre_fase = fase.nombre
            fase.delete()
            
            # Reordenar las fases restantes
            fases_restantes = plan.fases.all().order_by('orden')
            for idx, fase_restante in enumerate(fases_restantes, start=1):
                fase_restante.orden = idx
                fase_restante.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Fase "{nombre_fase}" eliminada exitosamente.'
            })
            
        except Exception as e:
            logger.error(f"Error al eliminar fase: {e}")
            return JsonResponse({'success': False, 'error': f'Error al eliminar la fase: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def crear_cita_desde_fase(request, plan_id, fase_id):
    """Crea una cita específica para una fase de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener la fase
    fase = get_object_or_404(FaseTratamiento, id=fase_id, plan=plan)
    
    # Validar que el plan esté en un estado que permita crear citas
    if request.method == 'POST':
        # Verificar que el plan tenga consentimiento firmado y presupuesto aceptado
        tiene_consentimiento_firmado = plan.consentimientos.filter(estado='firmado').exists()
        presupuesto_aceptado = plan.presupuesto_aceptado
        
        if not tiene_consentimiento_firmado:
            return JsonResponse({
                'success': False, 
                'error': 'No se pueden crear citas. El tratamiento debe tener al menos un consentimiento informado firmado.'
            }, status=400)
        
        if not presupuesto_aceptado:
            return JsonResponse({
                'success': False, 
                'error': 'No se pueden crear citas. El presupuesto del tratamiento debe estar aceptado.'
            }, status=400)
        
        # Verificar que el plan esté en un estado válido para crear citas
        if plan.estado not in ['aprobado', 'en_progreso']:
            return JsonResponse({
                'success': False, 
                'error': f'No se pueden crear citas para un tratamiento en estado "{plan.get_estado_display()}". El tratamiento debe estar aprobado o en progreso.'
            }, status=400)
        
        # Si el plan está aprobado pero no en progreso, cambiarlo automáticamente a en_progreso
        if plan.estado == 'aprobado':
            plan.estado = 'en_progreso'
            plan.save()
    
    if request.method == 'POST':
        fecha_hora_str = request.POST.get('fecha_hora', '')
        tipo_servicio_id = request.POST.get('tipo_servicio', '').strip()
        notas = request.POST.get('notas', '').strip()
        
        if not fecha_hora_str:
            return JsonResponse({'success': False, 'error': 'Debe seleccionar una fecha y hora.'}, status=400)
        
        try:
            # Convertir fecha y hacerla timezone-aware
            fecha_hora_naive = datetime.fromisoformat(fecha_hora_str)
            fecha_hora = timezone.make_aware(fecha_hora_naive)
            
            # Validar que la fecha no sea en el pasado
            ahora = timezone.now()
            if fecha_hora < ahora:
                return JsonResponse({'success': False, 'error': 'No se pueden crear citas en fechas pasadas.'}, status=400)
            
            # Obtener tipo de servicio si se proporcionó
            tipo_servicio = None
            if tipo_servicio_id:
                try:
                    tipo_servicio = TipoServicio.objects.get(id=tipo_servicio_id, activo=True)
                except TipoServicio.DoesNotExist:
                    pass
            
            # Validar horario del dentista
            dentista = plan.dentista
            dia_semana = fecha_hora.weekday()
            hora_cita = fecha_hora.time()
            
            horarios_dia = HorarioDentista.objects.filter(
                dentista=dentista,
                dia_semana=dia_semana,
                activo=True
            )
            
            if not horarios_dia.exists():
                dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
                return JsonResponse({'success': False, 'error': f'El dentista no trabaja los {dias_nombres[dia_semana]}.'}, status=400)
            
            hora_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita < horario.hora_fin:
                    hora_valida = True
                    break
            
            if not hora_valida:
                return JsonResponse({'success': False, 'error': 'La hora seleccionada no está dentro del horario de trabajo del dentista.'}, status=400)
            
            # Validar duración y solapamiento
            duracion_minutos = tipo_servicio.duracion_estimada if tipo_servicio and tipo_servicio.duracion_estimada else 30
            fecha_hora_fin = fecha_hora + timedelta(minutes=duracion_minutos)
            hora_fin_cita = fecha_hora_fin.time()
            
            duracion_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita and hora_fin_cita <= horario.hora_fin:
                    duracion_valida = True
                    break
            
            if not duracion_valida:
                return JsonResponse({'success': False, 'error': f'La duración del servicio ({duracion_minutos} minutos) no cabe en el horario seleccionado.'}, status=400)
            
            # Verificar solapamiento con otras citas
            citas_existentes = Cita.objects.filter(
                dentista=dentista,
                fecha_hora__date=fecha_hora.date(),
                estado__in=['disponible', 'reservada', 'confirmada', 'en_progreso']
            )
            
            for cita_existente in citas_existentes:
                fecha_hora_existente_fin = cita_existente.fecha_hora
                if cita_existente.tipo_servicio and cita_existente.tipo_servicio.duracion_estimada:
                    fecha_hora_existente_fin += timedelta(minutes=cita_existente.tipo_servicio.duracion_estimada)
                else:
                    fecha_hora_existente_fin += timedelta(minutes=30)
                
                if (fecha_hora < fecha_hora_existente_fin and fecha_hora_fin > cita_existente.fecha_hora):
                    return JsonResponse({'success': False, 'error': f'La cita se solapa con otra cita existente a las {cita_existente.fecha_hora.strftime("%H:%M")}.'}, status=400)
            
            # Obtener precio del servicio
            precio_cobrado = None
            if tipo_servicio:
                precio_cobrado = tipo_servicio.precio_base
            
            # Crear la cita asociada al plan y fase
            cita = Cita.objects.create(
                fecha_hora=fecha_hora,
                tipo_servicio=tipo_servicio,
                tipo_consulta=tipo_servicio.nombre if tipo_servicio else f"Cita - {fase.nombre}",
                precio_cobrado=precio_cobrado,
                notas=notas or f"Cita de la fase: {fase.nombre}",
                dentista=dentista,
                cliente=plan.cliente,
                paciente_nombre=plan.cliente.nombre_completo,
                paciente_email=plan.cliente.email,
                paciente_telefono=plan.cliente.telefono,
                estado='reservada',
                creada_por=perfil,
                plan_tratamiento=plan,
                fase_tratamiento=fase
            )
            
            # Enviar notificaciones (WhatsApp Y SMS) si tiene teléfono
            # Usar el teléfono del cliente del plan
            telefono_cliente = plan.cliente.telefono if plan.cliente else None
            if telefono_cliente:
                try:
                    from citas.mensajeria_service import enviar_notificaciones_cita
                    resultado = enviar_notificaciones_cita(cita, telefono_override=telefono_cliente)
                    canales_enviados = []
                    if resultado['whatsapp']['enviado']:
                        canales_enviados.append('WhatsApp')
                    if resultado['sms']['enviado']:
                        canales_enviados.append('SMS')
                    if canales_enviados:
                        logger.info(f"Notificaciones enviadas por {', '.join(canales_enviados)} para cita {cita.id} del plan {plan.id}")
                except Exception as e:
                    logger.error(f"Error al enviar notificaciones para cita {cita.id}: {e}")
            
            mensaje = f'Cita creada exitosamente para la fase "{fase.nombre}" el {fecha_hora.strftime("%d/%m/%Y a las %H:%M")}.'
            if plan.cliente.telefono:
                mensaje += ' Se ha enviado un SMS de confirmación.'
            
            return JsonResponse({
                'success': True,
                'message': mensaje,
                'cita_id': cita.id
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en el formato de fecha: {e}'}, status=400)
        except Exception as e:
            logger.error(f"Error al crear cita desde fase: {e}")
            return JsonResponse({'success': False, 'error': f'Error al crear la cita: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def cancelar_plan_tratamiento(request, plan_id):
    """Cancela un plan (DENTISTA puede cancelar sus planes)"""
    from django.http import JsonResponse
    
    try:
        perfil = Perfil.objects.get(user=request.user)
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
        messages.error(request, 'No tienes permisos.')
        return redirect('login')
    
    # Obtener plan con permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    else:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
        messages.error(request, 'No tienes permisos.')
        return redirect('panel_trabajador')
    
    if not plan.puede_ser_cancelado_por(perfil):
        error_msg = f'No puedes cancelar este plan en estado "{plan.get_estado_display()}".'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg}, status=400)
        messages.error(request, error_msg)
        return redirect('detalle_plan_tratamiento', plan_id=plan.id)
    
    if request.method == 'POST':
        motivo_cancelacion = request.POST.get('motivo_cancelacion', '').strip()
        
        if not motivo_cancelacion:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'El motivo de cancelación es obligatorio.'}, status=400)
            messages.error(request, 'El motivo de cancelación es obligatorio.')
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
        
        try:
            plan.estado = 'cancelado'
            plan.motivo_cancelacion = motivo_cancelacion
            plan.fecha_cancelacion = timezone.now()
            plan.save()
            
            success_msg = f'El tratamiento "{plan.nombre}" ha sido cancelado exitosamente. Quedará registrado como cancelado en el historial del cliente.'
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': success_msg
                })
            
            messages.success(request, success_msg)
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
        except Exception as e:
            logger.error(f"Error al cancelar plan {plan_id}: {e}")
            error_msg = 'Error al cancelar el plan. Por favor, intenta nuevamente.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg}, status=500)
            messages.error(request, error_msg)
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
    
    # GET: Mostrar formulario (si se accede directamente)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    return render(request, 'citas/planes_tratamiento/cancelar_plan_tratamiento.html', {'plan': plan, 'perfil': perfil})


@login_required
def crear_plan_desde_odontograma(request, odontograma_id):
    """Crea un plan de tratamiento basado en un odontograma"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos.')
        return redirect('login')
    
    if not (perfil.es_dentista() or perfil.es_administrativo()):
        messages.error(request, 'Solo dentistas y administrativos pueden crear planes.')
        return redirect('panel_trabajador')
    
    # Obtener odontograma
    if perfil.es_administrativo():
        odontograma = get_object_or_404(Odontograma, id=odontograma_id)
    else:
        odontograma = get_object_or_404(Odontograma, id=odontograma_id, dentista=perfil)
    
    # Redirigir a crear plan con datos pre-llenados
    return redirect(f"{reverse('crear_plan_tratamiento')}?odontograma_id={odontograma_id}&cliente_id={odontograma.cliente.id if odontograma.cliente else ''}")


@login_required
def registrar_pago_tratamiento(request, plan_id):
    """Registra un pago parcial para un plan de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    if request.method == 'POST':
        try:
            monto_str = request.POST.get('monto', '0').strip()
            fecha_pago_str = request.POST.get('fecha_pago', '')
            metodo_pago = request.POST.get('metodo_pago', '').strip()
            numero_comprobante = request.POST.get('numero_comprobante', '').strip()
            notas = request.POST.get('notas', '').strip()
            cita_id = request.POST.get('cita_id', '').strip()
            
            if not monto_str or float(monto_str) <= 0:
                return JsonResponse({'success': False, 'error': 'El monto debe ser mayor a 0.'}, status=400)
            
            if not fecha_pago_str:
                return JsonResponse({'success': False, 'error': 'La fecha de pago es requerida.'}, status=400)
            
            if not metodo_pago or metodo_pago not in ['efectivo', 'transferencia', 'tarjeta', 'cheque']:
                return JsonResponse({'success': False, 'error': 'El método de pago es requerido.'}, status=400)
            
            # Procesar monto
            monto_str = re.sub(r'[^\d.]', '', monto_str) or '0'
            monto = float(monto_str)
            
            # Validar que el monto no exceda el saldo pendiente
            saldo_pendiente = plan.saldo_pendiente
            if monto > saldo_pendiente:
                return JsonResponse({
                    'success': False,
                    'error': f'El monto excede el saldo pendiente (${saldo_pendiente:,.0f}).'
                }, status=400)
            
            # Procesar fecha
            fecha_pago = datetime.strptime(fecha_pago_str, '%Y-%m-%d').date()
            
            # Obtener cita si se proporcionó
            cita = None
            if cita_id:
                try:
                    cita = Cita.objects.get(id=cita_id, plan_tratamiento=plan)
                except Cita.DoesNotExist:
                    pass
            
            # Crear el pago
            pago = PagoTratamiento.objects.create(
                plan_tratamiento=plan,
                monto=monto,
                fecha_pago=fecha_pago,
                metodo_pago=metodo_pago,
                numero_comprobante=numero_comprobante if numero_comprobante else None,
                notas=notas if notas else None,
                cita=cita,
                registrado_por=perfil
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Pago de ${monto:,.0f} registrado exitosamente.',
                'pago': {
                    'id': pago.id,
                    'monto': str(pago.monto),
                    'fecha_pago': pago.fecha_pago.strftime('%d/%m/%Y'),
                    'metodo_pago': pago.get_metodo_pago_display(),
                },
                'total_pagado': str(plan.total_pagado),
                'saldo_pendiente': str(plan.saldo_pendiente),
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en los datos: {str(e)}'}, status=400)
        except Exception as e:
            logger.error(f"Error al registrar pago: {e}")
            return JsonResponse({'success': False, 'error': f'Error al registrar el pago: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def eliminar_pago_tratamiento(request, plan_id, pago_id):
    """Elimina un pago registrado (solo administrativos)"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({'success': False, 'error': 'Solo los administrativos pueden eliminar pagos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    plan = get_object_or_404(PlanTratamiento, id=plan_id)
    pago = get_object_or_404(PagoTratamiento, id=pago_id, plan_tratamiento=plan)
    
    if request.method == 'POST':
        try:
            monto = pago.monto
            pago.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Pago de ${monto:,.0f} eliminado exitosamente.',
                'total_pagado': str(plan.total_pagado),
                'saldo_pendiente': str(plan.saldo_pendiente),
            })
            
        except Exception as e:
            logger.error(f"Error al eliminar pago: {e}")
            return JsonResponse({'success': False, 'error': f'Error al eliminar el pago: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def enviar_documentos_tratamiento(request, plan_id):
    """Vista para enviar presupuesto y consentimientos de un tratamiento por correo"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not (perfil.es_administrativo() or perfil.es_dentista()):
            return JsonResponse({'success': False, 'error': 'No tienes permisos para enviar documentos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan de tratamiento
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    else:
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    # Verificar que el cliente tenga email
    if not plan.cliente.email:
        return JsonResponse({'success': False, 'error': 'El cliente no tiene un email registrado.'}, status=400)
    
    try:
        from django.core.mail import EmailMultiAlternatives
        from django.conf import settings
        from io import BytesIO
        from datetime import datetime
        
        # Obtener información de la clínica
        try:
            from configuracion.models import InformacionClinica
            info_clinica = InformacionClinica.obtener()
            nombre_clinica = info_clinica.nombre_clinica
            email_clinica = info_clinica.email or getattr(settings, 'EMAIL_FROM', 'noreply@clinica.com')
        except:
            nombre_clinica = "Clínica Dental"
            email_clinica = getattr(settings, 'EMAIL_FROM', 'noreply@clinica.com')
        
        # Generar PDF del presupuesto
        presupuesto_response = exportar_presupuesto_pdf(request, plan.id)
        presupuesto_pdf = presupuesto_response.content
        
        # Obtener consentimientos pendientes o no firmados
        consentimientos = plan.consentimientos.filter(estado__in=['pendiente', 'firmado']).order_by('-fecha_creacion')
        
        # Crear el email
        asunto = f"Documentos del Tratamiento - {plan.nombre} - {nombre_clinica}"
        
        mensaje_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: linear-gradient(135deg, #14b8a6 0%, #0d9488 100%); min-height: 100vh;">
            <div style="max-width: 600px; margin: 0 auto; padding: 40px 20px;">
                <!-- Header con logo -->
                <div style="background: white; border-radius: 12px 12px 0 0; padding: 30px; text-align: center; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    <div style="display: inline-flex; align-items: center; gap: 12px; margin-bottom: 10px;">
                        <div style="width: 50px; height: 50px; background: linear-gradient(135deg, #14b8a6, #0d9488); border-radius: 50%; display: flex; align-items: center; justify-content: center; color: white; font-size: 1.8rem;">
                            <span style="font-size: 1.8rem;">🦷</span>
                        </div>
                        <h1 style="margin: 0; color: #0f766e; font-size: 1.75rem; font-weight: 700;">Clínica San Felipe</h1>
                    </div>
                    <p style="margin: 0; color: #64748b; font-size: 0.9rem;">Victoria, Región de la Araucanía</p>
                </div>
                
                <!-- Contenido del correo -->
                <div style="background: white; padding: 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    <h2 style="color: #1e293b; margin-top: 0; font-size: 1.5rem;">Estimado/a {plan.cliente.nombre_completo},</h2>
                    
                    <p style="color: #475569; line-height: 1.6; font-size: 1rem;">Le enviamos los documentos relacionados con su tratamiento: <strong style="color: #0f766e;">{plan.nombre}</strong></p>
                    
                    <div style="background: linear-gradient(135deg, #f0fdfa, #e0f2f1); padding: 20px; border-radius: 8px; margin: 25px 0; border-left: 4px solid #14b8a6;">
                        <h3 style="margin-top: 0; color: #0f766e; font-size: 1.1rem; margin-bottom: 12px;">Documentos adjuntos:</h3>
                        <ul style="margin: 0; padding-left: 20px; color: #475569;">
                            <li style="margin-bottom: 8px;"><strong style="color: #0f766e;">Presupuesto del Tratamiento</strong></li>
        """
        
        if consentimientos.exists():
            mensaje_html += f"<li style='margin-bottom: 8px;'><strong style='color: #0f766e;'>Consentimientos Informados</strong> ({consentimientos.count()} documento(s))</li>"
        
        mensaje_html += f"""
                        </ul>
                    </div>
                    
                    <p style="color: #475569; line-height: 1.6;">Por favor, revise cuidadosamente estos documentos. Si tiene alguna consulta, no dude en contactarnos.</p>
                    
                    <div style="background: #fef3c7; padding: 18px; border-radius: 8px; margin: 25px 0; border-left: 4px solid #f59e0b;">
                        <p style="margin: 0 0 12px 0; color: #92400e; font-weight: 600; font-size: 1rem;"><strong>Importante:</strong> Para proceder con el tratamiento, necesitamos que:</p>
                        <ul style="margin: 0; padding-left: 20px; color: #92400e;">
                            <li style="margin-bottom: 6px;">Acepte el presupuesto</li>
                            <li style="margin-bottom: 6px;">Firme el consentimiento informado</li>
                        </ul>
                    </div>
                    
                    <p style="color: #475569; line-height: 1.6;">Puede hacerlo desde su panel de cliente en nuestra página web o contactándonos directamente.</p>
                </div>
                
                <!-- Footer -->
                <div style="background: white; border-radius: 0 0 12px 12px; padding: 25px 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border-top: 1px solid #e0f2f1;">
                    <p style="margin: 0; color: #64748b; font-size: 0.9rem; line-height: 1.6;">
                        Saludos cordiales,<br>
                        <strong style="color: #0f766e; font-size: 1rem;">{nombre_clinica}</strong>
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        mensaje_texto = f"""
        Estimado/a {plan.cliente.nombre_completo},
        
        Le enviamos los documentos relacionados con su tratamiento: {plan.nombre}
        
        Documentos adjuntos:
        - Presupuesto del Tratamiento
        - Consentimientos Informados
        
        Por favor, revise cuidadosamente estos documentos. Si tiene alguna consulta, no dude en contactarnos.
        
        Importante: Para proceder con el tratamiento, necesitamos que:
        - Acepte el presupuesto
        - Firme el consentimiento informado
        
        Puede hacerlo desde su panel de cliente en nuestra página web o contactándonos directamente.
        
        Saludos cordiales,
        {nombre_clinica}
        """
        
        # Crear el email
        email = EmailMultiAlternatives(
            asunto,
            mensaje_texto,
            email_clinica,
            [plan.cliente.email],
        )
        email.attach_alternative(mensaje_html, "text/html")
        
        # Adjuntar presupuesto
        filename_presupuesto = f"presupuesto_{plan.cliente.nombre_completo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        email.attach(filename_presupuesto, presupuesto_pdf, 'application/pdf')
        
        # Adjuntar consentimientos
        for consentimiento in consentimientos:
            try:
                consentimiento_response = exportar_consentimiento_pdf(request, consentimiento.id)
                consentimiento_pdf = consentimiento_response.content
                filename_consentimiento = f"consentimiento_{consentimiento.titulo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
                email.attach(filename_consentimiento, consentimiento_pdf, 'application/pdf')
            except Exception as e:
                logger.error(f"Error al generar PDF del consentimiento {consentimiento.id}: {e}")
                # Continuar con los demás documentos
        
        # Enviar email
        email.send()
        
        # Actualizar fecha de envío en documentos relacionados
        from historial_clinico.models import DocumentoCliente
        documentos_actualizados = 0
        for consentimiento in consentimientos:
            # Actualizar el documento relacionado si existe
            try:
                documento = DocumentoCliente.objects.filter(
                    plan_tratamiento=plan,
                    tipo='consentimiento',
                    cliente=plan.cliente
                ).filter(
                    titulo__icontains=consentimiento.titulo
                ).first()
                
                if documento:
                    documento.enviado_por_correo = True
                    documento.fecha_envio = timezone.now()
                    documento.email_destinatario = plan.cliente.email
                    documento.save(update_fields=['enviado_por_correo', 'fecha_envio', 'email_destinatario'])
                    documentos_actualizados += 1
            except Exception as e:
                logger.warning(f"No se pudo actualizar documento para consentimiento {consentimiento.id}: {e}")
        
        # Actualizar también el documento del presupuesto si existe
        try:
            doc_presupuesto = DocumentoCliente.objects.filter(
                plan_tratamiento=plan,
                tipo='presupuesto',
                cliente=plan.cliente
            ).first()
            
            if doc_presupuesto:
                doc_presupuesto.enviado_por_correo = True
                doc_presupuesto.fecha_envio = timezone.now()
                doc_presupuesto.email_destinatario = plan.cliente.email
                doc_presupuesto.save(update_fields=['enviado_por_correo', 'fecha_envio', 'email_destinatario'])
                documentos_actualizados += 1
        except Exception as e:
            logger.warning(f"No se pudo actualizar documento de presupuesto: {e}")
        
        return JsonResponse({
            'success': True,
            'message': f'Documentos enviados exitosamente a {plan.cliente.email}. Se enviaron {1 + consentimientos.count()} documento(s).'
        })
        
    except Exception as e:
        logger.error(f"Error al enviar documentos del tratamiento: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'error': f'Error al enviar los documentos: {str(e)}'}, status=500)

    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener la fase
    fase = get_object_or_404(FaseTratamiento, id=fase_id, plan=plan)
    
    # Verificar que el plan puede ser editado
    if not plan.puede_ser_editado_por(perfil):
        return JsonResponse({'success': False, 'error': f'No puedes eliminar fases de este plan en estado "{plan.get_estado_display()}".'}, status=400)
    
    # Verificar si la fase tiene citas asociadas
    citas_asociadas = Cita.objects.filter(fase_tratamiento=fase).count()
    if citas_asociadas > 0:
        return JsonResponse({
            'success': False,
            'error': f'No se puede eliminar la fase porque tiene {citas_asociadas} cita(s) asociada(s).'
        }, status=400)
    
    if request.method == 'POST':
        try:
            nombre_fase = fase.nombre
            fase.delete()
            
            # Reordenar las fases restantes
            fases_restantes = plan.fases.all().order_by('orden')
            for idx, fase_restante in enumerate(fases_restantes, start=1):
                fase_restante.orden = idx
                fase_restante.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Fase "{nombre_fase}" eliminada exitosamente.'
            })
            
        except Exception as e:
            logger.error(f"Error al eliminar fase: {e}")
            return JsonResponse({'success': False, 'error': f'Error al eliminar la fase: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def crear_cita_desde_fase(request, plan_id, fase_id):
    """Crea una cita específica para una fase de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener la fase
    fase = get_object_or_404(FaseTratamiento, id=fase_id, plan=plan)
    
    # Validar que el plan esté en un estado que permita crear citas
    if request.method == 'POST':
        # Verificar que el plan tenga consentimiento firmado y presupuesto aceptado
        tiene_consentimiento_firmado = plan.consentimientos.filter(estado='firmado').exists()
        presupuesto_aceptado = plan.presupuesto_aceptado
        
        if not tiene_consentimiento_firmado:
            return JsonResponse({
                'success': False, 
                'error': 'No se pueden crear citas. El tratamiento debe tener al menos un consentimiento informado firmado.'
            }, status=400)
        
        if not presupuesto_aceptado:
            return JsonResponse({
                'success': False, 
                'error': 'No se pueden crear citas. El presupuesto del tratamiento debe estar aceptado.'
            }, status=400)
        
        # Verificar que el plan esté en un estado válido para crear citas
        if plan.estado not in ['aprobado', 'en_progreso']:
            return JsonResponse({
                'success': False, 
                'error': f'No se pueden crear citas para un tratamiento en estado "{plan.get_estado_display()}". El tratamiento debe estar aprobado o en progreso.'
            }, status=400)
        
        # Si el plan está aprobado pero no en progreso, cambiarlo automáticamente a en_progreso
        if plan.estado == 'aprobado':
            plan.estado = 'en_progreso'
            plan.save()
    
    if request.method == 'POST':
        fecha_hora_str = request.POST.get('fecha_hora', '')
        tipo_servicio_id = request.POST.get('tipo_servicio', '').strip()
        notas = request.POST.get('notas', '').strip()
        
        if not fecha_hora_str:
            return JsonResponse({'success': False, 'error': 'Debe seleccionar una fecha y hora.'}, status=400)
        
        try:
            # Convertir fecha y hacerla timezone-aware
            fecha_hora_naive = datetime.fromisoformat(fecha_hora_str)
            fecha_hora = timezone.make_aware(fecha_hora_naive)
            
            # Validar que la fecha no sea en el pasado
            ahora = timezone.now()
            if fecha_hora < ahora:
                return JsonResponse({'success': False, 'error': 'No se pueden crear citas en fechas pasadas.'}, status=400)
            
            # Obtener tipo de servicio si se proporcionó
            tipo_servicio = None
            if tipo_servicio_id:
                try:
                    tipo_servicio = TipoServicio.objects.get(id=tipo_servicio_id, activo=True)
                except TipoServicio.DoesNotExist:
                    pass
            
            # Validar horario del dentista
            dentista = plan.dentista
            dia_semana = fecha_hora.weekday()
            hora_cita = fecha_hora.time()
            
            horarios_dia = HorarioDentista.objects.filter(
                dentista=dentista,
                dia_semana=dia_semana,
                activo=True
            )
            
            if not horarios_dia.exists():
                dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
                return JsonResponse({'success': False, 'error': f'El dentista no trabaja los {dias_nombres[dia_semana]}.'}, status=400)
            
            hora_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita < horario.hora_fin:
                    hora_valida = True
                    break
            
            if not hora_valida:
                return JsonResponse({'success': False, 'error': 'La hora seleccionada no está dentro del horario de trabajo del dentista.'}, status=400)
            
            # Validar duración y solapamiento
            duracion_minutos = tipo_servicio.duracion_estimada if tipo_servicio and tipo_servicio.duracion_estimada else 30
            fecha_hora_fin = fecha_hora + timedelta(minutes=duracion_minutos)
            hora_fin_cita = fecha_hora_fin.time()
            
            duracion_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita and hora_fin_cita <= horario.hora_fin:
                    duracion_valida = True
                    break
            
            if not duracion_valida:
                return JsonResponse({'success': False, 'error': f'La duración del servicio ({duracion_minutos} minutos) no cabe en el horario seleccionado.'}, status=400)
            
            # Verificar solapamiento con otras citas
            citas_existentes = Cita.objects.filter(
                dentista=dentista,
                fecha_hora__date=fecha_hora.date(),
                estado__in=['disponible', 'reservada', 'confirmada', 'en_progreso']
            )
            
            for cita_existente in citas_existentes:
                fecha_hora_existente_fin = cita_existente.fecha_hora
                if cita_existente.tipo_servicio and cita_existente.tipo_servicio.duracion_estimada:
                    fecha_hora_existente_fin += timedelta(minutes=cita_existente.tipo_servicio.duracion_estimada)
                else:
                    fecha_hora_existente_fin += timedelta(minutes=30)
                
                if (fecha_hora < fecha_hora_existente_fin and fecha_hora_fin > cita_existente.fecha_hora):
                    return JsonResponse({'success': False, 'error': f'La cita se solapa con otra cita existente a las {cita_existente.fecha_hora.strftime("%H:%M")}.'}, status=400)
            
            # Obtener precio del servicio
            precio_cobrado = None
            if tipo_servicio:
                precio_cobrado = tipo_servicio.precio_base
            
            # Crear la cita asociada al plan y fase
            cita = Cita.objects.create(
                fecha_hora=fecha_hora,
                tipo_servicio=tipo_servicio,
                tipo_consulta=tipo_servicio.nombre if tipo_servicio else f"Cita - {fase.nombre}",
                precio_cobrado=precio_cobrado,
                notas=notas or f"Cita de la fase: {fase.nombre}",
                dentista=dentista,
                cliente=plan.cliente,
                paciente_nombre=plan.cliente.nombre_completo,
                paciente_email=plan.cliente.email,
                paciente_telefono=plan.cliente.telefono,
                estado='reservada',
                creada_por=perfil,
                plan_tratamiento=plan,
                fase_tratamiento=fase
            )
            
            # Enviar notificaciones (WhatsApp Y SMS) si tiene teléfono
            # Usar el teléfono del cliente del plan
            telefono_cliente = plan.cliente.telefono if plan.cliente else None
            if telefono_cliente:
                try:
                    from citas.mensajeria_service import enviar_notificaciones_cita
                    resultado = enviar_notificaciones_cita(cita, telefono_override=telefono_cliente)
                    canales_enviados = []
                    if resultado['whatsapp']['enviado']:
                        canales_enviados.append('WhatsApp')
                    if resultado['sms']['enviado']:
                        canales_enviados.append('SMS')
                    if canales_enviados:
                        logger.info(f"Notificaciones enviadas por {', '.join(canales_enviados)} para cita {cita.id} del plan {plan.id}")
                except Exception as e:
                    logger.error(f"Error al enviar notificaciones para cita {cita.id}: {e}")
            
            mensaje = f'Cita creada exitosamente para la fase "{fase.nombre}" el {fecha_hora.strftime("%d/%m/%Y a las %H:%M")}.'
            if plan.cliente.telefono:
                mensaje += ' Se ha enviado un SMS de confirmación.'
            
            return JsonResponse({
                'success': True,
                'message': mensaje,
                'cita_id': cita.id
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en el formato de fecha: {e}'}, status=400)
        except Exception as e:
            logger.error(f"Error al crear cita desde fase: {e}")
            return JsonResponse({'success': False, 'error': f'Error al crear la cita: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def cancelar_plan_tratamiento(request, plan_id):
    """Cancela un plan (DENTISTA puede cancelar sus planes)"""
    from django.http import JsonResponse
    
    try:
        perfil = Perfil.objects.get(user=request.user)
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
        messages.error(request, 'No tienes permisos.')
        return redirect('login')
    
    # Obtener plan con permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    else:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
        messages.error(request, 'No tienes permisos.')
        return redirect('panel_trabajador')
    
    if not plan.puede_ser_cancelado_por(perfil):
        error_msg = f'No puedes cancelar este plan en estado "{plan.get_estado_display()}".'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg}, status=400)
        messages.error(request, error_msg)
        return redirect('detalle_plan_tratamiento', plan_id=plan.id)
    
    if request.method == 'POST':
        motivo_cancelacion = request.POST.get('motivo_cancelacion', '').strip()
        
        if not motivo_cancelacion:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'El motivo de cancelación es obligatorio.'}, status=400)
            messages.error(request, 'El motivo de cancelación es obligatorio.')
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
        
        try:
            plan.estado = 'cancelado'
            plan.motivo_cancelacion = motivo_cancelacion
            plan.fecha_cancelacion = timezone.now()
            plan.save()
            
            success_msg = f'El tratamiento "{plan.nombre}" ha sido cancelado exitosamente. Quedará registrado como cancelado en el historial del cliente.'
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': success_msg
                })
            
            messages.success(request, success_msg)
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
        except Exception as e:
            logger.error(f"Error al cancelar plan {plan_id}: {e}")
            error_msg = 'Error al cancelar el plan. Por favor, intenta nuevamente.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg}, status=500)
            messages.error(request, error_msg)
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
    
    # GET: Mostrar formulario (si se accede directamente)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    return render(request, 'citas/planes_tratamiento/cancelar_plan_tratamiento.html', {'plan': plan, 'perfil': perfil})


@login_required
def crear_plan_desde_odontograma(request, odontograma_id):
    """Crea un plan de tratamiento basado en un odontograma"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos.')
        return redirect('login')
    
    if not (perfil.es_dentista() or perfil.es_administrativo()):
        messages.error(request, 'Solo dentistas y administrativos pueden crear planes.')
        return redirect('panel_trabajador')
    
    # Obtener odontograma
    if perfil.es_administrativo():
        odontograma = get_object_or_404(Odontograma, id=odontograma_id)
    else:
        odontograma = get_object_or_404(Odontograma, id=odontograma_id, dentista=perfil)
    
    # Redirigir a crear plan con datos pre-llenados
    return redirect(f"{reverse('crear_plan_tratamiento')}?odontograma_id={odontograma_id}&cliente_id={odontograma.cliente.id if odontograma.cliente else ''}")


@login_required
def registrar_pago_tratamiento(request, plan_id):
    """Registra un pago parcial para un plan de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    if request.method == 'POST':
        try:
            monto_str = request.POST.get('monto', '0').strip()
            fecha_pago_str = request.POST.get('fecha_pago', '')
            metodo_pago = request.POST.get('metodo_pago', '').strip()
            numero_comprobante = request.POST.get('numero_comprobante', '').strip()
            notas = request.POST.get('notas', '').strip()
            cita_id = request.POST.get('cita_id', '').strip()
            
            if not monto_str or float(monto_str) <= 0:
                return JsonResponse({'success': False, 'error': 'El monto debe ser mayor a 0.'}, status=400)
            
            if not fecha_pago_str:
                return JsonResponse({'success': False, 'error': 'La fecha de pago es requerida.'}, status=400)
            
            if not metodo_pago or metodo_pago not in ['efectivo', 'transferencia', 'tarjeta', 'cheque']:
                return JsonResponse({'success': False, 'error': 'El método de pago es requerido.'}, status=400)
            
            # Procesar monto
            monto_str = re.sub(r'[^\d.]', '', monto_str) or '0'
            monto = float(monto_str)
            
            # Validar que el monto no exceda el saldo pendiente
            saldo_pendiente = plan.saldo_pendiente
            if monto > saldo_pendiente:
                return JsonResponse({
                    'success': False,
                    'error': f'El monto excede el saldo pendiente (${saldo_pendiente:,.0f}).'
                }, status=400)
            
            # Procesar fecha
            fecha_pago = datetime.strptime(fecha_pago_str, '%Y-%m-%d').date()
            
            # Obtener cita si se proporcionó
            cita = None
            if cita_id:
                try:
                    cita = Cita.objects.get(id=cita_id, plan_tratamiento=plan)
                except Cita.DoesNotExist:
                    pass
            
            # Crear el pago
            pago = PagoTratamiento.objects.create(
                plan_tratamiento=plan,
                monto=monto,
                fecha_pago=fecha_pago,
                metodo_pago=metodo_pago,
                numero_comprobante=numero_comprobante if numero_comprobante else None,
                notas=notas if notas else None,
                cita=cita,
                registrado_por=perfil
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Pago de ${monto:,.0f} registrado exitosamente.',
                'pago': {
                    'id': pago.id,
                    'monto': str(pago.monto),
                    'fecha_pago': pago.fecha_pago.strftime('%d/%m/%Y'),
                    'metodo_pago': pago.get_metodo_pago_display(),
                },
                'total_pagado': str(plan.total_pagado),
                'saldo_pendiente': str(plan.saldo_pendiente),
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en los datos: {str(e)}'}, status=400)
        except Exception as e:
            logger.error(f"Error al registrar pago: {e}")
            return JsonResponse({'success': False, 'error': f'Error al registrar el pago: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def eliminar_pago_tratamiento(request, plan_id, pago_id):
    """Elimina un pago registrado (solo administrativos)"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({'success': False, 'error': 'Solo los administrativos pueden eliminar pagos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    plan = get_object_or_404(PlanTratamiento, id=plan_id)
    pago = get_object_or_404(PagoTratamiento, id=pago_id, plan_tratamiento=plan)
    
    if request.method == 'POST':
        try:
            monto = pago.monto
            pago.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Pago de ${monto:,.0f} eliminado exitosamente.',
                'total_pagado': str(plan.total_pagado),
                'saldo_pendiente': str(plan.saldo_pendiente),
            })
            
        except Exception as e:
            logger.error(f"Error al eliminar pago: {e}")
            return JsonResponse({'success': False, 'error': f'Error al eliminar el pago: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def enviar_documentos_tratamiento(request, plan_id):
    """Vista para enviar presupuesto y consentimientos de un tratamiento por correo"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not (perfil.es_administrativo() or perfil.es_dentista()):
            return JsonResponse({'success': False, 'error': 'No tienes permisos para enviar documentos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan de tratamiento
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    else:
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    # Verificar que el cliente tenga email
    if not plan.cliente.email:
        return JsonResponse({'success': False, 'error': 'El cliente no tiene un email registrado.'}, status=400)
    
    try:
        from django.core.mail import EmailMultiAlternatives
        from django.conf import settings
        from io import BytesIO
        from datetime import datetime
        
        # Obtener información de la clínica
        try:
            from configuracion.models import InformacionClinica
            info_clinica = InformacionClinica.obtener()
            nombre_clinica = info_clinica.nombre_clinica
            email_clinica = info_clinica.email or getattr(settings, 'EMAIL_FROM', 'noreply@clinica.com')
        except:
            nombre_clinica = "Clínica Dental"
            email_clinica = getattr(settings, 'EMAIL_FROM', 'noreply@clinica.com')
        
        # Generar PDF del presupuesto
        presupuesto_response = exportar_presupuesto_pdf(request, plan.id)
        presupuesto_pdf = presupuesto_response.content
        
        # Obtener consentimientos pendientes o no firmados
        consentimientos = plan.consentimientos.filter(estado__in=['pendiente', 'firmado']).order_by('-fecha_creacion')
        
        # Crear el email
        asunto = f"Documentos del Tratamiento - {plan.nombre} - {nombre_clinica}"
        
        mensaje_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: linear-gradient(135deg, #14b8a6 0%, #0d9488 100%); min-height: 100vh;">
            <div style="max-width: 600px; margin: 0 auto; padding: 40px 20px;">
                <!-- Header con logo -->
                <div style="background: white; border-radius: 12px 12px 0 0; padding: 30px; text-align: center; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    <div style="display: inline-flex; align-items: center; gap: 12px; margin-bottom: 10px;">
                        <div style="width: 50px; height: 50px; background: linear-gradient(135deg, #14b8a6, #0d9488); border-radius: 50%; display: flex; align-items: center; justify-content: center; color: white; font-size: 1.8rem;">
                            <span style="font-size: 1.8rem;">🦷</span>
                        </div>
                        <h1 style="margin: 0; color: #0f766e; font-size: 1.75rem; font-weight: 700;">Clínica San Felipe</h1>
                    </div>
                    <p style="margin: 0; color: #64748b; font-size: 0.9rem;">Victoria, Región de la Araucanía</p>
                </div>
                
                <!-- Contenido del correo -->
                <div style="background: white; padding: 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    <h2 style="color: #1e293b; margin-top: 0; font-size: 1.5rem;">Estimado/a {plan.cliente.nombre_completo},</h2>
                    
                    <p style="color: #475569; line-height: 1.6; font-size: 1rem;">Le enviamos los documentos relacionados con su tratamiento: <strong style="color: #0f766e;">{plan.nombre}</strong></p>
                    
                    <div style="background: linear-gradient(135deg, #f0fdfa, #e0f2f1); padding: 20px; border-radius: 8px; margin: 25px 0; border-left: 4px solid #14b8a6;">
                        <h3 style="margin-top: 0; color: #0f766e; font-size: 1.1rem; margin-bottom: 12px;">Documentos adjuntos:</h3>
                        <ul style="margin: 0; padding-left: 20px; color: #475569;">
                            <li style="margin-bottom: 8px;"><strong style="color: #0f766e;">Presupuesto del Tratamiento</strong></li>
        """
        
        if consentimientos.exists():
            mensaje_html += f"<li style='margin-bottom: 8px;'><strong style='color: #0f766e;'>Consentimientos Informados</strong> ({consentimientos.count()} documento(s))</li>"
        
        mensaje_html += f"""
                        </ul>
                    </div>
                    
                    <p style="color: #475569; line-height: 1.6;">Por favor, revise cuidadosamente estos documentos. Si tiene alguna consulta, no dude en contactarnos.</p>
                    
                    <div style="background: #fef3c7; padding: 18px; border-radius: 8px; margin: 25px 0; border-left: 4px solid #f59e0b;">
                        <p style="margin: 0 0 12px 0; color: #92400e; font-weight: 600; font-size: 1rem;"><strong>Importante:</strong> Para proceder con el tratamiento, necesitamos que:</p>
                        <ul style="margin: 0; padding-left: 20px; color: #92400e;">
                            <li style="margin-bottom: 6px;">Acepte el presupuesto</li>
                            <li style="margin-bottom: 6px;">Firme el consentimiento informado</li>
                        </ul>
                    </div>
                    
                    <p style="color: #475569; line-height: 1.6;">Puede hacerlo desde su panel de cliente en nuestra página web o contactándonos directamente.</p>
                </div>
                
                <!-- Footer -->
                <div style="background: white; border-radius: 0 0 12px 12px; padding: 25px 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border-top: 1px solid #e0f2f1;">
                    <p style="margin: 0; color: #64748b; font-size: 0.9rem; line-height: 1.6;">
                        Saludos cordiales,<br>
                        <strong style="color: #0f766e; font-size: 1rem;">{nombre_clinica}</strong>
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        mensaje_texto = f"""
        Estimado/a {plan.cliente.nombre_completo},
        
        Le enviamos los documentos relacionados con su tratamiento: {plan.nombre}
        
        Documentos adjuntos:
        - Presupuesto del Tratamiento
        - Consentimientos Informados
        
        Por favor, revise cuidadosamente estos documentos. Si tiene alguna consulta, no dude en contactarnos.
        
        Importante: Para proceder con el tratamiento, necesitamos que:
        - Acepte el presupuesto
        - Firme el consentimiento informado
        
        Puede hacerlo desde su panel de cliente en nuestra página web o contactándonos directamente.
        
        Saludos cordiales,
        {nombre_clinica}
        """
        
        # Crear el email
        email = EmailMultiAlternatives(
            asunto,
            mensaje_texto,
            email_clinica,
            [plan.cliente.email],
        )
        email.attach_alternative(mensaje_html, "text/html")
        
        # Adjuntar presupuesto
        filename_presupuesto = f"presupuesto_{plan.cliente.nombre_completo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        email.attach(filename_presupuesto, presupuesto_pdf, 'application/pdf')
        
        # Adjuntar consentimientos
        for consentimiento in consentimientos:
            try:
                consentimiento_response = exportar_consentimiento_pdf(request, consentimiento.id)
                consentimiento_pdf = consentimiento_response.content
                filename_consentimiento = f"consentimiento_{consentimiento.titulo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
                email.attach(filename_consentimiento, consentimiento_pdf, 'application/pdf')
            except Exception as e:
                logger.error(f"Error al generar PDF del consentimiento {consentimiento.id}: {e}")
                # Continuar con los demás documentos
        
        # Enviar email
        email.send()
        
        # Actualizar fecha de envío en documentos relacionados
        from historial_clinico.models import DocumentoCliente
        documentos_actualizados = 0
        for consentimiento in consentimientos:
            # Actualizar el documento relacionado si existe
            try:
                documento = DocumentoCliente.objects.filter(
                    plan_tratamiento=plan,
                    tipo='consentimiento',
                    cliente=plan.cliente
                ).filter(
                    titulo__icontains=consentimiento.titulo
                ).first()
                
                if documento:
                    documento.enviado_por_correo = True
                    documento.fecha_envio = timezone.now()
                    documento.email_destinatario = plan.cliente.email
                    documento.save(update_fields=['enviado_por_correo', 'fecha_envio', 'email_destinatario'])
                    documentos_actualizados += 1
            except Exception as e:
                logger.warning(f"No se pudo actualizar documento para consentimiento {consentimiento.id}: {e}")
        
        # Actualizar también el documento del presupuesto si existe
        try:
            doc_presupuesto = DocumentoCliente.objects.filter(
                plan_tratamiento=plan,
                tipo='presupuesto',
                cliente=plan.cliente
            ).first()
            
            if doc_presupuesto:
                doc_presupuesto.enviado_por_correo = True
                doc_presupuesto.fecha_envio = timezone.now()
                doc_presupuesto.email_destinatario = plan.cliente.email
                doc_presupuesto.save(update_fields=['enviado_por_correo', 'fecha_envio', 'email_destinatario'])
                documentos_actualizados += 1
        except Exception as e:
            logger.warning(f"No se pudo actualizar documento de presupuesto: {e}")
        
        return JsonResponse({
            'success': True,
            'message': f'Documentos enviados exitosamente a {plan.cliente.email}. Se enviaron {1 + consentimientos.count()} documento(s).'
        })
        
    except Exception as e:
        logger.error(f"Error al enviar documentos del tratamiento: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'error': f'Error al enviar los documentos: {str(e)}'}, status=500)

    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener la fase
    fase = get_object_or_404(FaseTratamiento, id=fase_id, plan=plan)
    
    # Verificar que el plan puede ser editado
    if not plan.puede_ser_editado_por(perfil):
        return JsonResponse({'success': False, 'error': f'No puedes eliminar fases de este plan en estado "{plan.get_estado_display()}".'}, status=400)
    
    # Verificar si la fase tiene citas asociadas
    citas_asociadas = Cita.objects.filter(fase_tratamiento=fase).count()
    if citas_asociadas > 0:
        return JsonResponse({
            'success': False,
            'error': f'No se puede eliminar la fase porque tiene {citas_asociadas} cita(s) asociada(s).'
        }, status=400)
    
    if request.method == 'POST':
        try:
            nombre_fase = fase.nombre
            fase.delete()
            
            # Reordenar las fases restantes
            fases_restantes = plan.fases.all().order_by('orden')
            for idx, fase_restante in enumerate(fases_restantes, start=1):
                fase_restante.orden = idx
                fase_restante.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Fase "{nombre_fase}" eliminada exitosamente.'
            })
            
        except Exception as e:
            logger.error(f"Error al eliminar fase: {e}")
            return JsonResponse({'success': False, 'error': f'Error al eliminar la fase: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def crear_cita_desde_fase(request, plan_id, fase_id):
    """Crea una cita específica para una fase de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener la fase
    fase = get_object_or_404(FaseTratamiento, id=fase_id, plan=plan)
    
    # Validar que el plan esté en un estado que permita crear citas
    if request.method == 'POST':
        # Verificar que el plan tenga consentimiento firmado y presupuesto aceptado
        tiene_consentimiento_firmado = plan.consentimientos.filter(estado='firmado').exists()
        presupuesto_aceptado = plan.presupuesto_aceptado
        
        if not tiene_consentimiento_firmado:
            return JsonResponse({
                'success': False, 
                'error': 'No se pueden crear citas. El tratamiento debe tener al menos un consentimiento informado firmado.'
            }, status=400)
        
        if not presupuesto_aceptado:
            return JsonResponse({
                'success': False, 
                'error': 'No se pueden crear citas. El presupuesto del tratamiento debe estar aceptado.'
            }, status=400)
        
        # Verificar que el plan esté en un estado válido para crear citas
        if plan.estado not in ['aprobado', 'en_progreso']:
            return JsonResponse({
                'success': False, 
                'error': f'No se pueden crear citas para un tratamiento en estado "{plan.get_estado_display()}". El tratamiento debe estar aprobado o en progreso.'
            }, status=400)
        
        # Si el plan está aprobado pero no en progreso, cambiarlo automáticamente a en_progreso
        if plan.estado == 'aprobado':
            plan.estado = 'en_progreso'
            plan.save()
    
    if request.method == 'POST':
        fecha_hora_str = request.POST.get('fecha_hora', '')
        tipo_servicio_id = request.POST.get('tipo_servicio', '').strip()
        notas = request.POST.get('notas', '').strip()
        
        if not fecha_hora_str:
            return JsonResponse({'success': False, 'error': 'Debe seleccionar una fecha y hora.'}, status=400)
        
        try:
            # Convertir fecha y hacerla timezone-aware
            fecha_hora_naive = datetime.fromisoformat(fecha_hora_str)
            fecha_hora = timezone.make_aware(fecha_hora_naive)
            
            # Validar que la fecha no sea en el pasado
            ahora = timezone.now()
            if fecha_hora < ahora:
                return JsonResponse({'success': False, 'error': 'No se pueden crear citas en fechas pasadas.'}, status=400)
            
            # Obtener tipo de servicio si se proporcionó
            tipo_servicio = None
            if tipo_servicio_id:
                try:
                    tipo_servicio = TipoServicio.objects.get(id=tipo_servicio_id, activo=True)
                except TipoServicio.DoesNotExist:
                    pass
            
            # Validar horario del dentista
            dentista = plan.dentista
            dia_semana = fecha_hora.weekday()
            hora_cita = fecha_hora.time()
            
            horarios_dia = HorarioDentista.objects.filter(
                dentista=dentista,
                dia_semana=dia_semana,
                activo=True
            )
            
            if not horarios_dia.exists():
                dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
                return JsonResponse({'success': False, 'error': f'El dentista no trabaja los {dias_nombres[dia_semana]}.'}, status=400)
            
            hora_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita < horario.hora_fin:
                    hora_valida = True
                    break
            
            if not hora_valida:
                return JsonResponse({'success': False, 'error': 'La hora seleccionada no está dentro del horario de trabajo del dentista.'}, status=400)
            
            # Validar duración y solapamiento
            duracion_minutos = tipo_servicio.duracion_estimada if tipo_servicio and tipo_servicio.duracion_estimada else 30
            fecha_hora_fin = fecha_hora + timedelta(minutes=duracion_minutos)
            hora_fin_cita = fecha_hora_fin.time()
            
            duracion_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita and hora_fin_cita <= horario.hora_fin:
                    duracion_valida = True
                    break
            
            if not duracion_valida:
                return JsonResponse({'success': False, 'error': f'La duración del servicio ({duracion_minutos} minutos) no cabe en el horario seleccionado.'}, status=400)
            
            # Verificar solapamiento con otras citas
            citas_existentes = Cita.objects.filter(
                dentista=dentista,
                fecha_hora__date=fecha_hora.date(),
                estado__in=['disponible', 'reservada', 'confirmada', 'en_progreso']
            )
            
            for cita_existente in citas_existentes:
                fecha_hora_existente_fin = cita_existente.fecha_hora
                if cita_existente.tipo_servicio and cita_existente.tipo_servicio.duracion_estimada:
                    fecha_hora_existente_fin += timedelta(minutes=cita_existente.tipo_servicio.duracion_estimada)
                else:
                    fecha_hora_existente_fin += timedelta(minutes=30)
                
                if (fecha_hora < fecha_hora_existente_fin and fecha_hora_fin > cita_existente.fecha_hora):
                    return JsonResponse({'success': False, 'error': f'La cita se solapa con otra cita existente a las {cita_existente.fecha_hora.strftime("%H:%M")}.'}, status=400)
            
            # Obtener precio del servicio
            precio_cobrado = None
            if tipo_servicio:
                precio_cobrado = tipo_servicio.precio_base
            
            # Crear la cita asociada al plan y fase
            cita = Cita.objects.create(
                fecha_hora=fecha_hora,
                tipo_servicio=tipo_servicio,
                tipo_consulta=tipo_servicio.nombre if tipo_servicio else f"Cita - {fase.nombre}",
                precio_cobrado=precio_cobrado,
                notas=notas or f"Cita de la fase: {fase.nombre}",
                dentista=dentista,
                cliente=plan.cliente,
                paciente_nombre=plan.cliente.nombre_completo,
                paciente_email=plan.cliente.email,
                paciente_telefono=plan.cliente.telefono,
                estado='reservada',
                creada_por=perfil,
                plan_tratamiento=plan,
                fase_tratamiento=fase
            )
            
            # Enviar notificaciones (WhatsApp Y SMS) si tiene teléfono
            # Usar el teléfono del cliente del plan
            telefono_cliente = plan.cliente.telefono if plan.cliente else None
            if telefono_cliente:
                try:
                    from citas.mensajeria_service import enviar_notificaciones_cita
                    resultado = enviar_notificaciones_cita(cita, telefono_override=telefono_cliente)
                    canales_enviados = []
                    if resultado['whatsapp']['enviado']:
                        canales_enviados.append('WhatsApp')
                    if resultado['sms']['enviado']:
                        canales_enviados.append('SMS')
                    if canales_enviados:
                        logger.info(f"Notificaciones enviadas por {', '.join(canales_enviados)} para cita {cita.id} del plan {plan.id}")
                except Exception as e:
                    logger.error(f"Error al enviar notificaciones para cita {cita.id}: {e}")
            
            mensaje = f'Cita creada exitosamente para la fase "{fase.nombre}" el {fecha_hora.strftime("%d/%m/%Y a las %H:%M")}.'
            if plan.cliente.telefono:
                mensaje += ' Se ha enviado un SMS de confirmación.'
            
            return JsonResponse({
                'success': True,
                'message': mensaje,
                'cita_id': cita.id
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en el formato de fecha: {e}'}, status=400)
        except Exception as e:
            logger.error(f"Error al crear cita desde fase: {e}")
            return JsonResponse({'success': False, 'error': f'Error al crear la cita: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def cancelar_plan_tratamiento(request, plan_id):
    """Cancela un plan (DENTISTA puede cancelar sus planes)"""
    from django.http import JsonResponse
    
    try:
        perfil = Perfil.objects.get(user=request.user)
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
        messages.error(request, 'No tienes permisos.')
        return redirect('login')
    
    # Obtener plan con permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    else:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
        messages.error(request, 'No tienes permisos.')
        return redirect('panel_trabajador')
    
    if not plan.puede_ser_cancelado_por(perfil):
        error_msg = f'No puedes cancelar este plan en estado "{plan.get_estado_display()}".'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg}, status=400)
        messages.error(request, error_msg)
        return redirect('detalle_plan_tratamiento', plan_id=plan.id)
    
    if request.method == 'POST':
        motivo_cancelacion = request.POST.get('motivo_cancelacion', '').strip()
        
        if not motivo_cancelacion:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'El motivo de cancelación es obligatorio.'}, status=400)
            messages.error(request, 'El motivo de cancelación es obligatorio.')
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
        
        try:
            plan.estado = 'cancelado'
            plan.motivo_cancelacion = motivo_cancelacion
            plan.fecha_cancelacion = timezone.now()
            plan.save()
            
            success_msg = f'El tratamiento "{plan.nombre}" ha sido cancelado exitosamente. Quedará registrado como cancelado en el historial del cliente.'
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': success_msg
                })
            
            messages.success(request, success_msg)
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
        except Exception as e:
            logger.error(f"Error al cancelar plan {plan_id}: {e}")
            error_msg = 'Error al cancelar el plan. Por favor, intenta nuevamente.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg}, status=500)
            messages.error(request, error_msg)
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
    
    # GET: Mostrar formulario (si se accede directamente)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    return render(request, 'citas/planes_tratamiento/cancelar_plan_tratamiento.html', {'plan': plan, 'perfil': perfil})


@login_required
def crear_plan_desde_odontograma(request, odontograma_id):
    """Crea un plan de tratamiento basado en un odontograma"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos.')
        return redirect('login')
    
    if not (perfil.es_dentista() or perfil.es_administrativo()):
        messages.error(request, 'Solo dentistas y administrativos pueden crear planes.')
        return redirect('panel_trabajador')
    
    # Obtener odontograma
    if perfil.es_administrativo():
        odontograma = get_object_or_404(Odontograma, id=odontograma_id)
    else:
        odontograma = get_object_or_404(Odontograma, id=odontograma_id, dentista=perfil)
    
    # Redirigir a crear plan con datos pre-llenados
    return redirect(f"{reverse('crear_plan_tratamiento')}?odontograma_id={odontograma_id}&cliente_id={odontograma.cliente.id if odontograma.cliente else ''}")


@login_required
def registrar_pago_tratamiento(request, plan_id):
    """Registra un pago parcial para un plan de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    if request.method == 'POST':
        try:
            monto_str = request.POST.get('monto', '0').strip()
            fecha_pago_str = request.POST.get('fecha_pago', '')
            metodo_pago = request.POST.get('metodo_pago', '').strip()
            numero_comprobante = request.POST.get('numero_comprobante', '').strip()
            notas = request.POST.get('notas', '').strip()
            cita_id = request.POST.get('cita_id', '').strip()
            
            if not monto_str or float(monto_str) <= 0:
                return JsonResponse({'success': False, 'error': 'El monto debe ser mayor a 0.'}, status=400)
            
            if not fecha_pago_str:
                return JsonResponse({'success': False, 'error': 'La fecha de pago es requerida.'}, status=400)
            
            if not metodo_pago or metodo_pago not in ['efectivo', 'transferencia', 'tarjeta', 'cheque']:
                return JsonResponse({'success': False, 'error': 'El método de pago es requerido.'}, status=400)
            
            # Procesar monto
            monto_str = re.sub(r'[^\d.]', '', monto_str) or '0'
            monto = float(monto_str)
            
            # Validar que el monto no exceda el saldo pendiente
            saldo_pendiente = plan.saldo_pendiente
            if monto > saldo_pendiente:
                return JsonResponse({
                    'success': False,
                    'error': f'El monto excede el saldo pendiente (${saldo_pendiente:,.0f}).'
                }, status=400)
            
            # Procesar fecha
            fecha_pago = datetime.strptime(fecha_pago_str, '%Y-%m-%d').date()
            
            # Obtener cita si se proporcionó
            cita = None
            if cita_id:
                try:
                    cita = Cita.objects.get(id=cita_id, plan_tratamiento=plan)
                except Cita.DoesNotExist:
                    pass
            
            # Crear el pago
            pago = PagoTratamiento.objects.create(
                plan_tratamiento=plan,
                monto=monto,
                fecha_pago=fecha_pago,
                metodo_pago=metodo_pago,
                numero_comprobante=numero_comprobante if numero_comprobante else None,
                notas=notas if notas else None,
                cita=cita,
                registrado_por=perfil
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Pago de ${monto:,.0f} registrado exitosamente.',
                'pago': {
                    'id': pago.id,
                    'monto': str(pago.monto),
                    'fecha_pago': pago.fecha_pago.strftime('%d/%m/%Y'),
                    'metodo_pago': pago.get_metodo_pago_display(),
                },
                'total_pagado': str(plan.total_pagado),
                'saldo_pendiente': str(plan.saldo_pendiente),
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en los datos: {str(e)}'}, status=400)
        except Exception as e:
            logger.error(f"Error al registrar pago: {e}")
            return JsonResponse({'success': False, 'error': f'Error al registrar el pago: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def eliminar_pago_tratamiento(request, plan_id, pago_id):
    """Elimina un pago registrado (solo administrativos)"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({'success': False, 'error': 'Solo los administrativos pueden eliminar pagos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    plan = get_object_or_404(PlanTratamiento, id=plan_id)
    pago = get_object_or_404(PagoTratamiento, id=pago_id, plan_tratamiento=plan)
    
    if request.method == 'POST':
        try:
            monto = pago.monto
            pago.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Pago de ${monto:,.0f} eliminado exitosamente.',
                'total_pagado': str(plan.total_pagado),
                'saldo_pendiente': str(plan.saldo_pendiente),
            })
            
        except Exception as e:
            logger.error(f"Error al eliminar pago: {e}")
            return JsonResponse({'success': False, 'error': f'Error al eliminar el pago: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def enviar_documentos_tratamiento(request, plan_id):
    """Vista para enviar presupuesto y consentimientos de un tratamiento por correo"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not (perfil.es_administrativo() or perfil.es_dentista()):
            return JsonResponse({'success': False, 'error': 'No tienes permisos para enviar documentos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan de tratamiento
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    else:
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    # Verificar que el cliente tenga email
    if not plan.cliente.email:
        return JsonResponse({'success': False, 'error': 'El cliente no tiene un email registrado.'}, status=400)
    
    try:
        from django.core.mail import EmailMultiAlternatives
        from django.conf import settings
        from io import BytesIO
        from datetime import datetime
        
        # Obtener información de la clínica
        try:
            from configuracion.models import InformacionClinica
            info_clinica = InformacionClinica.obtener()
            nombre_clinica = info_clinica.nombre_clinica
            email_clinica = info_clinica.email or getattr(settings, 'EMAIL_FROM', 'noreply@clinica.com')
        except:
            nombre_clinica = "Clínica Dental"
            email_clinica = getattr(settings, 'EMAIL_FROM', 'noreply@clinica.com')
        
        # Generar PDF del presupuesto
        presupuesto_response = exportar_presupuesto_pdf(request, plan.id)
        presupuesto_pdf = presupuesto_response.content
        
        # Obtener consentimientos pendientes o no firmados
        consentimientos = plan.consentimientos.filter(estado__in=['pendiente', 'firmado']).order_by('-fecha_creacion')
        
        # Crear el email
        asunto = f"Documentos del Tratamiento - {plan.nombre} - {nombre_clinica}"
        
        mensaje_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: linear-gradient(135deg, #14b8a6 0%, #0d9488 100%); min-height: 100vh;">
            <div style="max-width: 600px; margin: 0 auto; padding: 40px 20px;">
                <!-- Header con logo -->
                <div style="background: white; border-radius: 12px 12px 0 0; padding: 30px; text-align: center; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    <div style="display: inline-flex; align-items: center; gap: 12px; margin-bottom: 10px;">
                        <div style="width: 50px; height: 50px; background: linear-gradient(135deg, #14b8a6, #0d9488); border-radius: 50%; display: flex; align-items: center; justify-content: center; color: white; font-size: 1.8rem;">
                            <span style="font-size: 1.8rem;">🦷</span>
                        </div>
                        <h1 style="margin: 0; color: #0f766e; font-size: 1.75rem; font-weight: 700;">Clínica San Felipe</h1>
                    </div>
                    <p style="margin: 0; color: #64748b; font-size: 0.9rem;">Victoria, Región de la Araucanía</p>
                </div>
                
                <!-- Contenido del correo -->
                <div style="background: white; padding: 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    <h2 style="color: #1e293b; margin-top: 0; font-size: 1.5rem;">Estimado/a {plan.cliente.nombre_completo},</h2>
                    
                    <p style="color: #475569; line-height: 1.6; font-size: 1rem;">Le enviamos los documentos relacionados con su tratamiento: <strong style="color: #0f766e;">{plan.nombre}</strong></p>
                    
                    <div style="background: linear-gradient(135deg, #f0fdfa, #e0f2f1); padding: 20px; border-radius: 8px; margin: 25px 0; border-left: 4px solid #14b8a6;">
                        <h3 style="margin-top: 0; color: #0f766e; font-size: 1.1rem; margin-bottom: 12px;">Documentos adjuntos:</h3>
                        <ul style="margin: 0; padding-left: 20px; color: #475569;">
                            <li style="margin-bottom: 8px;"><strong style="color: #0f766e;">Presupuesto del Tratamiento</strong></li>
        """
        
        if consentimientos.exists():
            mensaje_html += f"<li style='margin-bottom: 8px;'><strong style='color: #0f766e;'>Consentimientos Informados</strong> ({consentimientos.count()} documento(s))</li>"
        
        mensaje_html += f"""
                        </ul>
                    </div>
                    
                    <p style="color: #475569; line-height: 1.6;">Por favor, revise cuidadosamente estos documentos. Si tiene alguna consulta, no dude en contactarnos.</p>
                    
                    <div style="background: #fef3c7; padding: 18px; border-radius: 8px; margin: 25px 0; border-left: 4px solid #f59e0b;">
                        <p style="margin: 0 0 12px 0; color: #92400e; font-weight: 600; font-size: 1rem;"><strong>Importante:</strong> Para proceder con el tratamiento, necesitamos que:</p>
                        <ul style="margin: 0; padding-left: 20px; color: #92400e;">
                            <li style="margin-bottom: 6px;">Acepte el presupuesto</li>
                            <li style="margin-bottom: 6px;">Firme el consentimiento informado</li>
                        </ul>
                    </div>
                    
                    <p style="color: #475569; line-height: 1.6;">Puede hacerlo desde su panel de cliente en nuestra página web o contactándonos directamente.</p>
                </div>
                
                <!-- Footer -->
                <div style="background: white; border-radius: 0 0 12px 12px; padding: 25px 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border-top: 1px solid #e0f2f1;">
                    <p style="margin: 0; color: #64748b; font-size: 0.9rem; line-height: 1.6;">
                        Saludos cordiales,<br>
                        <strong style="color: #0f766e; font-size: 1rem;">{nombre_clinica}</strong>
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        mensaje_texto = f"""
        Estimado/a {plan.cliente.nombre_completo},
        
        Le enviamos los documentos relacionados con su tratamiento: {plan.nombre}
        
        Documentos adjuntos:
        - Presupuesto del Tratamiento
        - Consentimientos Informados
        
        Por favor, revise cuidadosamente estos documentos. Si tiene alguna consulta, no dude en contactarnos.
        
        Importante: Para proceder con el tratamiento, necesitamos que:
        - Acepte el presupuesto
        - Firme el consentimiento informado
        
        Puede hacerlo desde su panel de cliente en nuestra página web o contactándonos directamente.
        
        Saludos cordiales,
        {nombre_clinica}
        """
        
        # Crear el email
        email = EmailMultiAlternatives(
            asunto,
            mensaje_texto,
            email_clinica,
            [plan.cliente.email],
        )
        email.attach_alternative(mensaje_html, "text/html")
        
        # Adjuntar presupuesto
        filename_presupuesto = f"presupuesto_{plan.cliente.nombre_completo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        email.attach(filename_presupuesto, presupuesto_pdf, 'application/pdf')
        
        # Adjuntar consentimientos
        for consentimiento in consentimientos:
            try:
                consentimiento_response = exportar_consentimiento_pdf(request, consentimiento.id)
                consentimiento_pdf = consentimiento_response.content
                filename_consentimiento = f"consentimiento_{consentimiento.titulo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
                email.attach(filename_consentimiento, consentimiento_pdf, 'application/pdf')
            except Exception as e:
                logger.error(f"Error al generar PDF del consentimiento {consentimiento.id}: {e}")
                # Continuar con los demás documentos
        
        # Enviar email
        email.send()
        
        # Actualizar fecha de envío en documentos relacionados
        from historial_clinico.models import DocumentoCliente
        documentos_actualizados = 0
        for consentimiento in consentimientos:
            # Actualizar el documento relacionado si existe
            try:
                documento = DocumentoCliente.objects.filter(
                    plan_tratamiento=plan,
                    tipo='consentimiento',
                    cliente=plan.cliente
                ).filter(
                    titulo__icontains=consentimiento.titulo
                ).first()
                
                if documento:
                    documento.enviado_por_correo = True
                    documento.fecha_envio = timezone.now()
                    documento.email_destinatario = plan.cliente.email
                    documento.save(update_fields=['enviado_por_correo', 'fecha_envio', 'email_destinatario'])
                    documentos_actualizados += 1
            except Exception as e:
                logger.warning(f"No se pudo actualizar documento para consentimiento {consentimiento.id}: {e}")
        
        # Actualizar también el documento del presupuesto si existe
        try:
            doc_presupuesto = DocumentoCliente.objects.filter(
                plan_tratamiento=plan,
                tipo='presupuesto',
                cliente=plan.cliente
            ).first()
            
            if doc_presupuesto:
                doc_presupuesto.enviado_por_correo = True
                doc_presupuesto.fecha_envio = timezone.now()
                doc_presupuesto.email_destinatario = plan.cliente.email
                doc_presupuesto.save(update_fields=['enviado_por_correo', 'fecha_envio', 'email_destinatario'])
                documentos_actualizados += 1
        except Exception as e:
            logger.warning(f"No se pudo actualizar documento de presupuesto: {e}")
        
        return JsonResponse({
            'success': True,
            'message': f'Documentos enviados exitosamente a {plan.cliente.email}. Se enviaron {1 + consentimientos.count()} documento(s).'
        })
        
    except Exception as e:
        logger.error(f"Error al enviar documentos del tratamiento: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'error': f'Error al enviar los documentos: {str(e)}'}, status=500)

    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener la fase
    fase = get_object_or_404(FaseTratamiento, id=fase_id, plan=plan)
    
    # Verificar que el plan puede ser editado
    if not plan.puede_ser_editado_por(perfil):
        return JsonResponse({'success': False, 'error': f'No puedes eliminar fases de este plan en estado "{plan.get_estado_display()}".'}, status=400)
    
    # Verificar si la fase tiene citas asociadas
    citas_asociadas = Cita.objects.filter(fase_tratamiento=fase).count()
    if citas_asociadas > 0:
        return JsonResponse({
            'success': False,
            'error': f'No se puede eliminar la fase porque tiene {citas_asociadas} cita(s) asociada(s).'
        }, status=400)
    
    if request.method == 'POST':
        try:
            nombre_fase = fase.nombre
            fase.delete()
            
            # Reordenar las fases restantes
            fases_restantes = plan.fases.all().order_by('orden')
            for idx, fase_restante in enumerate(fases_restantes, start=1):
                fase_restante.orden = idx
                fase_restante.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Fase "{nombre_fase}" eliminada exitosamente.'
            })
            
        except Exception as e:
            logger.error(f"Error al eliminar fase: {e}")
            return JsonResponse({'success': False, 'error': f'Error al eliminar la fase: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def crear_cita_desde_fase(request, plan_id, fase_id):
    """Crea una cita específica para una fase de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener la fase
    fase = get_object_or_404(FaseTratamiento, id=fase_id, plan=plan)
    
    # Validar que el plan esté en un estado que permita crear citas
    if request.method == 'POST':
        # Verificar que el plan tenga consentimiento firmado y presupuesto aceptado
        tiene_consentimiento_firmado = plan.consentimientos.filter(estado='firmado').exists()
        presupuesto_aceptado = plan.presupuesto_aceptado
        
        if not tiene_consentimiento_firmado:
            return JsonResponse({
                'success': False, 
                'error': 'No se pueden crear citas. El tratamiento debe tener al menos un consentimiento informado firmado.'
            }, status=400)
        
        if not presupuesto_aceptado:
            return JsonResponse({
                'success': False, 
                'error': 'No se pueden crear citas. El presupuesto del tratamiento debe estar aceptado.'
            }, status=400)
        
        # Verificar que el plan esté en un estado válido para crear citas
        if plan.estado not in ['aprobado', 'en_progreso']:
            return JsonResponse({
                'success': False, 
                'error': f'No se pueden crear citas para un tratamiento en estado "{plan.get_estado_display()}". El tratamiento debe estar aprobado o en progreso.'
            }, status=400)
        
        # Si el plan está aprobado pero no en progreso, cambiarlo automáticamente a en_progreso
        if plan.estado == 'aprobado':
            plan.estado = 'en_progreso'
            plan.save()
    
    if request.method == 'POST':
        fecha_hora_str = request.POST.get('fecha_hora', '')
        tipo_servicio_id = request.POST.get('tipo_servicio', '').strip()
        notas = request.POST.get('notas', '').strip()
        
        if not fecha_hora_str:
            return JsonResponse({'success': False, 'error': 'Debe seleccionar una fecha y hora.'}, status=400)
        
        try:
            # Convertir fecha y hacerla timezone-aware
            fecha_hora_naive = datetime.fromisoformat(fecha_hora_str)
            fecha_hora = timezone.make_aware(fecha_hora_naive)
            
            # Validar que la fecha no sea en el pasado
            ahora = timezone.now()
            if fecha_hora < ahora:
                return JsonResponse({'success': False, 'error': 'No se pueden crear citas en fechas pasadas.'}, status=400)
            
            # Obtener tipo de servicio si se proporcionó
            tipo_servicio = None
            if tipo_servicio_id:
                try:
                    tipo_servicio = TipoServicio.objects.get(id=tipo_servicio_id, activo=True)
                except TipoServicio.DoesNotExist:
                    pass
            
            # Validar horario del dentista
            dentista = plan.dentista
            dia_semana = fecha_hora.weekday()
            hora_cita = fecha_hora.time()
            
            horarios_dia = HorarioDentista.objects.filter(
                dentista=dentista,
                dia_semana=dia_semana,
                activo=True
            )
            
            if not horarios_dia.exists():
                dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
                return JsonResponse({'success': False, 'error': f'El dentista no trabaja los {dias_nombres[dia_semana]}.'}, status=400)
            
            hora_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita < horario.hora_fin:
                    hora_valida = True
                    break
            
            if not hora_valida:
                return JsonResponse({'success': False, 'error': 'La hora seleccionada no está dentro del horario de trabajo del dentista.'}, status=400)
            
            # Validar duración y solapamiento
            duracion_minutos = tipo_servicio.duracion_estimada if tipo_servicio and tipo_servicio.duracion_estimada else 30
            fecha_hora_fin = fecha_hora + timedelta(minutes=duracion_minutos)
            hora_fin_cita = fecha_hora_fin.time()
            
            duracion_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita and hora_fin_cita <= horario.hora_fin:
                    duracion_valida = True
                    break
            
            if not duracion_valida:
                return JsonResponse({'success': False, 'error': f'La duración del servicio ({duracion_minutos} minutos) no cabe en el horario seleccionado.'}, status=400)
            
            # Verificar solapamiento con otras citas
            citas_existentes = Cita.objects.filter(
                dentista=dentista,
                fecha_hora__date=fecha_hora.date(),
                estado__in=['disponible', 'reservada', 'confirmada', 'en_progreso']
            )
            
            for cita_existente in citas_existentes:
                fecha_hora_existente_fin = cita_existente.fecha_hora
                if cita_existente.tipo_servicio and cita_existente.tipo_servicio.duracion_estimada:
                    fecha_hora_existente_fin += timedelta(minutes=cita_existente.tipo_servicio.duracion_estimada)
                else:
                    fecha_hora_existente_fin += timedelta(minutes=30)
                
                if (fecha_hora < fecha_hora_existente_fin and fecha_hora_fin > cita_existente.fecha_hora):
                    return JsonResponse({'success': False, 'error': f'La cita se solapa con otra cita existente a las {cita_existente.fecha_hora.strftime("%H:%M")}.'}, status=400)
            
            # Obtener precio del servicio
            precio_cobrado = None
            if tipo_servicio:
                precio_cobrado = tipo_servicio.precio_base
            
            # Crear la cita asociada al plan y fase
            cita = Cita.objects.create(
                fecha_hora=fecha_hora,
                tipo_servicio=tipo_servicio,
                tipo_consulta=tipo_servicio.nombre if tipo_servicio else f"Cita - {fase.nombre}",
                precio_cobrado=precio_cobrado,
                notas=notas or f"Cita de la fase: {fase.nombre}",
                dentista=dentista,
                cliente=plan.cliente,
                paciente_nombre=plan.cliente.nombre_completo,
                paciente_email=plan.cliente.email,
                paciente_telefono=plan.cliente.telefono,
                estado='reservada',
                creada_por=perfil,
                plan_tratamiento=plan,
                fase_tratamiento=fase
            )
            
            # Enviar notificaciones (WhatsApp Y SMS) si tiene teléfono
            # Usar el teléfono del cliente del plan
            telefono_cliente = plan.cliente.telefono if plan.cliente else None
            if telefono_cliente:
                try:
                    from citas.mensajeria_service import enviar_notificaciones_cita
                    resultado = enviar_notificaciones_cita(cita, telefono_override=telefono_cliente)
                    canales_enviados = []
                    if resultado['whatsapp']['enviado']:
                        canales_enviados.append('WhatsApp')
                    if resultado['sms']['enviado']:
                        canales_enviados.append('SMS')
                    if canales_enviados:
                        logger.info(f"Notificaciones enviadas por {', '.join(canales_enviados)} para cita {cita.id} del plan {plan.id}")
                except Exception as e:
                    logger.error(f"Error al enviar notificaciones para cita {cita.id}: {e}")
            
            mensaje = f'Cita creada exitosamente para la fase "{fase.nombre}" el {fecha_hora.strftime("%d/%m/%Y a las %H:%M")}.'
            if plan.cliente.telefono:
                mensaje += ' Se ha enviado un SMS de confirmación.'
            
            return JsonResponse({
                'success': True,
                'message': mensaje,
                'cita_id': cita.id
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en el formato de fecha: {e}'}, status=400)
        except Exception as e:
            logger.error(f"Error al crear cita desde fase: {e}")
            return JsonResponse({'success': False, 'error': f'Error al crear la cita: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def cancelar_plan_tratamiento(request, plan_id):
    """Cancela un plan (DENTISTA puede cancelar sus planes)"""
    from django.http import JsonResponse
    
    try:
        perfil = Perfil.objects.get(user=request.user)
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
        messages.error(request, 'No tienes permisos.')
        return redirect('login')
    
    # Obtener plan con permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    else:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
        messages.error(request, 'No tienes permisos.')
        return redirect('panel_trabajador')
    
    if not plan.puede_ser_cancelado_por(perfil):
        error_msg = f'No puedes cancelar este plan en estado "{plan.get_estado_display()}".'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg}, status=400)
        messages.error(request, error_msg)
        return redirect('detalle_plan_tratamiento', plan_id=plan.id)
    
    if request.method == 'POST':
        motivo_cancelacion = request.POST.get('motivo_cancelacion', '').strip()
        
        if not motivo_cancelacion:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'El motivo de cancelación es obligatorio.'}, status=400)
            messages.error(request, 'El motivo de cancelación es obligatorio.')
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
        
        try:
            plan.estado = 'cancelado'
            plan.motivo_cancelacion = motivo_cancelacion
            plan.fecha_cancelacion = timezone.now()
            plan.save()
            
            success_msg = f'El tratamiento "{plan.nombre}" ha sido cancelado exitosamente. Quedará registrado como cancelado en el historial del cliente.'
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': success_msg
                })
            
            messages.success(request, success_msg)
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
        except Exception as e:
            logger.error(f"Error al cancelar plan {plan_id}: {e}")
            error_msg = 'Error al cancelar el plan. Por favor, intenta nuevamente.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg}, status=500)
            messages.error(request, error_msg)
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
    
    # GET: Mostrar formulario (si se accede directamente)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    return render(request, 'citas/planes_tratamiento/cancelar_plan_tratamiento.html', {'plan': plan, 'perfil': perfil})


@login_required
def crear_plan_desde_odontograma(request, odontograma_id):
    """Crea un plan de tratamiento basado en un odontograma"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos.')
        return redirect('login')
    
    if not (perfil.es_dentista() or perfil.es_administrativo()):
        messages.error(request, 'Solo dentistas y administrativos pueden crear planes.')
        return redirect('panel_trabajador')
    
    # Obtener odontograma
    if perfil.es_administrativo():
        odontograma = get_object_or_404(Odontograma, id=odontograma_id)
    else:
        odontograma = get_object_or_404(Odontograma, id=odontograma_id, dentista=perfil)
    
    # Redirigir a crear plan con datos pre-llenados
    return redirect(f"{reverse('crear_plan_tratamiento')}?odontograma_id={odontograma_id}&cliente_id={odontograma.cliente.id if odontograma.cliente else ''}")


@login_required
def registrar_pago_tratamiento(request, plan_id):
    """Registra un pago parcial para un plan de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    if request.method == 'POST':
        try:
            monto_str = request.POST.get('monto', '0').strip()
            fecha_pago_str = request.POST.get('fecha_pago', '')
            metodo_pago = request.POST.get('metodo_pago', '').strip()
            numero_comprobante = request.POST.get('numero_comprobante', '').strip()
            notas = request.POST.get('notas', '').strip()
            cita_id = request.POST.get('cita_id', '').strip()
            
            if not monto_str or float(monto_str) <= 0:
                return JsonResponse({'success': False, 'error': 'El monto debe ser mayor a 0.'}, status=400)
            
            if not fecha_pago_str:
                return JsonResponse({'success': False, 'error': 'La fecha de pago es requerida.'}, status=400)
            
            if not metodo_pago or metodo_pago not in ['efectivo', 'transferencia', 'tarjeta', 'cheque']:
                return JsonResponse({'success': False, 'error': 'El método de pago es requerido.'}, status=400)
            
            # Procesar monto
            monto_str = re.sub(r'[^\d.]', '', monto_str) or '0'
            monto = float(monto_str)
            
            # Validar que el monto no exceda el saldo pendiente
            saldo_pendiente = plan.saldo_pendiente
            if monto > saldo_pendiente:
                return JsonResponse({
                    'success': False,
                    'error': f'El monto excede el saldo pendiente (${saldo_pendiente:,.0f}).'
                }, status=400)
            
            # Procesar fecha
            fecha_pago = datetime.strptime(fecha_pago_str, '%Y-%m-%d').date()
            
            # Obtener cita si se proporcionó
            cita = None
            if cita_id:
                try:
                    cita = Cita.objects.get(id=cita_id, plan_tratamiento=plan)
                except Cita.DoesNotExist:
                    pass
            
            # Crear el pago
            pago = PagoTratamiento.objects.create(
                plan_tratamiento=plan,
                monto=monto,
                fecha_pago=fecha_pago,
                metodo_pago=metodo_pago,
                numero_comprobante=numero_comprobante if numero_comprobante else None,
                notas=notas if notas else None,
                cita=cita,
                registrado_por=perfil
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Pago de ${monto:,.0f} registrado exitosamente.',
                'pago': {
                    'id': pago.id,
                    'monto': str(pago.monto),
                    'fecha_pago': pago.fecha_pago.strftime('%d/%m/%Y'),
                    'metodo_pago': pago.get_metodo_pago_display(),
                },
                'total_pagado': str(plan.total_pagado),
                'saldo_pendiente': str(plan.saldo_pendiente),
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en los datos: {str(e)}'}, status=400)
        except Exception as e:
            logger.error(f"Error al registrar pago: {e}")
            return JsonResponse({'success': False, 'error': f'Error al registrar el pago: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def eliminar_pago_tratamiento(request, plan_id, pago_id):
    """Elimina un pago registrado (solo administrativos)"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({'success': False, 'error': 'Solo los administrativos pueden eliminar pagos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    plan = get_object_or_404(PlanTratamiento, id=plan_id)
    pago = get_object_or_404(PagoTratamiento, id=pago_id, plan_tratamiento=plan)
    
    if request.method == 'POST':
        try:
            monto = pago.monto
            pago.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Pago de ${monto:,.0f} eliminado exitosamente.',
                'total_pagado': str(plan.total_pagado),
                'saldo_pendiente': str(plan.saldo_pendiente),
            })
            
        except Exception as e:
            logger.error(f"Error al eliminar pago: {e}")
            return JsonResponse({'success': False, 'error': f'Error al eliminar el pago: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def enviar_documentos_tratamiento(request, plan_id):
    """Vista para enviar presupuesto y consentimientos de un tratamiento por correo"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not (perfil.es_administrativo() or perfil.es_dentista()):
            return JsonResponse({'success': False, 'error': 'No tienes permisos para enviar documentos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan de tratamiento
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    else:
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    # Verificar que el cliente tenga email
    if not plan.cliente.email:
        return JsonResponse({'success': False, 'error': 'El cliente no tiene un email registrado.'}, status=400)
    
    try:
        from django.core.mail import EmailMultiAlternatives
        from django.conf import settings
        from io import BytesIO
        from datetime import datetime
        
        # Obtener información de la clínica
        try:
            from configuracion.models import InformacionClinica
            info_clinica = InformacionClinica.obtener()
            nombre_clinica = info_clinica.nombre_clinica
            email_clinica = info_clinica.email or getattr(settings, 'EMAIL_FROM', 'noreply@clinica.com')
        except:
            nombre_clinica = "Clínica Dental"
            email_clinica = getattr(settings, 'EMAIL_FROM', 'noreply@clinica.com')
        
        # Generar PDF del presupuesto
        presupuesto_response = exportar_presupuesto_pdf(request, plan.id)
        presupuesto_pdf = presupuesto_response.content
        
        # Obtener consentimientos pendientes o no firmados
        consentimientos = plan.consentimientos.filter(estado__in=['pendiente', 'firmado']).order_by('-fecha_creacion')
        
        # Crear el email
        asunto = f"Documentos del Tratamiento - {plan.nombre} - {nombre_clinica}"
        
        mensaje_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: linear-gradient(135deg, #14b8a6 0%, #0d9488 100%); min-height: 100vh;">
            <div style="max-width: 600px; margin: 0 auto; padding: 40px 20px;">
                <!-- Header con logo -->
                <div style="background: white; border-radius: 12px 12px 0 0; padding: 30px; text-align: center; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    <div style="display: inline-flex; align-items: center; gap: 12px; margin-bottom: 10px;">
                        <div style="width: 50px; height: 50px; background: linear-gradient(135deg, #14b8a6, #0d9488); border-radius: 50%; display: flex; align-items: center; justify-content: center; color: white; font-size: 1.8rem;">
                            <span style="font-size: 1.8rem;">🦷</span>
                        </div>
                        <h1 style="margin: 0; color: #0f766e; font-size: 1.75rem; font-weight: 700;">Clínica San Felipe</h1>
                    </div>
                    <p style="margin: 0; color: #64748b; font-size: 0.9rem;">Victoria, Región de la Araucanía</p>
                </div>
                
                <!-- Contenido del correo -->
                <div style="background: white; padding: 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    <h2 style="color: #1e293b; margin-top: 0; font-size: 1.5rem;">Estimado/a {plan.cliente.nombre_completo},</h2>
                    
                    <p style="color: #475569; line-height: 1.6; font-size: 1rem;">Le enviamos los documentos relacionados con su tratamiento: <strong style="color: #0f766e;">{plan.nombre}</strong></p>
                    
                    <div style="background: linear-gradient(135deg, #f0fdfa, #e0f2f1); padding: 20px; border-radius: 8px; margin: 25px 0; border-left: 4px solid #14b8a6;">
                        <h3 style="margin-top: 0; color: #0f766e; font-size: 1.1rem; margin-bottom: 12px;">Documentos adjuntos:</h3>
                        <ul style="margin: 0; padding-left: 20px; color: #475569;">
                            <li style="margin-bottom: 8px;"><strong style="color: #0f766e;">Presupuesto del Tratamiento</strong></li>
        """
        
        if consentimientos.exists():
            mensaje_html += f"<li style='margin-bottom: 8px;'><strong style='color: #0f766e;'>Consentimientos Informados</strong> ({consentimientos.count()} documento(s))</li>"
        
        mensaje_html += f"""
                        </ul>
                    </div>
                    
                    <p style="color: #475569; line-height: 1.6;">Por favor, revise cuidadosamente estos documentos. Si tiene alguna consulta, no dude en contactarnos.</p>
                    
                    <div style="background: #fef3c7; padding: 18px; border-radius: 8px; margin: 25px 0; border-left: 4px solid #f59e0b;">
                        <p style="margin: 0 0 12px 0; color: #92400e; font-weight: 600; font-size: 1rem;"><strong>Importante:</strong> Para proceder con el tratamiento, necesitamos que:</p>
                        <ul style="margin: 0; padding-left: 20px; color: #92400e;">
                            <li style="margin-bottom: 6px;">Acepte el presupuesto</li>
                            <li style="margin-bottom: 6px;">Firme el consentimiento informado</li>
                        </ul>
                    </div>
                    
                    <p style="color: #475569; line-height: 1.6;">Puede hacerlo desde su panel de cliente en nuestra página web o contactándonos directamente.</p>
                </div>
                
                <!-- Footer -->
                <div style="background: white; border-radius: 0 0 12px 12px; padding: 25px 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border-top: 1px solid #e0f2f1;">
                    <p style="margin: 0; color: #64748b; font-size: 0.9rem; line-height: 1.6;">
                        Saludos cordiales,<br>
                        <strong style="color: #0f766e; font-size: 1rem;">{nombre_clinica}</strong>
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        mensaje_texto = f"""
        Estimado/a {plan.cliente.nombre_completo},
        
        Le enviamos los documentos relacionados con su tratamiento: {plan.nombre}
        
        Documentos adjuntos:
        - Presupuesto del Tratamiento
        - Consentimientos Informados
        
        Por favor, revise cuidadosamente estos documentos. Si tiene alguna consulta, no dude en contactarnos.
        
        Importante: Para proceder con el tratamiento, necesitamos que:
        - Acepte el presupuesto
        - Firme el consentimiento informado
        
        Puede hacerlo desde su panel de cliente en nuestra página web o contactándonos directamente.
        
        Saludos cordiales,
        {nombre_clinica}
        """
        
        # Crear el email
        email = EmailMultiAlternatives(
            asunto,
            mensaje_texto,
            email_clinica,
            [plan.cliente.email],
        )
        email.attach_alternative(mensaje_html, "text/html")
        
        # Adjuntar presupuesto
        filename_presupuesto = f"presupuesto_{plan.cliente.nombre_completo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        email.attach(filename_presupuesto, presupuesto_pdf, 'application/pdf')
        
        # Adjuntar consentimientos
        for consentimiento in consentimientos:
            try:
                consentimiento_response = exportar_consentimiento_pdf(request, consentimiento.id)
                consentimiento_pdf = consentimiento_response.content
                filename_consentimiento = f"consentimiento_{consentimiento.titulo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
                email.attach(filename_consentimiento, consentimiento_pdf, 'application/pdf')
            except Exception as e:
                logger.error(f"Error al generar PDF del consentimiento {consentimiento.id}: {e}")
                # Continuar con los demás documentos
        
        # Enviar email
        email.send()
        
        # Actualizar fecha de envío en documentos relacionados
        from historial_clinico.models import DocumentoCliente
        documentos_actualizados = 0
        for consentimiento in consentimientos:
            # Actualizar el documento relacionado si existe
            try:
                documento = DocumentoCliente.objects.filter(
                    plan_tratamiento=plan,
                    tipo='consentimiento',
                    cliente=plan.cliente
                ).filter(
                    titulo__icontains=consentimiento.titulo
                ).first()
                
                if documento:
                    documento.enviado_por_correo = True
                    documento.fecha_envio = timezone.now()
                    documento.email_destinatario = plan.cliente.email
                    documento.save(update_fields=['enviado_por_correo', 'fecha_envio', 'email_destinatario'])
                    documentos_actualizados += 1
            except Exception as e:
                logger.warning(f"No se pudo actualizar documento para consentimiento {consentimiento.id}: {e}")
        
        # Actualizar también el documento del presupuesto si existe
        try:
            doc_presupuesto = DocumentoCliente.objects.filter(
                plan_tratamiento=plan,
                tipo='presupuesto',
                cliente=plan.cliente
            ).first()
            
            if doc_presupuesto:
                doc_presupuesto.enviado_por_correo = True
                doc_presupuesto.fecha_envio = timezone.now()
                doc_presupuesto.email_destinatario = plan.cliente.email
                doc_presupuesto.save(update_fields=['enviado_por_correo', 'fecha_envio', 'email_destinatario'])
                documentos_actualizados += 1
        except Exception as e:
            logger.warning(f"No se pudo actualizar documento de presupuesto: {e}")
        
        return JsonResponse({
            'success': True,
            'message': f'Documentos enviados exitosamente a {plan.cliente.email}. Se enviaron {1 + consentimientos.count()} documento(s).'
        })
        
    except Exception as e:
        logger.error(f"Error al enviar documentos del tratamiento: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'error': f'Error al enviar los documentos: {str(e)}'}, status=500)

    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener la fase
    fase = get_object_or_404(FaseTratamiento, id=fase_id, plan=plan)
    
    # Verificar que el plan puede ser editado
    if not plan.puede_ser_editado_por(perfil):
        return JsonResponse({'success': False, 'error': f'No puedes eliminar fases de este plan en estado "{plan.get_estado_display()}".'}, status=400)
    
    # Verificar si la fase tiene citas asociadas
    citas_asociadas = Cita.objects.filter(fase_tratamiento=fase).count()
    if citas_asociadas > 0:
        return JsonResponse({
            'success': False,
            'error': f'No se puede eliminar la fase porque tiene {citas_asociadas} cita(s) asociada(s).'
        }, status=400)
    
    if request.method == 'POST':
        try:
            nombre_fase = fase.nombre
            fase.delete()
            
            # Reordenar las fases restantes
            fases_restantes = plan.fases.all().order_by('orden')
            for idx, fase_restante in enumerate(fases_restantes, start=1):
                fase_restante.orden = idx
                fase_restante.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Fase "{nombre_fase}" eliminada exitosamente.'
            })
            
        except Exception as e:
            logger.error(f"Error al eliminar fase: {e}")
            return JsonResponse({'success': False, 'error': f'Error al eliminar la fase: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def crear_cita_desde_fase(request, plan_id, fase_id):
    """Crea una cita específica para una fase de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener la fase
    fase = get_object_or_404(FaseTratamiento, id=fase_id, plan=plan)
    
    # Validar que el plan esté en un estado que permita crear citas
    if request.method == 'POST':
        # Verificar que el plan tenga consentimiento firmado y presupuesto aceptado
        tiene_consentimiento_firmado = plan.consentimientos.filter(estado='firmado').exists()
        presupuesto_aceptado = plan.presupuesto_aceptado
        
        if not tiene_consentimiento_firmado:
            return JsonResponse({
                'success': False, 
                'error': 'No se pueden crear citas. El tratamiento debe tener al menos un consentimiento informado firmado.'
            }, status=400)
        
        if not presupuesto_aceptado:
            return JsonResponse({
                'success': False, 
                'error': 'No se pueden crear citas. El presupuesto del tratamiento debe estar aceptado.'
            }, status=400)
        
        # Verificar que el plan esté en un estado válido para crear citas
        if plan.estado not in ['aprobado', 'en_progreso']:
            return JsonResponse({
                'success': False, 
                'error': f'No se pueden crear citas para un tratamiento en estado "{plan.get_estado_display()}". El tratamiento debe estar aprobado o en progreso.'
            }, status=400)
        
        # Si el plan está aprobado pero no en progreso, cambiarlo automáticamente a en_progreso
        if plan.estado == 'aprobado':
            plan.estado = 'en_progreso'
            plan.save()
    
    if request.method == 'POST':
        fecha_hora_str = request.POST.get('fecha_hora', '')
        tipo_servicio_id = request.POST.get('tipo_servicio', '').strip()
        notas = request.POST.get('notas', '').strip()
        
        if not fecha_hora_str:
            return JsonResponse({'success': False, 'error': 'Debe seleccionar una fecha y hora.'}, status=400)
        
        try:
            # Convertir fecha y hacerla timezone-aware
            fecha_hora_naive = datetime.fromisoformat(fecha_hora_str)
            fecha_hora = timezone.make_aware(fecha_hora_naive)
            
            # Validar que la fecha no sea en el pasado
            ahora = timezone.now()
            if fecha_hora < ahora:
                return JsonResponse({'success': False, 'error': 'No se pueden crear citas en fechas pasadas.'}, status=400)
            
            # Obtener tipo de servicio si se proporcionó
            tipo_servicio = None
            if tipo_servicio_id:
                try:
                    tipo_servicio = TipoServicio.objects.get(id=tipo_servicio_id, activo=True)
                except TipoServicio.DoesNotExist:
                    pass
            
            # Validar horario del dentista
            dentista = plan.dentista
            dia_semana = fecha_hora.weekday()
            hora_cita = fecha_hora.time()
            
            horarios_dia = HorarioDentista.objects.filter(
                dentista=dentista,
                dia_semana=dia_semana,
                activo=True
            )
            
            if not horarios_dia.exists():
                dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
                return JsonResponse({'success': False, 'error': f'El dentista no trabaja los {dias_nombres[dia_semana]}.'}, status=400)
            
            hora_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita < horario.hora_fin:
                    hora_valida = True
                    break
            
            if not hora_valida:
                return JsonResponse({'success': False, 'error': 'La hora seleccionada no está dentro del horario de trabajo del dentista.'}, status=400)
            
            # Validar duración y solapamiento
            duracion_minutos = tipo_servicio.duracion_estimada if tipo_servicio and tipo_servicio.duracion_estimada else 30
            fecha_hora_fin = fecha_hora + timedelta(minutes=duracion_minutos)
            hora_fin_cita = fecha_hora_fin.time()
            
            duracion_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita and hora_fin_cita <= horario.hora_fin:
                    duracion_valida = True
                    break
            
            if not duracion_valida:
                return JsonResponse({'success': False, 'error': f'La duración del servicio ({duracion_minutos} minutos) no cabe en el horario seleccionado.'}, status=400)
            
            # Verificar solapamiento con otras citas
            citas_existentes = Cita.objects.filter(
                dentista=dentista,
                fecha_hora__date=fecha_hora.date(),
                estado__in=['disponible', 'reservada', 'confirmada', 'en_progreso']
            )
            
            for cita_existente in citas_existentes:
                fecha_hora_existente_fin = cita_existente.fecha_hora
                if cita_existente.tipo_servicio and cita_existente.tipo_servicio.duracion_estimada:
                    fecha_hora_existente_fin += timedelta(minutes=cita_existente.tipo_servicio.duracion_estimada)
                else:
                    fecha_hora_existente_fin += timedelta(minutes=30)
                
                if (fecha_hora < fecha_hora_existente_fin and fecha_hora_fin > cita_existente.fecha_hora):
                    return JsonResponse({'success': False, 'error': f'La cita se solapa con otra cita existente a las {cita_existente.fecha_hora.strftime("%H:%M")}.'}, status=400)
            
            # Obtener precio del servicio
            precio_cobrado = None
            if tipo_servicio:
                precio_cobrado = tipo_servicio.precio_base
            
            # Crear la cita asociada al plan y fase
            cita = Cita.objects.create(
                fecha_hora=fecha_hora,
                tipo_servicio=tipo_servicio,
                tipo_consulta=tipo_servicio.nombre if tipo_servicio else f"Cita - {fase.nombre}",
                precio_cobrado=precio_cobrado,
                notas=notas or f"Cita de la fase: {fase.nombre}",
                dentista=dentista,
                cliente=plan.cliente,
                paciente_nombre=plan.cliente.nombre_completo,
                paciente_email=plan.cliente.email,
                paciente_telefono=plan.cliente.telefono,
                estado='reservada',
                creada_por=perfil,
                plan_tratamiento=plan,
                fase_tratamiento=fase
            )
            
            # Enviar notificaciones (WhatsApp Y SMS) si tiene teléfono
            # Usar el teléfono del cliente del plan
            telefono_cliente = plan.cliente.telefono if plan.cliente else None
            if telefono_cliente:
                try:
                    from citas.mensajeria_service import enviar_notificaciones_cita
                    resultado = enviar_notificaciones_cita(cita, telefono_override=telefono_cliente)
                    canales_enviados = []
                    if resultado['whatsapp']['enviado']:
                        canales_enviados.append('WhatsApp')
                    if resultado['sms']['enviado']:
                        canales_enviados.append('SMS')
                    if canales_enviados:
                        logger.info(f"Notificaciones enviadas por {', '.join(canales_enviados)} para cita {cita.id} del plan {plan.id}")
                except Exception as e:
                    logger.error(f"Error al enviar notificaciones para cita {cita.id}: {e}")
            
            mensaje = f'Cita creada exitosamente para la fase "{fase.nombre}" el {fecha_hora.strftime("%d/%m/%Y a las %H:%M")}.'
            if plan.cliente.telefono:
                mensaje += ' Se ha enviado un SMS de confirmación.'
            
            return JsonResponse({
                'success': True,
                'message': mensaje,
                'cita_id': cita.id
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en el formato de fecha: {e}'}, status=400)
        except Exception as e:
            logger.error(f"Error al crear cita desde fase: {e}")
            return JsonResponse({'success': False, 'error': f'Error al crear la cita: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def cancelar_plan_tratamiento(request, plan_id):
    """Cancela un plan (DENTISTA puede cancelar sus planes)"""
    from django.http import JsonResponse
    
    try:
        perfil = Perfil.objects.get(user=request.user)
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
        messages.error(request, 'No tienes permisos.')
        return redirect('login')
    
    # Obtener plan con permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    else:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
        messages.error(request, 'No tienes permisos.')
        return redirect('panel_trabajador')
    
    if not plan.puede_ser_cancelado_por(perfil):
        error_msg = f'No puedes cancelar este plan en estado "{plan.get_estado_display()}".'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg}, status=400)
        messages.error(request, error_msg)
        return redirect('detalle_plan_tratamiento', plan_id=plan.id)
    
    if request.method == 'POST':
        motivo_cancelacion = request.POST.get('motivo_cancelacion', '').strip()
        
        if not motivo_cancelacion:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'El motivo de cancelación es obligatorio.'}, status=400)
            messages.error(request, 'El motivo de cancelación es obligatorio.')
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
        
        try:
            plan.estado = 'cancelado'
            plan.motivo_cancelacion = motivo_cancelacion
            plan.fecha_cancelacion = timezone.now()
            plan.save()
            
            success_msg = f'El tratamiento "{plan.nombre}" ha sido cancelado exitosamente. Quedará registrado como cancelado en el historial del cliente.'
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': success_msg
                })
            
            messages.success(request, success_msg)
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
        except Exception as e:
            logger.error(f"Error al cancelar plan {plan_id}: {e}")
            error_msg = 'Error al cancelar el plan. Por favor, intenta nuevamente.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg}, status=500)
            messages.error(request, error_msg)
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
    
    # GET: Mostrar formulario (si se accede directamente)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    return render(request, 'citas/planes_tratamiento/cancelar_plan_tratamiento.html', {'plan': plan, 'perfil': perfil})


@login_required
def crear_plan_desde_odontograma(request, odontograma_id):
    """Crea un plan de tratamiento basado en un odontograma"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos.')
        return redirect('login')
    
    if not (perfil.es_dentista() or perfil.es_administrativo()):
        messages.error(request, 'Solo dentistas y administrativos pueden crear planes.')
        return redirect('panel_trabajador')
    
    # Obtener odontograma
    if perfil.es_administrativo():
        odontograma = get_object_or_404(Odontograma, id=odontograma_id)
    else:
        odontograma = get_object_or_404(Odontograma, id=odontograma_id, dentista=perfil)
    
    # Redirigir a crear plan con datos pre-llenados
    return redirect(f"{reverse('crear_plan_tratamiento')}?odontograma_id={odontograma_id}&cliente_id={odontograma.cliente.id if odontograma.cliente else ''}")


@login_required
def registrar_pago_tratamiento(request, plan_id):
    """Registra un pago parcial para un plan de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    if request.method == 'POST':
        try:
            monto_str = request.POST.get('monto', '0').strip()
            fecha_pago_str = request.POST.get('fecha_pago', '')
            metodo_pago = request.POST.get('metodo_pago', '').strip()
            numero_comprobante = request.POST.get('numero_comprobante', '').strip()
            notas = request.POST.get('notas', '').strip()
            cita_id = request.POST.get('cita_id', '').strip()
            
            if not monto_str or float(monto_str) <= 0:
                return JsonResponse({'success': False, 'error': 'El monto debe ser mayor a 0.'}, status=400)
            
            if not fecha_pago_str:
                return JsonResponse({'success': False, 'error': 'La fecha de pago es requerida.'}, status=400)
            
            if not metodo_pago or metodo_pago not in ['efectivo', 'transferencia', 'tarjeta', 'cheque']:
                return JsonResponse({'success': False, 'error': 'El método de pago es requerido.'}, status=400)
            
            # Procesar monto
            monto_str = re.sub(r'[^\d.]', '', monto_str) or '0'
            monto = float(monto_str)
            
            # Validar que el monto no exceda el saldo pendiente
            saldo_pendiente = plan.saldo_pendiente
            if monto > saldo_pendiente:
                return JsonResponse({
                    'success': False,
                    'error': f'El monto excede el saldo pendiente (${saldo_pendiente:,.0f}).'
                }, status=400)
            
            # Procesar fecha
            fecha_pago = datetime.strptime(fecha_pago_str, '%Y-%m-%d').date()
            
            # Obtener cita si se proporcionó
            cita = None
            if cita_id:
                try:
                    cita = Cita.objects.get(id=cita_id, plan_tratamiento=plan)
                except Cita.DoesNotExist:
                    pass
            
            # Crear el pago
            pago = PagoTratamiento.objects.create(
                plan_tratamiento=plan,
                monto=monto,
                fecha_pago=fecha_pago,
                metodo_pago=metodo_pago,
                numero_comprobante=numero_comprobante if numero_comprobante else None,
                notas=notas if notas else None,
                cita=cita,
                registrado_por=perfil
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Pago de ${monto:,.0f} registrado exitosamente.',
                'pago': {
                    'id': pago.id,
                    'monto': str(pago.monto),
                    'fecha_pago': pago.fecha_pago.strftime('%d/%m/%Y'),
                    'metodo_pago': pago.get_metodo_pago_display(),
                },
                'total_pagado': str(plan.total_pagado),
                'saldo_pendiente': str(plan.saldo_pendiente),
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en los datos: {str(e)}'}, status=400)
        except Exception as e:
            logger.error(f"Error al registrar pago: {e}")
            return JsonResponse({'success': False, 'error': f'Error al registrar el pago: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def eliminar_pago_tratamiento(request, plan_id, pago_id):
    """Elimina un pago registrado (solo administrativos)"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({'success': False, 'error': 'Solo los administrativos pueden eliminar pagos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    plan = get_object_or_404(PlanTratamiento, id=plan_id)
    pago = get_object_or_404(PagoTratamiento, id=pago_id, plan_tratamiento=plan)
    
    if request.method == 'POST':
        try:
            monto = pago.monto
            pago.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Pago de ${monto:,.0f} eliminado exitosamente.',
                'total_pagado': str(plan.total_pagado),
                'saldo_pendiente': str(plan.saldo_pendiente),
            })
            
        except Exception as e:
            logger.error(f"Error al eliminar pago: {e}")
            return JsonResponse({'success': False, 'error': f'Error al eliminar el pago: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def enviar_documentos_tratamiento(request, plan_id):
    """Vista para enviar presupuesto y consentimientos de un tratamiento por correo"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not (perfil.es_administrativo() or perfil.es_dentista()):
            return JsonResponse({'success': False, 'error': 'No tienes permisos para enviar documentos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan de tratamiento
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    else:
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    # Verificar que el cliente tenga email
    if not plan.cliente.email:
        return JsonResponse({'success': False, 'error': 'El cliente no tiene un email registrado.'}, status=400)
    
    try:
        from django.core.mail import EmailMultiAlternatives
        from django.conf import settings
        from io import BytesIO
        from datetime import datetime
        
        # Obtener información de la clínica
        try:
            from configuracion.models import InformacionClinica
            info_clinica = InformacionClinica.obtener()
            nombre_clinica = info_clinica.nombre_clinica
            email_clinica = info_clinica.email or getattr(settings, 'EMAIL_FROM', 'noreply@clinica.com')
        except:
            nombre_clinica = "Clínica Dental"
            email_clinica = getattr(settings, 'EMAIL_FROM', 'noreply@clinica.com')
        
        # Generar PDF del presupuesto
        presupuesto_response = exportar_presupuesto_pdf(request, plan.id)
        presupuesto_pdf = presupuesto_response.content
        
        # Obtener consentimientos pendientes o no firmados
        consentimientos = plan.consentimientos.filter(estado__in=['pendiente', 'firmado']).order_by('-fecha_creacion')
        
        # Crear el email
        asunto = f"Documentos del Tratamiento - {plan.nombre} - {nombre_clinica}"
        
        mensaje_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: linear-gradient(135deg, #14b8a6 0%, #0d9488 100%); min-height: 100vh;">
            <div style="max-width: 600px; margin: 0 auto; padding: 40px 20px;">
                <!-- Header con logo -->
                <div style="background: white; border-radius: 12px 12px 0 0; padding: 30px; text-align: center; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    <div style="display: inline-flex; align-items: center; gap: 12px; margin-bottom: 10px;">
                        <div style="width: 50px; height: 50px; background: linear-gradient(135deg, #14b8a6, #0d9488); border-radius: 50%; display: flex; align-items: center; justify-content: center; color: white; font-size: 1.8rem;">
                            <span style="font-size: 1.8rem;">🦷</span>
                        </div>
                        <h1 style="margin: 0; color: #0f766e; font-size: 1.75rem; font-weight: 700;">Clínica San Felipe</h1>
                    </div>
                    <p style="margin: 0; color: #64748b; font-size: 0.9rem;">Victoria, Región de la Araucanía</p>
                </div>
                
                <!-- Contenido del correo -->
                <div style="background: white; padding: 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    <h2 style="color: #1e293b; margin-top: 0; font-size: 1.5rem;">Estimado/a {plan.cliente.nombre_completo},</h2>
                    
                    <p style="color: #475569; line-height: 1.6; font-size: 1rem;">Le enviamos los documentos relacionados con su tratamiento: <strong style="color: #0f766e;">{plan.nombre}</strong></p>
                    
                    <div style="background: linear-gradient(135deg, #f0fdfa, #e0f2f1); padding: 20px; border-radius: 8px; margin: 25px 0; border-left: 4px solid #14b8a6;">
                        <h3 style="margin-top: 0; color: #0f766e; font-size: 1.1rem; margin-bottom: 12px;">Documentos adjuntos:</h3>
                        <ul style="margin: 0; padding-left: 20px; color: #475569;">
                            <li style="margin-bottom: 8px;"><strong style="color: #0f766e;">Presupuesto del Tratamiento</strong></li>
        """
        
        if consentimientos.exists():
            mensaje_html += f"<li style='margin-bottom: 8px;'><strong style='color: #0f766e;'>Consentimientos Informados</strong> ({consentimientos.count()} documento(s))</li>"
        
        mensaje_html += f"""
                        </ul>
                    </div>
                    
                    <p style="color: #475569; line-height: 1.6;">Por favor, revise cuidadosamente estos documentos. Si tiene alguna consulta, no dude en contactarnos.</p>
                    
                    <div style="background: #fef3c7; padding: 18px; border-radius: 8px; margin: 25px 0; border-left: 4px solid #f59e0b;">
                        <p style="margin: 0 0 12px 0; color: #92400e; font-weight: 600; font-size: 1rem;"><strong>Importante:</strong> Para proceder con el tratamiento, necesitamos que:</p>
                        <ul style="margin: 0; padding-left: 20px; color: #92400e;">
                            <li style="margin-bottom: 6px;">Acepte el presupuesto</li>
                            <li style="margin-bottom: 6px;">Firme el consentimiento informado</li>
                        </ul>
                    </div>
                    
                    <p style="color: #475569; line-height: 1.6;">Puede hacerlo desde su panel de cliente en nuestra página web o contactándonos directamente.</p>
                </div>
                
                <!-- Footer -->
                <div style="background: white; border-radius: 0 0 12px 12px; padding: 25px 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border-top: 1px solid #e0f2f1;">
                    <p style="margin: 0; color: #64748b; font-size: 0.9rem; line-height: 1.6;">
                        Saludos cordiales,<br>
                        <strong style="color: #0f766e; font-size: 1rem;">{nombre_clinica}</strong>
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        mensaje_texto = f"""
        Estimado/a {plan.cliente.nombre_completo},
        
        Le enviamos los documentos relacionados con su tratamiento: {plan.nombre}
        
        Documentos adjuntos:
        - Presupuesto del Tratamiento
        - Consentimientos Informados
        
        Por favor, revise cuidadosamente estos documentos. Si tiene alguna consulta, no dude en contactarnos.
        
        Importante: Para proceder con el tratamiento, necesitamos que:
        - Acepte el presupuesto
        - Firme el consentimiento informado
        
        Puede hacerlo desde su panel de cliente en nuestra página web o contactándonos directamente.
        
        Saludos cordiales,
        {nombre_clinica}
        """
        
        # Crear el email
        email = EmailMultiAlternatives(
            asunto,
            mensaje_texto,
            email_clinica,
            [plan.cliente.email],
        )
        email.attach_alternative(mensaje_html, "text/html")
        
        # Adjuntar presupuesto
        filename_presupuesto = f"presupuesto_{plan.cliente.nombre_completo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        email.attach(filename_presupuesto, presupuesto_pdf, 'application/pdf')
        
        # Adjuntar consentimientos
        for consentimiento in consentimientos:
            try:
                consentimiento_response = exportar_consentimiento_pdf(request, consentimiento.id)
                consentimiento_pdf = consentimiento_response.content
                filename_consentimiento = f"consentimiento_{consentimiento.titulo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
                email.attach(filename_consentimiento, consentimiento_pdf, 'application/pdf')
            except Exception as e:
                logger.error(f"Error al generar PDF del consentimiento {consentimiento.id}: {e}")
                # Continuar con los demás documentos
        
        # Enviar email
        email.send()
        
        # Actualizar fecha de envío en documentos relacionados
        from historial_clinico.models import DocumentoCliente
        documentos_actualizados = 0
        for consentimiento in consentimientos:
            # Actualizar el documento relacionado si existe
            try:
                documento = DocumentoCliente.objects.filter(
                    plan_tratamiento=plan,
                    tipo='consentimiento',
                    cliente=plan.cliente
                ).filter(
                    titulo__icontains=consentimiento.titulo
                ).first()
                
                if documento:
                    documento.enviado_por_correo = True
                    documento.fecha_envio = timezone.now()
                    documento.email_destinatario = plan.cliente.email
                    documento.save(update_fields=['enviado_por_correo', 'fecha_envio', 'email_destinatario'])
                    documentos_actualizados += 1
            except Exception as e:
                logger.warning(f"No se pudo actualizar documento para consentimiento {consentimiento.id}: {e}")
        
        # Actualizar también el documento del presupuesto si existe
        try:
            doc_presupuesto = DocumentoCliente.objects.filter(
                plan_tratamiento=plan,
                tipo='presupuesto',
                cliente=plan.cliente
            ).first()
            
            if doc_presupuesto:
                doc_presupuesto.enviado_por_correo = True
                doc_presupuesto.fecha_envio = timezone.now()
                doc_presupuesto.email_destinatario = plan.cliente.email
                doc_presupuesto.save(update_fields=['enviado_por_correo', 'fecha_envio', 'email_destinatario'])
                documentos_actualizados += 1
        except Exception as e:
            logger.warning(f"No se pudo actualizar documento de presupuesto: {e}")
        
        return JsonResponse({
            'success': True,
            'message': f'Documentos enviados exitosamente a {plan.cliente.email}. Se enviaron {1 + consentimientos.count()} documento(s).'
        })
        
    except Exception as e:
        logger.error(f"Error al enviar documentos del tratamiento: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'error': f'Error al enviar los documentos: {str(e)}'}, status=500)

    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener la fase
    fase = get_object_or_404(FaseTratamiento, id=fase_id, plan=plan)
    
    # Verificar que el plan puede ser editado
    if not plan.puede_ser_editado_por(perfil):
        return JsonResponse({'success': False, 'error': f'No puedes eliminar fases de este plan en estado "{plan.get_estado_display()}".'}, status=400)
    
    # Verificar si la fase tiene citas asociadas
    citas_asociadas = Cita.objects.filter(fase_tratamiento=fase).count()
    if citas_asociadas > 0:
        return JsonResponse({
            'success': False,
            'error': f'No se puede eliminar la fase porque tiene {citas_asociadas} cita(s) asociada(s).'
        }, status=400)
    
    if request.method == 'POST':
        try:
            nombre_fase = fase.nombre
            fase.delete()
            
            # Reordenar las fases restantes
            fases_restantes = plan.fases.all().order_by('orden')
            for idx, fase_restante in enumerate(fases_restantes, start=1):
                fase_restante.orden = idx
                fase_restante.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Fase "{nombre_fase}" eliminada exitosamente.'
            })
            
        except Exception as e:
            logger.error(f"Error al eliminar fase: {e}")
            return JsonResponse({'success': False, 'error': f'Error al eliminar la fase: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def crear_cita_desde_fase(request, plan_id, fase_id):
    """Crea una cita específica para una fase de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener la fase
    fase = get_object_or_404(FaseTratamiento, id=fase_id, plan=plan)
    
    # Validar que el plan esté en un estado que permita crear citas
    if request.method == 'POST':
        # Verificar que el plan tenga consentimiento firmado y presupuesto aceptado
        tiene_consentimiento_firmado = plan.consentimientos.filter(estado='firmado').exists()
        presupuesto_aceptado = plan.presupuesto_aceptado
        
        if not tiene_consentimiento_firmado:
            return JsonResponse({
                'success': False, 
                'error': 'No se pueden crear citas. El tratamiento debe tener al menos un consentimiento informado firmado.'
            }, status=400)
        
        if not presupuesto_aceptado:
            return JsonResponse({
                'success': False, 
                'error': 'No se pueden crear citas. El presupuesto del tratamiento debe estar aceptado.'
            }, status=400)
        
        # Verificar que el plan esté en un estado válido para crear citas
        if plan.estado not in ['aprobado', 'en_progreso']:
            return JsonResponse({
                'success': False, 
                'error': f'No se pueden crear citas para un tratamiento en estado "{plan.get_estado_display()}". El tratamiento debe estar aprobado o en progreso.'
            }, status=400)
        
        # Si el plan está aprobado pero no en progreso, cambiarlo automáticamente a en_progreso
        if plan.estado == 'aprobado':
            plan.estado = 'en_progreso'
            plan.save()
    
    if request.method == 'POST':
        fecha_hora_str = request.POST.get('fecha_hora', '')
        tipo_servicio_id = request.POST.get('tipo_servicio', '').strip()
        notas = request.POST.get('notas', '').strip()
        
        if not fecha_hora_str:
            return JsonResponse({'success': False, 'error': 'Debe seleccionar una fecha y hora.'}, status=400)
        
        try:
            # Convertir fecha y hacerla timezone-aware
            fecha_hora_naive = datetime.fromisoformat(fecha_hora_str)
            fecha_hora = timezone.make_aware(fecha_hora_naive)
            
            # Validar que la fecha no sea en el pasado
            ahora = timezone.now()
            if fecha_hora < ahora:
                return JsonResponse({'success': False, 'error': 'No se pueden crear citas en fechas pasadas.'}, status=400)
            
            # Obtener tipo de servicio si se proporcionó
            tipo_servicio = None
            if tipo_servicio_id:
                try:
                    tipo_servicio = TipoServicio.objects.get(id=tipo_servicio_id, activo=True)
                except TipoServicio.DoesNotExist:
                    pass
            
            # Validar horario del dentista
            dentista = plan.dentista
            dia_semana = fecha_hora.weekday()
            hora_cita = fecha_hora.time()
            
            horarios_dia = HorarioDentista.objects.filter(
                dentista=dentista,
                dia_semana=dia_semana,
                activo=True
            )
            
            if not horarios_dia.exists():
                dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
                return JsonResponse({'success': False, 'error': f'El dentista no trabaja los {dias_nombres[dia_semana]}.'}, status=400)
            
            hora_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita < horario.hora_fin:
                    hora_valida = True
                    break
            
            if not hora_valida:
                return JsonResponse({'success': False, 'error': 'La hora seleccionada no está dentro del horario de trabajo del dentista.'}, status=400)
            
            # Validar duración y solapamiento
            duracion_minutos = tipo_servicio.duracion_estimada if tipo_servicio and tipo_servicio.duracion_estimada else 30
            fecha_hora_fin = fecha_hora + timedelta(minutes=duracion_minutos)
            hora_fin_cita = fecha_hora_fin.time()
            
            duracion_valida = False
            for horario in horarios_dia:
                if horario.hora_inicio <= hora_cita and hora_fin_cita <= horario.hora_fin:
                    duracion_valida = True
                    break
            
            if not duracion_valida:
                return JsonResponse({'success': False, 'error': f'La duración del servicio ({duracion_minutos} minutos) no cabe en el horario seleccionado.'}, status=400)
            
            # Verificar solapamiento con otras citas
            citas_existentes = Cita.objects.filter(
                dentista=dentista,
                fecha_hora__date=fecha_hora.date(),
                estado__in=['disponible', 'reservada', 'confirmada', 'en_progreso']
            )
            
            for cita_existente in citas_existentes:
                fecha_hora_existente_fin = cita_existente.fecha_hora
                if cita_existente.tipo_servicio and cita_existente.tipo_servicio.duracion_estimada:
                    fecha_hora_existente_fin += timedelta(minutes=cita_existente.tipo_servicio.duracion_estimada)
                else:
                    fecha_hora_existente_fin += timedelta(minutes=30)
                
                if (fecha_hora < fecha_hora_existente_fin and fecha_hora_fin > cita_existente.fecha_hora):
                    return JsonResponse({'success': False, 'error': f'La cita se solapa con otra cita existente a las {cita_existente.fecha_hora.strftime("%H:%M")}.'}, status=400)
            
            # Obtener precio del servicio
            precio_cobrado = None
            if tipo_servicio:
                precio_cobrado = tipo_servicio.precio_base
            
            # Crear la cita asociada al plan y fase
            cita = Cita.objects.create(
                fecha_hora=fecha_hora,
                tipo_servicio=tipo_servicio,
                tipo_consulta=tipo_servicio.nombre if tipo_servicio else f"Cita - {fase.nombre}",
                precio_cobrado=precio_cobrado,
                notas=notas or f"Cita de la fase: {fase.nombre}",
                dentista=dentista,
                cliente=plan.cliente,
                paciente_nombre=plan.cliente.nombre_completo,
                paciente_email=plan.cliente.email,
                paciente_telefono=plan.cliente.telefono,
                estado='reservada',
                creada_por=perfil,
                plan_tratamiento=plan,
                fase_tratamiento=fase
            )
            
            # Enviar notificaciones (WhatsApp Y SMS) si tiene teléfono
            # Usar el teléfono del cliente del plan
            telefono_cliente = plan.cliente.telefono if plan.cliente else None
            if telefono_cliente:
                try:
                    from citas.mensajeria_service import enviar_notificaciones_cita
                    resultado = enviar_notificaciones_cita(cita, telefono_override=telefono_cliente)
                    canales_enviados = []
                    if resultado['whatsapp']['enviado']:
                        canales_enviados.append('WhatsApp')
                    if resultado['sms']['enviado']:
                        canales_enviados.append('SMS')
                    if canales_enviados:
                        logger.info(f"Notificaciones enviadas por {', '.join(canales_enviados)} para cita {cita.id} del plan {plan.id}")
                except Exception as e:
                    logger.error(f"Error al enviar notificaciones para cita {cita.id}: {e}")
            
            mensaje = f'Cita creada exitosamente para la fase "{fase.nombre}" el {fecha_hora.strftime("%d/%m/%Y a las %H:%M")}.'
            if plan.cliente.telefono:
                mensaje += ' Se ha enviado un SMS de confirmación.'
            
            return JsonResponse({
                'success': True,
                'message': mensaje,
                'cita_id': cita.id
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en el formato de fecha: {e}'}, status=400)
        except Exception as e:
            logger.error(f"Error al crear cita desde fase: {e}")
            return JsonResponse({'success': False, 'error': f'Error al crear la cita: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def cancelar_plan_tratamiento(request, plan_id):
    """Cancela un plan (DENTISTA puede cancelar sus planes)"""
    from django.http import JsonResponse
    
    try:
        perfil = Perfil.objects.get(user=request.user)
    except Perfil.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
        messages.error(request, 'No tienes permisos.')
        return redirect('login')
    
    # Obtener plan con permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    else:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
        messages.error(request, 'No tienes permisos.')
        return redirect('panel_trabajador')
    
    if not plan.puede_ser_cancelado_por(perfil):
        error_msg = f'No puedes cancelar este plan en estado "{plan.get_estado_display()}".'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg}, status=400)
        messages.error(request, error_msg)
        return redirect('detalle_plan_tratamiento', plan_id=plan.id)
    
    if request.method == 'POST':
        motivo_cancelacion = request.POST.get('motivo_cancelacion', '').strip()
        
        if not motivo_cancelacion:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'El motivo de cancelación es obligatorio.'}, status=400)
            messages.error(request, 'El motivo de cancelación es obligatorio.')
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
        
        try:
            plan.estado = 'cancelado'
            plan.motivo_cancelacion = motivo_cancelacion
            plan.fecha_cancelacion = timezone.now()
            plan.save()
            
            success_msg = f'El tratamiento "{plan.nombre}" ha sido cancelado exitosamente. Quedará registrado como cancelado en el historial del cliente.'
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': success_msg
                })
            
            messages.success(request, success_msg)
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
        except Exception as e:
            logger.error(f"Error al cancelar plan {plan_id}: {e}")
            error_msg = 'Error al cancelar el plan. Por favor, intenta nuevamente.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg}, status=500)
            messages.error(request, error_msg)
            return redirect('detalle_plan_tratamiento', plan_id=plan.id)
    
    # GET: Mostrar formulario (si se accede directamente)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    return render(request, 'citas/planes_tratamiento/cancelar_plan_tratamiento.html', {'plan': plan, 'perfil': perfil})


@login_required
def crear_plan_desde_odontograma(request, odontograma_id):
    """Crea un plan de tratamiento basado en un odontograma"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            messages.error(request, 'Tu cuenta está desactivada.')
            return redirect('login')
    except Perfil.DoesNotExist:
        messages.error(request, 'No tienes permisos.')
        return redirect('login')
    
    if not (perfil.es_dentista() or perfil.es_administrativo()):
        messages.error(request, 'Solo dentistas y administrativos pueden crear planes.')
        return redirect('panel_trabajador')
    
    # Obtener odontograma
    if perfil.es_administrativo():
        odontograma = get_object_or_404(Odontograma, id=odontograma_id)
    else:
        odontograma = get_object_or_404(Odontograma, id=odontograma_id, dentista=perfil)
    
    # Redirigir a crear plan con datos pre-llenados
    return redirect(f"{reverse('crear_plan_tratamiento')}?odontograma_id={odontograma_id}&cliente_id={odontograma.cliente.id if odontograma.cliente else ''}")


@login_required
def registrar_pago_tratamiento(request, plan_id):
    """Registra un pago parcial para un plan de tratamiento"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.activo:
            return JsonResponse({'success': False, 'error': 'Tu cuenta está desactivada.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan con verificación de permisos
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    elif perfil.es_dentista():
        pacientes_dentista = perfil.get_pacientes_asignados()
        clientes_ids = [p['id'] for p in pacientes_dentista if 'id' in p and isinstance(p['id'], int)]
        plan = get_object_or_404(
            PlanTratamiento,
            id=plan_id,
            dentista=perfil,
            cliente_id__in=clientes_ids
        )
    else:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    if request.method == 'POST':
        try:
            monto_str = request.POST.get('monto', '0').strip()
            fecha_pago_str = request.POST.get('fecha_pago', '')
            metodo_pago = request.POST.get('metodo_pago', '').strip()
            numero_comprobante = request.POST.get('numero_comprobante', '').strip()
            notas = request.POST.get('notas', '').strip()
            cita_id = request.POST.get('cita_id', '').strip()
            
            if not monto_str or float(monto_str) <= 0:
                return JsonResponse({'success': False, 'error': 'El monto debe ser mayor a 0.'}, status=400)
            
            if not fecha_pago_str:
                return JsonResponse({'success': False, 'error': 'La fecha de pago es requerida.'}, status=400)
            
            if not metodo_pago or metodo_pago not in ['efectivo', 'transferencia', 'tarjeta', 'cheque']:
                return JsonResponse({'success': False, 'error': 'El método de pago es requerido.'}, status=400)
            
            # Procesar monto
            monto_str = re.sub(r'[^\d.]', '', monto_str) or '0'
            monto = float(monto_str)
            
            # Validar que el monto no exceda el saldo pendiente
            saldo_pendiente = plan.saldo_pendiente
            if monto > saldo_pendiente:
                return JsonResponse({
                    'success': False,
                    'error': f'El monto excede el saldo pendiente (${saldo_pendiente:,.0f}).'
                }, status=400)
            
            # Procesar fecha
            fecha_pago = datetime.strptime(fecha_pago_str, '%Y-%m-%d').date()
            
            # Obtener cita si se proporcionó
            cita = None
            if cita_id:
                try:
                    cita = Cita.objects.get(id=cita_id, plan_tratamiento=plan)
                except Cita.DoesNotExist:
                    pass
            
            # Crear el pago
            pago = PagoTratamiento.objects.create(
                plan_tratamiento=plan,
                monto=monto,
                fecha_pago=fecha_pago,
                metodo_pago=metodo_pago,
                numero_comprobante=numero_comprobante if numero_comprobante else None,
                notas=notas if notas else None,
                cita=cita,
                registrado_por=perfil
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Pago de ${monto:,.0f} registrado exitosamente.',
                'pago': {
                    'id': pago.id,
                    'monto': str(pago.monto),
                    'fecha_pago': pago.fecha_pago.strftime('%d/%m/%Y'),
                    'metodo_pago': pago.get_metodo_pago_display(),
                },
                'total_pagado': str(plan.total_pagado),
                'saldo_pendiente': str(plan.saldo_pendiente),
            })
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Error en los datos: {str(e)}'}, status=400)
        except Exception as e:
            logger.error(f"Error al registrar pago: {e}")
            return JsonResponse({'success': False, 'error': f'Error al registrar el pago: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def eliminar_pago_tratamiento(request, plan_id, pago_id):
    """Elimina un pago registrado (solo administrativos)"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return JsonResponse({'success': False, 'error': 'Solo los administrativos pueden eliminar pagos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    plan = get_object_or_404(PlanTratamiento, id=plan_id)
    pago = get_object_or_404(PagoTratamiento, id=pago_id, plan_tratamiento=plan)
    
    if request.method == 'POST':
        try:
            monto = pago.monto
            pago.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Pago de ${monto:,.0f} eliminado exitosamente.',
                'total_pagado': str(plan.total_pagado),
                'saldo_pendiente': str(plan.saldo_pendiente),
            })
            
        except Exception as e:
            logger.error(f"Error al eliminar pago: {e}")
            return JsonResponse({'success': False, 'error': f'Error al eliminar el pago: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


@login_required
def enviar_documentos_tratamiento(request, plan_id):
    """Vista para enviar presupuesto y consentimientos de un tratamiento por correo"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not (perfil.es_administrativo() or perfil.es_dentista()):
            return JsonResponse({'success': False, 'error': 'No tienes permisos para enviar documentos.'}, status=403)
    except Perfil.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No tienes permisos.'}, status=403)
    
    # Obtener el plan de tratamiento
    if perfil.es_administrativo():
        plan = get_object_or_404(PlanTratamiento, id=plan_id)
    else:
        plan = get_object_or_404(PlanTratamiento, id=plan_id, dentista=perfil)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    
    # Verificar que el cliente tenga email
    if not plan.cliente.email:
        return JsonResponse({'success': False, 'error': 'El cliente no tiene un email registrado.'}, status=400)
    
    try:
        from django.core.mail import EmailMultiAlternatives
        from django.conf import settings
        from io import BytesIO
        from datetime import datetime
        
        # Obtener información de la clínica
        try:
            from configuracion.models import InformacionClinica
            info_clinica = InformacionClinica.obtener()
            nombre_clinica = info_clinica.nombre_clinica
            email_clinica = info_clinica.email or getattr(settings, 'EMAIL_FROM', 'noreply@clinica.com')
        except:
            nombre_clinica = "Clínica Dental"
            email_clinica = getattr(settings, 'EMAIL_FROM', 'noreply@clinica.com')
        
        # Generar PDF del presupuesto
        presupuesto_response = exportar_presupuesto_pdf(request, plan.id)
        presupuesto_pdf = presupuesto_response.content
        
        # Obtener consentimientos pendientes o no firmados
        consentimientos = plan.consentimientos.filter(estado__in=['pendiente', 'firmado']).order_by('-fecha_creacion')
        
        # Crear el email
        asunto = f"Documentos del Tratamiento - {plan.nombre} - {nombre_clinica}"
        
        mensaje_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: linear-gradient(135deg, #14b8a6 0%, #0d9488 100%); min-height: 100vh;">
            <div style="max-width: 600px; margin: 0 auto; padding: 40px 20px;">
                <!-- Header con logo -->
                <div style="background: white; border-radius: 12px 12px 0 0; padding: 30px; text-align: center; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    <div style="display: inline-flex; align-items: center; gap: 12px; margin-bottom: 10px;">
                        <div style="width: 50px; height: 50px; background: linear-gradient(135deg, #14b8a6, #0d9488); border-radius: 50%; display: flex; align-items: center; justify-content: center; color: white; font-size: 1.8rem;">
                            <span style="font-size: 1.8rem;">🦷</span>
                        </div>
                        <h1 style="margin: 0; color: #0f766e; font-size: 1.75rem; font-weight: 700;">Clínica San Felipe</h1>
                    </div>
                    <p style="margin: 0; color: #64748b; font-size: 0.9rem;">Victoria, Región de la Araucanía</p>
                </div>
                
                <!-- Contenido del correo -->
                <div style="background: white; padding: 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    <h2 style="color: #1e293b; margin-top: 0; font-size: 1.5rem;">Estimado/a {plan.cliente.nombre_completo},</h2>
                    
                    <p style="color: #475569; line-height: 1.6; font-size: 1rem;">Le enviamos los documentos relacionados con su tratamiento: <strong style="color: #0f766e;">{plan.nombre}</strong></p>
                    
                    <div style="background: linear-gradient(135deg, #f0fdfa, #e0f2f1); padding: 20px; border-radius: 8px; margin: 25px 0; border-left: 4px solid #14b8a6;">
                        <h3 style="margin-top: 0; color: #0f766e; font-size: 1.1rem; margin-bottom: 12px;">Documentos adjuntos:</h3>
                        <ul style="margin: 0; padding-left: 20px; color: #475569;">
                            <li style="margin-bottom: 8px;"><strong style="color: #0f766e;">Presupuesto del Tratamiento</strong></li>
        """
        
        if consentimientos.exists():
            mensaje_html += f"<li style='margin-bottom: 8px;'><strong style='color: #0f766e;'>Consentimientos Informados</strong> ({consentimientos.count()} documento(s))</li>"
        
        mensaje_html += f"""
                        </ul>
                    </div>
                    
                    <p style="color: #475569; line-height: 1.6;">Por favor, revise cuidadosamente estos documentos. Si tiene alguna consulta, no dude en contactarnos.</p>
                    
                    <div style="background: #fef3c7; padding: 18px; border-radius: 8px; margin: 25px 0; border-left: 4px solid #f59e0b;">
                        <p style="margin: 0 0 12px 0; color: #92400e; font-weight: 600; font-size: 1rem;"><strong>Importante:</strong> Para proceder con el tratamiento, necesitamos que:</p>
                        <ul style="margin: 0; padding-left: 20px; color: #92400e;">
                            <li style="margin-bottom: 6px;">Acepte el presupuesto</li>
                            <li style="margin-bottom: 6px;">Firme el consentimiento informado</li>
                        </ul>
                    </div>
                    
                    <p style="color: #475569; line-height: 1.6;">Puede hacerlo desde su panel de cliente en nuestra página web o contactándonos directamente.</p>
                </div>
                
                <!-- Footer -->
                <div style="background: white; border-radius: 0 0 12px 12px; padding: 25px 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); border-top: 1px solid #e0f2f1;">
                    <p style="margin: 0; color: #64748b; font-size: 0.9rem; line-height: 1.6;">
                        Saludos cordiales,<br>
                        <strong style="color: #0f766e; font-size: 1rem;">{nombre_clinica}</strong>
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        mensaje_texto = f"""
        Estimado/a {plan.cliente.nombre_completo},
        
        Le enviamos los documentos relacionados con su tratamiento: {plan.nombre}
        
        Documentos adjuntos:
        - Presupuesto del Tratamiento
        - Consentimientos Informados
        
        Por favor, revise cuidadosamente estos documentos. Si tiene alguna consulta, no dude en contactarnos.
        
        Importante: Para proceder con el tratamiento, necesitamos que:
        - Acepte el presupuesto
        - Firme el consentimiento informado
        
        Puede hacerlo desde su panel de cliente en nuestra página web o contactándonos directamente.
        
        Saludos cordiales,
        {nombre_clinica}
        """
        
        # Crear el email
        email = EmailMultiAlternatives(
            asunto,
            mensaje_texto,
            email_clinica,
            [plan.cliente.email],
        )
        email.attach_alternative(mensaje_html, "text/html")
        
        # Adjuntar presupuesto
        filename_presupuesto = f"presupuesto_{plan.cliente.nombre_completo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        email.attach(filename_presupuesto, presupuesto_pdf, 'application/pdf')
        
        # Adjuntar consentimientos
        for consentimiento in consentimientos:
            try:
                consentimiento_response = exportar_consentimiento_pdf(request, consentimiento.id)
                consentimiento_pdf = consentimiento_response.content
                filename_consentimiento = f"consentimiento_{consentimiento.titulo.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
                email.attach(filename_consentimiento, consentimiento_pdf, 'application/pdf')
            except Exception as e:
                logger.error(f"Error al generar PDF del consentimiento {consentimiento.id}: {e}")
                # Continuar con los demás documentos
        
        # Enviar email
        email.send()
        
        # Actualizar fecha de envío en documentos relacionados
        from historial_clinico.models import DocumentoCliente
        documentos_actualizados = 0
        for consentimiento in consentimientos:
            # Actualizar el documento relacionado si existe
            try:
                documento = DocumentoCliente.objects.filter(
                    plan_tratamiento=plan,
                    tipo='consentimiento',
                    cliente=plan.cliente
                ).filter(
                    titulo__icontains=consentimiento.titulo
                ).first()
                
                if documento:
                    documento.enviado_por_correo = True
                    documento.fecha_envio = timezone.now()
                    documento.email_destinatario = plan.cliente.email
                    documento.save(update_fields=['enviado_por_correo', 'fecha_envio', 'email_destinatario'])
                    documentos_actualizados += 1
            except Exception as e:
                logger.warning(f"No se pudo actualizar documento para consentimiento {consentimiento.id}: {e}")
        
        # Actualizar también el documento del presupuesto si existe
        try:
            doc_presupuesto = DocumentoCliente.objects.filter(
                plan_tratamiento=plan,
                tipo='presupuesto',
                cliente=plan.cliente
            ).first()
            
            if doc_presupuesto:
                doc_presupuesto.enviado_por_correo = True
                doc_presupuesto.fecha_envio = timezone.now()
                doc_presupuesto.email_destinatario = plan.cliente.email
                doc_presupuesto.save(update_fields=['enviado_por_correo', 'fecha_envio', 'email_destinatario'])
                documentos_actualizados += 1
        except Exception as e:
            logger.warning(f"No se pudo actualizar documento de presupuesto: {e}")
        
        return JsonResponse({
            'success': True,
            'message': f'Documentos enviados exitosamente a {plan.cliente.email}. Se enviaron {1 + consentimientos.count()} documento(s).'
        })
        
    except Exception as e:
        logger.error(f"Error al enviar documentos del tratamiento: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'error': f'Error al enviar los documentos: {str(e)}'}, status=500)
