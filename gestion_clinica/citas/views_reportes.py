from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Count, Sum, Q, Avg, F
from django.utils import timezone
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

from .models import Cita, TipoServicio
from pacientes.models import Cliente
from inventario.models import Insumo, MovimientoInsumo
from personal.models import Perfil
from proveedores.models import Proveedor, SolicitudInsumo, Pedido
from historial_clinico.models import PlanTratamiento
from finanzas.models import IngresoManual, EgresoManual
from .models_auditoria import registrar_auditoria

# Color turquesa para Excel: #14B8A6 (primary) y #0D9488 (dark)
TURQUESA_PRIMARY = "14B8A6"
TURQUESA_DARK = "0D9488"
TURQUESA_LIGHT = "2DD4BF"
TURQUESA_BG = "F0FDFA"

@login_required
def reportes(request):
    """Vista principal de Reportes - Solo descargas Excel"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para acceder a los reportes.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    # Obtener fechas para estadísticas
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    fin_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    hace_7_dias = hoy - timedelta(days=7)
    
    # Estadísticas generales
    total_citas = Cita.objects.count()
    total_clientes = Cliente.objects.count()
    total_insumos = Insumo.objects.count()
    total_personal = Perfil.objects.filter(activo=True).count()
    total_proveedores = Proveedor.objects.filter(activo=True).count()
    total_solicitudes = SolicitudInsumo.objects.count()
    total_planes = PlanTratamiento.objects.count()
    total_servicios = TipoServicio.objects.filter(activo=True).count()
    
    # Estadísticas de citas
    citas_hoy = Cita.objects.filter(fecha_hora__date=hoy).count()
    citas_semana = Cita.objects.filter(fecha_hora__date__gte=hace_7_dias).count()
    citas_mes = Cita.objects.filter(
        fecha_hora__date__gte=inicio_mes,
        fecha_hora__date__lte=fin_mes
    ).count()
    citas_completadas_mes = Cita.objects.filter(
        estado='completada',
        fecha_hora__date__gte=inicio_mes,
        fecha_hora__date__lte=fin_mes
    ).count()
    citas_pendientes = Cita.objects.filter(estado__in=['reservada', 'confirmada']).count()
    
    # Estadísticas de clientes
    clientes_nuevos_mes = Cliente.objects.filter(
        fecha_registro__date__gte=inicio_mes,
        fecha_registro__date__lte=fin_mes
    ).count()
    clientes_activos = Cliente.objects.filter(activo=True).count()
    
    # Insumos con stock bajo
    insumos_bajo_stock = Insumo.objects.filter(
        cantidad_actual__lte=F('cantidad_minima')
    ).count()
    insumos_disponibles = Insumo.objects.filter(estado='disponible').count()
    
    # Finanzas
    total_ingresos = IngresoManual.objects.aggregate(Sum('monto'))['monto__sum'] or 0
    total_egresos = EgresoManual.objects.aggregate(Sum('monto'))['monto__sum'] or 0
    ingresos_mes = IngresoManual.objects.filter(
        fecha__gte=inicio_mes,
        fecha__lte=fin_mes
    ).aggregate(Sum('monto'))['monto__sum'] or 0
    egresos_mes = EgresoManual.objects.filter(
        fecha__gte=inicio_mes,
        fecha__lte=fin_mes
    ).aggregate(Sum('monto'))['monto__sum'] or 0
    balance_mes = ingresos_mes - egresos_mes
    
    # Planes de tratamiento
    planes_en_progreso = PlanTratamiento.objects.filter(estado='en_progreso').count()
    planes_completados = PlanTratamiento.objects.filter(estado='completado').count()
    
    # Solicitudes
    solicitudes_pendientes = SolicitudInsumo.objects.filter(estado='pendiente').count()
    solicitudes_enviadas = SolicitudInsumo.objects.filter(estado='enviada').count()
    
    context = {
        'perfil': perfil,
        'es_admin': True,
        # Totales generales
        'total_citas': total_citas,
        'total_clientes': total_clientes,
        'total_insumos': total_insumos,
        'total_personal': total_personal,
        'total_proveedores': total_proveedores,
        'total_solicitudes': total_solicitudes,
        'total_planes': total_planes,
        'total_servicios': total_servicios,
        # Estadísticas de citas
        'citas_hoy': citas_hoy,
        'citas_semana': citas_semana,
        'citas_mes': citas_mes,
        'citas_completadas_mes': citas_completadas_mes,
        'citas_pendientes': citas_pendientes,
        # Estadísticas de clientes
        'clientes_nuevos_mes': clientes_nuevos_mes,
        'clientes_activos': clientes_activos,
        # Insumos
        'insumos_bajo_stock': insumos_bajo_stock,
        'insumos_disponibles': insumos_disponibles,
        # Finanzas
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'ingresos_mes': ingresos_mes,
        'egresos_mes': egresos_mes,
        'balance_mes': balance_mes,
        # Planes
        'planes_en_progreso': planes_en_progreso,
        'planes_completados': planes_completados,
        # Solicitudes
        'solicitudes_pendientes': solicitudes_pendientes,
        'solicitudes_enviadas': solicitudes_enviadas,
        # Fechas
        'mes_actual': inicio_mes.strftime('%B %Y'),
    }
    
    return render(request, 'citas/reportes/reportes.html', context)


# ========== FUNCIONES DE EXPORTACIÓN A EXCEL CON DISEÑO TURQUESA ==========

def _crear_encabezado_excel(ws, titulo, subtitulo=None):
    """Crea un encabezado profesional con diseño turquesa"""
    # Título principal
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=16, color=TURQUESA_DARK)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:J1')
    
    # Subtítulo con fecha
    fecha_gen = datetime.now().strftime("%d/%m/%Y %H:%M")
    if subtitulo:
        ws['A2'] = f"{subtitulo} - Generado el {fecha_gen}"
    else:
        ws['A2'] = f"Generado el {fecha_gen}"
    ws['A2'].font = Font(size=11, color="666666")
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:J2')
    
    # Espacio
    return 4  # Retorna la fila donde empezar los datos


def _aplicar_estilo_turquesa(ws, fila_inicio, headers):
    """Aplica estilo turquesa a los encabezados"""
    header_fill = PatternFill(start_color=TURQUESA_PRIMARY, end_color=TURQUESA_DARK, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=fila_inicio, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    return fila_inicio + 1, border  # Retorna la siguiente fila y el estilo de borde


@login_required
def exportar_excel_citas(request):
    """Exporta todas las citas a Excel con diseño turquesa"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    except Exception as e:
        logger.error(f"Error al verificar permisos en exportar_excel_citas: {str(e)}")
        return HttpResponse('Error al procesar la solicitud', status=500)
    
    # Registrar en auditoría
    registrar_auditoria(
        usuario=perfil,
        accion='exportar',
        modulo='citas',
        descripcion='Exportación de citas a Excel',
        detalles='Archivo Excel generado con todas las citas',
        request=request
    )
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Citas"
        
        # Encabezado
        fila_datos = _crear_encabezado_excel(ws, "REPORTE DE CITAS", "Clínica Dental")
        
        # Encabezados de columna
        headers = ['ID', 'Fecha', 'Hora', 'Tipo Consulta', 'Estado', 'Paciente', 'Email', 'Teléfono', 'Dentista', 'Notas']
        fila_actual, border = _aplicar_estilo_turquesa(ws, fila_datos, headers)
        
        # Datos
        citas = Cita.objects.select_related('dentista').all().order_by('-fecha_hora')
        for cita in citas:
            try:
                ws.cell(row=fila_actual, column=1, value=cita.id)
                ws.cell(row=fila_actual, column=2, value=cita.fecha_hora.strftime('%d/%m/%Y') if cita.fecha_hora else 'N/A')
                ws.cell(row=fila_actual, column=3, value=cita.fecha_hora.strftime('%H:%M') if cita.fecha_hora else 'N/A')
                ws.cell(row=fila_actual, column=4, value=str(cita.tipo_consulta) if cita.tipo_consulta else 'N/A')
                ws.cell(row=fila_actual, column=5, value=cita.get_estado_display() if cita.estado else 'N/A')
                ws.cell(row=fila_actual, column=6, value=str(cita.paciente_nombre) if cita.paciente_nombre else 'Sin asignar')
                ws.cell(row=fila_actual, column=7, value=str(cita.paciente_email) if cita.paciente_email else '')
                ws.cell(row=fila_actual, column=8, value=str(cita.paciente_telefono) if cita.paciente_telefono else '')
                ws.cell(row=fila_actual, column=9, value=cita.dentista.nombre_completo if cita.dentista and hasattr(cita.dentista, 'nombre_completo') else 'Sin asignar')
                ws.cell(row=fila_actual, column=10, value=str(cita.notas) if cita.notas else '')
                
                # Aplicar bordes y estilo alternado
                for col in range(1, 11):
                    cell = ws.cell(row=fila_actual, column=col)
                    cell.border = border
                    if fila_actual % 2 == 0:
                        cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                
                fila_actual += 1
            except Exception as e:
                logger.error(f"Error al procesar cita ID {cita.id}: {str(e)}")
                continue
        
        # Ajustar anchos
        column_widths = [8, 12, 10, 20, 12, 25, 30, 15, 25, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Altura de fila de encabezado
        ws.row_dimensions[fila_datos].height = 25
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=citas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        wb.save(response)
        
        return response
    except Exception as e:
        logger.error(f"Error al generar Excel de citas: {str(e)}")
        return HttpResponse('Error al generar el archivo Excel', status=500)


@login_required
def exportar_excel_clientes(request):
    """Exporta todos los clientes a Excel con diseño turquesa"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    except Exception as e:
        logger.error(f"Error al verificar permisos en exportar_excel_clientes: {str(e)}")
        return HttpResponse('Error al procesar la solicitud', status=500)
    
    # Registrar en auditoría
    registrar_auditoria(
        usuario=perfil,
        accion='exportar',
        modulo='clientes',
        descripcion='Exportación de clientes a Excel',
        detalles='Archivo Excel generado con todos los clientes',
        request=request
    )
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        fila_datos = _crear_encabezado_excel(ws, "REPORTE DE CLIENTES", "Clínica Dental")
        
        headers = ['ID', 'Nombre Completo', 'Email', 'Teléfono', 'Fecha Registro', 'Activo', 'Dentista Asignado', 'Total Citas', 'Notas']
        fila_actual, border = _aplicar_estilo_turquesa(ws, fila_datos, headers)
        
        clientes = Cliente.objects.select_related('dentista_asignado').annotate(num_citas=Count('citas')).order_by('nombre_completo')
        for cliente in clientes:
            try:
                ws.cell(row=fila_actual, column=1, value=cliente.id)
                ws.cell(row=fila_actual, column=2, value=str(cliente.nombre_completo) if cliente.nombre_completo else 'N/A')
                ws.cell(row=fila_actual, column=3, value=str(cliente.email) if cliente.email else '')
                ws.cell(row=fila_actual, column=4, value=str(cliente.telefono) if cliente.telefono else '')
                ws.cell(row=fila_actual, column=5, value=cliente.fecha_registro.strftime('%d/%m/%Y %H:%M') if cliente.fecha_registro else 'N/A')
                ws.cell(row=fila_actual, column=6, value='Sí' if cliente.activo else 'No')
                ws.cell(row=fila_actual, column=7, value=cliente.dentista_asignado.nombre_completo if cliente.dentista_asignado and hasattr(cliente.dentista_asignado, 'nombre_completo') else 'Sin asignar')
                ws.cell(row=fila_actual, column=8, value=cliente.num_citas or 0)
                ws.cell(row=fila_actual, column=9, value=str(cliente.notas) if cliente.notas else '')
                
                for col in range(1, 10):
                    cell = ws.cell(row=fila_actual, column=col)
                    cell.border = border
                    if fila_actual % 2 == 0:
                        cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                
                fila_actual += 1
            except Exception as e:
                logger.error(f"Error al procesar cliente ID {cliente.id}: {str(e)}")
                continue
        
        column_widths = [8, 30, 35, 15, 20, 10, 25, 12, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        ws.row_dimensions[fila_datos].height = 25
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=clientes_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        wb.save(response)
        
        return response
    except Exception as e:
        logger.error(f"Error al generar Excel de clientes: {str(e)}")
        return HttpResponse('Error al generar el archivo Excel', status=500)


@login_required
def exportar_excel_insumos(request):
    """Exporta todos los insumos a Excel con diseño turquesa"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    except Exception as e:
        logger.error(f"Error al verificar permisos en exportar_excel_insumos: {str(e)}")
        return HttpResponse('Error al procesar la solicitud', status=500)
    
    # Registrar en auditoría
    registrar_auditoria(
        usuario=perfil,
        accion='exportar',
        modulo='inventario',
        descripcion='Exportación de insumos a Excel',
        detalles='Archivo Excel generado con todos los insumos',
        request=request
    )
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Insumos"
        
        fila_datos = _crear_encabezado_excel(ws, "REPORTE DE INSUMOS", "Clínica Dental")
        
        headers = ['ID', 'Nombre', 'Categoría', 'Cantidad Actual', 'Cantidad Mínima', 'Unidad', 'Precio Unitario', 'Estado', 'Proveedor', 'Ubicación']
        fila_actual, border = _aplicar_estilo_turquesa(ws, fila_datos, headers)
        
        insumos = Insumo.objects.select_related('proveedor_principal').all().order_by('nombre')
        for insumo in insumos:
            try:
                ws.cell(row=fila_actual, column=1, value=insumo.id)
                ws.cell(row=fila_actual, column=2, value=str(insumo.nombre) if insumo.nombre else 'Sin nombre')
                ws.cell(row=fila_actual, column=3, value=insumo.get_categoria_display() if insumo.categoria else 'N/A')
                ws.cell(row=fila_actual, column=4, value=insumo.cantidad_actual if insumo.cantidad_actual is not None else 0)
                ws.cell(row=fila_actual, column=5, value=insumo.cantidad_minima if insumo.cantidad_minima is not None else 0)
                ws.cell(row=fila_actual, column=6, value=str(insumo.unidad_medida) if insumo.unidad_medida else 'N/A')
                ws.cell(row=fila_actual, column=7, value=float(insumo.precio_unitario) if insumo.precio_unitario else 0)
                ws.cell(row=fila_actual, column=8, value=insumo.get_estado_display() if insumo.estado else 'N/A')
                ws.cell(row=fila_actual, column=9, value=insumo.proveedor_principal.nombre if insumo.proveedor_principal and hasattr(insumo.proveedor_principal, 'nombre') else (str(insumo.proveedor_texto) if insumo.proveedor_texto else 'Sin proveedor'))
                ws.cell(row=fila_actual, column=10, value=str(insumo.ubicacion) if insumo.ubicacion else '')
                
                for col in range(1, 11):
                    cell = ws.cell(row=fila_actual, column=col)
                    cell.border = border
                    if fila_actual % 2 == 0:
                        cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                
                fila_actual += 1
            except Exception as e:
                logger.error(f"Error al procesar insumo ID {insumo.id}: {str(e)}")
                continue
        
        column_widths = [8, 30, 20, 15, 15, 12, 15, 15, 25, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        ws.row_dimensions[fila_datos].height = 25
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=insumos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        wb.save(response)
        
        return response
    except Exception as e:
        logger.error(f"Error al generar Excel de insumos: {str(e)}")
        return HttpResponse('Error al generar el archivo Excel', status=500)


@login_required
def exportar_excel_finanzas(request):
    """Exporta reporte financiero a Excel con diseño turquesa"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    except Exception as e:
        logger.error(f"Error al verificar permisos en exportar_excel_finanzas: {str(e)}")
        return HttpResponse('Error al procesar la solicitud', status=500)
    
    # Verificar si se solicita solo ingresos o egresos
    tipo_exportacion = request.GET.get('tipo', '').lower()
    
    # Registrar en auditoría
    detalles_exportacion = f'Tipo: {tipo_exportacion if tipo_exportacion else "Completo"}'
    registrar_auditoria(
        usuario=perfil,
        accion='exportar',
        modulo='finanzas',
        descripcion='Exportación de finanzas a Excel',
        detalles=detalles_exportacion,
        request=request
    )
    
    try:
        wb = Workbook()
        
        # Si se solicita solo ingresos, generar solo esa hoja
        if tipo_exportacion == 'ingresos':
            ws = wb.active
            ws.title = "Ingresos"
            fila_datos = _crear_encabezado_excel(ws, "LISTADO DE INGRESOS", "Clínica Dental")
            
            headers = ['Tipo', 'Fecha', 'Monto', 'Descripción', 'Cliente/Dentista', 'Creado por']
            fila_actual, border = _aplicar_estilo_turquesa(ws, fila_datos, headers)
            
            # Obtener citas completadas (excluir las eliminadas del historial)
            citas_completadas = Cita.objects.filter(
                estado='completada'
            ).exclude(
                precio_cobrado__isnull=True,
                tipo_servicio__isnull=True
            ).select_related('tipo_servicio', 'cliente', 'dentista').order_by('-fecha_hora')
            
            # Agregar TODAS las citas completadas (sin límite)
            for cita in citas_completadas:
                try:
                    monto_cita = 0
                    if cita.precio_cobrado:
                        monto_cita = float(cita.precio_cobrado)
                    elif cita.tipo_servicio and cita.tipo_servicio.precio_base:
                        monto_cita = float(cita.tipo_servicio.precio_base)
                    
                    if monto_cita > 0:  # Solo incluir si tiene monto
                        ws.cell(row=fila_actual, column=1, value='CITA')
                        ws.cell(row=fila_actual, column=2, value=cita.fecha_hora.strftime('%d/%m/%Y %H:%M') if cita.fecha_hora else 'N/A')
                        ws.cell(row=fila_actual, column=3, value=monto_cita)
                        descripcion = cita.tipo_servicio.nombre if cita.tipo_servicio else (cita.tipo_consulta or 'Cita')
                        ws.cell(row=fila_actual, column=4, value=str(descripcion))
                        cliente_info = ''
                        if cita.cliente:
                            cliente_info = cita.cliente.nombre_completo or 'Sin cliente'
                        if cita.dentista:
                            cliente_info += f' - Dr/a. {cita.dentista.nombre_completo}' if cliente_info else f'Dr/a. {cita.dentista.nombre_completo}'
                        ws.cell(row=fila_actual, column=5, value=cliente_info or 'N/A')
                        ws.cell(row=fila_actual, column=6, value='Sistema')
                        
                        for col in range(1, 7):
                            cell = ws.cell(row=fila_actual, column=col)
                            cell.border = border
                            if fila_actual % 2 == 0:
                                cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                        
                        fila_actual += 1
                except Exception as e:
                    logger.error(f"Error al procesar cita ID {cita.id}: {str(e)}")
                    continue
            
            # Agregar TODOS los ingresos manuales (sin límite)
            ingresos_manuales = IngresoManual.objects.select_related('creado_por').order_by('-fecha', '-creado_el')
            for ingreso in ingresos_manuales:
                try:
                    ws.cell(row=fila_actual, column=1, value='MANUAL')
                    ws.cell(row=fila_actual, column=2, value=ingreso.fecha.strftime('%d/%m/%Y') if ingreso.fecha else 'N/A')
                    ws.cell(row=fila_actual, column=3, value=float(ingreso.monto) if ingreso.monto else 0)
                    ws.cell(row=fila_actual, column=4, value=str(ingreso.descripcion) if ingreso.descripcion else '')
                    ws.cell(row=fila_actual, column=5, value='-')
                    ws.cell(row=fila_actual, column=6, value=ingreso.creado_por.nombre_completo if ingreso.creado_por and hasattr(ingreso.creado_por, 'nombre_completo') else 'N/A')
                    
                    for col in range(1, 7):
                        cell = ws.cell(row=fila_actual, column=col)
                        cell.border = border
                        if fila_actual % 2 == 0:
                            cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                    
                    fila_actual += 1
                except Exception as e:
                    logger.error(f"Error al procesar ingreso ID {ingreso.id}: {str(e)}")
                    continue
            
            # Agregar fila de totales
            fila_totales = fila_actual + 1
            ws.cell(row=fila_totales, column=1, value='TOTAL').font = Font(bold=True)
            ws.cell(row=fila_totales, column=2, value='')
            total_ingresos = sum([float(ws.cell(row=i, column=3).value or 0) for i in range(fila_datos + 1, fila_actual)])
            ws.cell(row=fila_totales, column=3, value=total_ingresos).font = Font(bold=True)
            ws.cell(row=fila_totales, column=4, value='')
            ws.cell(row=fila_totales, column=5, value='')
            ws.cell(row=fila_totales, column=6, value='')
            
            for col in range(1, 7):
                cell = ws.cell(row=fila_totales, column=col)
                cell.border = border
                cell.fill = PatternFill(start_color=TURQUESA_PRIMARY, end_color=TURQUESA_PRIMARY, fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            column_widths = [10, 16, 15, 40, 30, 25]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            ws.row_dimensions[fila_datos].height = 25
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=ingresos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
            wb.save(response)
            return response
        
        # Si se solicita solo egresos, generar solo esa hoja
        elif tipo_exportacion == 'egresos':
            ws = wb.active
            ws.title = "Egresos"
            fila_datos = _crear_encabezado_excel(ws, "LISTADO DE EGRESOS", "Clínica Dental")
            
            headers = ['Tipo', 'Fecha', 'Monto', 'Descripción', 'Cantidad', 'Creado por']
            fila_actual, border = _aplicar_estilo_turquesa(ws, fila_datos, headers)
            
            # Obtener movimientos de insumos tipo "entrada" (compras) como egresos
            movimientos_entrada = MovimientoInsumo.objects.filter(tipo='entrada').select_related('insumo', 'realizado_por').order_by('-fecha_movimiento')
            
            # Obtener solicitudes de insumos marcadas como egreso automático
            solicitudes_egreso = SolicitudInsumo.objects.filter(
                monto_egreso__isnull=False
            ).select_related('insumo', 'solicitado_por').order_by('-fecha_solicitud')
            
            # Obtener egresos manuales
            egresos_manuales = EgresoManual.objects.select_related('creado_por').order_by('-fecha', '-creado_el')
            
            # Agregar TODAS las compras (sin límite)
            for movimiento in movimientos_entrada:
                try:
                    if movimiento.insumo and movimiento.insumo.precio_unitario:
                        monto_movimiento = float(movimiento.insumo.precio_unitario) * movimiento.cantidad
                        if monto_movimiento > 0:
                            ws.cell(row=fila_actual, column=1, value='COMPRA')
                            ws.cell(row=fila_actual, column=2, value=movimiento.fecha_movimiento.strftime('%d/%m/%Y %H:%M') if movimiento.fecha_movimiento else 'N/A')
                            ws.cell(row=fila_actual, column=3, value=monto_movimiento)
                            ws.cell(row=fila_actual, column=4, value=movimiento.insumo.nombre or 'Sin nombre')
                            ws.cell(row=fila_actual, column=5, value=f"{movimiento.cantidad} {movimiento.insumo.unidad_medida or ''}")
                            creado_por = movimiento.realizado_por.nombre_completo if movimiento.realizado_por and hasattr(movimiento.realizado_por, 'nombre_completo') else 'Sistema'
                            ws.cell(row=fila_actual, column=6, value=creado_por)
                            
                            for col in range(1, 7):
                                cell = ws.cell(row=fila_actual, column=col)
                                cell.border = border
                                if fila_actual % 2 == 0:
                                    cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                            
                            fila_actual += 1
                except Exception as e:
                    logger.error(f"Error al procesar movimiento ID {movimiento.id}: {str(e)}")
                    continue
            
            # Agregar TODAS las solicitudes (sin límite)
            for solicitud in solicitudes_egreso:
                try:
                    if solicitud.monto_egreso:
                        ws.cell(row=fila_actual, column=1, value='SOLICITUD')
                        ws.cell(row=fila_actual, column=2, value=solicitud.fecha_solicitud.strftime('%d/%m/%Y %H:%M') if solicitud.fecha_solicitud else 'N/A')
                        ws.cell(row=fila_actual, column=3, value=float(solicitud.monto_egreso))
                        ws.cell(row=fila_actual, column=4, value=solicitud.insumo.nombre if solicitud.insumo else 'Sin insumo')
                        ws.cell(row=fila_actual, column=5, value=f"{solicitud.cantidad_solicitada} {solicitud.insumo.unidad_medida if solicitud.insumo else ''}")
                        creado_por = solicitud.solicitado_por.nombre_completo if solicitud.solicitado_por and hasattr(solicitud.solicitado_por, 'nombre_completo') else 'Sistema'
                        ws.cell(row=fila_actual, column=6, value=creado_por)
                        
                        for col in range(1, 7):
                            cell = ws.cell(row=fila_actual, column=col)
                            cell.border = border
                            if fila_actual % 2 == 0:
                                cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                        
                        fila_actual += 1
                except Exception as e:
                    logger.error(f"Error al procesar solicitud ID {solicitud.id}: {str(e)}")
                    continue
            
            # Agregar TODOS los egresos manuales (sin límite)
            for egreso in egresos_manuales:
                try:
                    ws.cell(row=fila_actual, column=1, value='MANUAL')
                    ws.cell(row=fila_actual, column=2, value=egreso.fecha.strftime('%d/%m/%Y') if egreso.fecha else 'N/A')
                    ws.cell(row=fila_actual, column=3, value=float(egreso.monto) if egreso.monto else 0)
                    ws.cell(row=fila_actual, column=4, value=str(egreso.descripcion) if egreso.descripcion else '')
                    ws.cell(row=fila_actual, column=5, value='-')
                    ws.cell(row=fila_actual, column=6, value=egreso.creado_por.nombre_completo if egreso.creado_por and hasattr(egreso.creado_por, 'nombre_completo') else 'N/A')
                    
                    for col in range(1, 7):
                        cell = ws.cell(row=fila_actual, column=col)
                        cell.border = border
                        if fila_actual % 2 == 0:
                            cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                    
                    fila_actual += 1
                except Exception as e:
                    logger.error(f"Error al procesar egreso ID {egreso.id}: {str(e)}")
                    continue
            
            # Agregar fila de totales
            fila_totales = fila_actual + 1
            ws.cell(row=fila_totales, column=1, value='TOTAL').font = Font(bold=True)
            ws.cell(row=fila_totales, column=2, value='')
            total_egresos = sum([float(ws.cell(row=i, column=3).value or 0) for i in range(fila_datos + 1, fila_actual)])
            ws.cell(row=fila_totales, column=3, value=total_egresos).font = Font(bold=True)
            ws.cell(row=fila_totales, column=4, value='')
            ws.cell(row=fila_totales, column=5, value='')
            ws.cell(row=fila_totales, column=6, value='')
            
            for col in range(1, 7):
                cell = ws.cell(row=fila_totales, column=col)
                cell.border = border
                cell.fill = PatternFill(start_color=TURQUESA_PRIMARY, end_color=TURQUESA_PRIMARY, fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            column_widths = [12, 16, 15, 40, 15, 25]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            ws.row_dimensions[fila_datos].height = 25
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=egresos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
            wb.save(response)
            return response
        
        # Si no se especifica tipo, generar reporte completo
        # Hoja 1: Resumen
        ws1 = wb.active
        ws1.title = "Resumen"
        
        fila_datos = _crear_encabezado_excel(ws1, "REPORTE FINANCIERO", "Clínica Dental")
        
        # Estadísticas financieras (usando la misma lógica que gestor_finanzas)
        hoy = timezone.now().date()
        inicio_mes = hoy.replace(day=1)
        fin_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        # ===== INGRESOS =====
        # Obtener citas completadas (excluir las eliminadas del historial)
        citas_completadas = Cita.objects.filter(
            estado='completada'
        ).exclude(
            precio_cobrado__isnull=True,
            tipo_servicio__isnull=True
        ).select_related('tipo_servicio')
        
        # Calcular ingresos totales (citas + manuales)
        total_ingresos = 0
        for cita in citas_completadas:
            if cita.precio_cobrado:
                total_ingresos += float(cita.precio_cobrado)
            elif cita.tipo_servicio and cita.tipo_servicio.precio_base:
                total_ingresos += float(cita.tipo_servicio.precio_base)
        
        # Sumar ingresos manuales
        ingresos_manuales = IngresoManual.objects.all()
        for ingreso_manual in ingresos_manuales:
            total_ingresos += float(ingreso_manual.monto)
        
        # Ingresos del mes
        citas_mes = citas_completadas.filter(
            fecha_hora__date__gte=inicio_mes,
            fecha_hora__date__lte=fin_mes
        )
        ingresos_mes = 0
        for cita in citas_mes:
            if cita.precio_cobrado:
                ingresos_mes += float(cita.precio_cobrado)
            elif cita.tipo_servicio and cita.tipo_servicio.precio_base:
                ingresos_mes += float(cita.tipo_servicio.precio_base)
        
        # Sumar ingresos manuales del mes
        ingresos_manuales_mes = ingresos_manuales.filter(
            fecha__gte=inicio_mes,
            fecha__lte=fin_mes
        )
        for ingreso_manual in ingresos_manuales_mes:
            ingresos_mes += float(ingreso_manual.monto)
        
        # ===== EGRESOS =====
        # Obtener movimientos de insumos tipo "entrada" (compras) como egresos
        movimientos_entrada = MovimientoInsumo.objects.filter(tipo='entrada').select_related('insumo')
        
        # Obtener solicitudes de insumos marcadas como egreso automático
        solicitudes_egreso = SolicitudInsumo.objects.filter(
            monto_egreso__isnull=False
        )
        
        # Obtener egresos manuales
        egresos_manuales = EgresoManual.objects.all()
        
        # Calcular egresos totales
        total_egresos = 0
        for movimiento in movimientos_entrada:
            if movimiento.insumo and movimiento.insumo.precio_unitario:
                total_egresos += float(movimiento.insumo.precio_unitario) * movimiento.cantidad
        
        # Sumar egresos automáticos de solicitudes
        for solicitud in solicitudes_egreso:
            if solicitud.monto_egreso:
                total_egresos += float(solicitud.monto_egreso)
        
        # Sumar egresos manuales
        for egreso_manual in egresos_manuales:
            total_egresos += float(egreso_manual.monto)
        
        # Egresos del mes
        movimientos_mes = movimientos_entrada.filter(
            fecha_movimiento__date__gte=inicio_mes,
            fecha_movimiento__date__lte=fin_mes
        )
        egresos_mes = 0
        for movimiento in movimientos_mes:
            if movimiento.insumo and movimiento.insumo.precio_unitario:
                egresos_mes += float(movimiento.insumo.precio_unitario) * movimiento.cantidad
        
        # Sumar egresos automáticos del mes
        solicitudes_mes = solicitudes_egreso.filter(
            fecha_solicitud__date__gte=inicio_mes,
            fecha_solicitud__date__lte=fin_mes
        )
        for solicitud in solicitudes_mes:
            if solicitud.monto_egreso:
                egresos_mes += float(solicitud.monto_egreso)
        
        # Sumar egresos manuales del mes
        egresos_manuales_mes = egresos_manuales.filter(
            fecha__gte=inicio_mes,
            fecha__lte=fin_mes
        )
        for egreso_manual in egresos_manuales_mes:
            egresos_mes += float(egreso_manual.monto)
        
        # Balance
        balance_total = total_ingresos - total_egresos
        balance_mes = ingresos_mes - egresos_mes
        
        # Encabezados
        headers = ['Concepto', 'Valor']
        fila_actual, border = _aplicar_estilo_turquesa(ws1, fila_datos, headers)
        
        # Datos
        datos_financieros = [
            ('Total Ingresos', f'${float(total_ingresos):,.0f}'),
            ('Total Egresos', f'${float(total_egresos):,.0f}'),
            ('Balance Total', f'${float(balance_total):,.0f}'),
            ('Ingresos del Mes', f'${float(ingresos_mes):,.0f}'),
            ('Egresos del Mes', f'${float(egresos_mes):,.0f}'),
            ('Balance del Mes', f'${float(balance_mes):,.0f}'),
        ]
        
        for concepto, valor in datos_financieros:
            ws1.cell(row=fila_actual, column=1, value=concepto)
            ws1.cell(row=fila_actual, column=2, value=valor)
            
            for col in range(1, 3):
                cell = ws1.cell(row=fila_actual, column=col)
                cell.border = border
                if fila_actual % 2 == 0:
                    cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
            
            fila_actual += 1
        
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 20
        ws1.row_dimensions[fila_datos].height = 25
        
        # Hoja 2: Ingresos (Citas + Manuales)
        ws2 = wb.create_sheet("Ingresos")
        fila_datos2 = _crear_encabezado_excel(ws2, "INGRESOS", "Clínica Dental")
        
        headers2 = ['Tipo', 'Fecha', 'Monto', 'Descripción', 'Cliente/Dentista', 'Creado por']
        fila_actual2, border2 = _aplicar_estilo_turquesa(ws2, fila_datos2, headers2)
        
        # Agregar citas completadas (sin límite para reporte completo)
        for cita in citas_completadas.order_by('-fecha_hora'):
            try:
                monto_cita = 0
                if cita.precio_cobrado:
                    monto_cita = float(cita.precio_cobrado)
                elif cita.tipo_servicio and cita.tipo_servicio.precio_base:
                    monto_cita = float(cita.tipo_servicio.precio_base)
                
                if monto_cita > 0:  # Solo incluir si tiene monto
                    ws2.cell(row=fila_actual2, column=1, value='CITA')
                    ws2.cell(row=fila_actual2, column=2, value=cita.fecha_hora.strftime('%d/%m/%Y %H:%M') if cita.fecha_hora else 'N/A')
                    ws2.cell(row=fila_actual2, column=3, value=monto_cita)
                    descripcion = cita.tipo_servicio.nombre if cita.tipo_servicio else (cita.tipo_consulta or 'Cita')
                    ws2.cell(row=fila_actual2, column=4, value=str(descripcion))
                    cliente_info = ''
                    if cita.cliente:
                        cliente_info = cita.cliente.nombre_completo or 'Sin cliente'
                    if cita.dentista:
                        cliente_info += f' - Dr/a. {cita.dentista.nombre_completo}' if cliente_info else f'Dr/a. {cita.dentista.nombre_completo}'
                    ws2.cell(row=fila_actual2, column=5, value=cliente_info or 'N/A')
                    ws2.cell(row=fila_actual2, column=6, value='Sistema')
                    
                    for col in range(1, 7):
                        cell = ws2.cell(row=fila_actual2, column=col)
                        cell.border = border2
                        if fila_actual2 % 2 == 0:
                            cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                    
                    fila_actual2 += 1
            except Exception as e:
                logger.error(f"Error al procesar cita ID {cita.id}: {str(e)}")
                continue
        
        # Agregar ingresos manuales
        for ingreso in ingresos_manuales.select_related('creado_por').order_by('-fecha', '-creado_el'):
            try:
                ws2.cell(row=fila_actual2, column=1, value='MANUAL')
                ws2.cell(row=fila_actual2, column=2, value=ingreso.fecha.strftime('%d/%m/%Y') if ingreso.fecha else 'N/A')
                ws2.cell(row=fila_actual2, column=3, value=float(ingreso.monto) if ingreso.monto else 0)
                ws2.cell(row=fila_actual2, column=4, value=str(ingreso.descripcion) if ingreso.descripcion else '')
                ws2.cell(row=fila_actual2, column=5, value='-')
                ws2.cell(row=fila_actual2, column=6, value=ingreso.creado_por.nombre_completo if ingreso.creado_por and hasattr(ingreso.creado_por, 'nombre_completo') else 'N/A')
                
                for col in range(1, 7):
                    cell = ws2.cell(row=fila_actual2, column=col)
                    cell.border = border2
                    if fila_actual2 % 2 == 0:
                        cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                
                fila_actual2 += 1
            except Exception as e:
                logger.error(f"Error al procesar ingreso ID {ingreso.id}: {str(e)}")
                continue
        
        column_widths2 = [10, 16, 15, 40, 30, 25]
        for i, width in enumerate(column_widths2, 1):
            ws2.column_dimensions[get_column_letter(i)].width = width
        ws2.row_dimensions[fila_datos2].height = 25
        
        # Hoja 3: Egresos (Compras + Solicitudes + Manuales)
        ws3 = wb.create_sheet("Egresos")
        fila_datos3 = _crear_encabezado_excel(ws3, "EGRESOS", "Clínica Dental")
        
        headers3 = ['Tipo', 'Fecha', 'Monto', 'Descripción', 'Cantidad', 'Creado por']
        fila_actual3, border3 = _aplicar_estilo_turquesa(ws3, fila_datos3, headers3)
        
        # Agregar compras de inventario
        for movimiento in movimientos_entrada.order_by('-fecha_movimiento')[:100]:  # Limitar a 100 más recientes
            try:
                if movimiento.insumo and movimiento.insumo.precio_unitario:
                    monto_movimiento = float(movimiento.insumo.precio_unitario) * movimiento.cantidad
                    if monto_movimiento > 0:
                        ws3.cell(row=fila_actual3, column=1, value='COMPRA')
                        ws3.cell(row=fila_actual3, column=2, value=movimiento.fecha_movimiento.strftime('%d/%m/%Y %H:%M') if movimiento.fecha_movimiento else 'N/A')
                        ws3.cell(row=fila_actual3, column=3, value=monto_movimiento)
                        ws3.cell(row=fila_actual3, column=4, value=movimiento.insumo.nombre or 'Sin nombre')
                        ws3.cell(row=fila_actual3, column=5, value=f"{movimiento.cantidad} {movimiento.insumo.unidad_medida or ''}")
                        creado_por = movimiento.realizado_por.nombre_completo if movimiento.realizado_por and hasattr(movimiento.realizado_por, 'nombre_completo') else 'Sistema'
                        ws3.cell(row=fila_actual3, column=6, value=creado_por)
                        
                        for col in range(1, 7):
                            cell = ws3.cell(row=fila_actual3, column=col)
                            cell.border = border3
                            if fila_actual3 % 2 == 0:
                                cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                        
                        fila_actual3 += 1
            except Exception as e:
                logger.error(f"Error al procesar movimiento ID {movimiento.id}: {str(e)}")
                continue
        
        # Agregar solicitudes de proveedores
        for solicitud in solicitudes_egreso.order_by('-fecha_solicitud')[:100]:
            try:
                if solicitud.monto_egreso:
                    ws3.cell(row=fila_actual3, column=1, value='SOLICITUD')
                    ws3.cell(row=fila_actual3, column=2, value=solicitud.fecha_solicitud.strftime('%d/%m/%Y %H:%M') if solicitud.fecha_solicitud else 'N/A')
                    ws3.cell(row=fila_actual3, column=3, value=float(solicitud.monto_egreso))
                    ws3.cell(row=fila_actual3, column=4, value=solicitud.insumo.nombre if solicitud.insumo else 'Sin insumo')
                    ws3.cell(row=fila_actual3, column=5, value=f"{solicitud.cantidad_solicitada} {solicitud.insumo.unidad_medida if solicitud.insumo else ''}")
                    creado_por = solicitud.solicitado_por.nombre_completo if solicitud.solicitado_por and hasattr(solicitud.solicitado_por, 'nombre_completo') else 'Sistema'
                    ws3.cell(row=fila_actual3, column=6, value=creado_por)
                    
                    for col in range(1, 7):
                        cell = ws3.cell(row=fila_actual3, column=col)
                        cell.border = border3
                        if fila_actual3 % 2 == 0:
                            cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                    
                    fila_actual3 += 1
            except Exception as e:
                logger.error(f"Error al procesar solicitud ID {solicitud.id}: {str(e)}")
                continue
        
        # Agregar egresos manuales
        for egreso in egresos_manuales.select_related('creado_por').order_by('-fecha', '-creado_el'):
            try:
                ws3.cell(row=fila_actual3, column=1, value='MANUAL')
                ws3.cell(row=fila_actual3, column=2, value=egreso.fecha.strftime('%d/%m/%Y') if egreso.fecha else 'N/A')
                ws3.cell(row=fila_actual3, column=3, value=float(egreso.monto) if egreso.monto else 0)
                ws3.cell(row=fila_actual3, column=4, value=str(egreso.descripcion) if egreso.descripcion else '')
                ws3.cell(row=fila_actual3, column=5, value='-')
                ws3.cell(row=fila_actual3, column=6, value=egreso.creado_por.nombre_completo if egreso.creado_por and hasattr(egreso.creado_por, 'nombre_completo') else 'N/A')
                
                for col in range(1, 7):
                    cell = ws3.cell(row=fila_actual3, column=col)
                    cell.border = border3
                    if fila_actual3 % 2 == 0:
                        cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                
                fila_actual3 += 1
            except Exception as e:
                logger.error(f"Error al procesar egreso ID {egreso.id}: {str(e)}")
                continue
        
        column_widths3 = [12, 16, 15, 40, 15, 25]
        for i, width in enumerate(column_widths3, 1):
            ws3.column_dimensions[get_column_letter(i)].width = width
        ws3.row_dimensions[fila_datos3].height = 25
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=finanzas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        wb.save(response)
        
        return response
    except Exception as e:
        logger.error(f"Error al generar Excel de finanzas: {str(e)}")
        return HttpResponse('Error al generar el archivo Excel', status=500)


@login_required
def exportar_excel_proveedores(request):
    """Exporta todos los proveedores a Excel con diseño turquesa"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    except Exception as e:
        logger.error(f"Error al verificar permisos en exportar_excel_proveedores: {str(e)}")
        return HttpResponse('Error al procesar la solicitud', status=500)
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Proveedores"
        
        fila_datos = _crear_encabezado_excel(ws, "REPORTE DE PROVEEDORES", "Clínica Dental")
        
        headers = ['ID', 'Nombre', 'RUT', 'Email', 'Teléfono', 'Dirección', 'Contacto', 'Sitio Web', 'Activo', 'Total Solicitudes']
        fila_actual, border = _aplicar_estilo_turquesa(ws, fila_datos, headers)
        
        proveedores = Proveedor.objects.annotate(
            total_solicitudes=Count('solicitudes')
        ).order_by('nombre')
        
        for proveedor in proveedores:
            try:
                ws.cell(row=fila_actual, column=1, value=proveedor.id)
                ws.cell(row=fila_actual, column=2, value=str(proveedor.nombre) if proveedor.nombre else 'Sin nombre')
                ws.cell(row=fila_actual, column=3, value=str(proveedor.rut) if proveedor.rut else '')
                ws.cell(row=fila_actual, column=4, value=str(proveedor.email) if proveedor.email else '')
                ws.cell(row=fila_actual, column=5, value=str(proveedor.telefono) if proveedor.telefono else '')
                ws.cell(row=fila_actual, column=6, value=str(proveedor.direccion) if proveedor.direccion else '')
                ws.cell(row=fila_actual, column=7, value=str(proveedor.contacto_nombre) if proveedor.contacto_nombre else '')
                ws.cell(row=fila_actual, column=8, value=str(proveedor.sitio_web) if proveedor.sitio_web else '')
                ws.cell(row=fila_actual, column=9, value='Sí' if proveedor.activo else 'No')
                ws.cell(row=fila_actual, column=10, value=proveedor.total_solicitudes or 0)
                
                for col in range(1, 11):
                    cell = ws.cell(row=fila_actual, column=col)
                    cell.border = border
                    if fila_actual % 2 == 0:
                        cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                
                fila_actual += 1
            except Exception as e:
                logger.error(f"Error al procesar proveedor ID {proveedor.id}: {str(e)}")
                continue
        
        column_widths = [8, 30, 15, 30, 15, 30, 20, 25, 10, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        ws.row_dimensions[fila_datos].height = 25
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=proveedores_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        wb.save(response)
        
        return response
    except Exception as e:
        logger.error(f"Error al generar Excel de proveedores: {str(e)}")
        return HttpResponse('Error al generar el archivo Excel', status=500)


@login_required
def exportar_excel_solicitudes(request):
    """Exporta todas las solicitudes a Excel con diseño turquesa"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    except Exception as e:
        logger.error(f"Error al verificar permisos en exportar_excel_solicitudes: {str(e)}")
        return HttpResponse('Error al procesar la solicitud', status=500)
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Solicitudes"
        
        fila_datos = _crear_encabezado_excel(ws, "REPORTE DE SOLICITUDES", "Clínica Dental")
        
        headers = ['ID', 'Fecha Solicitud', 'Proveedor', 'Insumo', 'Cantidad', 'Unidad', 'Fecha Entrega', 'Estado', 'Precio Unitario', 'Monto Total']
        fila_actual, border = _aplicar_estilo_turquesa(ws, fila_datos, headers)
        
        solicitudes = SolicitudInsumo.objects.select_related('proveedor', 'insumo', 'solicitado_por').order_by('-fecha_solicitud')
        
        for solicitud in solicitudes:
            try:
                ws.cell(row=fila_actual, column=1, value=solicitud.id)
                ws.cell(row=fila_actual, column=2, value=solicitud.fecha_solicitud.strftime('%d/%m/%Y %H:%M') if solicitud.fecha_solicitud else 'N/A')
                ws.cell(row=fila_actual, column=3, value=solicitud.proveedor.nombre if solicitud.proveedor else 'Sin proveedor')
                ws.cell(row=fila_actual, column=4, value=solicitud.insumo.nombre if solicitud.insumo else 'Sin insumo')
                ws.cell(row=fila_actual, column=5, value=solicitud.cantidad_solicitada or 0)
                ws.cell(row=fila_actual, column=6, value=solicitud.insumo.unidad_medida if solicitud.insumo and solicitud.insumo.unidad_medida else 'N/A')
                ws.cell(row=fila_actual, column=7, value=solicitud.fecha_entrega_esperada.strftime('%d/%m/%Y') if solicitud.fecha_entrega_esperada else 'N/A')
                ws.cell(row=fila_actual, column=8, value=solicitud.get_estado_display() if solicitud.estado else 'N/A')
                ws.cell(row=fila_actual, column=9, value=float(solicitud.precio_unitario) if solicitud.precio_unitario else 0)
                ws.cell(row=fila_actual, column=10, value=float(solicitud.monto_egreso) if solicitud.monto_egreso else 0)
                
                for col in range(1, 11):
                    cell = ws.cell(row=fila_actual, column=col)
                    cell.border = border
                    if fila_actual % 2 == 0:
                        cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                
                fila_actual += 1
            except Exception as e:
                logger.error(f"Error al procesar solicitud ID {solicitud.id}: {str(e)}")
                continue
        
        column_widths = [8, 18, 25, 25, 12, 10, 15, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        ws.row_dimensions[fila_datos].height = 25
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=solicitudes_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        wb.save(response)
        
        return response
    except Exception as e:
        logger.error(f"Error al generar Excel de solicitudes: {str(e)}")
        return HttpResponse('Error al generar el archivo Excel', status=500)


@login_required
def exportar_excel_personal(request):
    """Exporta todo el personal a Excel con diseño turquesa"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    except Exception as e:
        logger.error(f"Error al verificar permisos en exportar_excel_personal: {str(e)}")
        return HttpResponse('Error al procesar la solicitud', status=500)
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Personal"
        
        fila_datos = _crear_encabezado_excel(ws, "REPORTE DE PERSONAL", "Clínica Dental")
        
        headers = ['ID', 'Nombre Completo', 'Email', 'Teléfono', 'Rol', 'Especialidad', 'Activo', 'Fecha Ingreso', 'Total Citas']
        fila_actual, border = _aplicar_estilo_turquesa(ws, fila_datos, headers)
        
        personal = Perfil.objects.annotate(
            total_citas=Count('citas_asignadas')
        ).order_by('rol', 'nombre_completo')
        
        for persona in personal:
            try:
                ws.cell(row=fila_actual, column=1, value=persona.id)
                ws.cell(row=fila_actual, column=2, value=str(persona.nombre_completo) if persona.nombre_completo else 'Sin nombre')
                ws.cell(row=fila_actual, column=3, value=str(persona.email) if persona.email else '')
                ws.cell(row=fila_actual, column=4, value=str(persona.telefono) if persona.telefono else '')
                ws.cell(row=fila_actual, column=5, value=persona.get_rol_display() if persona.rol else 'N/A')
                ws.cell(row=fila_actual, column=6, value=str(persona.especialidad) if persona.especialidad else '')
                ws.cell(row=fila_actual, column=7, value='Sí' if persona.activo else 'No')
                ws.cell(row=fila_actual, column=8, value=persona.fecha_ingreso.strftime('%d/%m/%Y') if persona.fecha_ingreso else 'N/A')
                ws.cell(row=fila_actual, column=9, value=persona.total_citas or 0)
                
                for col in range(1, 10):
                    cell = ws.cell(row=fila_actual, column=col)
                    cell.border = border
                    if fila_actual % 2 == 0:
                        cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                
                fila_actual += 1
            except Exception as e:
                logger.error(f"Error al procesar personal ID {persona.id}: {str(e)}")
                continue
        
        column_widths = [8, 30, 30, 15, 15, 20, 10, 15, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        ws.row_dimensions[fila_datos].height = 25
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=personal_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        wb.save(response)
        
        return response
    except Exception as e:
        logger.error(f"Error al generar Excel de personal: {str(e)}")
        return HttpResponse('Error al generar el archivo Excel', status=500)


@login_required
def exportar_excel_servicios(request):
    """Exporta todos los servicios a Excel con diseño turquesa"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    except Exception as e:
        logger.error(f"Error al verificar permisos en exportar_excel_servicios: {str(e)}")
        return HttpResponse('Error al procesar la solicitud', status=500)
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Servicios"
        
        fila_datos = _crear_encabezado_excel(ws, "REPORTE DE SERVICIOS", "Clínica Dental")
        
        headers = ['ID', 'Nombre', 'Categoría', 'Descripción', 'Precio Base', 'Duración (min)', 'Requiere Dentista', 'Activo', 'Fecha Creación']
        fila_actual, border = _aplicar_estilo_turquesa(ws, fila_datos, headers)
        
        servicios = TipoServicio.objects.order_by('categoria', 'nombre')
        
        for servicio in servicios:
            try:
                ws.cell(row=fila_actual, column=1, value=servicio.id)
                ws.cell(row=fila_actual, column=2, value=str(servicio.nombre) if servicio.nombre else 'Sin nombre')
                ws.cell(row=fila_actual, column=3, value=servicio.get_categoria_display() if servicio.categoria else 'N/A')
                ws.cell(row=fila_actual, column=4, value=str(servicio.descripcion) if servicio.descripcion else '')
                ws.cell(row=fila_actual, column=5, value=float(servicio.precio_base) if servicio.precio_base else 0)
                ws.cell(row=fila_actual, column=6, value=servicio.duracion_estimada if servicio.duracion_estimada else 'N/A')
                ws.cell(row=fila_actual, column=7, value='Sí' if servicio.requiere_dentista else 'No')
                ws.cell(row=fila_actual, column=8, value='Sí' if servicio.activo else 'No')
                ws.cell(row=fila_actual, column=9, value=servicio.creado_el.strftime('%d/%m/%Y') if servicio.creado_el else 'N/A')
                
                for col in range(1, 10):
                    cell = ws.cell(row=fila_actual, column=col)
                    cell.border = border
                    if fila_actual % 2 == 0:
                        cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                
                fila_actual += 1
            except Exception as e:
                logger.error(f"Error al procesar servicio ID {servicio.id}: {str(e)}")
                continue
        
        column_widths = [8, 30, 20, 40, 15, 15, 15, 10, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        ws.row_dimensions[fila_datos].height = 25
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=servicios_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        wb.save(response)
        
        return response
    except Exception as e:
        logger.error(f"Error al generar Excel de servicios: {str(e)}")
        return HttpResponse('Error al generar el archivo Excel', status=500)


@login_required
def exportar_excel_planes_tratamiento(request):
    """Exporta todos los planes de tratamiento a Excel con diseño turquesa"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            return HttpResponse('No tienes permisos', status=403)
    except Perfil.DoesNotExist:
        return HttpResponse('No autorizado', status=403)
    except Exception as e:
        logger.error(f"Error al verificar permisos en exportar_excel_planes_tratamiento: {str(e)}")
        return HttpResponse('Error al procesar la solicitud', status=500)
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Planes Tratamiento"
        
        fila_datos = _crear_encabezado_excel(ws, "REPORTE DE PLANES DE TRATAMIENTO", "Clínica Dental")
        
        headers = ['ID', 'Nombre', 'Cliente', 'Dentista', 'Estado', 'Presupuesto Total', 'Descuento', 'Precio Final', 'Progreso %', 'Fecha Creación']
        fila_actual, border = _aplicar_estilo_turquesa(ws, fila_datos, headers)
        
        planes = PlanTratamiento.objects.select_related('cliente', 'dentista').order_by('-creado_el')
        
        for plan in planes:
            try:
                ws.cell(row=fila_actual, column=1, value=plan.id)
                ws.cell(row=fila_actual, column=2, value=str(plan.nombre) if plan.nombre else 'Sin nombre')
                ws.cell(row=fila_actual, column=3, value=plan.cliente.nombre_completo if plan.cliente and hasattr(plan.cliente, 'nombre_completo') else 'Sin cliente')
                ws.cell(row=fila_actual, column=4, value=plan.dentista.nombre_completo if plan.dentista and hasattr(plan.dentista, 'nombre_completo') else 'Sin dentista')
                ws.cell(row=fila_actual, column=5, value=plan.get_estado_display() if plan.estado else 'N/A')
                ws.cell(row=fila_actual, column=6, value=float(plan.presupuesto_total) if plan.presupuesto_total else 0)
                ws.cell(row=fila_actual, column=7, value=float(plan.descuento) if plan.descuento else 0)
                ws.cell(row=fila_actual, column=8, value=float(plan.precio_final) if plan.precio_final else 0)
                ws.cell(row=fila_actual, column=9, value=f"{plan.progreso_porcentaje}%" if plan.progreso_porcentaje is not None else "0%")
                ws.cell(row=fila_actual, column=10, value=plan.creado_el.strftime('%d/%m/%Y') if plan.creado_el else 'N/A')
                
                for col in range(1, 11):
                    cell = ws.cell(row=fila_actual, column=col)
                    cell.border = border
                    if fila_actual % 2 == 0:
                        cell.fill = PatternFill(start_color=TURQUESA_BG, end_color=TURQUESA_BG, fill_type="solid")
                
                fila_actual += 1
            except Exception as e:
                logger.error(f"Error al procesar plan ID {plan.id}: {str(e)}")
                continue
        
        column_widths = [8, 30, 25, 25, 20, 18, 12, 15, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        ws.row_dimensions[fila_datos].height = 25
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=planes_tratamiento_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        wb.save(response)
        
        return response
    except Exception as e:
        logger.error(f"Error al generar Excel de planes de tratamiento: {str(e)}")
        return HttpResponse('Error al generar el archivo Excel', status=500)


@login_required
def estadisticas(request):
    """Vista de Estadísticas con dashboards y gráficos importantes para administradores"""
    try:
        perfil = Perfil.objects.get(user=request.user)
        if not perfil.es_administrativo():
            messages.error(request, 'No tienes permisos para acceder a las estadísticas.')
            return redirect('panel_trabajador')
    except Perfil.DoesNotExist:
        return redirect('login')
    
    # Obtener fechas para filtros (pueden venir de parámetros GET)
    hoy = timezone.now().date()
    
    # Si hay parámetros de fecha, usarlos; sino usar valores por defecto
    fecha_desde_param = request.GET.get('fecha_desde', '')
    fecha_hasta_param = request.GET.get('fecha_hasta', '')
    
    if fecha_desde_param and fecha_hasta_param:
        try:
            fecha_desde = datetime.strptime(fecha_desde_param, '%Y-%m-%d').date()
            fecha_hasta = datetime.strptime(fecha_hasta_param, '%Y-%m-%d').date()
        except ValueError:
            fecha_desde = hoy.replace(day=1)
            fecha_hasta = hoy
    else:
        # Valores por defecto: mes actual
        fecha_desde = hoy.replace(day=1)
        fecha_hasta = hoy
    
    inicio_mes = fecha_desde if fecha_desde == hoy.replace(day=1) else fecha_desde
    fin_mes = fecha_hasta
    inicio_año = hoy.replace(month=1, day=1)
    hace_30_dias = hoy - timedelta(days=30)
    hace_7_dias = hoy - timedelta(days=7)
    
    # ===== ESTADÍSTICAS GENERALES =====
    total_citas = Cita.objects.count()
    total_clientes = Cliente.objects.count()
    clientes_activos = Cliente.objects.filter(activo=True).count()
    total_personal = Perfil.objects.filter(activo=True).count()
    total_insumos = Insumo.objects.count()
    
    # Citas por estado
    citas_disponibles = Cita.objects.filter(estado='disponible').count()
    citas_reservadas = Cita.objects.filter(estado='reservada').count()
    citas_completadas = Cita.objects.filter(estado='completada').count()
    citas_canceladas = Cita.objects.filter(estado='cancelada').count()
    
    # Citas del período seleccionado
    citas_mes = Cita.objects.filter(
        fecha_hora__date__gte=fecha_desde,
        fecha_hora__date__lte=fecha_hasta
    ).count()
    
    citas_semana = Cita.objects.filter(
        fecha_hora__date__gte=hace_7_dias
    ).count()
    
    citas_hoy = Cita.objects.filter(fecha_hora__date=hoy).count()
    
    citas_completadas_mes = Cita.objects.filter(
        estado='completada',
        fecha_hora__date__gte=inicio_mes,
        fecha_hora__date__lte=fin_mes
    ).count()
    
    # Insumos con stock bajo
    insumos_bajo_stock = Insumo.objects.filter(
        cantidad_actual__lte=F('cantidad_minima')
    ).count()
    
    # Dentistas más activos
    dentistas_activos = Perfil.objects.filter(
        rol='dentista',
        activo=True
    ).annotate(
        num_citas=Count('citas_asignadas')
    ).order_by('-num_citas')[:5]
    
    # Clientes nuevos en el período seleccionado
    clientes_nuevos_mes = Cliente.objects.filter(
        fecha_registro__date__gte=fecha_desde,
        fecha_registro__date__lte=fecha_hasta
    ).count()
    
    # Tipos de consulta más frecuentes
    tipos_consulta = Cita.objects.exclude(
        tipo_consulta__isnull=True
    ).exclude(
        tipo_consulta=''
    ).values('tipo_consulta').annotate(
        total=Count('id')
    ).order_by('-total')[:5]
    
    # Citas por día de la semana (últimos 30 días)
    citas_por_dia = []
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    for i in range(7):
        dia = dias_semana[i]
        count = Cita.objects.filter(
            fecha_hora__date__gte=hace_30_dias,
            fecha_hora__week_day=i+2  # Django week_day: 1=domingo, 2=lunes...
        ).count()
        citas_por_dia.append({'dia': dia, 'total': count})
    
    # Citas por mes (últimos 6 meses)
    citas_por_mes = []
    meses_nombres = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    for i in range(6):
        mes_fecha = hoy - timedelta(days=30*i)
        inicio = mes_fecha.replace(day=1)
        fin = (inicio + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        count = Cita.objects.filter(
            fecha_hora__date__gte=inicio,
            fecha_hora__date__lte=fin
        ).count()
        citas_por_mes.insert(0, {
            'mes': meses_nombres[inicio.month - 1],
            'total': count
        })
    
    # Finanzas (usando período seleccionado)
    total_ingresos = IngresoManual.objects.aggregate(Sum('monto'))['monto__sum'] or 0
    total_egresos = EgresoManual.objects.aggregate(Sum('monto'))['monto__sum'] or 0
    ingresos_mes = IngresoManual.objects.filter(fecha__gte=fecha_desde, fecha__lte=fecha_hasta).aggregate(Sum('monto'))['monto__sum'] or 0
    egresos_mes = EgresoManual.objects.filter(fecha__gte=fecha_desde, fecha__lte=fecha_hasta).aggregate(Sum('monto'))['monto__sum'] or 0
    balance_mes = ingresos_mes - egresos_mes
    
    # Ingresos y egresos por mes (últimos 6 meses)
    ingresos_por_mes = []
    egresos_por_mes = []
    for i in range(6):
        mes_fecha = hoy - timedelta(days=30*i)
        inicio = mes_fecha.replace(day=1)
        fin = (inicio + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        ingreso = IngresoManual.objects.filter(fecha__gte=inicio, fecha__lte=fin).aggregate(Sum('monto'))['monto__sum'] or 0
        egreso = EgresoManual.objects.filter(fecha__gte=inicio, fecha__lte=fin).aggregate(Sum('monto'))['monto__sum'] or 0
        ingresos_por_mes.insert(0, float(ingreso))
        egresos_por_mes.insert(0, float(egreso))
    
    # Planes de tratamiento
    total_planes = PlanTratamiento.objects.count()
    planes_en_progreso = PlanTratamiento.objects.filter(estado='en_progreso').count()
    planes_completados = PlanTratamiento.objects.filter(estado='completado').count()
    planes_pendientes = PlanTratamiento.objects.filter(estado='pendiente_aprobacion').count()
    
    # Valor total de planes de tratamiento
    valor_total_planes = PlanTratamiento.objects.aggregate(Sum('presupuesto_total'))['presupuesto_total__sum'] or 0
    
    # Solicitudes de insumos
    total_solicitudes = SolicitudInsumo.objects.count()
    solicitudes_pendientes = SolicitudInsumo.objects.filter(estado='pendiente').count()
    solicitudes_enviadas = SolicitudInsumo.objects.filter(estado='enviada').count()
    solicitudes_recibidas = SolicitudInsumo.objects.filter(estado='recibida').count()
    
    # Top 5 insumos más solicitados
    insumos_mas_solicitados = SolicitudInsumo.objects.values('insumo__nombre').annotate(
        total=Count('id')
    ).order_by('-total')[:5]
    
    # ===== MÉTRICAS ADICIONALES PARA DASHBOARD MEJORADO =====
    
    # Comparativas con mes anterior
    mes_anterior_inicio = (inicio_mes - timedelta(days=1)).replace(day=1)
    mes_anterior_fin = inicio_mes - timedelta(days=1)
    
    citas_mes_anterior = Cita.objects.filter(
        fecha_hora__date__gte=mes_anterior_inicio,
        fecha_hora__date__lte=mes_anterior_fin
    ).count()
    
    ingresos_mes_anterior = IngresoManual.objects.filter(
        fecha__gte=mes_anterior_inicio,
        fecha__lte=mes_anterior_fin
    ).aggregate(Sum('monto'))['monto__sum'] or 0
    
    clientes_nuevos_mes_anterior = Cliente.objects.filter(
        fecha_registro__date__gte=mes_anterior_inicio,
        fecha_registro__date__lte=mes_anterior_fin
    ).count()
    
    # Tasa de cancelación (canceladas del período / total citas del período)
    citas_canceladas_mes = Cita.objects.filter(
        estado='cancelada',
        fecha_hora__date__gte=fecha_desde,
        fecha_hora__date__lte=fecha_hasta
    ).count()
    tasa_cancelacion = (citas_canceladas_mes / citas_mes * 100) if citas_mes > 0 else 0
    
    # Tasa de ocupación (citas completadas vs total)
    tasa_ocupacion = (citas_completadas_mes / citas_mes * 100) if citas_mes > 0 else 0
    
    # Ingresos por dentista (basado en citas completadas del mes)
    ingresos_por_dentista = []
    dentistas_con_ingresos = Perfil.objects.filter(
        rol='dentista',
        activo=True
    ).annotate(
        num_citas=Count('citas_asignadas', filter=Q(citas_asignadas__estado='completada', citas_asignadas__fecha_hora__date__gte=fecha_desde, citas_asignadas__fecha_hora__date__lte=fecha_hasta))
    ).filter(num_citas__gt=0).order_by('-num_citas')[:5]
    
    for dentista in dentistas_con_ingresos:
        # Estimación: cada cita completada genera un ingreso promedio
        ingresos_por_dentista.append({
            'nombre': dentista.nombre_completo,
            'citas': dentista.num_citas,
            'ingreso_estimado': dentista.num_citas * 50000  # Estimación promedio
        })
    
    # Valor total del inventario
    valor_total_inventario = sum([
        (insumo.cantidad_actual * float(insumo.precio_unitario or 0))
        for insumo in Insumo.objects.all()
    ])
    
    # Rotación de insumos (movimientos del período)
    movimientos_mes = MovimientoInsumo.objects.filter(
        fecha_movimiento__date__gte=fecha_desde,
        fecha_movimiento__date__lte=fecha_hasta
    ).count()
    
    # Top 5 insumos más usados (por movimientos de salida)
    insumos_mas_usados = MovimientoInsumo.objects.filter(
        tipo='salida',
        fecha_movimiento__date__gte=fecha_desde,
        fecha_movimiento__date__lte=fecha_hasta
    ).values('insumo__nombre').annotate(
        total_salidas=Sum('cantidad')
    ).order_by('-total_salidas')[:5]
    
    # Análisis de proveedores
    total_proveedores = Proveedor.objects.filter(activo=True).count()
    proveedores_mas_usados = Proveedor.objects.annotate(
        total_solicitudes=Count('solicitudes')
    ).filter(activo=True).order_by('-total_solicitudes')[:5]
    
    # Porcentajes de cambio
    cambio_citas = ((citas_mes - citas_mes_anterior) / citas_mes_anterior * 100) if citas_mes_anterior > 0 else 0
    cambio_ingresos = ((ingresos_mes - ingresos_mes_anterior) / ingresos_mes_anterior * 100) if ingresos_mes_anterior > 0 else 0
    cambio_clientes = ((clientes_nuevos_mes - clientes_nuevos_mes_anterior) / clientes_nuevos_mes_anterior * 100) if clientes_nuevos_mes_anterior > 0 else 0
    
    context = {
        'perfil': perfil,
        'es_admin': True,
        # Estadísticas generales
        'total_citas': total_citas,
        'total_clientes': total_clientes,
        'clientes_activos': clientes_activos,
        'total_personal': total_personal,
        'total_insumos': total_insumos,
        # Citas
        'citas_disponibles': citas_disponibles,
        'citas_reservadas': citas_reservadas,
        'citas_completadas': citas_completadas,
        'citas_canceladas': citas_canceladas,
        'citas_mes': citas_mes,
        'citas_semana': citas_semana,
        'citas_hoy': citas_hoy,
        'citas_completadas_mes': citas_completadas_mes,
        # Insumos
        'insumos_bajo_stock': insumos_bajo_stock,
        # Dentistas
        'dentistas_activos': dentistas_activos,
        # Clientes
        'clientes_nuevos_mes': clientes_nuevos_mes,
        # Gráficos
        'tipos_consulta': list(tipos_consulta),
        'citas_por_dia': citas_por_dia,
        'citas_por_mes': citas_por_mes,
        # Finanzas
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'ingresos_mes': ingresos_mes,
        'egresos_mes': egresos_mes,
        'balance_mes': balance_mes,
        'ingresos_por_mes': ingresos_por_mes,
        'egresos_por_mes': egresos_por_mes,
        # Planes de tratamiento
        'total_planes': total_planes,
        'planes_en_progreso': planes_en_progreso,
        'planes_completados': planes_completados,
        'planes_pendientes': planes_pendientes,
        'valor_total_planes': valor_total_planes,
        # Solicitudes
        'total_solicitudes': total_solicitudes,
        'solicitudes_pendientes': solicitudes_pendientes,
        'solicitudes_enviadas': solicitudes_enviadas,
        'solicitudes_recibidas': solicitudes_recibidas,
        'insumos_mas_solicitados': list(insumos_mas_solicitados),
        # Métricas adicionales
        'citas_mes_anterior': citas_mes_anterior,
        'ingresos_mes_anterior': ingresos_mes_anterior,
        'clientes_nuevos_mes_anterior': clientes_nuevos_mes_anterior,
        'tasa_cancelacion': round(tasa_cancelacion, 1),
        'tasa_ocupacion': round(tasa_ocupacion, 1),
        'citas_canceladas_mes': citas_canceladas_mes,
        'ingresos_por_dentista': ingresos_por_dentista,
        'valor_total_inventario': valor_total_inventario,
        'movimientos_mes': movimientos_mes,
        'insumos_mas_usados': list(insumos_mas_usados),
        'total_proveedores': total_proveedores,
        'proveedores_mas_usados': proveedores_mas_usados,
        'cambio_citas': round(cambio_citas, 1),
        'cambio_ingresos': round(cambio_ingresos, 1),
        'cambio_clientes': round(cambio_clientes, 1),
        # Fechas
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'mes_actual': fecha_desde.strftime('%B %Y') if fecha_desde else inicio_mes.strftime('%B %Y'),
        'hoy': hoy.strftime('%d/%m/%Y'),
    }
    
    return render(request, 'citas/estadisticas/estadisticas.html', context)

